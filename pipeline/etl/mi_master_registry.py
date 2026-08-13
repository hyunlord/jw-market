from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
import re
from pathlib import Path
from typing import Any, Mapping
import unicodedata

from openpyxl import load_workbook
import yaml

from pipeline.etl.lib.storage import get_mi_master_path


SOURCE_SHEET = "시장정의 & Target"
DEFAULT_RULES_PATH = (
    Path(__file__).resolve().parent / "config" / "mi_master_rules.yaml"
)


@dataclass(frozen=True, slots=True)
class MarketSheet:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    source_type: str
    catalog_source_type: str
    atc_codes: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class TargetBrand:
    brand_name: str
    jw_product_name: str
    strategic_market_id: str
    ml_id: str
    cd_id: str
    source_type: str
    is_target: bool
    layer3_aliases: tuple[str, ...]
    source_note: str


@dataclass(frozen=True, slots=True)
class MiMasterRegistry:
    market_sheets: tuple[MarketSheet, ...]
    market_by_id: dict[str, dict[str, str]]
    market_definition_columns: dict[str, tuple[int, ...]]
    analyze_matrix: dict[str, dict[str, bool]]
    cd_specs: tuple[dict[str, Any], ...]
    direct_competition_by_cd_id: dict[str, tuple[str, ...]]
    detail_sheets: tuple[str, ...]
    target_brands: tuple[TargetBrand, ...]


def _compact(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("에이플러스", "a+")
    return re.sub(r"[^0-9a-z가-힣+]+", "", text)


def _parts(value: Any) -> set[str]:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("에이플러스", "a+")
    return {
        compact
        for token in re.split(r"[\s/,·&]+", text)
        if (compact := _compact(token))
    }


def _header_row(sheet: Any) -> int | None:
    # Detail sheets place their tabular header below the title band. The
    # workbook-wide inventory starts at row 1 and is intentionally not a
    # strategic-market detail sheet.
    for row_id in range(3, 13):
        values = [
            unicodedata.normalize(
                "NFKC",
                str(sheet.cell(row_id, column_id).value or ""),
            ).upper()
            for column_id in range(1, 27)
        ]
        has_atc = any("ATC" in value for value in values)
        has_entity = any(
            token in value
            for value in values
            for token in ("MOLECULE", "성분", "PRODUCT", "제품")
        )
        if has_atc and has_entity:
            return row_id
    return None


def _source_type(values: list[Any]) -> str:
    joined = " ".join(str(value or "") for value in values).lower()
    if "ubist" in joined and "iqvia" in joined:
        return "BOTH"
    if "ubist" in joined:
        return "UBIST"
    if "iqvia" in joined:
        return "IQVIA"
    raise ValueError(f"MI Master source type is not declared: {values!r}")


def _atc_codes(values: list[Any]) -> tuple[str, ...]:
    codes: list[str] = []
    for value in values:
        for code in re.findall(r"\b[A-Z]\d{2}[A-Z]\d\b", str(value or "").upper()):
            if code not in codes:
                codes.append(code)
    return tuple(codes)


def _target_columns(sheet: Any) -> dict[int, str]:
    values: dict[int, str] = {}
    empty_run = 0
    for column_id in range(3, 101):
        value = str(sheet.cell(6, column_id).value or "").strip()
        if value:
            values[column_id] = value
            empty_run = 0
        else:
            empty_run += 1
            if empty_run >= 5 and values:
                break
    return values


def _sheet_for_target(
    target_name: str,
    sheet_names: list[str],
) -> str:
    target_parts = _parts(target_name)
    target_compact = _compact(target_name)
    scored: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        sheet_parts = _parts(sheet_name)
        exact = len(target_parts & sheet_parts)
        compact_match = int(
            target_compact in _compact(sheet_name)
            or _compact(sheet_name) in target_compact
        )
        score = exact * 10 + compact_match
        if score:
            scored.append((score, sheet_name))
    if not scored:
        raise ValueError(f"no detail sheet matches target product {target_name!r}")
    best_score = max(score for score, _ in scored)
    best = [name for score, name in scored if score == best_score]
    if len(best) != 1:
        raise ValueError(
            f"ambiguous detail sheet for target product {target_name!r}: {best}"
        )
    return best[0]


def _ordered_sheets(
    sheet_names: list[str],
    rules: Mapping[str, Any],
) -> list[str]:
    ordered = list(sheet_names)
    for move in rules.get("strategy_order", []):
        sheet_name = str(move["sheet_name"])
        before = str(move["before"])
        if sheet_name not in ordered or before not in ordered:
            raise ValueError(f"invalid strategy order rule: {move!r}")
        ordered.remove(sheet_name)
        ordered.insert(ordered.index(before), sheet_name)
    return ordered


def _display_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("/", " ")).strip().replace(
        "A+",
        "에이플러스",
    )


def _target_brand_names(
    product_name: str,
    *,
    sheet_name: str,
    rules: Mapping[str, Any],
) -> tuple[str, ...]:
    brand_names = tuple(
        part.strip()
        for part in product_name.split("/")
        if part.strip()
    )
    overrides = dict(rules.get("target_brand_name_overrides", {}))
    sheet_overrides = dict(overrides.get(sheet_name, {}))
    return tuple(
        str(sheet_overrides.get(brand_name, brand_name))
        for brand_name in brand_names
    )


def _target_brand_aliases(
    brand_name: str,
    rules: Mapping[str, Any],
) -> tuple[str, ...]:
    aliases = dict(rules.get("target_brand_aliases", {})).get(brand_name, [])
    return tuple(str(alias) for alias in aliases)


def _jw_product_fields(
    brand_name: str,
    *,
    sheet_name: str,
    rules: Mapping[str, Any],
) -> tuple[str, str]:
    overrides = dict(rules.get("jw_product_overrides", {}))
    override = dict(overrides.get(brand_name, {}))
    source_notes = dict(rules.get("jw_product_source_notes", {}))
    return (
        str(override.get("name", brand_name)),
        str(
            override.get(
                "source_note",
                source_notes.get(sheet_name, "sheet split"),
            )
        ),
    )


def _headers(sheet: Any, row_id: int) -> tuple[str, ...]:
    return tuple(
        unicodedata.normalize(
            "NFKC",
            str(sheet.cell(row_id, column_id).value or ""),
        ).strip()
        for column_id in range(1, 27)
    )


def _analysis_axes(
    definition_sheet: Any,
    detail_sheet: Any,
    header_row: int,
    columns: tuple[int, ...],
    overrides: Mapping[str, Any],
) -> dict[str, bool]:
    headers = " | ".join(_headers(detail_sheet, header_row)).upper()

    def marked(row_id: int) -> bool:
        return any(
            str(definition_sheet.cell(row_id, column_id).value or "").strip()
            for column_id in columns
        )

    axes = {
        "class": marked(14),
        "molecule": marked(15)
        or "MOLECULE" in headers
        or "성분" in headers,
        "dosage_form": marked(17),
        "strength_pack": marked(18)
        or "STRENGTH" in headers
        or "성분용량" in headers
        or "PACK DESC" in headers,
        "nhi_type": marked(19)
        or "NHI TYPE" in headers
        or "급여/비급여" in headers,
        "ox_gx": any(
            token in str(
                definition_sheet.cell(19, column_id).value or ""
            ).upper()
            for column_id in columns
            for token in ("OX", "BIOSIMILAR")
        ),
        "fish_oil": any(
            "FISH" in str(
                definition_sheet.cell(19, column_id).value or ""
            ).upper()
            for column_id in columns
        ),
    }
    axes.update(
        {
            str(key): bool(value)
            for key, value in dict(overrides).items()
        }
    )
    return axes


@lru_cache(maxsize=8)
def load_mi_master_rules(path: Path = DEFAULT_RULES_PATH) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        rules = yaml.safe_load(handle) or {}
    if int(rules.get("version", 0)) != 1:
        raise ValueError(f"unsupported MI Master rules version: {rules.get('version')!r}")
    return rules


def discover_mi_master_registry(
    workbook_path: Path,
    *,
    rules_path: Path = DEFAULT_RULES_PATH,
) -> MiMasterRegistry:
    rules = load_mi_master_rules(rules_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"required sheet not found: {SOURCE_SHEET!r}")
        definition_sheet = workbook[SOURCE_SHEET]
        header_rows = {
            sheet_name: header_row
            for sheet_name in workbook.sheetnames
            if sheet_name != SOURCE_SHEET
            if (header_row := _header_row(workbook[sheet_name])) is not None
        }
        if not header_rows:
            raise ValueError("MI Master has no discoverable detail sheets")

        target_columns = _target_columns(definition_sheet)
        target_sheet = {
            column_id: _sheet_for_target(product_name, list(header_rows))
            for column_id, product_name in target_columns.items()
        }
        columns_by_sheet: dict[str, list[int]] = {
            sheet_name: [] for sheet_name in header_rows
        }
        for column_id, sheet_name in target_sheet.items():
            columns_by_sheet[sheet_name].append(column_id)
        unmatched = [
            sheet_name
            for sheet_name, columns in columns_by_sheet.items()
            if not columns
        ]
        if unmatched:
            raise ValueError(f"detail sheets have no target columns: {unmatched}")

        ordered = _ordered_sheets(list(header_rows), rules)
        market_by_id: dict[str, dict[str, str]] = {}
        market_definition_columns: dict[str, tuple[int, ...]] = {}
        market_sheets: list[MarketSheet] = []
        analyze_matrix: dict[str, dict[str, bool]] = {}
        strategy_by_sheet: dict[str, tuple[str, str]] = {}
        axis_overrides = rules.get("analysis_axis_overrides", {})
        catalog_source_overrides = dict(
            rules.get("catalog_source_type_overrides", {})
        )

        for index, sheet_name in enumerate(ordered, start=1):
            strategic_id = f"strategy_{index:03d}"
            ml_id = f"ml_{index:03d}"
            columns = tuple(columns_by_sheet[sheet_name])
            source = _source_type(
                [definition_sheet.cell(10, column_id).value for column_id in columns]
            )
            market_by_id[strategic_id] = {
                "sheet_name": sheet_name,
                "source_type": source,
            }
            market_definition_columns[strategic_id] = columns
            market_sheets.append(
                MarketSheet(
                    strategic_id,
                    sheet_name,
                    header_rows[sheet_name],
                    source,
                    str(catalog_source_overrides.get(sheet_name, source)),
                    _atc_codes(
                        [
                            definition_sheet.cell(7, column_id).value
                            for column_id in columns
                        ]
                    ),
                )
            )
            analyze_matrix[ml_id] = _analysis_axes(
                definition_sheet,
                workbook[sheet_name],
                header_rows[sheet_name],
                columns,
                dict(axis_overrides.get(sheet_name, {})),
            )
            strategy_by_sheet[sheet_name] = (strategic_id, ml_id)

        collapsed = {
            str(item["sheet_name"])
            for item in rules.get("cd_collapses", [])
        }
        cd_name_overrides = dict(rules.get("cd_name_overrides", {}))
        cd_specs: list[dict[str, Any]] = []
        cd_id_by_column: dict[int, str] = {}
        direct_competition_by_cd_id: dict[str, tuple[str, ...]] = {}
        for column_id, product_name in sorted(target_columns.items()):
            sheet_name = target_sheet[column_id]
            columns = tuple(columns_by_sheet[sheet_name])
            if sheet_name in collapsed:
                existing = next(
                    (
                        spec
                        for spec in cd_specs
                        if spec["strategic_market_id"]
                        == strategy_by_sheet[sheet_name][0]
                    ),
                    None,
                )
                if existing is not None:
                    cd_id_by_column[column_id] = str(existing["cd_id"])
                    continue
                spec_columns = columns
                name = sheet_name
            else:
                spec_columns = (column_id,)
                name = str(
                    cd_name_overrides.get(
                        sheet_name,
                        _display_name(product_name),
                    )
                )
            cd_index = len(cd_specs) + 1
            cd_id = f"cd_{cd_index:03d}"
            strategic_id, ml_id = strategy_by_sheet[sheet_name]
            cd_specs.append(
                {
                    "cd_id": cd_id,
                    "name": name,
                    "ml_id": ml_id,
                    "cd_filter_id": f"cdf_{cd_index:03d}",
                    "strategic_market_id": strategic_id,
                    "column_ids": spec_columns,
                }
            )
            direct_competition_by_cd_id[cd_id] = tuple(
                str(value).strip()
                for spec_column in spec_columns
                for row_id in range(48, 51)
                if (
                    value := definition_sheet.cell(row_id, spec_column).value
                ) is not None
                and str(value).strip()
            )
            for spec_column in spec_columns:
                cd_id_by_column[spec_column] = cd_id

        target_brands: list[TargetBrand] = []
        first_brand_by_market: set[str] = set()
        strategy_index = {
            market.strategic_market_id: index
            for index, market in enumerate(market_sheets)
        }
        for column_id, product_name in sorted(
            target_columns.items(),
            key=lambda item: (
                strategy_index[strategy_by_sheet[target_sheet[item[0]]][0]],
                item[0],
            ),
        ):
            sheet_name = target_sheet[column_id]
            strategic_id, ml_id = strategy_by_sheet[sheet_name]
            for brand_name in _target_brand_names(
                str(product_name),
                sheet_name=sheet_name,
                rules=rules,
            ):
                jw_product_name, source_note = _jw_product_fields(
                    brand_name,
                    sheet_name=sheet_name,
                    rules=rules,
                )
                target_brands.append(
                    TargetBrand(
                        brand_name=brand_name,
                        jw_product_name=jw_product_name,
                        strategic_market_id=strategic_id,
                        ml_id=ml_id,
                        cd_id=cd_id_by_column[column_id],
                        source_type=market_by_id[strategic_id]["source_type"],
                        is_target=strategic_id not in first_brand_by_market,
                        layer3_aliases=_target_brand_aliases(brand_name, rules),
                        source_note=source_note,
                    )
                )
                first_brand_by_market.add(strategic_id)

        market_definition_columns = dict(
            sorted(
                market_definition_columns.items(),
                key=lambda item: min(item[1]),
            )
        )
        return MiMasterRegistry(
            market_sheets=tuple(market_sheets),
            market_by_id=market_by_id,
            market_definition_columns=market_definition_columns,
            analyze_matrix=analyze_matrix,
            cd_specs=tuple(cd_specs),
            direct_competition_by_cd_id=direct_competition_by_cd_id,
            detail_sheets=tuple(ordered),
            target_brands=tuple(target_brands),
        )
    finally:
        workbook.close()


@lru_cache(maxsize=1)
def default_mi_master_registry() -> MiMasterRegistry:
    return discover_mi_master_registry(get_mi_master_path())


def _lookup(path: str, record: Mapping[str, Any], context: Mapping[str, Any]) -> Any:
    current: Any = {"record": record, **context}
    for part in path.split("."):
        if not isinstance(current, Mapping):
            return None
        current = current.get(part)
    return current


def _matches(
    expected: Mapping[str, Any],
    record: Mapping[str, Any],
    context: Mapping[str, Any],
) -> bool:
    for key, expected_value in expected.items():
        actual = context.get(key, record.get(key))
        if _compact(actual) != _compact(expected_value):
            return False
    return True


def apply_record_rules(
    record: Mapping[str, Any],
    *,
    stage: str,
    context: Mapping[str, Any],
    rules_path: Path = DEFAULT_RULES_PATH,
) -> dict[str, Any]:
    updated = dict(record)
    for rule in load_mi_master_rules(rules_path).get("record_rules", []):
        if str(rule.get("stage")) != stage:
            continue
        if not _matches(dict(rule.get("match", {})), updated, context):
            continue
        actions = dict(rule.get("actions", {}))
        for field, value in dict(actions.get("set", {})).items():
            updated[str(field)] = value
        for field, source in dict(actions.get("copy", {})).items():
            updated[str(field)] = _lookup(str(source), updated, context)
        for pair in actions.get("null_if_equal", []):
            left, right = str(pair[0]), str(pair[1])
            if updated.get(left) == updated.get(right):
                updated[left] = None
        for field, sources in dict(actions.get("first_present", {})).items():
            updated[str(field)] = next(
                (
                    value
                    for source in sources
                    if (value := _lookup(str(source), updated, context))
                    not in (None, "")
                ),
                None,
            )
    return updated
