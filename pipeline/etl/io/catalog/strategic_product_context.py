from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.etl.io.catalog import master_drug, strategic_brand
from pipeline.etl.io.catalog._common import is_empty_row
from pipeline.etl.io.catalog.strategic_product_text import clean_text, extract_atc_code

PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_CATALOG_PATH = PROJECT_ROOT / "pipeline" / "etl" / "config" / "master_column_mapping_catalog.md"

def load_context_by_brand_id() -> dict[str, dict[str, Any]]:
    xlsx_path = master_drug.resolve_input_file(master_drug.DEFAULT_INPUT_FILE)
    metadata_catalog = master_drug.load_column_metadata_catalog(DEFAULT_CATALOG_PATH)

    contexts: dict[str, dict[str, Any]] = {}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for ml_index, config in enumerate(master_drug.MARKET_SHEETS, start=1):
            ws = wb[config.sheet_name]
            headers = master_drug._headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                master_drug.explicit_lookup_join(row_items)
                if config.strategic_market_id == "strategy_008"
                else {}
            )
            for source_row_id, values in row_items:
                if master_drug.is_empty_row(values):
                    continue
                standard_values, extras = master_drug.apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                fields = strategic_brand.strategic_fields(standard_values, extras)
                brand_id = f"sb_{ml_index:03d}_{source_row_id:05d}"
                contexts[brand_id] = {
                    "strategic_market_id": config.strategic_market_id,
                    "source_row_id": source_row_id,
                    "atc4_code": extract_atc_code(fields.get("atc4_code")),
                    "product_name": strategic_brand.make_name(
                        standard_values,
                        config.strategic_market_id,
                        source_row_id,
                    ),
                    "manufacturer": fields.get("제조사"),
                    "seller": fields.get("판매사"),
                    "pack_desc": clean_text(standard_values.get("pack_desc")) or clean_text(extras.get("product_pack")),
                    "strength": fields.get("strength_pack"),
                    "molecule": fields.get("molecule"),
                    "class": fields.get("class"),
                    "dosage_form": fields.get("dosage_form"),
                    "nhi_type": fields.get("nhi_type"),
                    "ox_gx": fields.get("ox_gx"),
                    "fish_oil": fields.get("fish_oil"),
                }
    finally:
        wb.close()
    return contexts
