"""
prototype_10_master_mapping_table_to_parquet.py
================================================
MI Master manual mapping table -> Parquet.

Phase 09d policy:
- Canonical row-generation logic:
  /Users/rexxa/github/jw-market/etl/master_loader.py::_manual_mapping_records
- Canonical market-sheet scan / mapping helpers:
  /Users/rexxa/github/jw-market/etl/master_loader.py and
  /Users/rexxa/github/jw-market/etl/master_transform.py
- Canonical schema:
  /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_mapping_table
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.

Usage, in Step D after user review:
    python3 scripts/prototype_10_master_mapping_table_to_parquet.py
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


from pipeline.etl.lib.storage import get_mi_master_path
from pipeline.etl.io.catalog._common import (
    STANDARD_PREFIX,
    _extra_key,
    _header_lookup,
    _lookup_key,
    _lookup_position_value,
    _lookup_source_value,
    _position_value,
    _single_lookup_key,
    apply_column_mapping,
    build_raw_row_payload,
    cell_text,
    dumps_json,
    explicit_lookup_join,
    is_empty_row,
    load_column_metadata_catalog as load_column_metadata_catalog_common,
    make_header_keys,
    normalize_header,
    to_jsonable,
    utc_now_text,
    write_records_parquet,
)


DEFAULT_INPUT_FILE = get_mi_master_path()
MASTER_ROOT = DEFAULT_INPUT_FILE.parent
DEFAULT_CATALOG_PATH = Path("docs/reference/master_column_mapping_catalog.md")
DEFAULT_OUTPUT_FILE = Path("parquet/master_mapping_table/master_mapping_table.parquet")

STANDARD_PREFIX = "drug_extra_json."
# 4/22 기준 5932행에서 260518 기준 5956행으로 24행이 늘었다.
# diff 확인 결과 시장정의/Target 계열의 정상 추가분이라, 검증을 완화하지 않고
# 새 원본 버전에 맞춘 strict count로 고정한다. 행수 검사를 제거하는 대안은
# mapping 누락을 조기에 잡지 못하므로 기각했다.
EXPECTED_ROW_COUNT = 5956
ZERO_MAPPING_MARKETS = {"strategy_006", "strategy_007", "strategy_009"}

MASTER_MAPPING_TABLE_COLUMNS = (
    "mapping_id",
    "strategic_market_id",
    "source_value",
    "target_column",
    "target_value",
    "mapping_type",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)


@dataclass(frozen=True)
class MarketSheetConfig:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    source_type: str


MARKET_SHEETS: tuple[MarketSheetConfig, ...] = (
    MarketSheetConfig("strategy_001", "라베칸 라베칸듀오", 5, "UBIST"),
    MarketSheetConfig("strategy_002", "제이클", 5, "IQVIA"),
    MarketSheetConfig("strategy_003", "가드렛 가드메트", 5, "IQVIA"),
    MarketSheetConfig("strategy_004", "타발리스", 5, "IQVIA"),
    MarketSheetConfig("strategy_005", "시그마트", 5, "UBIST"),
    MarketSheetConfig("strategy_006", "리바로 리바로젯", 4, "UBIST"),
    MarketSheetConfig("strategy_007", "리바로페노", 4, "UBIST"),
    MarketSheetConfig("strategy_008", "리바로하이 리바로브이", 5, "UBIST"),
    MarketSheetConfig("strategy_009", "트루패스 피나스타 제이다트", 5, "UBIST"),
    MarketSheetConfig("strategy_010", "뉴트로진 모빌리아", 5, "IQVIA"),
    MarketSheetConfig("strategy_011", "악템라", 5, "IQVIA"),
    MarketSheetConfig("strategy_012", "페린젝트 베노훼럼", 5, "IQVIA"),
    MarketSheetConfig("strategy_013", "헴리브라", 5, "IQVIA"),
    MarketSheetConfig("strategy_014", "위너프 위너프A+", 5, "IQVIA"),
    MarketSheetConfig("strategy_015", "엔커버", 7, "IQVIA"),
    MarketSheetConfig("strategy_016", "플라주오피", 5, "IQVIA"),
)

EXPECTED_MARKET_STATS = {
    "strategy_001": {
        "sheet_name": "라베칸 라베칸듀오",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 637,
        "excluded_rows": 0,
        "staging_rows": 358,
        "manual_specs": 3,
        "mapping_rows": 716,
    },
    "strategy_002": {
        "sheet_name": "제이클",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 950,
        "excluded_rows": 0,
        "staging_rows": 45,
        "manual_specs": 3,
        "mapping_rows": 135,
    },
    "strategy_003": {
        "sheet_name": "가드렛 가드메트",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 913,
        "excluded_rows": 0,
        "staging_rows": 82,
        "manual_specs": 2,
        "mapping_rows": 164,
    },
    "strategy_004": {
        "sheet_name": "타발리스",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 985,
        "excluded_rows": 0,
        "staging_rows": 10,
        "manual_specs": 1,
        "mapping_rows": 9,
    },
    "strategy_005": {
        "sheet_name": "시그마트",
        "header_row": 5,
        "raw_rows_scanned": 294,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 294,
        "manual_specs": 4,
        "mapping_rows": 1176,
    },
    "strategy_006": {
        "sheet_name": "리바로 리바로젯",
        "header_row": 4,
        "raw_rows_scanned": 1295,
        "empty_rows": 200,
        "excluded_rows": 48,
        "staging_rows": 1047,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_007": {
        "sheet_name": "리바로페노",
        "header_row": 4,
        "raw_rows_scanned": 996,
        "empty_rows": 385,
        "excluded_rows": 494,
        "staging_rows": 117,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_008": {
        "sheet_name": "리바로하이 리바로브이",
        "header_row": 5,
        "raw_rows_scanned": 1096,
        "empty_rows": 15,
        "excluded_rows": 0,
        "staging_rows": 1081,
        "manual_specs": 5,
        "mapping_rows": 1676,
    },
    "strategy_009": {
        "sheet_name": "트루패스 피나스타 제이다트",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 589,
        "excluded_rows": 1,
        "staging_rows": 405,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_010": {
        "sheet_name": "뉴트로진 모빌리아",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 985,
        "excluded_rows": 0,
        "staging_rows": 10,
        "manual_specs": 4,
        "mapping_rows": 38,
    },
    "strategy_011": {
        "sheet_name": "악템라",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 969,
        "excluded_rows": 0,
        "staging_rows": 26,
        "manual_specs": 3,
        "mapping_rows": 78,
    },
    "strategy_012": {
        "sheet_name": "페린젝트 베노훼럼",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 919,
        "excluded_rows": 0,
        "staging_rows": 76,
        "manual_specs": 5,
        "mapping_rows": 198,
    },
    "strategy_013": {
        "sheet_name": "헴리브라",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 981,
        "excluded_rows": 1,
        "staging_rows": 13,
        "manual_specs": 2,
        "mapping_rows": 26,
    },
    "strategy_014": {
        "sheet_name": "위너프 위너프A+",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 664,
        "excluded_rows": 8,
        "staging_rows": 323,
        "manual_specs": 6,
        "mapping_rows": 1696,
    },
    "strategy_015": {
        "sheet_name": "엔커버",
        "header_row": 7,
        "raw_rows_scanned": 993,
        "empty_rows": 989,
        "excluded_rows": 0,
        "staging_rows": 4,
        "manual_specs": 1,
        "mapping_rows": 4,
    },
    "strategy_016": {
        "sheet_name": "플라주오피",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 943,
        "excluded_rows": 31,
        "staging_rows": 21,
        "manual_specs": 2,
        "mapping_rows": 40,
    },
}

EXPECTED_MARKET_DISTRIBUTION = {
    "strategy_001": 716,
    "strategy_002": 135,
    "strategy_003": 164,
    "strategy_004": 9,
    "strategy_005": 1176,
    "strategy_008": 1676,
    "strategy_010": 38,
    "strategy_011": 78,
    "strategy_012": 198,
    "strategy_013": 26,
    "strategy_014": 1696,
    "strategy_015": 4,
    "strategy_016": 40,
}

EXPECTED_MAPPING_TYPE_DISTRIBUTION = {
    "class_recode": 1376,
    "manual_mapping": 2009,
    "molecule_recode": 1984,
    "nhi_overlay": 490,
    "strength_recode": 97,
}


@dataclass
class MarketMappingStats:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    max_row: int
    max_col: int
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    excluded_rows: int = 0
    staging_rows: int = 0
    manual_specs: int = 0
    mapping_rows: int = 0


def resolve_input_file(path: Path) -> Path:
    if path.exists():
        return path
    candidates = sorted(file for file in MASTER_ROOT.glob("*.xlsx") if not file.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError(f"No Master xlsx found under {MASTER_ROOT}")
    return candidates[-1]


def is_excluded_row(values: list[Any] | tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if "제외" in text and not text.startswith("비제외"):
            return True
    return False


def make_mapping_id(
    strategic_market_id: str,
    source_column: object,
    source_value: object,
    target_column: object,
    target_value: object,
    source_row_id: int | None = None,
) -> str:
    parts = [
        strategic_market_id,
        str(source_row_id or ""),
        str(source_column or ""),
        str(source_value or ""),
        str(target_column or ""),
        str(target_value or ""),
    ]
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"{strategic_market_id}:{digest}"


def mapping_type_for(target_column: str, source_column: object) -> str:
    lowered = f"{target_column} {source_column or ''}".lower()
    if "molecule" in lowered or "성분" in lowered:
        return "molecule_recode"
    if "class" in lowered:
        return "class_recode"
    if "strength" in lowered or "규격" in lowered:
        return "strength_recode"
    if "nhi" in lowered or "급여" in lowered:
        return "nhi_overlay"
    return "manual_mapping"


def raw_value_for_mapping(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    spec: dict[str, Any],
) -> Any:
    lookup = _header_lookup(headers, values)
    overlay_target = normalize_header(spec.get("overlay_target"))
    source_column = normalize_header(spec.get("source_column"))
    if overlay_target and overlay_target in lookup:
        return lookup.get(overlay_target)
    return _lookup_source_value(lookup, source_column)


def _headers_from_sheet(ws: Any, header_row: int) -> list[Any]:
    return list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))


def _manual_mapping_records(
    strategic_market_id: str,
    source_sheet: str,
    source_file_version: str,
    source_row_id: int,
    headers: list[Any],
    values: tuple[Any, ...],
    metadata: dict[str, dict[str, Any]],
    standard_values: dict[str, Any],
    extras: dict[str, Any],
    ingested_at: str,
    inclusion_standard_values: dict[str, Any] | None = None,
    inclusion_extras: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    inclusion_standard_values = inclusion_standard_values or standard_values
    inclusion_extras = inclusion_extras or extras
    for target_column, spec in metadata.items():
        metadata_type = str(spec.get("type") or "")
        if not metadata_type.startswith("manual"):
            continue
        source_column = spec.get("source_column")
        if target_column.startswith("drug_extra_json."):
            extra_key = target_column.split(".", 1)[1]
            target_value = extras.get(extra_key)
            inclusion_target_value = inclusion_extras.get(extra_key)
        else:
            target_value = standard_values.get(target_column)
            inclusion_target_value = inclusion_standard_values.get(target_column)
        source_value = raw_value_for_mapping(headers, values, spec)
        if source_value is None and inclusion_target_value is None:
            continue
        records.append(
            {
                "mapping_id": make_mapping_id(
                    strategic_market_id,
                    source_column,
                    source_value,
                    target_column,
                    target_value,
                    source_row_id=source_row_id,
                ),
                "strategic_market_id": strategic_market_id,
                "source_value": None if source_value is None else str(source_value),
                "target_column": target_column,
                "target_value": None if target_value is None else str(target_value),
                "mapping_type": mapping_type_for(target_column, source_column),
                "source_sheet": source_sheet,
                "source_file_version": source_file_version,
                "ingested_at": ingested_at,
            }
        )
    return records


def load_mapping_records(
    xlsx_path: Path,
    catalog_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], list[MarketMappingStats]]:
    metadata_catalog = load_column_metadata_catalog(catalog_path)
    timestamp = ingested_at or utc_now_text()
    records: list[dict[str, Any]] = []
    stats: list[MarketMappingStats] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for config in MARKET_SHEETS:
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}; sheets={wb.sheetnames}")
            ws = wb[config.sheet_name]
            headers = _headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            manual_specs = sum(
                1 for spec in metadata.values() if str(spec.get("type") or "").startswith("manual")
            )
            market_stats = MarketMappingStats(
                strategic_market_id=config.strategic_market_id,
                sheet_name=config.sheet_name,
                header_row=config.header_row,
                max_row=ws.max_row,
                max_col=ws.max_column,
                manual_specs=manual_specs,
            )

            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                explicit_lookup_join(row_items) if config.strategic_market_id == "strategy_008" else {}
            )

            for source_row_id, values in row_items:
                market_stats.raw_rows_scanned += 1
                if is_empty_row(values):
                    market_stats.empty_rows += 1
                    continue
                if is_excluded_row(values):
                    market_stats.excluded_rows += 1
                    continue

                market_stats.staging_rows += 1
                standard_values, extras = apply_column_mapping(headers, values, metadata)
                inclusion_standard_values = standard_values
                if source_row_id in explicit_overrides:
                    standard_values = dict(standard_values)
                    standard_values.update(explicit_overrides[source_row_id])
                mapping_rows = _manual_mapping_records(
                    config.strategic_market_id,
                    config.sheet_name,
                    xlsx_path.name,
                    source_row_id,
                    headers,
                    values,
                    metadata,
                    standard_values,
                    extras,
                    timestamp,
                    inclusion_standard_values=inclusion_standard_values,
                )
                records.extend(mapping_rows)
                market_stats.mapping_rows += len(mapping_rows)

            stats.append(market_stats)
    finally:
        wb.close()

    return records, stats


def _count_by(records: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        value = str(record.get(key) or "")
        counts[value] = counts.get(value, 0) + 1
    return counts


def validate_records(records: list[dict[str, Any]], stats: list[MarketMappingStats]) -> None:
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"mapping row count must be {EXPECTED_ROW_COUNT}, found {len(records)}")

    mapping_ids = [record["mapping_id"] for record in records]
    if len(set(mapping_ids)) != EXPECTED_ROW_COUNT:
        duplicates = sorted({value for value in mapping_ids if mapping_ids.count(value) > 1})
        raise ValueError(f"mapping_id must be unique, duplicate examples={duplicates[:10]}")

    stats_by_market = {item.strategic_market_id: item for item in stats}
    expected_market_ids = {config.strategic_market_id for config in MARKET_SHEETS}
    if set(stats_by_market) != expected_market_ids:
        raise ValueError(
            f"market stats mismatch: expected={sorted(expected_market_ids)}, actual={sorted(stats_by_market)}"
        )

    for market_id, expected in EXPECTED_MARKET_STATS.items():
        actual = stats_by_market[market_id]
        for field in (
            "sheet_name",
            "header_row",
            "raw_rows_scanned",
            "empty_rows",
            "excluded_rows",
            "staging_rows",
            "manual_specs",
            "mapping_rows",
        ):
            actual_value = getattr(actual, field)
            expected_value = expected[field]
            if actual_value != expected_value:
                raise ValueError(
                    f"{market_id} {field} mismatch: expected={expected_value}, actual={actual_value}"
                )

    zero_mapping_actual = {item.strategic_market_id for item in stats if item.mapping_rows == 0}
    if zero_mapping_actual != ZERO_MAPPING_MARKETS:
        raise ValueError(
            f"zero mapping market mismatch: expected={sorted(ZERO_MAPPING_MARKETS)}, "
            f"actual={sorted(zero_mapping_actual)}"
        )

    market_distribution = _count_by(records, "strategic_market_id")
    if market_distribution != EXPECTED_MARKET_DISTRIBUTION:
        raise ValueError(
            f"market distribution mismatch: expected={EXPECTED_MARKET_DISTRIBUTION}, "
            f"actual={market_distribution}"
        )

    type_distribution = _count_by(records, "mapping_type")
    if type_distribution != EXPECTED_MAPPING_TYPE_DISTRIBUTION:
        raise ValueError(
            f"mapping_type distribution mismatch: expected={EXPECTED_MAPPING_TYPE_DISTRIBUTION}, "
            f"actual={type_distribution}"
        )

    expected_columns = set(MASTER_MAPPING_TABLE_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )
        if not record["mapping_id"]:
            raise ValueError(f"row {index} has blank mapping_id")
        if not record["target_column"]:
            raise ValueError(f"row {index} has blank target_column: {record}")
        if record["mapping_type"] not in EXPECTED_MAPPING_TYPE_DISTRIBUTION:
            raise ValueError(f"row {index} unexpected mapping_type: {record['mapping_type']}")


def print_summary(
    records: list[dict[str, Any]],
    stats: list[MarketMappingStats],
    output_file: Path,
) -> None:
    print("Phase 09d master_mapping_table load")
    print(f"mapping_rows: {len(records)}")
    print(f"unique_mapping_id: {len({record['mapping_id'] for record in records})}")
    print(f"mapping_type_distribution: {_count_by(records, 'mapping_type')}")
    print("market_distribution:")
    for item in stats:
        print(
            f"  {item.strategic_market_id} {item.sheet_name}: "
            f"raw={item.raw_rows_scanned}, empty={item.empty_rows}, "
            f"excluded={item.excluded_rows}, staging={item.staging_rows}, "
            f"manual_specs={item.manual_specs}, mapping={item.mapping_rows}"
        )
    if output_file.exists():
        print(f"output_file: {output_file} ({output_file.stat().st_size:,} bytes)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load MI Master mapping_table parquet.")
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--catalog-path", type=Path, default=DEFAULT_CATALOG_PATH)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = resolve_input_file(args.input_file)
    records, stats = load_mapping_records(xlsx_path, args.catalog_path)
    validate_records(records, stats)
    write_parquet(records, args.output_file)
    print_summary(records, stats, args.output_file)
    print("validate_records: PASS")


if __name__ == "__main__":
    main()

def load_column_metadata_catalog(path: Path) -> dict[str, dict[str, dict[str, Any]]]:
    return load_column_metadata_catalog_common(path, {config.strategic_market_id for config in MARKET_SHEETS})


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    write_records_parquet(records, MASTER_MAPPING_TABLE_COLUMNS, output_file, stringify=True)

