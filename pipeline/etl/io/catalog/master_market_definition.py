"""
prototype_07_master_market_definition_to_parquet.py
===================================================
MI Master "시장정의 & Target" sheet -> Parquet.

Phase 09a policy:
- Canonical logic: /Users/rexxa/github/jw-market/etl/master_market_definition.py
- Canonical schema: /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_market_definition
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.
- Q&A sheet and JW major product list are intentionally not parsed in 09a.

Usage, in Step D after user review:
    python3 scripts/prototype_07_master_market_definition_to_parquet.py
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


from pipeline.etl.lib.storage import get_mi_master_path
from pipeline.etl.io.catalog._common import (
    STANDARD_PREFIX,
    _extra_key,
    _header_lookup,
    _lookup_key,
    _lookup_position_value,
    _lookup_source_value,
    _position_value,
    _single_lookup_key,
    apply_column_mapping,
    build_raw_row_payload,
    cell_text,
    dumps_json,
    explicit_lookup_join,
    is_empty_row,
    load_column_metadata_catalog as load_column_metadata_catalog_common,
    make_header_keys,
    normalize_header,
    to_jsonable,
    utc_now_text,
    write_records_parquet,
)


DEFAULT_INPUT_FILE = get_mi_master_path()
DEFAULT_OUTPUT_FILE = Path("parquet/master_market_definition/master_market_definition.parquet")
SOURCE_SHEET = "시장정의 & Target"

MASTER_MARKET_DEFINITION_COLUMNS = (
    "strategic_market_id",
    "market_name",
    "source_type",
    "market_atc_codes_json",
    "full_market_atc4_codes_json",
    "direct_competition_brands_json",
    "description",
    "analysis_levels_json",
    "analysis_level_funnel",
    "analysis_level_etc",
    "target_customer_priority_json",
    "raw_row_json",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)

MARKET_BY_ID = {
    "strategy_001": {"sheet_name": "라베칸 라베칸듀오", "source_type": "UBIST"},
    "strategy_002": {"sheet_name": "제이클", "source_type": "IQVIA"},
    "strategy_003": {"sheet_name": "가드렛 가드메트", "source_type": "IQVIA"},
    "strategy_004": {"sheet_name": "타발리스", "source_type": "IQVIA"},
    "strategy_005": {"sheet_name": "시그마트", "source_type": "IQVIA"},
    "strategy_006": {"sheet_name": "리바로 리바로젯", "source_type": "UBIST"},
    "strategy_007": {"sheet_name": "리바로페노", "source_type": "UBIST"},
    "strategy_008": {"sheet_name": "리바로하이 리바로브이", "source_type": "UBIST"},
    "strategy_009": {"sheet_name": "트루패스 피나스타 제이다트", "source_type": "UBIST"},
    "strategy_010": {"sheet_name": "뉴트로진 모빌리아", "source_type": "IQVIA"},
    "strategy_011": {"sheet_name": "악템라", "source_type": "IQVIA"},
    "strategy_012": {"sheet_name": "페린젝트 베노훼럼", "source_type": "IQVIA"},
    "strategy_013": {"sheet_name": "헴리브라", "source_type": "IQVIA"},
    "strategy_014": {"sheet_name": "위너프 위너프A+", "source_type": "IQVIA"},
    "strategy_015": {"sheet_name": "엔커버", "source_type": "IQVIA"},
    "strategy_016": {"sheet_name": "플라주오피", "source_type": "IQVIA"},
}

# 1-based column indexes in sheet "시장정의 & Target".
MARKET_DEFINITION_COLUMNS: dict[str, tuple[int, ...]] = {
    "strategy_001": (3,),
    "strategy_002": (4,),
    "strategy_003": (5,),
    "strategy_004": (6,),
    "strategy_005": (7,),
    "strategy_006": (8,),
    "strategy_007": (9,),
    "strategy_008": (10, 11),
    "strategy_009": (12, 13),
    "strategy_010": (14, 15),
    "strategy_011": (16,),
    "strategy_012": (17, 18),
    "strategy_013": (19,),
    "strategy_015": (20,),
    "strategy_014": (21,),
    "strategy_016": (22,),
}

MARKET_DESCRIPTIONS = {
    "strategy_015": "IQVIA 기준 하모닐란과 엔커버 2개의 PRODUCT NAME KOR 에 대해 PACK DESC 를 하위분류로 4가지로 분석",
}

ANALYSIS_LEVEL_ROWS = {
    14: "Class",
    15: "Molecule",
    16: "Brand",
    17: "Dosage Form",
    18: "Strength",
    19: "Etc",
}
FULL_MARKET_ROWS = range(22, 45)
DIRECT_COMPETITION_ROWS = range(48, 51)
TARGET_CUSTOMER_ROWS = range(54, 58)
METRIC_ROWS = range(61, 65)

EXPECTED_STRATEGIC_MARKET_IDS = tuple(MARKET_DEFINITION_COLUMNS.keys())

ATC_CODE_PATTERN = r"[A-Z]\d{0,2}[A-Z](?:\d{0,2})?"
ATC_BRACKET_RE = re.compile(rf"\[({ATC_CODE_PATTERN})\]")
ATC_PLAIN_RE = re.compile(rf"^({ATC_CODE_PATTERN})\b")


def atc_code_from_text(value: object) -> str | None:
    text = cell_text(value)
    if not text:
        return None
    bracket_match = ATC_BRACKET_RE.search(text)
    if bracket_match:
        return bracket_match.group(1)
    plain_match = ATC_PLAIN_RE.match(text)
    if plain_match:
        return plain_match.group(1)
    return None


def find_atc_column(ws) -> int | None:
    for column_id in range(1, 12):
        for row_id in range(3, 9):
            value = ws.cell(row_id, column_id).value
            if value and "ATC" in str(value).upper():
                return column_id
    return None


def extract_atc_from_market_sheet(wb, sheet_name: str) -> list[str]:
    """Extract normalized ATC codes from a market sheet's ATC column."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    atc_column = find_atc_column(ws)
    if atc_column is None:
        return []

    codes: set[str] = set()
    for (value,) in ws.iter_rows(
        min_row=6,
        max_row=ws.max_row,
        min_col=atc_column,
        max_col=atc_column,
        values_only=True,
    ):
        code = atc_code_from_text(value)
        if code:
            codes.add(code)
    return sorted(codes)


def blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_MARKET_DEFINITION_COLUMNS}


def source_type_from_values(values: list[str | None], fallback: str) -> str:
    joined = " ".join(value or "" for value in values).lower()
    if "ubist" in joined and "iqvia" in joined:
        return "BOTH"
    if "ubist" in joined:
        return "UBIST"
    if "iqvia" in joined:
        return "IQVIA"
    return fallback


def values_for_rows(ws, rows: range, columns: tuple[int, ...]) -> list[dict[str, Any]]:
    values: list[dict[str, Any]] = []
    for row_id in rows:
        row_label = cell_text(ws.cell(row_id, 2).value) or cell_text(ws.cell(row_id, 1).value)
        for column_id in columns:
            value = ws.cell(row_id, column_id).value
            if value is None or str(value).strip() == "":
                continue
            values.append(
                {
                    "row_id": row_id,
                    "label": row_label,
                    "column_id": column_id,
                    "product_name": cell_text(ws.cell(6, column_id).value),
                    "value": to_jsonable(value),
                }
            )
    return values


def analysis_levels(ws, columns: tuple[int, ...]) -> dict[str, Any]:
    levels: dict[str, Any] = {}
    for row_id, level_name in ANALYSIS_LEVEL_ROWS.items():
        by_product: dict[str, Any] = {}
        values: list[Any] = []
        for column_id in columns:
            product_name = cell_text(ws.cell(6, column_id).value) or f"col_{column_id}"
            value = to_jsonable(ws.cell(row_id, column_id).value)
            by_product[product_name] = value
            if value not in (None, ""):
                values.append(value)
        levels[level_name] = {
            "by_product": by_product,
            "values": values,
        }
    levels["Metrics"] = values_for_rows(ws, METRIC_ROWS, columns)
    return levels


def target_customer_priority(ws, columns: tuple[int, ...]) -> list[dict[str, Any]]:
    return values_for_rows(ws, TARGET_CUSTOMER_ROWS, columns)


def raw_payload_for_market(ws, columns: tuple[int, ...]) -> dict[str, Any]:
    payload_columns = []
    for column_id in columns:
        values = []
        for row_id in range(5, 65):
            label = cell_text(ws.cell(row_id, 2).value) or cell_text(ws.cell(row_id, 1).value)
            cell_value = ws.cell(row_id, column_id).value
            if label is None and (cell_value is None or str(cell_value).strip() == ""):
                continue
            values.append(
                {
                    "row_id": row_id,
                    "label": label,
                    "value": to_jsonable(cell_value),
                }
            )
        payload_columns.append(
            {
                "column_id": column_id,
                "team": cell_text(ws.cell(5, column_id).value),
                "product_name": cell_text(ws.cell(6, column_id).value),
                "values": values,
            }
        )
    return {"source_sheet": SOURCE_SHEET, "columns": payload_columns}


def iter_market_definition_rows(xlsx_path: Path, ingested_at: str | None = None) -> Iterable[dict[str, Any]]:
    timestamp = ingested_at or utc_now_text()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            raise ValueError(f"required sheet not found: {SOURCE_SHEET!r}; sheets={wb.sheetnames}")
        ws = wb[SOURCE_SHEET]

        for strategic_market_id, columns in MARKET_DEFINITION_COLUMNS.items():
            config = MARKET_BY_ID[strategic_market_id]
            source_values = [cell_text(ws.cell(10, column).value) for column in columns]
            market_atc_codes = extract_atc_from_market_sheet(wb, config["sheet_name"])
            full_market_values = values_for_rows(ws, FULL_MARKET_ROWS, columns)
            direct_competition_values = values_for_rows(ws, DIRECT_COMPETITION_ROWS, columns)
            levels = analysis_levels(ws, columns)
            customer_priority = target_customer_priority(ws, columns)
            etc_values = levels.get("Etc", {}).get("values", [])

            record = blank_record()
            record.update(
                {
                    "strategic_market_id": strategic_market_id,
                    "market_name": config["sheet_name"],
                    "source_type": source_type_from_values(source_values, config["source_type"]),
                    "market_atc_codes_json": dumps_json(market_atc_codes),
                    "full_market_atc4_codes_json": dumps_json(full_market_values),
                    "direct_competition_brands_json": dumps_json(direct_competition_values),
                    "description": MARKET_DESCRIPTIONS.get(strategic_market_id),
                    "analysis_levels_json": dumps_json(levels),
                    "analysis_level_funnel": "O" if strategic_market_id == "strategy_001" else None,
                    "analysis_level_etc": "; ".join(str(value).strip() for value in etc_values if value),
                    "target_customer_priority_json": dumps_json(customer_priority),
                    "raw_row_json": dumps_json(raw_payload_for_market(ws, columns)),
                    "source_sheet": SOURCE_SHEET,
                    "source_file_version": xlsx_path.name,
                    "ingested_at": timestamp,
                }
            )
            yield record
    finally:
        wb.close()


def validate_records(records: list[dict[str, Any]]) -> None:
    if len(records) != 16:
        raise ValueError(f"market_definition row count must be 16, found {len(records)}")

    ids = [record["strategic_market_id"] for record in records]
    if len(set(ids)) != 16:
        raise ValueError(f"strategic_market_id must be unique, found duplicate ids: {ids}")
    if tuple(ids) != EXPECTED_STRATEGIC_MARKET_IDS:
        raise ValueError(
            "strategic_market_id order/mapping mismatch: "
            f"expected={EXPECTED_STRATEGIC_MARKET_IDS}, actual={tuple(ids)}"
        )

    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - set(MASTER_MARKET_DEFINITION_COLUMNS))
        missing_columns = sorted(set(MASTER_MARKET_DEFINITION_COLUMNS) - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )
        for column in (
            "market_atc_codes_json",
            "full_market_atc4_codes_json",
            "direct_competition_brands_json",
            "analysis_levels_json",
            "target_customer_priority_json",
            "raw_row_json",
        ):
            json.loads(record[column])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("--output-file", default=str(DEFAULT_OUTPUT_FILE))
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_file = Path(args.output_file)
    if not input_file.exists():
        sys.exit(f"ERROR: input file not found: {input_file}")

    print("=" * 72)
    print("MI Master market_definition -> Parquet")
    print("=" * 72)
    print(f"  input file:   {input_file}")
    print(f"  source sheet: {SOURCE_SHEET}")
    print(f"  output file:  {output_file}")
    print(f"  columns:      {len(MASTER_MARKET_DEFINITION_COLUMNS)} DDL columns")
    print("  helpers:      none")

    records = list(iter_market_definition_rows(input_file))
    validate_records(records)
    write_parquet(records, output_file)

    print("\nResult")
    print(f"  rows:                   {len(records)}")
    print(f"  unique strategic ids:   {len({row['strategic_market_id'] for row in records})}")
    print(f"  output size:            {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  ingested_at:            {records[0]['ingested_at'] if records else None}")
    print("\nDone")


if __name__ == "__main__":
    main()

def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    write_records_parquet(records, MASTER_MARKET_DEFINITION_COLUMNS, output_file, compression_level=3)

