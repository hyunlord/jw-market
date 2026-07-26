from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.etl.io.catalog.master import drug as master_drug
from pipeline.etl.io.catalog._lib.common import is_empty_row, read_parquet_rows
from pipeline.etl.io.catalog.brand.strategic_brand_logic import (
    _class_source_indexes,
    assign_cd_id,
    classify_exclusion_cells,
    contains_excluded,
    dumps_json_array,
    extract_atc_code,
    make_name,
    normalize_for_match,
    null_if_excluded,
    source_version_from_ml_market,
    strategic_fields,
)
from pipeline.etl.io.catalog.brand.strategic_atc4_expansion import (
    RawAtc4Brand,
    load_raw_atc4_brands,
)
from pipeline.etl.io.catalog.brand.strategic_brand_schema import EXPECTED_COLUMNS, MERGE_NAME_BY_NAME
from pipeline.etl.io.catalog.brand.strategic_brand_schema import validate_records
from pipeline.etl.mi_master_registry import apply_record_rules

PROJECT_ROOT = Path(__file__).resolve().parents[5]
DEFAULT_CATALOG_PATH = PROJECT_ROOT / "pipeline" / "etl" / "config" / "master_column_mapping_catalog.md"


def utc_now_datetime() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _source_row_id_from_brand_id(brand_id: Any) -> int:
    try:
        return int(str(brand_id).rsplit("_", 1)[1])
    except (IndexError, TypeError, ValueError):
        return 0


def _apply_source_rules(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    standard_values: dict[str, Any],
    sheet_name: str,
) -> dict[str, Any]:
    raw = {
        str(header).strip(): value
        for header, value in zip(headers, values)
        if header is not None
    }
    return apply_record_rules(
        standard_values,
        stage="strategic_brand_source",
        context={"sheet_name": sheet_name, "raw": raw},
    )


def _assign_cd_id_for_atc4_codes(
    *,
    ml_id: str,
    atc4_codes: list[str],
    cd_markets_for_ml: dict[str, list[dict[str, Any]]],
    filter_by_id: dict[str, dict[str, Any]],
) -> tuple[str | None, list[str]]:
    candidates: list[str] = []
    for atc4_code in atc4_codes:
        match_context = {
            "ml_id": ml_id,
            "atc4_code": atc4_code,
            "class": None,
            "molecule": None,
            "dosage_form": None,
            "nhi_type": None,
        }
        _cd_id, row_candidates = assign_cd_id(match_context, cd_markets_for_ml, filter_by_id)
        for candidate in row_candidates:
            if candidate not in candidates:
                candidates.append(candidate)
    if len(candidates) == 1:
        return candidates[0], candidates
    return None, candidates


def _expanded_record(
    *,
    ml_index: int,
    ml_id: str,
    seq: int,
    raw_brand: RawAtc4Brand,
    source_file_version: str,
    ingested_at: datetime,
    cd_markets_for_ml: dict[str, list[dict[str, Any]]],
    filter_by_id: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], list[str]]:
    atc4_codes = sorted(raw_brand.atc4_codes)
    cd_id, candidates = _assign_cd_id_for_atc4_codes(
        ml_id=ml_id,
        atc4_codes=atc4_codes,
        cd_markets_for_ml=cd_markets_for_ml,
        filter_by_id=filter_by_id,
    )
    record = {
        "brand_id": f"sb_{ml_index:03d}_atc4_{seq:05d}",
        "name": raw_brand.name,
        "merge_name": MERGE_NAME_BY_NAME.get(raw_brand.name, raw_brand.name),
        "ml_id": ml_id,
        "cd_id": cd_id,
        "is_excluded": False,
        "is_class_excluded": False,
        "allowed_atc4_codes_json": dumps_json_array(atc4_codes),
        "class": None,
        "class_1": None,
        "class_2": None,
        "molecule": None,
        "dosage_form": None,
        "strength_pack": None,
        "nhi_type": None,
        "ox_gx": None,
        "fish_oil": None,
        "판매사": None,
        "제조사": None,
        "source_file_version": source_file_version,
        "ingested_at": ingested_at,
    }
    return {column: record.get(column) for column in EXPECTED_COLUMNS}, candidates

def load_strategic_brand_records(
    ml_market_path: Path,
    cd_filter_path: Path,
    cd_market_path: Path,
    input_file: Path | None = None,
    catalog_path: Path | None = None,
    ubist_dir: Path | None = None,
    iqvia_nsa_dir: Path | None = None,
    ingested_at: datetime | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    xlsx_path = master_drug.resolve_input_file(input_file or master_drug.DEFAULT_INPUT_FILE)
    metadata_catalog = master_drug.load_column_metadata_catalog(catalog_path or DEFAULT_CATALOG_PATH)

    ml_rows = read_parquet_rows(ml_market_path)
    cd_filter_rows = read_parquet_rows(cd_filter_path)
    cd_market_rows = read_parquet_rows(cd_market_path)
    ml_ids = {str(row["ml_id"]) for row in ml_rows}
    filter_by_id = {str(row["cd_filter_id"]): row for row in cd_filter_rows}
    cd_markets_for_ml: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cd_market_rows:
        cd_markets_for_ml[str(row["ml_id"])].append(row)

    timestamp = ingested_at or utc_now_datetime()
    source_file_version = source_version_from_ml_market(ml_rows)
    raw_atc4_brands, raw_atc4_stats = load_raw_atc4_brands(
        ml_rows,
        ubist_dir=ubist_dir or (PROJECT_ROOT / "output" / "ubist"),
        iqvia_nsa_dir=iqvia_nsa_dir or (PROJECT_ROOT / "output" / "iqvia_nsa"),
    )
    records: list[dict[str, Any]] = []
    gadrelet_rows: list[dict[str, Any]] = []
    stats = {
        "raw_rows_scanned": Counter(),
        "empty_rows": Counter(),
        "excluded_rows": Counter(),
        "included_rows": Counter(),
        "nullified_cells": Counter(),
        "overlap_rows": [],
        "unknown_name_rows": [],
        "sheet_included_rows": Counter(),
        "expanded_atc4_rows": Counter(),
        "expanded_atc4_new_rows": Counter(),
        "expanded_atc4_skipped_sheet_keys": Counter(),
        "raw_atc4_sources": raw_atc4_stats,
    }

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for ml_index, config in enumerate(master_drug.MARKET_SHEETS, start=1):
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}")
            ws = wb[config.sheet_name]
            headers = master_drug._headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                master_drug.explicit_lookup_join(row_items)
                if config.strategic_market_id == "strategy_008"
                else {}
            )
            class_indexes = _class_source_indexes(headers, metadata)
            ml_id = f"ml_{ml_index:03d}"
            if ml_id not in ml_ids:
                raise ValueError(f"{config.strategic_market_id} missing ml_market FK: {ml_id}")

            allowed_atc4_by_name: dict[str, set[str]] = defaultdict(set)
            sheet_name_keys: set[str] = set()
            for source_row_id, values in row_items:
                if master_drug.is_empty_row(values):
                    continue
                row_excluded, _class_excluded = classify_exclusion_cells(
                    headers,
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                )
                if row_excluded:
                    continue
                standard_values, extras = master_drug.apply_column_mapping(headers, values, metadata)
                standard_values = _apply_source_rules(
                    headers,
                    values,
                    standard_values,
                    config.sheet_name,
                )
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                name = make_name(
                    standard_values,
                    source_row_id,
                    sheet_name=config.sheet_name,
                )
                sheet_name_keys.add(normalize_for_match(name))
                fields = strategic_fields(
                    standard_values,
                    extras,
                    sheet_name=config.sheet_name,
                )
                atc4_code = extract_atc_code(fields.get("atc4_code"))
                if atc4_code:
                    allowed_atc4_by_name[normalize_for_match(name)].add(atc4_code)
            for raw_brand in raw_atc4_brands.get(ml_id, {}).values():
                sheet_key = normalize_for_match(raw_brand.name)
                if sheet_key in sheet_name_keys:
                    allowed_atc4_by_name[sheet_key].update(raw_brand.atc4_codes)

            for source_row_id, values in row_items:
                stats["raw_rows_scanned"][config.strategic_market_id] += 1
                if master_drug.is_empty_row(values):
                    stats["empty_rows"][config.strategic_market_id] += 1
                    continue

                excluded, class_excluded = classify_exclusion_cells(
                    headers,
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                )
                if excluded:
                    stats["excluded_rows"][config.strategic_market_id] += 1

                standard_values, extras = master_drug.apply_column_mapping(headers, values, metadata)
                standard_values = _apply_source_rules(
                    headers,
                    values,
                    standard_values,
                    config.sheet_name,
                )
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])

                name = make_name(
                    standard_values,
                    source_row_id,
                    sheet_name=config.sheet_name,
                )
                if name.startswith("unknown_row_"):
                    stats["unknown_name_rows"].append(
                        {
                            "strategic_market_id": config.strategic_market_id,
                            "source_row_id": source_row_id,
                        }
                    )
                fields = strategic_fields(
                    standard_values,
                    extras,
                    sheet_name=config.sheet_name,
                )
                match_context = dict(fields)
                match_context["ml_id"] = ml_id
                cd_id, candidates = assign_cd_id(match_context, cd_markets_for_ml, filter_by_id)
                if len(candidates) > 1:
                    stats["overlap_rows"].append(
                        {
                            "strategic_market_id": config.strategic_market_id,
                            "source_row_id": source_row_id,
                            "name": name,
                            "candidates": ",".join(candidates),
                        }
                    )

                record = {
                    "brand_id": f"sb_{ml_index:03d}_{source_row_id:05d}",
                    "name": name,
                    "merge_name": MERGE_NAME_BY_NAME.get(name, name),
                    "ml_id": ml_id,
                    "cd_id": cd_id,
                    "is_excluded": bool(excluded),
                    "is_class_excluded": bool(class_excluded),
                    "allowed_atc4_codes_json": dumps_json_array(list(allowed_atc4_by_name.get(normalize_for_match(name), set()))),
                    "class": fields["class"],
                    "class_1": fields["class_1"],
                    "class_2": fields["class_2"],
                    "molecule": fields["molecule"],
                    "dosage_form": fields["dosage_form"],
                    "strength_pack": fields["strength_pack"],
                    "nhi_type": fields["nhi_type"],
                    "ox_gx": fields["ox_gx"],
                    "fish_oil": fields["fish_oil"],
                    "판매사": fields["판매사"],
                    "제조사": fields["제조사"],
                    "source_file_version": source_file_version,
                    "ingested_at": timestamp,
                }

                for column, value in {
                    "class": standard_values.get("class_2") or standard_values.get("class"),
                    "class_1": standard_values.get("class"),
                    "class_2": standard_values.get("class_2"),
                    "molecule": standard_values.get("molecule"),
                    "dosage_form": standard_values.get("dosage_form"),
                    "strength_pack": standard_values.get("strength") or standard_values.get("pack_desc") or extras.get("product_pack"),
                    "nhi_type": standard_values.get("nhi_type"),
                    "ox_gx": standard_values.get("ox_gx") or extras.get("ox_gx") or extras.get("ox_gx_biosimilar"),
                    "fish_oil": extras.get("fish_oil_yn"),
                    "판매사": standard_values.get("seller"),
                    "제조사": standard_values.get("manufacturer"),
                }.items():
                    if contains_excluded(value):
                        stats["nullified_cells"][column] += 1
                records.append({column: record.get(column) for column in EXPECTED_COLUMNS})
                stats["included_rows"][config.strategic_market_id] += 1
                stats["sheet_included_rows"][config.strategic_market_id] += 1

                if config.strategic_market_id == "strategy_003":
                    gadrelet_rows.append(
                        {
                            "brand_id": record["brand_id"],
                            "source_row_id": source_row_id,
                            "atc4_code": fields["atc4_code"],
                            "molecule": fields["molecule"],
                            "class": fields["class"],
                            "dosage_form": fields["dosage_form"],
                            "strategic_brand_name": name,
                            "cd_id": cd_id,
                        }
                    )
            expanded_seq = 0
            for raw_key, raw_brand in sorted(raw_atc4_brands.get(ml_id, {}).items(), key=lambda item: item[0]):
                if normalize_for_match(raw_brand.name) in sheet_name_keys:
                    stats["expanded_atc4_skipped_sheet_keys"][config.strategic_market_id] += 1
                    continue
                expanded_seq += 1
                record, candidates = _expanded_record(
                    ml_index=ml_index,
                    ml_id=ml_id,
                    seq=expanded_seq,
                    raw_brand=raw_brand,
                    source_file_version=source_file_version,
                    ingested_at=timestamp,
                    cd_markets_for_ml=cd_markets_for_ml,
                    filter_by_id=filter_by_id,
                )
                if len(candidates) > 1:
                    stats["overlap_rows"].append(
                        {
                            "strategic_market_id": config.strategic_market_id,
                            "source_row_id": _source_row_id_from_brand_id(record["brand_id"]),
                            "name": record["name"],
                            "candidates": ",".join(candidates),
                        }
                    )
                records.append(record)
                stats["included_rows"][config.strategic_market_id] += 1
                stats["expanded_atc4_rows"][config.strategic_market_id] += 1
                stats["expanded_atc4_new_rows"][config.strategic_market_id] += 1
    finally:
        wb.close()

    summary = {
        "stats": stats,
        "gadrelet_rows": gadrelet_rows,
        "source_file_version": source_file_version,
    }
    validate_records(records, summary, ml_rows, cd_market_rows)
    return records, summary
