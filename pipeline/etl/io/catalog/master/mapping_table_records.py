from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.etl.io.catalog._lib.common import (
    _header_lookup,
    _lookup_source_value,
    apply_column_mapping,
    explicit_lookup_join,
    is_empty_row,
    load_column_metadata_catalog as load_column_metadata_catalog_common,
    normalize_header,
    utc_now_text,
)
from pipeline.etl.io.catalog.master.mapping_table_schema import (
    MARKET_SHEETS,
    MASTER_ROOT,
    MarketMappingStats,
)


def resolve_input_file(path: Path) -> Path:
    if path.exists():
        return path
    candidates = sorted(file for file in MASTER_ROOT.glob("*.xlsx") if not file.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError(f"No Master xlsx found under {MASTER_ROOT}")
    return candidates[-1]


def is_excluded_row(values: list[Any] | tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if "제외" in text and not text.startswith("비제외"):
            return True
    return False


def make_mapping_id(
    strategic_market_id: str,
    source_column: object,
    source_value: object,
    target_column: object,
    target_value: object,
    source_row_id: int | None = None,
) -> str:
    parts = [
        strategic_market_id,
        str(source_row_id or ""),
        str(source_column or ""),
        str(source_value or ""),
        str(target_column or ""),
        str(target_value or ""),
    ]
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"{strategic_market_id}:{digest}"


def mapping_type_for(target_column: str, source_column: object) -> str:
    lowered = f"{target_column} {source_column or ''}".lower()
    if "molecule" in lowered or "성분" in lowered:
        return "molecule_recode"
    if "class" in lowered:
        return "class_recode"
    if "strength" in lowered or "규격" in lowered:
        return "strength_recode"
    if "nhi" in lowered or "급여" in lowered:
        return "nhi_overlay"
    return "manual_mapping"


def raw_value_for_mapping(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    spec: dict[str, Any],
) -> Any:
    lookup = _header_lookup(headers, values)
    overlay_target = normalize_header(spec.get("overlay_target"))
    source_column = normalize_header(spec.get("source_column"))
    if overlay_target and overlay_target in lookup:
        return lookup.get(overlay_target)
    return _lookup_source_value(lookup, source_column)


def _headers_from_sheet(ws: Any, header_row: int) -> list[Any]:
    return list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))


def _manual_mapping_records(
    strategic_market_id: str,
    source_sheet: str,
    source_file_version: str,
    source_row_id: int,
    headers: list[Any],
    values: tuple[Any, ...],
    metadata: dict[str, dict[str, Any]],
    standard_values: dict[str, Any],
    extras: dict[str, Any],
    ingested_at: str,
    inclusion_standard_values: dict[str, Any] | None = None,
    inclusion_extras: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    inclusion_standard_values = inclusion_standard_values or standard_values
    inclusion_extras = inclusion_extras or extras
    for target_column, spec in metadata.items():
        metadata_type = str(spec.get("type") or "")
        if not metadata_type.startswith("manual"):
            continue
        source_column = spec.get("source_column")
        if target_column.startswith("drug_extra_json."):
            extra_key = target_column.split(".", 1)[1]
            target_value = extras.get(extra_key)
            inclusion_target_value = inclusion_extras.get(extra_key)
        else:
            target_value = standard_values.get(target_column)
            inclusion_target_value = inclusion_standard_values.get(target_column)
        source_value = raw_value_for_mapping(headers, values, spec)
        if source_value is None and inclusion_target_value is None:
            continue
        records.append(
            {
                "mapping_id": make_mapping_id(
                    strategic_market_id,
                    source_column,
                    source_value,
                    target_column,
                    target_value,
                    source_row_id=source_row_id,
                ),
                "strategic_market_id": strategic_market_id,
                "source_value": None if source_value is None else str(source_value),
                "target_column": target_column,
                "target_value": None if target_value is None else str(target_value),
                "mapping_type": mapping_type_for(target_column, source_column),
                "source_sheet": source_sheet,
                "source_file_version": source_file_version,
                "ingested_at": ingested_at,
            }
        )
    return records


def load_mapping_records(
    xlsx_path: Path,
    catalog_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], list[MarketMappingStats]]:
    metadata_catalog = load_column_metadata_catalog(catalog_path)
    timestamp = ingested_at or utc_now_text()
    records: list[dict[str, Any]] = []
    stats: list[MarketMappingStats] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for config in MARKET_SHEETS:
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}; sheets={wb.sheetnames}")
            ws = wb[config.sheet_name]
            headers = _headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            manual_specs = sum(
                1 for spec in metadata.values() if str(spec.get("type") or "").startswith("manual")
            )
            market_stats = MarketMappingStats(
                strategic_market_id=config.strategic_market_id,
                sheet_name=config.sheet_name,
                header_row=config.header_row,
                max_row=ws.max_row,
                max_col=ws.max_column,
                manual_specs=manual_specs,
            )

            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                explicit_lookup_join(row_items) if config.strategic_market_id == "strategy_008" else {}
            )

            for source_row_id, values in row_items:
                market_stats.raw_rows_scanned += 1
                if is_empty_row(values):
                    market_stats.empty_rows += 1
                    continue
                if is_excluded_row(values):
                    market_stats.excluded_rows += 1
                    continue

                market_stats.staging_rows += 1
                standard_values, extras = apply_column_mapping(headers, values, metadata)
                inclusion_standard_values = standard_values
                if source_row_id in explicit_overrides:
                    standard_values = dict(standard_values)
                    standard_values.update(explicit_overrides[source_row_id])
                mapping_rows = _manual_mapping_records(
                    config.strategic_market_id,
                    config.sheet_name,
                    xlsx_path.name,
                    source_row_id,
                    headers,
                    values,
                    metadata,
                    standard_values,
                    extras,
                    timestamp,
                    inclusion_standard_values=inclusion_standard_values,
                )
                records.extend(mapping_rows)
                market_stats.mapping_rows += len(mapping_rows)

            stats.append(market_stats)
    finally:
        wb.close()

    return records, stats


def load_column_metadata_catalog(path: Path) -> dict[str, dict[str, dict[str, Any]]]:
    return load_column_metadata_catalog_common(path, {config.strategic_market_id for config in MARKET_SHEETS})
