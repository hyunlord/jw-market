from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.etl.io.catalog._lib.common import (
    apply_column_mapping,
    build_raw_row_payload,
    dumps_json,
    explicit_lookup_join,
    is_empty_row,
    load_column_metadata_catalog as load_column_metadata_catalog_common,
    normalize_header,
    utc_now_text,
)
from pipeline.etl.io.catalog._lib.exclusion_policy import classify_exclusion_cells as classify_exclusion_cells_by_policy
from pipeline.etl.io.catalog.master.drug_schema import (
    MARKET_SHEETS,
    MASTER_DRUG_COLUMNS,
    MASTER_ROOT,
    MarketDrugStats,
)


def resolve_input_file(path: Path) -> Path:
    if path.exists():
        # Match the current project loader path semantics: scripts/load_master.py
        # normally obtains this path from find_master_xlsx(), which returns the
        # actual filesystem directory entry. On macOS that preserves the
        # decomposed Korean filename in xlsx_path.name.
        for candidate in sorted(file for file in path.parent.glob("*.xlsx") if not file.name.startswith("~$")):
            try:
                if candidate.samefile(path):
                    return candidate
            except FileNotFoundError:
                continue
        return path
    candidates = sorted(file for file in MASTER_ROOT.glob("*.xlsx") if not file.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError(f"No Master xlsx found under {MASTER_ROOT}")
    return candidates[-1]


def _is_class_header(header: Any) -> bool:
    text = normalize_header(header)
    if not text:
        return False
    normalized = re.sub(r"[\s_-]+", "", text).lower()
    return normalized in {"class", "class1", "class2"} or normalized.startswith("class")


def _class_source_indexes(headers: list[Any] | tuple[Any, ...], metadata: dict[str, dict[str, Any]]) -> set[int]:
    indexes: set[int] = {idx for idx, header in enumerate(headers) if _is_class_header(header)}
    for target in ("class", "class_1", "class_2"):
        spec = metadata.get(target) or {}
        if spec.get("position") is not None:
            try:
                indexes.add(int(spec["position"]))
            except (TypeError, ValueError):
                pass
            continue
        source_column = normalize_header(spec.get("source_column"))
        if not source_column:
            continue
        for idx, header in enumerate(headers):
            text = normalize_header(header)
            if text and (text == source_column or text.startswith(source_column)):
                indexes.add(idx)
    return indexes


def is_excluded_row(
    values: list[Any] | tuple[Any, ...],
    class_indexes: set[int] | None = None,
    *,
    strategic_market_id: str | None = None,
    sheet_name: str | None = None,
) -> bool:
    class_indexes = set(class_indexes or ())
    row_excluded, _class_excluded = classify_exclusion_cells_by_policy(
        values,
        class_indexes=class_indexes,
        strategic_market_id=strategic_market_id,
        sheet_name=sheet_name,
    )
    return row_excluded


def _headers_from_sheet(ws: Any, header_row: int) -> list[Any]:
    return list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))


def _blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_DRUG_COLUMNS}


def load_drug_records(
    xlsx_path: Path,
    catalog_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], list[MarketDrugStats]]:
    metadata_catalog = load_column_metadata_catalog(catalog_path)
    timestamp = ingested_at or utc_now_text()
    records: list[dict[str, Any]] = []
    stats: list[MarketDrugStats] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for config in MARKET_SHEETS:
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}; sheets={wb.sheetnames}")
            ws = wb[config.sheet_name]
            headers = _headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            market_stats = MarketDrugStats(
                strategic_market_id=config.strategic_market_id,
                sheet_name=config.sheet_name,
                header_row=config.header_row,
                max_row=ws.max_row,
                max_col=ws.max_column,
            )
            drug_index = 0

            row_items = list(enumerate(
                ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                start=config.header_row + 1,
            ))
            explicit_overrides = (
                explicit_lookup_join(row_items) if config.strategic_market_id == "strategy_008" else {}
            )
            class_indexes = _class_source_indexes(headers, metadata)

            for source_row_id, values in row_items:
                market_stats.raw_rows_scanned += 1
                if is_empty_row(values):
                    market_stats.empty_rows += 1
                    continue
                raw_payload = build_raw_row_payload(headers, values, source_row_id)
                if is_excluded_row(
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                ):
                    market_stats.excluded_rows += 1
                    continue

                standard_values, extras = apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                drug_index += 1
                market_stats.staging_rows += 1

                record = _blank_record()
                record.update(
                    {
                        "strategic_market_id": config.strategic_market_id,
                        "market_name": config.sheet_name,
                        "source_type": config.source_type,
                        "drug_index": drug_index,
                        "drug_extra_json": dumps_json(extras),
                        "raw_row_json": dumps_json(raw_payload),
                        "column_metadata_json": dumps_json(metadata),
                        "source_sheet": config.sheet_name,
                        "source_file_version": xlsx_path.name,
                        "source_row_id": source_row_id,
                        "ingested_at": timestamp,
                    }
                )
                for key, value in standard_values.items():
                    if key in record:
                        record[key] = value
                records.append(record)
            stats.append(market_stats)
    finally:
        wb.close()

    return records, stats


def load_column_metadata_catalog(path: Path) -> dict[str, dict[str, dict[str, Any]]]:
    return load_column_metadata_catalog_common(path, {config.strategic_market_id for config in MARKET_SHEETS})
