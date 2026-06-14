"""MI Master market_definition -> Parquet (prototype_07)."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from pipeline.etl.io.catalog._lib.common import (
    cell_text,
    dumps_json,
    to_jsonable,
    utc_now_text,
    write_records_parquet,
)
from pipeline.etl.io.catalog.master.market_definition_schema import (
    ANALYSIS_LEVEL_ROWS,
    DEFAULT_INPUT_FILE,
    DEFAULT_OUTPUT_FILE,
    DIRECT_COMPETITION_ROWS,
    FULL_MARKET_ROWS,
    MARKET_BY_ID,
    MARKET_DEFINITION_COLUMNS,
    MARKET_DESCRIPTIONS,
    MASTER_MARKET_DEFINITION_COLUMNS,
    METRIC_ROWS,
    SOURCE_SHEET,
    TARGET_CUSTOMER_ROWS,
)
from pipeline.etl.io.catalog.master.market_definition_validation import validate_records

ATC_CODE_PATTERN = r"[A-Z]\d{0,2}[A-Z](?:\d{0,2})?"
ATC_BRACKET_RE = re.compile(rf"\[({ATC_CODE_PATTERN})\]")
ATC_PLAIN_RE = re.compile(rf"^({ATC_CODE_PATTERN})\b")


def atc_code_from_text(value: object) -> str | None:
    text = cell_text(value)
    if not text:
        return None
    bracket_match = ATC_BRACKET_RE.search(text)
    if bracket_match:
        return bracket_match.group(1)
    plain_match = ATC_PLAIN_RE.match(text)
    if plain_match:
        return plain_match.group(1)
    return None


def find_atc_column(ws) -> int | None:
    for column_id in range(1, 12):
        for row_id in range(3, 9):
            value = ws.cell(row_id, column_id).value
            if value and "ATC" in str(value).upper():
                return column_id
    return None


def extract_atc_from_market_sheet(wb, sheet_name: str) -> list[str]:
    """Extract normalized ATC codes from a market sheet's ATC column."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    atc_column = find_atc_column(ws)
    if atc_column is None:
        return []

    codes: set[str] = set()
    for (value,) in ws.iter_rows(
        min_row=6,
        max_row=ws.max_row,
        min_col=atc_column,
        max_col=atc_column,
        values_only=True,
    ):
        code = atc_code_from_text(value)
        if code:
            codes.add(code)
    return sorted(codes)


def blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_MARKET_DEFINITION_COLUMNS}


def source_type_from_values(values: list[str | None], fallback: str) -> str:
    joined = " ".join(value or "" for value in values).lower()
    if "ubist" in joined and "iqvia" in joined:
        return "BOTH"
    if "ubist" in joined:
        return "UBIST"
    if "iqvia" in joined:
        return "IQVIA"
    return fallback


def values_for_rows(ws, rows: range, columns: tuple[int, ...]) -> list[dict[str, Any]]:
    values: list[dict[str, Any]] = []
    for row_id in rows:
        row_label = cell_text(ws.cell(row_id, 2).value) or cell_text(ws.cell(row_id, 1).value)
        for column_id in columns:
            value = ws.cell(row_id, column_id).value
            if value is None or str(value).strip() == "":
                continue
            values.append(
                {
                    "row_id": row_id,
                    "label": row_label,
                    "column_id": column_id,
                    "product_name": cell_text(ws.cell(6, column_id).value),
                    "value": to_jsonable(value),
                }
            )
    return values


def analysis_levels(ws, columns: tuple[int, ...]) -> dict[str, Any]:
    levels: dict[str, Any] = {}
    for row_id, level_name in ANALYSIS_LEVEL_ROWS.items():
        by_product: dict[str, Any] = {}
        values: list[Any] = []
        for column_id in columns:
            product_name = cell_text(ws.cell(6, column_id).value) or f"col_{column_id}"
            value = to_jsonable(ws.cell(row_id, column_id).value)
            by_product[product_name] = value
            if value not in (None, ""):
                values.append(value)
        levels[level_name] = {
            "by_product": by_product,
            "values": values,
        }
    levels["Metrics"] = values_for_rows(ws, METRIC_ROWS, columns)
    return levels


def target_customer_priority(ws, columns: tuple[int, ...]) -> list[dict[str, Any]]:
    return values_for_rows(ws, TARGET_CUSTOMER_ROWS, columns)


def raw_payload_for_market(ws, columns: tuple[int, ...]) -> dict[str, Any]:
    payload_columns = []
    for column_id in columns:
        values = []
        for row_id in range(5, 65):
            label = cell_text(ws.cell(row_id, 2).value) or cell_text(ws.cell(row_id, 1).value)
            cell_value = ws.cell(row_id, column_id).value
            if label is None and (cell_value is None or str(cell_value).strip() == ""):
                continue
            values.append(
                {
                    "row_id": row_id,
                    "label": label,
                    "value": to_jsonable(cell_value),
                }
            )
        payload_columns.append(
            {
                "column_id": column_id,
                "team": cell_text(ws.cell(5, column_id).value),
                "product_name": cell_text(ws.cell(6, column_id).value),
                "values": values,
            }
        )
    return {"source_sheet": SOURCE_SHEET, "columns": payload_columns}


def iter_market_definition_rows(xlsx_path: Path, ingested_at: str | None = None) -> Iterable[dict[str, Any]]:
    timestamp = ingested_at or utc_now_text()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            raise ValueError(f"required sheet not found: {SOURCE_SHEET!r}; sheets={wb.sheetnames}")
        ws = wb[SOURCE_SHEET]

        for strategic_market_id, columns in MARKET_DEFINITION_COLUMNS.items():
            config = MARKET_BY_ID[strategic_market_id]
            source_values = [cell_text(ws.cell(10, column).value) for column in columns]
            market_atc_codes = extract_atc_from_market_sheet(wb, config["sheet_name"])
            full_market_values = values_for_rows(ws, FULL_MARKET_ROWS, columns)
            direct_competition_values = values_for_rows(ws, DIRECT_COMPETITION_ROWS, columns)
            levels = analysis_levels(ws, columns)
            customer_priority = target_customer_priority(ws, columns)
            etc_values = levels.get("Etc", {}).get("values", [])

            record = blank_record()
            record.update(
                {
                    "strategic_market_id": strategic_market_id,
                    "market_name": config["sheet_name"],
                    "source_type": source_type_from_values(source_values, config["source_type"]),
                    "market_atc_codes_json": dumps_json(market_atc_codes),
                    "full_market_atc4_codes_json": dumps_json(full_market_values),
                    "direct_competition_brands_json": dumps_json(direct_competition_values),
                    "description": MARKET_DESCRIPTIONS.get(strategic_market_id),
                    "analysis_levels_json": dumps_json(levels),
                    "analysis_level_funnel": "O" if strategic_market_id == "strategy_001" else None,
                    "analysis_level_etc": "; ".join(str(value).strip() for value in etc_values if value),
                    "target_customer_priority_json": dumps_json(customer_priority),
                    "raw_row_json": dumps_json(raw_payload_for_market(ws, columns)),
                    "source_sheet": SOURCE_SHEET,
                    "source_file_version": xlsx_path.name,
                    "ingested_at": timestamp,
                }
            )
            yield record
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("--output-file", default=str(DEFAULT_OUTPUT_FILE))
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_file = Path(args.output_file)
    if not input_file.exists():
        sys.exit(f"ERROR: input file not found: {input_file}")

    print("=" * 72)
    print("MI Master market_definition -> Parquet")
    print("=" * 72)
    print(f"  input file:   {input_file}")
    print(f"  source sheet: {SOURCE_SHEET}")
    print(f"  output file:  {output_file}")
    print(f"  columns:      {len(MASTER_MARKET_DEFINITION_COLUMNS)} DDL columns")
    print("  helpers:      none")

    records = list(iter_market_definition_rows(input_file))
    validate_records(records)
    write_parquet(records, output_file)

    print("\nResult")
    print(f"  rows:                   {len(records)}")
    print(f"  unique strategic ids:   {len({row['strategic_market_id'] for row in records})}")
    print(f"  output size:            {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  ingested_at:            {records[0]['ingested_at'] if records else None}")
    print("\nDone")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    write_records_parquet(records, MASTER_MARKET_DEFINITION_COLUMNS, output_file, compression_level=3)


if __name__ == "__main__":
    main()
