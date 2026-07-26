"""
prototype_08_master_qa_to_parquet.py
====================================
MI Master "Q&A" sheet -> Parquet.

Phase 09b policy:
- Canonical logic: /Users/rexxa/github/jw-market/etl/master_qa.py
- Canonical schema: /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_qa
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.

Usage, in Step D after user review:
    python3 scripts/prototype_08_master_qa_to_parquet.py
"""

from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


from pipeline.etl.lib.storage import get_mi_master_path
from pipeline.etl.io.catalog._lib.common import (
    STANDARD_PREFIX,
    _extra_key,
    _header_lookup,
    _lookup_key,
    _lookup_position_value,
    _lookup_source_value,
    _position_value,
    _single_lookup_key,
    apply_column_mapping,
    build_raw_row_payload,
    cell_text,
    dumps_json,
    explicit_lookup_join,
    is_empty_row,
    load_column_metadata_catalog as load_column_metadata_catalog_common,
    make_header_keys,
    normalize_header,
    to_jsonable,
    utc_now_text,
    write_records_parquet,
)
from pipeline.etl.io.catalog._lib.expected_counts import expected_int


DEFAULT_INPUT_FILE = get_mi_master_path()
DEFAULT_OUTPUT_FILE = Path("parquet/master_qa/master_qa.parquet")
SOURCE_SHEET = "Q&A"
HEADER_ROW = 2
EXPECTED_ROW_COUNT = expected_int("master_qa.row_count")

MASTER_QA_COLUMNS = (
    "qa_id",
    "strategic_market_id",
    "question_text",
    "answer_text",
    "application_actions_json",
    "source_remark",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)

EXPECTED_QA_IDS = tuple(f"qa_{index:04d}" for index in range(1, EXPECTED_ROW_COUNT + 1))

QA_HEADER_ALIASES = {
    "번호": "row_number",
    "no": "row_number",
    "no.": "row_number",
    "id": "row_number",
    "질문 유형": "question_type",
    "질문유형": "question_type",
    "문의 유형": "question_type",
    "문의유형": "question_type",
    "question type": "question_type",
    "question_type": "question_type",
    "시장": "market_name",
    "시장명": "market_name",
    "대상 시장": "market_name",
    "market": "market_name",
    "market name": "market_name",
    "market_name": "market_name",
    "질문": "question_text",
    "문의": "question_text",
    "question": "question_text",
    "question text": "question_text",
    "question_text": "question_text",
    "답변": "answer_text",
    "answer": "answer_text",
    "answer text": "answer_text",
    "answer_text": "answer_text",
    "비고": "source_remark",
    "마케팅 비고": "source_remark",
    "remark": "source_remark",
    "source remark": "source_remark",
    "source_remark": "source_remark",
}
REQUIRED_QA_FIELDS = (
    "question_type",
    "market_name",
    "question_text",
    "answer_text",
    "source_remark",
)
QA_RAW_FIELD_ORDER = {
    "row_number": 0,
    "question_type": 1,
    "market_name": 2,
    "question_text": 3,
    "answer_text": 4,
    "source_remark": 5,
}

MARKET_NAME_ALIASES = {
    "공통": None,
    "라베칸": "strategy_001",
    "라베칸 라베칸듀오": "strategy_001",
    "라베칸/라베칸듀오": "strategy_001",
    "제이클": "strategy_002",
    "가드렛 가드메트": "strategy_003",
    "가드렛/가드메트": "strategy_003",
    "타발리스": "strategy_004",
    "시그마트": "strategy_005",
    "리바로 리바로젯": "strategy_006",
    "리바로/리바로젯": "strategy_006",
    "리바로페노": "strategy_007",
    "리바로하이": "strategy_008",
    "리바로브이": "strategy_008",
    "리바로하이 리바로브이": "strategy_008",
    "트루패스": "strategy_009",
    "피나스타/제이다트": "strategy_009",
    "트루패스 피나스타 제이다트": "strategy_009",
    "뉴트로진": "strategy_010",
    "모빌리아": "strategy_010",
    "뉴트로진 모빌리아": "strategy_010",
    "악템라": "strategy_011",
    "페린젝트": "strategy_012",
    "베노훼럼": "strategy_012",
    "페린젝트 베노훼럼": "strategy_012",
    "헴리브라": "strategy_013",
    "위너프/A+": "strategy_014",
    "위너프 위너프A+": "strategy_014",
    "엔커버": "strategy_015",
    "플라주오피": "strategy_016",
}


@dataclass
class QALoadStats:
    sheet_name: str = SOURCE_SHEET
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    staging_rows: int = 0


def market_id_for_name(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return MARKET_NAME_ALIASES.get(text)


def blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_QA_COLUMNS}


def _normalize_qa_header(value: object) -> str | None:
    if value is None:
        return None
    text = unicodedata.normalize("NFC", str(value)).strip()
    text = text.removeprefix("\ufeff").strip().casefold()
    return text or None


def _qa_field_positions(headers: list[Any]) -> dict[str, int]:
    normalized_positions: dict[str, int] = {}
    field_positions: dict[str, int] = {}
    for position, header in enumerate(headers):
        normalized = _normalize_qa_header(header)
        if normalized is None:
            continue
        if normalized in normalized_positions:
            raise ValueError(
                "normalized header collision: "
                f"{normalized!r} at columns {normalized_positions[normalized] + 1} "
                f"and {position + 1}"
            )
        normalized_positions[normalized] = position

        field = QA_HEADER_ALIASES.get(normalized)
        if field is None:
            continue
        if field in field_positions:
            previous = field_positions[field]
            raise ValueError(
                "normalized header collision: "
                f"multiple headers map to {field!r} at columns {previous + 1} "
                f"and {position + 1}"
            )
        field_positions[field] = position

    missing = [field for field in REQUIRED_QA_FIELDS if field not in field_positions]
    if missing:
        raise ValueError(f"required Q&A headers missing: {missing}")
    return field_positions


def _qa_value(values: tuple[Any, ...], positions: dict[str, int], field: str) -> Any:
    position = positions[field]
    return values[position] if position < len(values) else None


def _canonical_raw_payload(
    headers: list[Any],
    values: tuple[Any, ...],
    source_row_id: int,
) -> dict[str, Any]:
    width = max(len(headers), len(values))
    padded_headers = headers + [None] * (width - len(headers))
    padded_values = list(values) + [None] * (width - len(values))
    cells = list(zip(padded_headers, padded_values))

    def order_key(cell: tuple[Any, Any]) -> tuple[int, str]:
        normalized = _normalize_qa_header(cell[0])
        field = QA_HEADER_ALIASES.get(normalized or "")
        return QA_RAW_FIELD_ORDER.get(field or "", 100), normalized or ""

    cells.sort(key=order_key)
    ordered_headers = [header for header, _ in cells]
    ordered_values = [value for _, value in cells]
    return build_raw_row_payload(ordered_headers, ordered_values, source_row_id=source_row_id)


def load_qa_records(
    xlsx_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], QALoadStats]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            # 260518 MI Master에는 Q&A 시트가 빠졌다. 이 단계의 역할은
            # 원본 Q&A sheet를 parquet로 옮기는 것뿐이고, 수프렙/TIRZEPATIDE
            # 같은 실제 override는 별도 override/config 경로에서 검증한다.
            # 시트 부재를 fatal로 두면 run_layer0_postfix가 원본 변경 하나로
            # 멈추므로, 빈 산출로 skip한다. 더미 Q&A를 생성하는 대안은 원본에
            # 없는 데이터를 만든다는 점에서 기각했다.
            return [], QALoadStats()

        ws = wb[SOURCE_SHEET]
        headers = list(
            next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
        )
        field_positions = _qa_field_positions(headers)
        timestamp = ingested_at or utc_now_text()
        stats = QALoadStats()
        records: list[dict[str, Any]] = []
        qa_index = 0

        for source_row_id, values in enumerate(
            ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
            start=HEADER_ROW + 1,
        ):
            stats.raw_rows_scanned += 1
            if is_empty_row(values):
                stats.empty_rows += 1
                continue

            qa_index += 1
            raw_payload = _canonical_raw_payload(headers, values, source_row_id)
            question_type = _qa_value(values, field_positions, "question_type")
            market_name = _qa_value(values, field_positions, "market_name")
            question_text = _qa_value(values, field_positions, "question_text")
            answer_text = _qa_value(values, field_positions, "answer_text")
            source_remark = _qa_value(values, field_positions, "source_remark")
            actions = {
                "question_type": question_type,
                "market_name": market_name,
                "raw_answer": answer_text,
                "raw_marketing_note": source_remark,
                "auto_apply_in_phase_2": False,
                "raw_row": raw_payload,
            }

            record = blank_record()
            record.update(
                {
                    "qa_id": f"qa_{qa_index:04d}",
                    "strategic_market_id": market_id_for_name(market_name),
                    "question_text": question_text,
                    "answer_text": answer_text,
                    "application_actions_json": dumps_json(actions),
                    "source_remark": source_remark,
                    "source_sheet": SOURCE_SHEET,
                    "source_file_version": xlsx_path.name,
                    "ingested_at": timestamp,
                }
            )
            records.append(record)
            stats.staging_rows += 1

        return records, stats
    finally:
        wb.close()


def validate_records(records: list[dict[str, Any]]) -> None:
    if not records:
        return
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"master_qa row count must be {EXPECTED_ROW_COUNT}, found {len(records)}")

    qa_ids = [record["qa_id"] for record in records]
    if tuple(qa_ids) != EXPECTED_QA_IDS:
        raise ValueError(f"qa_id sequence mismatch: expected={EXPECTED_QA_IDS}, actual={tuple(qa_ids)}")
    if len(set(qa_ids)) != EXPECTED_ROW_COUNT:
        raise ValueError(f"qa_id must be unique, found duplicates in {qa_ids}")

    expected_columns = set(MASTER_QA_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )

        actions = json.loads(record["application_actions_json"])
        expected_action_keys = {
            "question_type",
            "market_name",
            "raw_answer",
            "raw_marketing_note",
            "auto_apply_in_phase_2",
            "raw_row",
        }
        if set(actions) != expected_action_keys:
            raise ValueError(
                f"row {index} application_actions_json key mismatch: "
                f"expected={sorted(expected_action_keys)}, actual={sorted(actions)}"
            )
        if actions["auto_apply_in_phase_2"] is not False:
            raise ValueError(f"row {index} auto_apply_in_phase_2 must be false")
        raw_row = actions["raw_row"]
        if not isinstance(raw_row, dict) or "source_row_id" not in raw_row:
            raise ValueError(f"row {index} raw_row payload is invalid")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("--output-file", default=str(DEFAULT_OUTPUT_FILE))
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_file = Path(args.output_file)
    if not input_file.exists():
        sys.exit(f"ERROR: input file not found: {input_file}")

    print("=" * 72)
    print("MI Master Q&A -> Parquet")
    print("=" * 72)
    print(f"  input file:   {input_file}")
    print(f"  source sheet: {SOURCE_SHEET}")
    print(f"  output file:  {output_file}")
    print(f"  columns:      {len(MASTER_QA_COLUMNS)} DDL columns")
    print("  helpers:      none")

    records, stats = load_qa_records(input_file)
    validate_records(records)
    write_parquet(records, output_file)

    print("\nResult")
    print(f"  raw rows scanned: {stats.raw_rows_scanned}")
    print(f"  empty rows:       {stats.empty_rows}")
    print(f"  staging rows:     {stats.staging_rows}")
    print(f"  unique qa_ids:    {len({row['qa_id'] for row in records})}")
    print(f"  output size:      {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  ingested_at:      {records[0]['ingested_at'] if records else None}")
    print("\nDone")


if __name__ == "__main__":
    main()

def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    write_records_parquet(records, MASTER_QA_COLUMNS, output_file, compression_level=3)
