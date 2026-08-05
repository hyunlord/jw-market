#!/usr/bin/env python3
"""Load UBIST xlsx raw sales into hive-partitioned Parquet.

The source workbook layout is fixed across the 53 UBIST workbooks:
row 1 contains metric names, row 2 contains dimensions or periods, and
data begins at row 3. The loader streams rows into one Parquet file per
period partition so the full load does not need to materialize in memory.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import duckdb
import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

from pipeline.etl.lib.ops_utils import configure_logging, find_project_root
from pipeline.etl.lib.storage import get_data_path
from pipeline.etl.io.source_headers import normalize_source_header


LOGGER = configure_logging(__name__)
ROOT = find_project_root(Path(__file__).resolve())
UBIST_ROOT = get_data_path(
    bucket_env="MINIO_BUCKET_RAW_UBIST",
    bucket_default="jw-market-raw-ubist",
    local_default=ROOT / "data" / "UBIST",
)
TARGET_DIR = ROOT / "output" / "ubist"
KST = ZoneInfo("Asia/Seoul")
UBIST_LOAD_RETENTION_MONTHS = 6 * 12

CANONICAL_DIMENSIONS = [
    "제조사",
    "국내/외자",
    "판매사",
    "판매사2",
    "제품",
    "ATC",
    "브랜드",
    "약가",
    "성분",
    "성분용량",
    "일반/전문",
    "약품코드",
    "제형",
    "투여경로",
    "급여구분",
    "종별",
    "진료과",
    "연령",
    "성별",
]

PATENT_DIMENSIONS = [
    "물질특허만료일",
    "마지막특허만료일",
    "마지막특허특성",
    "PMS만료일",
    "약품허가일",
    "Generic",
]

METRIC_MAP = {
    "처방조제액(원)": "rx_amt",
    "처방건수_P": "rx_cnt",
    "처방량_P": "rx_qty",
}

PATENT_ALIASES = {
    "물질특허 만료일": "물질특허만료일",
    "마지막특허 만료일": "마지막특허만료일",
    "마지막특허 특성": "마지막특허특성",
}

GENERIC_VALUES = {
    "Original",
    "First Generic",
    "Generic",
    "개량신약(Super Generic)",
}

METRIC_COLUMNS = list(METRIC_MAP.values())
LINEAGE_COLUMNS = ["source_file", "source_folder", "source_sheet", "source_row_no", "ingested_at"]
STATIC_METADATA_COLUMNS = PATENT_DIMENSIONS
BUSINESS_GRAIN_COLUMNS = CANONICAL_DIMENSIONS + ["period_yyyymm"]
VALUE_COLUMNS = METRIC_COLUMNS + PATENT_DIMENSIONS
BUSINESS_METRIC_COLUMNS = BUSINESS_GRAIN_COLUMNS + VALUE_COLUMNS
COLUMNS = (
    CANONICAL_DIMENSIONS
    + PATENT_DIMENSIONS
    + METRIC_COLUMNS
    + ["period_yyyymm"]
    + LINEAGE_COLUMNS
)

STRING_COLUMNS = (
    CANONICAL_DIMENSIONS
    + PATENT_DIMENSIONS
    + ["period_yyyymm", "source_file", "source_folder", "source_sheet", "ingested_at"]
)

SCHEMA = pa.schema(
    [(col, pa.string()) for col in CANONICAL_DIMENSIONS + PATENT_DIMENSIONS]
    + [(metric, pa.float64()) for metric in METRIC_MAP.values()]
    + [
        ("period_yyyymm", pa.string()),
        ("source_file", pa.string()),
        ("source_folder", pa.string()),
        ("source_sheet", pa.string()),
        ("source_row_no", pa.int64()),
        ("ingested_at", pa.string()),
    ]
)


@dataclass
class SheetMapping:
    sheet_name: str
    dim_cols: list[tuple[int, str, str]]
    duplicate_cols: list[tuple[int, str, str]]
    metric_cols: list[tuple[int, str, str, str]]


@dataclass(frozen=True)
class WorkbookPeriodSummary:
    periods: tuple[str, ...]


@dataclass
class PartitionStats:
    row_count: int = 0
    source_files: set[str] = field(default_factory=set)


@dataclass
class DedupReport:
    period: str
    rows_before: int
    rows_after: int
    duplicate_groups: int
    duplicate_rows_removed: int
    conflict_groups: int
    conflict_rows: int
    conflicts: list[dict[str, object]] = field(default_factory=list)


def now_kst() -> str:
    return datetime.now(KST).isoformat(timespec="seconds")


def summarize_source(path: Path) -> WorkbookPeriodSummary:
    """Parse workbook headers for G3 validation without making load decisions."""
    periods: set[str] = set()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            try:
                header1 = next(rows_iter)
                header2 = next(rows_iter)
            except StopIteration:
                continue
            mapping = classify_sheet(sheet_name, header1, header2)
            periods.update(period for _, _, _, period in mapping.metric_cols)
    finally:
        workbook.close()
    return WorkbookPeriodSummary(periods=tuple(sorted(periods)))


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return unicodedata.normalize("NFC", text)


def to_string_or_none(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    return unicodedata.normalize("NFC", text) if text else None


def to_number_or_none(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"non-numeric UBIST metric: {value!r}") from exc


def parse_period(value: object) -> str:
    text = normalize_text(value)
    if not text:
        raise ValueError(f"unparseable period: {value!r}")
    match = re.match(r"^(\d{4})\s*년\s*(\d{1,2})\s*월$", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    raise ValueError(f"unparseable period: {value!r}")


def canonical_header(value: object) -> str | None:
    lookup = normalize_source_header(value)
    if not lookup:
        return None
    aliases = {normalize_source_header(source): target for source, target in PATENT_ALIASES.items()}
    canonical = {
        normalize_source_header(header): header
        for header in CANONICAL_DIMENSIONS + PATENT_DIMENSIONS
    }
    header = aliases.get(lookup) or canonical.get(lookup)
    if header:
        return header
    return None


def lookup_text(value: object) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    return re.sub(r"\s+", "", text).casefold()


def generic_lookup_keys(base: dict[str, object]) -> list[str]:
    keys: list[str] = []
    code = lookup_text(base.get("약품코드"))
    product = lookup_text(base.get("제품"))
    brand = lookup_text(base.get("브랜드"))
    if code:
        keys.append(f"code:{code}")
    if product:
        keys.append(f"product:{product}")
    if brand:
        keys.append(f"brand:{brand}")
    return keys


def normalize_generic_value(value: object) -> str | None:
    text = normalize_text(value)
    if text in GENERIC_VALUES:
        return text
    return None


def resolve_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if not path.is_absolute():
        path = ROOT / path
    if path.exists():
        return path

    candidates = [
        Path(unicodedata.normalize("NFC", str(path))),
        Path(unicodedata.normalize("NFD", str(path))),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    normalized_target = unicodedata.normalize("NFC", str(path))
    for candidate in ROOT.rglob(path.name):
        if unicodedata.normalize("NFC", str(candidate)) == normalized_target:
            return candidate

    raise FileNotFoundError(f"Path not found: {raw_path}")


def discover_xlsx(args: argparse.Namespace) -> list[Path]:
    paths: list[Path] = []
    if args.all or (getattr(args, "incremental", False) and not args.folder and not args.file):
        if not UBIST_ROOT.exists():
            raise FileNotFoundError(f"Missing UBIST root: {UBIST_ROOT}")
        # 엑셀 임시 잠금파일(~$) 제외 — 원본 아님.
        paths.extend(sorted(p for p in UBIST_ROOT.rglob("*.xlsx") if not p.name.startswith("~$")))
    if args.folder:
        folder = resolve_path(args.folder)
        # 엑셀 임시 잠금파일(~$) 제외 — 원본 아님.
        paths.extend(sorted(p for p in folder.rglob("*.xlsx") if not p.name.startswith("~$")))
    if args.file:
        paths.append(resolve_path(args.file))
    unique = sorted({p.resolve() for p in paths})
    if not unique:
        raise RuntimeError("No xlsx files selected. Use --all, --folder, or --file.")
    return unique


def source_folder_for(path: Path) -> str:
    try:
        return str(path.parent.relative_to(UBIST_ROOT))
    except ValueError:
        return str(path.parent)


def normalized_source_file(path_or_name: Path | str) -> str:
    name = path_or_name.name if isinstance(path_or_name, Path) else str(path_or_name)
    return unicodedata.normalize("NFC", name)


def classify_sheet(sheet_name: str, header1: tuple[object, ...], header2: tuple[object, ...]) -> SheetMapping:
    dim_cols: list[tuple[int, str, str]] = []
    duplicate_cols: list[tuple[int, str, str]] = []
    metric_cols: list[tuple[int, str, str, str]] = []
    seen_dims: set[str] = set()
    metric_headers = {normalize_source_header(metric): metric for metric in METRIC_MAP}

    for idx in range(max(len(header1), len(header2))):
        h1 = metric_headers.get(normalize_source_header(header1[idx] if idx < len(header1) else None))
        h2 = normalize_text(header2[idx] if idx < len(header2) else None)

        if h1 in METRIC_MAP:
            try:
                period = parse_period(h2)
            except ValueError:
                continue
            metric_cols.append((idx, h1, METRIC_MAP[h1], period))
            continue

        canonical = canonical_header(h2)
        if canonical:
            if canonical in seen_dims:
                duplicate_cols.append((idx, h2 or "", canonical))
            else:
                seen_dims.add(canonical)
                dim_cols.append((idx, h2 or "", canonical))

    if not metric_cols:
        raise RuntimeError(f"No metric columns found in sheet: {sheet_name}")
    return SheetMapping(sheet_name, dim_cols, duplicate_cols, metric_cols)


def row_has_identifier(base: dict[str, object]) -> bool:
    return any(base.get(col) for col in ("약품코드", "제품", "성분", "브랜드"))


def build_generic_lookup(
    xlsx_paths: list[Path],
    *,
    parquet_root: Path | None = None,
) -> dict[str, str]:
    """Build a source-derived Generic lookup from workbooks that include it.

    The 2026.03/04 UBIST workbooks do not carry the patent block, while the
    ingredient workbooks do.  We therefore derive Generic by stable identifiers
    before the row streaming pass and backfill rows that only have the core
    dimensions.
    """

    counts: dict[str, Counter[str]] = defaultdict(Counter)
    rows_seen = 0
    if parquet_root is not None and any(parquet_root.glob("year=*/month=*/data.parquet")):
        parquet_glob = str(parquet_root / "year=*" / "month=*" / "data.parquet")
        with duckdb.connect() as connection:
            existing = connection.execute(
                """
                SELECT "약품코드", "제품", "브랜드", "Generic", count(*) AS occurrences
                FROM read_parquet(?, union_by_name=true)
                WHERE "Generic" IS NOT NULL
                  AND trim(CAST("Generic" AS VARCHAR)) <> ''
                GROUP BY ALL
                """,
                [parquet_glob],
            ).fetchall()
        for code, product, brand, generic_value, occurrences in existing:
            generic = normalize_generic_value(generic_value)
            if not generic:
                continue
            base = {"약품코드": code, "제품": product, "브랜드": brand}
            for key in generic_lookup_keys(base):
                counts[key][generic] += int(occurrences)
            rows_seen += int(occurrences)
    for xlsx_path in xlsx_paths:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows_iter = worksheet.iter_rows(values_only=True)
                try:
                    header1 = next(rows_iter)
                    header2 = next(rows_iter)
                except StopIteration:
                    continue
                mapping = classify_sheet(sheet_name, header1, header2)
                if not any(canonical == "Generic" for _, _, canonical in mapping.dim_cols):
                    continue
                for row in rows_iter:
                    base = {canonical: to_string_or_none(row[idx] if idx < len(row) else None) for idx, _, canonical in mapping.dim_cols}
                    generic = normalize_generic_value(base.get("Generic"))
                    if not generic or not row_has_identifier(base):
                        continue
                    rows_seen += 1
                    for key in generic_lookup_keys(base):
                        counts[key][generic] += 1
        finally:
            workbook.close()

    lookup = {
        key: sorted(counter.items(), key=lambda item: (-item[1], item[0]))[0][0]
        for key, counter in counts.items()
        if counter
    }
    LOGGER.info("generic lookup built rows=%s keys=%s", f"{rows_seen:,}", f"{len(lookup):,}")
    return lookup


def fill_generic_from_lookup(base: dict[str, object], generic_lookup: dict[str, str] | None) -> None:
    if normalize_generic_value(base.get("Generic")):
        return
    if not generic_lookup:
        return
    for key in generic_lookup_keys(base):
        generic = generic_lookup.get(key)
        if generic:
            base["Generic"] = generic
            return


def iter_xlsx_rows(xlsx_path: Path, generic_lookup: dict[str, str] | None = None):
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            try:
                header1 = next(rows_iter)
                header2 = next(rows_iter)
            except StopIteration:
                continue

            mapping = classify_sheet(sheet_name, header1, header2)
            loaded_at = now_kst()

            for row_no, row in enumerate(rows_iter, start=3):
                base = {canonical: to_string_or_none(row[idx] if idx < len(row) else None) for idx, _, canonical in mapping.dim_cols}
                if not row_has_identifier(base):
                    continue
                fill_generic_from_lookup(base, generic_lookup)
                base.update(
                    {
                        "source_file": xlsx_path.name,
                        "source_folder": source_folder_for(xlsx_path),
                        "source_sheet": sheet_name,
                        "source_row_no": row_no,
                        "ingested_at": loaded_at,
                    }
                )

                period_metrics: dict[str, dict[str, float | None]] = defaultdict(dict)
                for idx, _, metric, period in mapping.metric_cols:
                    period_metrics[period][metric] = to_number_or_none(row[idx] if idx < len(row) else None)

                for period, metrics in period_metrics.items():
                    output = dict(base)
                    output.update(metrics)
                    output["period_yyyymm"] = period
                    yield period, output
    finally:
        workbook.close()


def count_source_rows_by_period(xlsx_path: Path) -> dict[str, int]:
    """Count the exact rows the loader would emit for each workbook period."""
    counts: Counter[str] = Counter()
    for period, _ in iter_xlsx_rows(xlsx_path):
        counts[period] += 1
    return dict(counts)


def prepare_frame(rows: list[dict[str, object]]) -> pd.DataFrame:
    frame = pd.DataFrame(rows)
    for col in COLUMNS:
        if col not in frame.columns:
            frame[col] = None
    for col in STRING_COLUMNS:
        frame[col] = frame[col].map(to_string_or_none)
    for metric in METRIC_MAP.values():
        frame[metric] = pd.to_numeric(frame[metric], errors="coerce")
    frame["source_row_no"] = pd.to_numeric(frame["source_row_no"], errors="coerce").fillna(0).astype("int64")
    return frame.reindex(columns=COLUMNS)


def _value_tuple_key(row: pd.Series) -> tuple[object, ...]:
    return tuple(_dedup_key_value(row[column]) for column in VALUE_COLUMNS)


def _has_metadata_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return bool(str(value).strip())


def _metadata_completeness_score(frame: pd.DataFrame) -> pd.Series:
    if not STATIC_METADATA_COLUMNS:
        return pd.Series(0, index=frame.index)
    return frame[STATIC_METADATA_COLUMNS].apply(lambda column: column.map(_has_metadata_value)).sum(axis=1)


def _dedup_key_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if hasattr(value, "item"):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    if isinstance(value, str):
        text = unicodedata.normalize("NFC", value.strip())
        return text or None
    return value


def deduplicate_business_grain(frame: pd.DataFrame, period: str) -> tuple[pd.DataFrame, DedupReport]:
    """Merge one UBIST business row independently of file names.

    The latest ``ingested_at`` wins when values differ. If the latest timestamp
    contains more than one value set, no deterministic winner exists and the
    load fails before publishing the partition.
    """
    if frame.empty:
        return frame, DedupReport(period, 0, 0, 0, 0, 0, 0)

    work = frame.reindex(columns=COLUMNS).copy()
    work["_value_tuple"] = work.apply(_value_tuple_key, axis=1)
    retained: list[pd.Series] = []
    conflicts: list[dict[str, object]] = []
    duplicate_groups = 0
    duplicate_rows_removed = 0
    conflict_rows = 0

    for _, group in work.groupby(BUSINESS_GRAIN_COLUMNS, dropna=False, sort=False):
        value_counts = group["_value_tuple"].value_counts(dropna=False)
        if len(value_counts) == 1 and len(group) > 1:
            duplicate_groups += 1
            duplicate_rows_removed += len(group) - 1

        latest_at = group["ingested_at"].max()
        latest = group[group["ingested_at"] == latest_at]
        if latest["_value_tuple"].nunique(dropna=False) > 1:
            identity = {
                column: _dedup_key_value(group.iloc[0][column])
                for column in BUSINESS_GRAIN_COLUMNS
            }
            raise RuntimeError(
                "UBIST row conflict has different values at the same ingest time: "
                f"period={period} ingested_at={latest_at} identity={identity}"
            )

        winner = latest.iloc[-1]
        retained.append(winner)
        if len(value_counts) > 1:
            conflict_rows += len(group)
            discarded = group.drop(index=winner.name)
            conflicts.append(
                {
                    "identity": {
                        column: _dedup_key_value(winner[column])
                        for column in BUSINESS_GRAIN_COLUMNS
                    },
                    "winner": {
                        column: _dedup_key_value(winner[column])
                        for column in VALUE_COLUMNS + LINEAGE_COLUMNS
                    },
                    "discarded": [
                        {
                            column: _dedup_key_value(row[column])
                            for column in VALUE_COLUMNS + LINEAGE_COLUMNS
                        }
                        for _, row in discarded.iterrows()
                    ],
                }
            )

    result = pd.DataFrame(retained).drop(columns=["_value_tuple"], errors="ignore")
    result = result.sort_values(BUSINESS_GRAIN_COLUMNS + ["ingested_at"], kind="mergesort").reset_index(drop=True)
    report = DedupReport(
        period=period,
        rows_before=len(frame),
        rows_after=len(result),
        duplicate_groups=duplicate_groups,
        duplicate_rows_removed=duplicate_rows_removed,
        conflict_groups=len(conflicts),
        conflict_rows=conflict_rows,
        conflicts=conflicts,
    )
    if report.conflict_groups:
        LOGGER.warning(
            "UBIST value conflicts resolved by ingested_at period=%s groups=%s rows=%s",
            period,
            report.conflict_groups,
            report.conflict_rows,
        )
    if report.duplicate_rows_removed:
        LOGGER.info(
            "UBIST business-grain dedup period=%s duplicate_groups=%s removed=%s rows_before=%s rows_after=%s",
            period,
            report.duplicate_groups,
            report.duplicate_rows_removed,
            report.rows_before,
            report.rows_after,
        )
    return result.reindex(columns=COLUMNS), report


def deduplicate_partition_file(
    path: Path,
    period: str,
    *,
    additional_paths: tuple[Path, ...] = (),
) -> DedupReport:
    temp_path = path.with_suffix(".dedup.tmp")
    quoted_columns = ", ".join(f'"{column}"' for column in COLUMNS)
    grain_columns = ", ".join(f'"{column}"' for column in BUSINESS_GRAIN_COLUMNS)
    value_columns = ", ".join(f'"{column}"' for column in VALUE_COLUMNS)
    output_order = ", ".join(
        [f'"{column}" ASC NULLS LAST' for column in BUSINESS_GRAIN_COLUMNS]
        + ['"ingested_at" ASC NULLS LAST', "_source_ordinal ASC"]
    )

    with tempfile.TemporaryDirectory(prefix="ubist-dedup-", dir=path.parent) as work_dir_name:
        work_dir = Path(work_dir_name)
        spill_dir = work_dir / "spill"
        spill_dir.mkdir()
        connection = duckdb.connect(str(work_dir / "dedup.duckdb"))
        try:
            connection.execute("SET memory_limit='4GB'")
            connection.execute("SET threads=2")
            connection.execute("SET preserve_insertion_order=false")
            connection.execute("SET temp_directory=?", [str(spill_dir)])
            connection.execute(
                f"""
                CREATE TABLE annotated AS
                SELECT
                  {quoted_columns},
                  row_number() OVER () AS _source_ordinal,
                  count(DISTINCT row({value_columns})) OVER (
                    PARTITION BY {grain_columns}
                  ) AS _value_variants,
                  count(*) OVER (
                    PARTITION BY {grain_columns}
                  ) AS _grain_size,
                  max("ingested_at") OVER (
                    PARTITION BY {grain_columns}
                  ) AS _latest_ingested_at
                FROM read_parquet(?)
                """,
                [[str(path), *(str(item) for item in additional_paths)]],
            )
            connection.execute(
                f"""
                CREATE TABLE ranked AS
                SELECT
                  *,
                  row_number() OVER (
                    PARTITION BY {grain_columns}
                    ORDER BY "ingested_at" DESC NULLS LAST, _source_ordinal DESC
                  ) AS _winner_rank
                FROM annotated
                """
            )
            ambiguous = connection.execute(
                f"""
                SELECT {grain_columns}, _latest_ingested_at
                FROM annotated
                WHERE "ingested_at" = _latest_ingested_at
                GROUP BY {grain_columns}, _latest_ingested_at
                HAVING count(DISTINCT row({value_columns})) > 1
                ORDER BY {grain_columns}, _latest_ingested_at
                LIMIT 1
                """
            ).fetchone()
            if ambiguous is not None:
                raise RuntimeError(
                    "UBIST row conflict has different values at the same ingest time: "
                    f"period={period} identity_and_time={ambiguous}"
                )

            conflict_frame = connection.execute(
                f"""
                SELECT {quoted_columns}
                FROM annotated
                WHERE _value_variants > 1
                ORDER BY _source_ordinal
                """
            ).df()
            if conflict_frame.empty:
                conflict_details: list[dict[str, object]] = []
            else:
                _, conflict_report = deduplicate_business_grain(conflict_frame, period)
                conflict_details = conflict_report.conflicts
            (
                rows_before,
                conflict_groups,
                conflict_rows,
                duplicate_groups,
                duplicate_rows_removed,
                rows_after,
            ) = connection.execute(
                f"""
                SELECT
                  (SELECT count(*) FROM annotated),
                  (SELECT count(DISTINCT row({grain_columns})) FROM annotated WHERE _value_variants > 1),
                  (SELECT count(*) FROM annotated WHERE _value_variants > 1),
                  (SELECT count(*) FROM ranked WHERE _winner_rank = 1 AND _value_variants = 1 AND _grain_size > 1),
                  (SELECT coalesce(sum(_grain_size - 1), 0) FROM ranked
                    WHERE _winner_rank = 1 AND _value_variants = 1),
                  (SELECT count(*) FROM ranked WHERE _winner_rank = 1)
                """
            ).fetchone()
            temp_sql = str(temp_path).replace("'", "''")
            connection.execute(
                f"""
                COPY (
                  SELECT {quoted_columns}
                  FROM ranked
                  WHERE _winner_rank = 1
                  ORDER BY {output_order}
                ) TO '{temp_sql}' (FORMAT PARQUET, COMPRESSION SNAPPY)
                """
            )
        finally:
            connection.close()

    report = DedupReport(
        period=period,
        rows_before=int(rows_before),
        rows_after=int(rows_after),
        duplicate_groups=int(duplicate_groups),
        duplicate_rows_removed=int(duplicate_rows_removed),
        conflict_groups=int(conflict_groups),
        conflict_rows=int(conflict_rows),
        conflicts=conflict_details,
    )
    temp_path.replace(path)
    if report.conflict_groups:
        LOGGER.warning(
            "UBIST value conflicts resolved by ingested_at period=%s groups=%s rows=%s",
            period,
            report.conflict_groups,
            report.conflict_rows,
        )
    if report.duplicate_rows_removed:
        LOGGER.info(
            "UBIST business-grain dedup period=%s duplicate_groups=%s removed=%s rows_before=%s rows_after=%s",
            period,
            report.duplicate_groups,
            report.duplicate_rows_removed,
            report.rows_before,
            report.rows_after,
        )
    return report


class PartitionWriter:
    def __init__(self, target_root: Path):
        self.target_root = target_root
        self.writers: dict[str, pq.ParquetWriter] = {}
        self.stats: dict[str, PartitionStats] = defaultdict(PartitionStats)
        self.loaded_existing: set[str] = set()

    def _path_for(self, period: str) -> Path:
        year, month = period.split("-")
        return self.target_root / f"year={year}" / f"month={month}" / "data.parquet"

    def _open_writer(self, period: str) -> pq.ParquetWriter:
        if period in self.writers:
            return self.writers[period]
        path = self._path_for(period)
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists() and period not in self.loaded_existing:
            incoming_path = path.with_name("data.incoming.parquet")
            if incoming_path.exists():
                incoming_path.unlink()
            writer = pq.ParquetWriter(incoming_path, SCHEMA, compression="snappy")
            stats = self.stats[period]
            stats.row_count += pq.ParquetFile(path).metadata.num_rows
            with duckdb.connect() as connection:
                stats.source_files.update(
                    row[0]
                    for row in connection.execute(
                        """
                        SELECT DISTINCT source_file
                        FROM read_parquet(?)
                        WHERE source_file IS NOT NULL
                        """,
                        [str(path)],
                    ).fetchall()
                )
            self.loaded_existing.add(period)
        else:
            writer = pq.ParquetWriter(path, SCHEMA, compression="snappy")
        self.writers[period] = writer
        return writer

    def write_rows(self, period: str, rows: list[dict[str, object]]) -> None:
        if not rows:
            return
        frame = prepare_frame(rows)
        table = pa.Table.from_pandas(frame, schema=SCHEMA, preserve_index=False)
        self._open_writer(period).write_table(table)
        stats = self.stats[period]
        stats.row_count += len(rows)
        stats.source_files.update(str(row["source_file"]) for row in rows)

    def close(self) -> None:
        for writer in self.writers.values():
            writer.close()
        self.writers.clear()


def flush_buffers(writer: PartitionWriter, buffers: dict[str, list[dict[str, object]]], *, final: bool = False) -> None:
    threshold = 25_000
    for period in list(buffers):
        if final or len(buffers[period]) >= threshold:
            rows = buffers.pop(period)
            writer.write_rows(period, rows)


def deduplicate_written_partitions(target: Path, stats: dict[str, PartitionStats]) -> list[DedupReport]:
    reports: list[DedupReport] = []
    for period in sorted(stats):
        path = PartitionWriter(target)._path_for(period)
        if not path.exists():
            continue
        incoming_path = path.with_name("data.incoming.parquet")
        additional_paths = (incoming_path,) if incoming_path.exists() else ()
        LOGGER.info("UBIST partition dedup start period=%s path=%s", period, path)
        report = deduplicate_partition_file(path, period, additional_paths=additional_paths)
        for additional_path in additional_paths:
            additional_path.unlink(missing_ok=True)
        reports.append(report)
        stats[period].row_count = pq.ParquetFile(path).metadata.num_rows
        with duckdb.connect() as connection:
            stats[period].source_files = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT DISTINCT source_file
                    FROM read_parquet(?)
                    WHERE source_file IS NOT NULL
                    """,
                    [str(path)],
                ).fetchall()
            }
    total_removed = sum(report.duplicate_rows_removed for report in reports)
    total_conflicts = sum(report.conflict_groups for report in reports)
    if total_removed or total_conflicts:
        LOGGER.info(
            "UBIST business-grain dedup complete removed=%s conflict_groups=%s",
            total_removed,
            total_conflicts,
        )
    return reports


def iter_included_xlsx_rows(
    xlsx_path: Path,
    generic_lookup: dict[str, str] | None,
    exclude_periods: frozenset[str],
    *,
    _row_source=iter_xlsx_rows,
):
    """Yield (period, row) from a UBIST workbook, skipping excluded periods.

    R-1 rehearsals pin selected months to canonical parquet sidecars. Those
    months must NOT be materialized by s1: if s1 wrote them, the downstream
    ``install_ubist_sidecars`` step would collide with its no-overwrite guard.
    Excluding them here leaves the partition for the sidecar step to create.
    A period never appears mid-file, so filtering per (period, row) is exact.
    """
    for period, row in _row_source(xlsx_path, generic_lookup):
        if period in exclude_periods:
            continue
        yield period, row


def load_to_parquet(
    xlsx_paths: list[Path],
    target: Path,
    *,
    mode: str,
    truncate: bool,
    previous_manifest: dict[str, object] | None = None,
    exclude_periods: frozenset[str] = frozenset(),
) -> dict[str, PartitionStats]:
    timestamp = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    tmp_target = target.parent / f"{target.name}.__tmp_{timestamp}"
    backup_target = target.parent / f"{target.name}.__backup_{timestamp}"

    if tmp_target.exists():
        shutil.rmtree(tmp_target)
    tmp_target.mkdir(parents=True)

    if mode == "append" and target.exists():
        shutil.copytree(target, tmp_target, dirs_exist_ok=True)
    elif mode != "replace":
        raise ValueError(f"Unsupported mode: {mode}")

    if exclude_periods:
        LOGGER.info("UBIST load excluding pinned sidecar periods=%s", sorted(exclude_periods))

    buffers: dict[str, list[dict[str, object]]] = defaultdict(list)
    writer = PartitionWriter(tmp_target)
    total_rows = 0
    generic_lookup = build_generic_lookup(xlsx_paths, parquet_root=tmp_target)
    try:
        for idx, xlsx_path in enumerate(xlsx_paths, start=1):
            LOGGER.info("[%s/%s] reading %s", idx, len(xlsx_paths), xlsx_path)
            for period, row in iter_included_xlsx_rows(xlsx_path, generic_lookup, exclude_periods):
                buffers[period].append(row)
                total_rows += 1
                if total_rows % 250_000 == 0:
                    LOGGER.info("rows=%s active_partitions=%s", f"{total_rows:,}", len(buffers))
                    flush_buffers(writer, buffers)
            flush_buffers(writer, buffers)
        flush_buffers(writer, buffers, final=True)
    finally:
        writer.close()

    leaked = exclude_periods & set(writer.stats)
    if leaked:
        # Fail closed: an excluded period must never be materialized by s1.
        raise RuntimeError(f"excluded UBIST periods leaked into load: {sorted(leaked)}")

    LOGGER.info(
        "UBIST stream load complete rows=%s partitions=%s; starting partition dedup",
        f"{total_rows:,}",
        len(writer.stats),
    )
    deduplicate_written_partitions(tmp_target, writer.stats)
    removed_periods = prune_ubist_partitions(tmp_target)
    if removed_periods:
        LOGGER.info(
            "UBIST load retention removed oldest partitions count=%s periods=%s",
            len(removed_periods),
            list(removed_periods),
        )
    write_manifest(tmp_target, writer.stats, xlsx_paths, previous_manifest=previous_manifest)

    if target.exists():
        if truncate or mode == "replace":
            target.rename(backup_target)
        else:
            raise RuntimeError(f"Target already exists: {target}")
    tmp_target.rename(target)
    if backup_target.exists():
        shutil.rmtree(backup_target)
    LOGGER.info("loaded rows=%s partitions=%s target=%s", f"{total_rows:,}", len(writer.stats), target)
    return writer.stats


def partition_path(period: str) -> str:
    year, month = period.split("-")
    return f"year={year}/month={month}/data.parquet"


def prune_ubist_partitions(
    target: Path,
    *,
    retention_months: int = UBIST_LOAD_RETENTION_MONTHS,
) -> tuple[str, ...]:
    """Keep the latest configured load partitions in a candidate tree."""
    partitions: list[tuple[str, Path]] = []
    for path in target.glob("year=*/month=*/data.parquet"):
        year = path.parent.parent.name.removeprefix("year=")
        month = path.parent.name.removeprefix("month=")
        if len(year) == 4 and len(month) == 2 and year.isdigit() and month.isdigit():
            partitions.append((f"{year}-{month}", path.parent))
    partitions.sort(key=lambda item: item[0])
    remove_count = max(0, len(partitions) - retention_months)
    removed = tuple(period for period, _ in partitions[:remove_count])
    for _, month_dir in partitions[:remove_count]:
        shutil.rmtree(month_dir)
        year_dir = month_dir.parent
        if not any(year_dir.iterdir()):
            year_dir.rmdir()
    return removed


def write_manifest(
    target: Path,
    stats: dict[str, PartitionStats],
    xlsx_paths: list[Path],
    *,
    previous_manifest: dict[str, object] | None = None,
) -> None:
    generated_at = now_kst()
    previous_partitions = {}
    if previous_manifest:
        for entry in previous_manifest.get("partitions", []):
            if isinstance(entry, dict) and entry.get("period_yyyymm"):
                period = str(entry["period_yyyymm"])
                if (target / partition_path(period)).exists():
                    previous_partitions[period] = dict(entry)
    partition_entries = previous_partitions
    for period in sorted(stats):
        partition_entries[period] = {
            "period_yyyymm": period,
            "path": partition_path(period),
            "row_count": stats[period].row_count,
            "source_files": sorted(stats[period].source_files),
            "loaded_at": generated_at,
        }
    source_file_count = len(
        {
            normalized_source_file(source)
            for entry in partition_entries.values()
            for source in entry.get("source_files", [])
        }
    )
    if not previous_manifest:
        source_file_count = len(xlsx_paths)
    manifest = {
        "schema_version": "1.0",
        "generated_at": generated_at,
        "storage": "parquet_hive_partition",
        "compression": "snappy",
        "metric_map": METRIC_MAP,
        "canonical_dimensions": CANONICAL_DIMENSIONS,
        "patent_dimensions": PATENT_DIMENSIONS,
        "dedup_identity": {
            "fact_grain_columns": BUSINESS_GRAIN_COLUMNS,
            "metric_columns": METRIC_COLUMNS,
            "patent_value_columns": PATENT_DIMENSIONS,
            "excluded_lineage_columns": LINEAGE_COLUMNS,
            "winner_policy": "latest ingested_at; fail when latest timestamp has different values",
            "filename_policy": "record lineage only; never select or skip rows",
            "conflict_policy": "record identity, winner, and discarded values",
        },
        "dimension_duplicate_policy": {
            "decision": "collapse_duplicate_semantic_headers",
            "kept": "first occurrence in the fixed dimension block",
            "dropped_after_sample_match": ["제조사", "국내/외자", "판매사", "ATC", "성분"],
            "retained_distinct": ["판매사2"],
        },
        "source_file_count": source_file_count,
        "partitions": [partition_entries[period] for period in sorted(partition_entries)],
    }
    target.mkdir(parents=True, exist_ok=True)
    (target / "_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def dry_run(xlsx_paths: list[Path], limit_rows: int = 3) -> None:
    print("# UBIST Parquet Loader Dry Run\n")
    for xlsx_path in xlsx_paths:
        print(f"## Source: {xlsx_path}\n")
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            header1 = next(rows_iter)
            header2 = next(rows_iter)
            mapping = classify_sheet(sheet_name, header1, header2)
            periods = sorted({period for _, _, _, period in mapping.metric_cols})
            metrics = sorted({metric for _, _, metric, _ in mapping.metric_cols})
            print(f"### Sheet: {sheet_name}")
            print(f"- rows: {worksheet.max_row:,}")
            print(f"- cols: {worksheet.max_column:,}")
            print(f"- kept dimensions: {len(mapping.dim_cols)}")
            for idx, raw, canonical in mapping.dim_cols:
                print(f"  - col {idx + 1}: {raw} -> {canonical}")
            print(f"- duplicate dimensions dropped: {len(mapping.duplicate_cols)}")
            for idx, raw, canonical in mapping.duplicate_cols:
                print(f"  - col {idx + 1}: {raw} -> {canonical}")
            print(f"- metrics: {metrics}")
            print(f"- periods: {periods[0]} .. {periods[-1]} ({len(periods)})")
            print("\n#### Sample Output Rows")
            samples = []
            for _, row in iter_xlsx_rows(xlsx_path):
                samples.append(row)
                if len(samples) >= limit_rows:
                    break
            frame = prepare_frame(samples)
            print(frame.head(limit_rows).to_string(index=False))
            print()


def read_manifest(target: Path) -> dict[str, object]:
    manifest_path = target / "_manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Missing UBIST manifest: {manifest_path}")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def run_incremental_ubist_load(
    *,
    target: Path,
    paths: list[Path] | None = None,
    folder: Path | None = None,
    file: Path | None = None,
    all_sources: bool = True,
    dry: bool = False,
) -> dict[str, PartitionStats]:
    args = argparse.Namespace(
        all=all_sources,
        folder=str(folder) if folder is not None else None,
        file=str(file) if file is not None else None,
        dry_run=dry,
        truncate=False,
        mode="append",
        target_dir=str(target),
        incremental=True,
    )
    xlsx_paths = [p.resolve() for p in paths] if paths is not None else discover_xlsx(args)
    if not xlsx_paths:
        raise RuntimeError("No xlsx files selected for UBIST incremental load")
    manifest = read_manifest(target)
    if dry:
        print("# UBIST Incremental Row Merge\n")
        print(f"- selected workbooks: {len(xlsx_paths)}")
        for path in xlsx_paths:
            print(f"- {path}")
        return {}
    return load_to_parquet(
        xlsx_paths,
        target,
        mode="append",
        truncate=True,
        previous_manifest=manifest,
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=False)
    source.add_argument("--folder", help="Folder containing UBIST xlsx files")
    source.add_argument("--file", help="Single UBIST xlsx file")
    source.add_argument("--all", action="store_true", help="Load all xlsx files below data/UBIST")
    parser.add_argument("--incremental", action="store_true", help="Merge all selected workbook rows into existing partitions")
    parser.add_argument("--dry-run", action="store_true", help="Analyze schema and sample rows without writing")
    parser.add_argument("--truncate", action="store_true", help="Replace the existing output/ubist target")
    parser.add_argument("--mode", choices=["replace", "append"], default="replace")
    parser.add_argument("--target-dir", default=str(TARGET_DIR))
    args = parser.parse_args(argv)
    if not args.incremental and not (args.folder or args.file or args.all):
        parser.error("one of --folder, --file, or --all is required unless --incremental is used")
    return args


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    try:
        xlsx_paths = discover_xlsx(args)
        if args.incremental:
            stats = run_incremental_ubist_load(
                target=Path(args.target_dir),
                paths=xlsx_paths,
                dry=args.dry_run,
            )
            if not args.dry_run:
                LOGGER.info("partition summary")
                for period in sorted(stats):
                    LOGGER.info("%s: %s", period, f"{stats[period].row_count:,}")
            return 0
        if args.dry_run:
            dry_run(xlsx_paths)
            return 0
        stats = load_to_parquet(xlsx_paths, Path(args.target_dir), mode=args.mode, truncate=args.truncate)
        LOGGER.info("partition summary")
        for period in sorted(stats):
            LOGGER.info("%s: %s", period, f"{stats[period].row_count:,}")
    except Exception as exc:
        LOGGER.error("ERROR: %s", exc)
        return 1
    return 0


def run_ubist_load(
    *,
    target: Path,
    mode: str = "replace",
    truncate: bool = True,
    paths: list[Path] | None = None,
    folder: Path | None = None,
    file: Path | None = None,
    all_sources: bool = True,
    exclude_periods: frozenset[str] = frozenset(),
) -> dict[str, PartitionStats]:
    args = argparse.Namespace(
        all=all_sources,
        folder=str(folder) if folder is not None else None,
        file=str(file) if file is not None else None,
        dry_run=False,
        truncate=truncate,
        mode=mode,
        target_dir=str(target),
    )
    xlsx_paths = [p.resolve() for p in paths] if paths is not None else discover_xlsx(args)
    return load_to_parquet(
        xlsx_paths, target, mode=mode, truncate=truncate, exclude_periods=exclude_periods
    )


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
