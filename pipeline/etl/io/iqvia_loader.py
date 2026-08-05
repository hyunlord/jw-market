#!/usr/bin/env python3
"""Load IQVIA NSA raw sources into MariaDB Layer 1 tables."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Iterator

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import pymysql

from pipeline.etl.io.iqvia_scope import (
    iqvia_record_in_scope,
    normalize_iqvia_atc4_codes,
    normalize_iqvia_quarters,
)
from pipeline.etl.io.source_headers import normalize_source_header
from pipeline.etl.lib.ops_utils import configure_logging, find_project_root, first_existing, retry
from pipeline.etl.lib.storage import get_data_path


LOGGER = configure_logging(__name__)
REPO_ROOT = find_project_root(Path(__file__).resolve())
IQVIA_ROOT = get_data_path(
    bucket_env="MINIO_BUCKET_RAW_IQVIA",
    bucket_default="jw-market-raw-iqvia",
    local_default=REPO_ROOT / "data" / "IQVIA",
)
AUDIT_DIR = REPO_ROOT / "audit" / "phase16c3_iqvia_mariadb"

NSA_TABLE = "iqvia_nsa_quarterly_raw"
DEFAULT_NSA_PARQUET_DIR = REPO_ROOT / "output" / "iqvia_nsa"
DEFAULT_RECORD_PARQUET_DIR = REPO_ROOT / "output" / "iqvia"
MONTH_NAME_TO_NUM = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

NSA_PARQUET_META_COLUMNS = (
    "audit_code",
    "audit_desc",
    "product_name",
    "product_name_kor",
    "pack_desc",
    "otc_ethical",
    "mfr_code",
    "mfr_name",
    "mfr_name_kor",
    "mft_type",
    "mfr_type_group",
    "atc1_code",
    "atc1_desc",
    "atc2_code",
    "atc2_desc",
    "atc3_code",
    "atc3_desc",
    "atc4_code",
    "atc4_desc",
    "nfc1_code",
    "nfc1_desc",
    "nfc2_code",
    "nfc2_desc",
    "nfc3_code",
    "nfc3_desc",
    "strength",
    "molecule_desc",
    "molecule_type",
    "nhi_type",
    "pack_launchdate",
    "product_launch_date",
    "herbal",
    "product_age",
    "pack_size",
    "pack_age",
)

NSA_PARQUET_METRICS = {
    "Values LC": "values_lc",
    "Units": "units",
    "Counting Units": "counting_units",
    "Dosage Units": "dosage_units",
    "Price": "price",
}

NSA_REQUIRED_STATIC_HEADERS = (
    "AUDIT CODE",
    "MFR CODE",
    "PRODUCT NAME",
    "PACK DESC",
)
NSA_PERIOD_HEADER = "DATA PERIOD"
NSA_OPTIONAL_STATIC_HEADERS = (
    "AUDIT DESC",
    "MFR NAME",
)
NSA_STATIC_HEADER_NAMES = {
    "AUDIT CODE",
    "AUDIT DESC",
    "PRODUCT NAME",
    "PRODUCT NAME KOR",
    "PACK DESC",
    "OTC/ETHICAL",
    "MFR CODE",
    "MFR NAME",
    "MFR NAME KOR",
    "MFR TYPE",
    "MFT TYPE",
    "MFR TYPE GROUP",
    "ATC 1",
    "ATC 1 CODE",
    "ATC 1 DESC",
    "ATC 2",
    "ATC 2 CODE",
    "ATC 2 DESC",
    "ATC 3",
    "ATC 3 CODE",
    "ATC 3 DESC",
    "ATC 4",
    "ATC 4 CODE",
    "ATC 4 DESC",
    "NFC 1",
    "NFC 1 CODE",
    "NFC 1 DESC",
    "NFC 2",
    "NFC 2 CODE",
    "NFC 2 DESC",
    "NFC 3",
    "NFC 3 CODE",
    "NFC 3 DESC",
    "STRENGTH",
    "MOLECULE DESC",
    "MOLECULE TYPE",
    "NHI TYPE",
    "PACK LAUNCH DATE",
    "PACK LAUNCHDATE",
    "PRODUCT LAUNCH DATE",
    "HERBAL",
    "PRODUCT AGE",
    "PACK SIZE",
    "PACK AGE",
}
NSA_HEADER_ALIASES = {
    "PACK DESCRIPTION": "PACK DESC",
}


class HeaderContractError(ValueError):
    """Raised before loading when an NSA source cannot satisfy its row contract."""


@dataclass(frozen=True)
class NsaDedupReport:
    period: str
    rows_before: int
    rows_after: int
    duplicate_groups: int
    duplicate_rows_removed: int
    conflict_groups: int
    conflict_rows: int


INSERT_SQL = {
    NSA_TABLE: f"""
        INSERT INTO {NSA_TABLE}
        (source_file, sheet_name, source_row_no, audit_code, audit_desc, mfr_code,
         mfr_name, period_yyyy, period_quarter, period_label, payload, source_master_version)
        VALUES
        (%(source_file)s, %(sheet_name)s, %(source_row_no)s, %(audit_code)s, %(audit_desc)s,
         %(mfr_code)s, %(mfr_name)s, %(period_yyyy)s, %(period_quarter)s,
         %(period_label)s, %(payload)s, %(source_master_version)s)
    """,
}


@dataclass(frozen=True)
class SourceSheet:
    path: Path
    sheet_name: str | None = None


@dataclass
class LoadStats:
    rows: int = 0
    files: int = 0
    sheets: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        # Stage/컨테이너(R-1 rehearsal, k8s) 환경에는 pipeline/docker/.env가 없고
        # MARIADB_* 환경변수만 secret으로 주입된다. general_config.load_env와
        # 동일하게 이때 env fallback을 허용한다(파일이 있으면 종전 파싱 유지).
        return {key: value for key, value in os.environ.items() if key.startswith("MARIADB_") or key == "HOST_PORT"}
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    # .env를 읽은 뒤에도 shell env가 있으면 그 값을 우선한다(로컬↔staging 동일 경로).
    for key in ("MARIADB_HOST", "MARIADB_PORT", "MARIADB_DATABASE", "MARIADB_USER", "MARIADB_PASSWORD", "HOST_PORT"):
        if os.environ.get(key):
            env[key] = os.environ[key]
    return env


@retry((pymysql.err.OperationalError, pymysql.err.InterfaceError), logger=LOGGER)
def connect(database: str | None = None) -> pymysql.connections.Connection:
    env_path = first_existing(REPO_ROOT / "pipeline" / "docker" / ".env", REPO_ROOT / "docker" / ".env")
    # .env 부재를 raise하지 않는다: load_env가 os.environ fallback을 반환하고,
    # 아래 user/password/host/database는 이미 os.getenv를 우선 사용한다.
    env = load_env(env_path)
    user = os.getenv("MARIADB_USER", env.get("MARIADB_USER", "jwapp"))
    password = os.getenv("MARIADB_PASSWORD") or (
        os.getenv("MARIADB_ROOT_PASSWORD") if user == "root" else env.get("MARIADB_PASSWORD")
    )
    if not password:
        raise RuntimeError("MariaDB password is not configured")
    return pymysql.connect(
        host=os.getenv("MARIADB_HOST", "127.0.0.1"),
        # k8s/R-1은 MARIADB_PORT(=3306)만 주입한다. general_config와 동일하게
        # MARIADB_PORT를 우선하고, 로컬 .env의 HOST_PORT(3307)는 fallback으로 둔다.
        port=int(
            os.getenv("MARIADB_PORT")
            or os.getenv("HOST_PORT")
            or env.get("MARIADB_PORT")
            or env.get("HOST_PORT", "3307")
        ),
        user=user,
        password=password,
        database=database or os.getenv("MARIADB_DATABASE", env.get("MARIADB_DATABASE", "jw_mart")),
        charset="utf8mb4",
        autocommit=False,
    )


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if pd.isna(value) if not isinstance(value, (list, dict, tuple, set)) else False:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def parse_nsa_metric(value: Any, metric: str) -> float | int | None:
    """Parse source numeric values without silently retaining invalid text."""
    cleaned = clean_value(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, bool):
        raise ValueError(f"non-numeric IQVIA NSA metric {metric}: {value!r}")
    if isinstance(cleaned, (int, float, Decimal)):
        return float(cleaned) if isinstance(cleaned, Decimal) else cleaned
    text = str(cleaned).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"non-numeric IQVIA NSA metric {metric}: {value!r}") from exc


def dedupe_keys(keys: list[str]) -> list[str]:
    counts: dict[str, int] = defaultdict(int)
    result: list[str] = []
    for key in keys:
        counts[key] += 1
        if counts[key] == 1:
            result.append(key)
        else:
            result.append(f"{key}_{counts[key]}")
    return result


def _display_source_header(value: object, idx: int) -> str:
    if value is None:
        return f"__blank_{idx}"
    text = unicodedata.normalize("NFKC", str(value))
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or f"__blank_{idx}"


def _nsa_header_maps() -> tuple[dict[str, str], dict[str, str]]:
    static = {
        normalize_source_header(header): header
        for header in NSA_STATIC_HEADER_NAMES | {NSA_PERIOD_HEADER}
    }
    static.update(
        {
            normalize_source_header(alias): canonical
            for alias, canonical in NSA_HEADER_ALIASES.items()
        }
    )
    metrics = {
        normalize_source_header(metric): metric
        for metric in NSA_PARQUET_METRICS
    }
    return static, metrics


def canonicalize_nsa_headers(
    headers_raw: Iterable[object], *, source: str
) -> list[str]:
    """Canonicalize and validate NSA headers before any source row is emitted."""
    static_map, metric_map = _nsa_header_maps()
    headers: list[str] = []
    unexpected: list[str] = []
    for idx, value in enumerate(headers_raw):
        display = _display_source_header(value, idx)
        lookup = normalize_source_header(value)
        canonical = static_map.get(lookup) if lookup else None
        if canonical is None and lookup:
            canonical = metric_map.get(lookup)
        if canonical is None and lookup:
            period_match = re.match(r"^(\d{1,2})/(\d{4})_(.+)$", display)
            if period_match:
                metric = metric_map.get(normalize_source_header(period_match.group(3)))
                if metric:
                    canonical = f"{int(period_match.group(1)):02d}/{period_match.group(2)}_{metric}"
        if canonical is None:
            canonical = display
            if not canonical.startswith("__blank_"):
                unexpected.append(canonical)
        headers.append(canonical)

    headers = dedupe_keys(headers)
    present = set(headers)
    missing_static = [header for header in NSA_REQUIRED_STATIC_HEADERS if header not in present]
    has_long_period = NSA_PERIOD_HEADER in present
    periods = nsa_period_columns(headers)
    if not has_long_period and not periods:
        missing_static.append(NSA_PERIOD_HEADER)
    if has_long_period:
        missing_static.extend(
            f"metric header {metric}"
            for metric in NSA_PARQUET_METRICS
            if metric not in present
        )
    else:
        for period, metric_columns in sorted(periods.items()):
            missing_static.extend(
                f"metric header {period}_{metric}"
                for metric in NSA_PARQUET_METRICS
                if metric not in metric_columns
            )
    if missing_static:
        raise HeaderContractError(
            f"IQVIA NSA required header missing source={source}: {', '.join(missing_static)}"
        )

    missing_optional = [header for header in NSA_OPTIONAL_STATIC_HEADERS if header not in present]
    if missing_optional:
        LOGGER.warning(
            "IQVIA NSA optional headers missing source=%s headers=%s",
            source,
            missing_optional,
        )
    if unexpected:
        LOGGER.warning(
            "IQVIA NSA unexpected headers retained as passthrough source=%s headers=%s",
            source,
            unexpected,
        )
    return headers


def row_to_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {k: clean_value(v) for k, v in row.items()}


def dumps_payload(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str)


def compute_product_key(audit_code: Any, product_name: Any, pack_desc: Any) -> str:
    text = f"{audit_code or ''}|{product_name or ''}|{pack_desc or ''}"
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:32]


def period_label_to_quarter(period_label: Any) -> str:
    text = str(period_label or "").strip()
    match = re.match(r"^(\d{4})-?Q([1-4])$", text, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"invalid IQVIA NSA period_label: {period_label!r}")
    return f"{match.group(1)}-Q{match.group(2)}"


def payload_lookup(payload: dict[str, Any], key: str) -> Any:
    normalized_key = key.lower()
    for source_key, value in payload.items():
        if str(source_key).lower() == normalized_key:
            return value
    return None


def nsa_record_to_parquet_row(record: dict[str, Any]) -> dict[str, Any]:
    payload = json.loads(str(record["payload"]))
    static = payload.get("static", {})
    period_values = payload.get("period_values", {})

    product_name = payload_lookup(static, "PRODUCT NAME")
    product_name_kor = payload_lookup(static, "PRODUCT NAME KOR")
    pack_desc = payload_lookup(static, "PACK DESC")
    audit_code = record.get("audit_code")
    row: dict[str, Any] = {
        "product_key": compute_product_key(audit_code, product_name, pack_desc),
        "source_file": record.get("source_file"),
        "sheet_name": record.get("sheet_name"),
        "source_row_no": record.get("source_row_no"),
        "period_label": period_label_to_quarter(record.get("period_label")),
    }

    explicit_values = {
        "audit_code": audit_code,
        "audit_desc": record.get("audit_desc"),
        "mfr_code": record.get("mfr_code"),
        "mfr_name": record.get("mfr_name"),
        "product_name": product_name,
        "product_name_kor": product_name_kor,
        "pack_desc": pack_desc,
    }
    static_keys = {
        "otc_ethical": "OTC/ETHICAL",
        "mfr_name_kor": "MFR NAME KOR",
        "mft_type": "MFT TYPE",
        "mfr_type_group": "MFR TYPE GROUP",
        "atc1_code": "ATC 1 CODE",
        "atc1_desc": "ATC 1 DESC",
        "atc2_code": "ATC 2 CODE",
        "atc2_desc": "ATC 2 DESC",
        "atc3_code": "ATC 3 CODE",
        "atc3_desc": "ATC 3 DESC",
        "atc4_code": "ATC 4 CODE",
        "atc4_desc": "ATC 4 DESC",
        "nfc1_code": "NFC 1 CODE",
        "nfc1_desc": "NFC 1 DESC",
        "nfc2_code": "NFC 2 CODE",
        "nfc2_desc": "NFC 2 DESC",
        "nfc3_code": "NFC 3 CODE",
        "nfc3_desc": "NFC 3 DESC",
        "strength": "STRENGTH",
        "molecule_desc": "MOLECULE DESC",
        "molecule_type": "MOLECULE TYPE",
        "nhi_type": "NHI TYPE",
        "pack_launchdate": "PACK LAUNCHDATE",
        "product_launch_date": "PRODUCT LAUNCH DATE",
        "herbal": "HERBAL",
        "product_age": "PRODUCT AGE",
        "pack_size": "PACK SIZE",
        "pack_age": "PACK AGE",
    }
    for column in NSA_PARQUET_META_COLUMNS:
        if column in explicit_values:
            row[column] = clean_value(explicit_values[column])
        else:
            row[column] = clean_value(payload_lookup(static, static_keys[column]))
    for metric, column in NSA_PARQUET_METRICS.items():
        row[column] = clean_value(period_values.get(metric))
    row["sources"] = str(record.get("source_file") or "")
    row["ingested_at"] = datetime.now(timezone.utc).isoformat()
    return row


def is_iqvia_source_file(path: Path, suffixes: set[str]) -> bool:
    """Return true for real IQVIA source files, excluding Office/macOS temp files."""
    name = path.name
    if name.startswith(("~$", "._")):
        return False
    return path.suffix.lower() in suffixes


def discover_files() -> list[Path]:
    root = IQVIA_ROOT / "NSA"
    if not root.exists():
        raise FileNotFoundError(f"Missing IQVIA NSA directory: {root}")
    return sorted(
        p
        for p in root.iterdir()
        if is_iqvia_source_file(p, {".csv", ".xlsx", ".xls"})
    )


def quarter_from_month(month: int) -> int | None:
    return {3: 1, 6: 2, 9: 3, 12: 4}.get(month)


def parse_yyyymm(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    m = re.search(r"(\d{4})[./-](\d{1,2})", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

    m = re.search(r"(\d{1,2})/(\d{4})", text)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"

    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

    for m in re.finditer(r"(?:^|[^A-Za-z])([A-Za-z]+)\.?\s*'?(\d{2,4})(?=$|[^0-9])", text):
        month = MONTH_NAME_TO_NUM.get(m.group(1).lower())
        if not month:
            continue
        if month:
            year = int(m.group(2))
            if year < 100:
                year += 2000
            return f"{year:04d}-{month:02d}"

    return None


def parse_period_from_filename(path: Path) -> str | None:
    return parse_yyyymm(path.stem)


def nsa_period_columns(headers: Iterable[str]) -> dict[str, dict[str, str]]:
    periods: dict[str, dict[str, str]] = defaultdict(dict)
    for header in headers:
        match = re.match(r"^(\d{1,2})/(\d{4})_(.+)$", str(header).strip())
        if not match:
            continue
        month = int(match.group(1))
        if quarter_from_month(month) is None:
            continue
        period = f"{int(match.group(2)):04d}-{month:02d}"
        periods[period][match.group(3)] = header
    return dict(periods)


def normalize_nsa_audit_desc(row: dict[str, Any]) -> Any:
    return row.get("AUDIT DESC") or row.get("AUDIT CODE")


def long_format_period_record(path: Path, sheet_name: str, source_row_no: int, raw: dict[str, Any], static_cols: list[str]) -> dict[str, Any] | None:
    period = parse_yyyymm(raw.get("DATA PERIOD"))
    if not period:
        return None
    year, month = period.split("-")
    quarter = quarter_from_month(int(month))
    if quarter is None:
        return None
    period_values = {metric: parse_nsa_metric(raw.get(metric), metric) for metric in NSA_PARQUET_METRICS}
    if all(value is None for value in period_values.values()):
        return None
    static = {key: raw.get(key) for key in static_cols}
    return {
        "source_file": path.name,
        "sheet_name": sheet_name,
        "source_row_no": source_row_no,
        "audit_code": clean_value(raw.get("AUDIT CODE")),
        "audit_desc": clean_value(normalize_nsa_audit_desc(raw)),
        "mfr_code": clean_value(raw.get("MFR CODE")),
        "mfr_name": clean_value(raw.get("MFR NAME")),
        "period_yyyy": int(year),
        "period_quarter": quarter,
        "period_label": f"{year}Q{quarter}",
        "payload": dumps_payload(
            {
                "__source": "NSA",
                "__raw_period": period,
                "__period_metric_columns": {metric: metric for metric in NSA_PARQUET_METRICS},
                "static": static,
                "period_values": period_values,
            }
        ),
        "source_master_version": None,
    }


def _nsa_key_value(value: Any) -> Any:
    value = clean_value(value)
    if isinstance(value, str):
        text = unicodedata.normalize("NFC", value.strip())
        return text or None
    return value


def _nsa_natural_key(record: dict[str, Any]) -> tuple[Any, ...] | None:
    payload = json.loads(str(record["payload"]))
    static = payload.get("static", {})
    key = (
        _nsa_key_value(record.get("audit_code")),
        _nsa_key_value(record.get("mfr_code")),
        _nsa_key_value(payload_lookup(static, "PRODUCT NAME")),
        _nsa_key_value(payload_lookup(static, "PACK DESC")),
        _nsa_key_value(record.get("period_label")),
    )
    return None if any(value is None for value in key) else key


def _nsa_semantic_signature(record: dict[str, Any]) -> str:
    payload = json.loads(str(record["payload"]))
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def _nsa_identity(record: dict[str, Any]) -> tuple[tuple[Any, ...] | None, str]:
    payload = json.loads(str(record["payload"]))
    static = payload.get("static", {})
    key = (
        _nsa_key_value(record.get("audit_code")),
        _nsa_key_value(record.get("mfr_code")),
        _nsa_key_value(payload_lookup(static, "PRODUCT NAME")),
        _nsa_key_value(payload_lookup(static, "PACK DESC")),
        _nsa_key_value(record.get("period_label")),
    )
    natural_key = None if any(value is None for value in key) else key
    canonical_payload = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return natural_key, hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()


def _nsa_source_recency(record: dict[str, Any], input_index: int) -> tuple[int, int, int, int, int]:
    source_file = str(record.get("source_file") or "")
    match = re.search(r"([A-Za-z]+)-(\d{1,2})-(\d{4})", source_file)
    if match:
        month = MONTH_NAME_TO_NUM.get(match.group(1).casefold(), 0)
        source_date = (int(match.group(3)), month, int(match.group(2)))
    else:
        source_date = (0, 0, 0)
    return (*source_date, int(record.get("source_row_no") or 0), input_index)


def deduplicate_nsa_records(
    records: Iterable[dict[str, Any]], period: str
) -> tuple[list[dict[str, Any]], NsaDedupReport]:
    """Collapse exact NSA facts while preserving every natural-key conflict.

    Exact facts keep the last occurrence, so callers that pass canonical files
    oldest-to-newest retain the latest-file lineage. A natural key with more
    than one payload signature is never collapsed.
    """
    source = list(records)
    groups: dict[tuple[Any, ...], list[tuple[int, dict[str, Any], str]]] = defaultdict(list)
    passthrough: list[tuple[int, dict[str, Any]]] = []
    for index, record in enumerate(source):
        natural_key = _nsa_natural_key(record)
        if natural_key is None:
            passthrough.append((index, record))
            continue
        groups[natural_key].append((index, record, _nsa_semantic_signature(record)))

    kept = list(passthrough)
    duplicate_groups = 0
    duplicate_rows_removed = 0
    conflict_groups = 0
    conflict_rows = 0
    for group in groups.values():
        signatures = {signature for _, _, signature in group}
        if len(signatures) > 1:
            conflict_groups += 1
            conflict_rows += len(group)
            kept.extend((index, record) for index, record, _ in group)
            continue
        if len(group) > 1:
            duplicate_groups += 1
            duplicate_rows_removed += len(group) - 1
        index, record, _ = max(
            group,
            key=lambda item: _nsa_source_recency(item[1], item[0]),
        )
        kept.append((index, record))

    result = [record for _, record in sorted(kept, key=lambda item: item[0])]
    report = NsaDedupReport(
        period=period,
        rows_before=len(source),
        rows_after=len(result),
        duplicate_groups=duplicate_groups,
        duplicate_rows_removed=duplicate_rows_removed,
        conflict_groups=conflict_groups,
        conflict_rows=conflict_rows,
    )
    if report.conflict_groups:
        LOGGER.warning(
            "IQVIA NSA natural-key conflicts preserved period=%s groups=%s rows=%s",
            period,
            report.conflict_groups,
            report.conflict_rows,
        )
    if report.duplicate_rows_removed:
        LOGGER.info(
            "IQVIA NSA exact-row dedup period=%s groups=%s removed=%s rows_before=%s rows_after=%s",
            period,
            report.duplicate_groups,
            report.duplicate_rows_removed,
            report.rows_before,
            report.rows_after,
        )
    return result, report


def iter_nsa_csv(
    path: Path,
    chunk_size: int = 2000,
    *,
    quarters: Iterable[str] | None = None,
    atc4_codes: Iterable[str] | None = None,
) -> Iterator[dict[str, Any]]:
    quarter_scope = normalize_iqvia_quarters(quarters)
    atc4_scope = normalize_iqvia_atc4_codes(atc4_codes)
    canonical_headers: list[str] | None = None
    for chunk in pd.read_csv(path, encoding="utf-8-sig", chunksize=chunk_size, dtype=object):
        if canonical_headers is None:
            canonical_headers = canonicalize_nsa_headers(chunk.columns, source=f"{path}:CSV")
        chunk.columns = canonical_headers
        periods = nsa_period_columns(chunk.columns)
        static_cols = [c for c in chunk.columns if not re.match(r"^\d{1,2}/\d{4}_", str(c)) and c not in NSA_PARQUET_METRICS]
        for offset, row in chunk.iterrows():
            raw = row_to_payload(row.to_dict())
            static = {k: raw.get(k) for k in static_cols}
            source_row_no = int(offset) + 2
            if not periods:
                record = long_format_period_record(path, "CSV", source_row_no, raw, static_cols)
                if record and iqvia_record_in_scope(
                    record,
                    quarters=quarter_scope,
                    atc4_codes=atc4_scope,
                ):
                    yield record
                continue
            for period, metric_cols in periods.items():
                period_values = {
                    metric: parse_nsa_metric(raw.get(col), metric)
                    for metric, col in metric_cols.items()
                }
                if all(v is None for v in period_values.values()):
                    continue
                year, month = period.split("-")
                quarter = quarter_from_month(int(month))
                if quarter is None:
                    continue
                record = {
                    "source_file": path.name,
                    "sheet_name": "CSV",
                    "source_row_no": source_row_no,
                    "audit_code": clean_value(raw.get("AUDIT CODE")),
                    "audit_desc": clean_value(normalize_nsa_audit_desc(raw)),
                    "mfr_code": clean_value(raw.get("MFR CODE")),
                    "mfr_name": clean_value(raw.get("MFR NAME")),
                    "period_yyyy": int(year),
                    "period_quarter": quarter,
                    "period_label": f"{year}Q{quarter}",
                    "payload": dumps_payload(
                        {
                            "__source": "NSA",
                            "__raw_period": period,
                            "__period_metric_columns": metric_cols,
                            "static": static,
                            "period_values": period_values,
                        }
                    ),
                    "source_master_version": None,
                }
                if iqvia_record_in_scope(
                    record,
                    quarters=quarter_scope,
                    atc4_codes=atc4_scope,
                ):
                    yield record


def iter_nsa_xlsx(
    path: Path,
    *,
    quarters: Iterable[str] | None = None,
    atc4_codes: Iterable[str] | None = None,
) -> Iterator[dict[str, Any]]:
    quarter_scope = normalize_iqvia_quarters(quarters)
    atc4_scope = normalize_iqvia_atc4_codes(atc4_codes)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows)
        except StopIteration:
            continue
        headers = canonicalize_nsa_headers(headers_raw, source=f"{path}:{ws.title}")
        periods = nsa_period_columns(headers)
        static_cols = [c for c in headers if not re.match(r"^\d{1,2}/\d{4}_", c) and c not in NSA_PARQUET_METRICS]
        for row_no, values in enumerate(rows, start=2):
            if all(v is None for v in values):
                continue
            raw = row_to_payload(dict(zip(headers, values)))
            static = {k: raw.get(k) for k in static_cols}
            if not periods:
                record = long_format_period_record(path, ws.title, row_no, raw, static_cols)
                if record and iqvia_record_in_scope(
                    record,
                    quarters=quarter_scope,
                    atc4_codes=atc4_scope,
                ):
                    yield record
                continue
            for period, metric_cols in periods.items():
                period_values = {metric: clean_value(raw.get(col)) for metric, col in metric_cols.items()}
                if all(v is None for v in period_values.values()):
                    continue
                year, month = period.split("-")
                quarter = quarter_from_month(int(month))
                if quarter is None:
                    continue
                record = {
                    "source_file": path.name,
                    "sheet_name": ws.title,
                    "source_row_no": row_no,
                    "audit_code": clean_value(raw.get("AUDIT CODE")),
                    "audit_desc": clean_value(normalize_nsa_audit_desc(raw)),
                    "mfr_code": clean_value(raw.get("MFR CODE")),
                    "mfr_name": clean_value(raw.get("MFR NAME")),
                    "period_yyyy": int(year),
                    "period_quarter": quarter,
                    "period_label": f"{year}Q{quarter}",
                    "payload": dumps_payload(
                        {
                            "__source": "NSA",
                            "__raw_period": period,
                            "__period_metric_columns": metric_cols,
                            "static": static,
                            "period_values": period_values,
                        }
                    ),
                    "source_master_version": None,
                }
                if iqvia_record_in_scope(
                    record,
                    quarters=quarter_scope,
                    atc4_codes=atc4_scope,
                ):
                    yield record


def source_sheets_for_resume(files: list[Path]) -> list[SourceSheet]:
    result: list[SourceSheet] = []
    for path in files:
        if path.suffix.lower() == ".csv":
            result.append(SourceSheet(path, "CSV"))
        else:
            wb = openpyxl.load_workbook(path, read_only=True)
            result.extend(SourceSheet(path, ws) for ws in wb.sheetnames)
    return result


def loaded_sheet_keys(conn: pymysql.connections.Connection, table: str) -> set[tuple[str, str | None]]:
    with conn.cursor() as cur:
        cur.execute(f"SELECT DISTINCT source_file, sheet_name FROM {table}")
        return {(row[0], row[1]) for row in cur.fetchall()}


def batch_insert(
    conn: pymysql.connections.Connection,
    target_table: str,
    records: Iterable[dict[str, Any]],
    batch_size: int,
    max_batch_bytes: int = 8_000_000,
) -> int:
    sql = INSERT_SQL[target_table]
    total = 0
    batch: list[dict[str, Any]] = []
    batch_bytes = 0
    with conn.cursor() as cursor:
        for record in records:
            batch.append(record)
            batch_bytes += len(str(record.get("payload", "")).encode("utf-8")) + 1024
            if len(batch) >= batch_size or batch_bytes >= max_batch_bytes:
                cursor.executemany(sql, batch)
                conn.commit()
                total += len(batch)
                LOGGER.info("[%s] %s rows committed", target_table, f"{total:,}")
                batch.clear()
                batch_bytes = 0
        if batch:
            cursor.executemany(sql, batch)
            conn.commit()
            total += len(batch)
            LOGGER.info("[%s] %s rows committed", target_table, f"{total:,}")
    return total


def iter_records(
    path: Path,
    *,
    quarters: Iterable[str] | None = None,
    atc4_codes: Iterable[str] | None = None,
) -> Iterator[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        yield from iter_nsa_csv(
            path,
            quarters=quarters,
            atc4_codes=atc4_codes,
        )
    else:
        yield from iter_nsa_xlsx(
            path,
            quarters=quarters,
            atc4_codes=atc4_codes,
        )


def canonical_nsa_files(files: Iterable[Path]) -> list[Path]:
    return sorted(
        path
        for path in files
        if path.suffix.lower() in {".csv", ".xlsx", ".xls"} and is_iqvia_source_file(path, {".csv", ".xlsx", ".xls"})
    )


def materialize_iqvia_nsa_parquet(files: list[Path], output_dir: Path) -> dict[str, int]:
    """Write canonical IQVIA NSA period parquet files for Layer0 consumers.

    The local NSA source set can arrive as CSV or workbook extracts.  Keep the
    source_file lineage column and materialize every discovered NSA source file.
    Exact row overlap is removed here; natural-key metric conflicts remain.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    records_by_period: dict[str, list[dict[str, Any]]] = defaultdict(list)

    canonical_files = canonical_nsa_files(files)
    for path in canonical_files:
        LOGGER.info("materializing NSA parquet from %s", path)
        for record in iter_records(path):
            records_by_period[period_label_to_quarter(record["period_label"])].append(record)

    written: dict[str, int] = {}
    for period, records in sorted(records_by_period.items()):
        deduped, _ = deduplicate_nsa_records(records, period)
        rows = [nsa_record_to_parquet_row(record) for record in deduped]
        out_path = output_dir / f"{period}.parquet"
        pd.DataFrame(rows).to_parquet(out_path, index=False)
        written[period] = len(rows)
        LOGGER.info("wrote %s: %s rows", out_path, f"{len(rows):,}")
    return written


def table_for_source() -> str:
    return NSA_TABLE


RECORD_PARQUET_COLUMNS = (
    "source_file",
    "sheet_name",
    "source_row_no",
    "audit_code",
    "audit_desc",
    "mfr_code",
    "mfr_name",
    "period_yyyy",
    "period_quarter",
    "period_label",
    "payload",
    "source_master_version",
)

RECORD_PARQUET_INT_COLUMNS = {
    "source_row_no",
    "period_yyyy",
    "period_quarter",
}


def record_parquet_schema() -> pa.Schema:
    fields = []
    for column in RECORD_PARQUET_COLUMNS:
        dtype = pa.int64() if column in RECORD_PARQUET_INT_COLUMNS else pa.string()
        fields.append(pa.field(column, dtype))
    return pa.schema(fields)


def normalize_record_for_parquet(record: dict[str, Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for column in RECORD_PARQUET_COLUMNS:
        value = clean_value(record.get(column))
        if column in RECORD_PARQUET_INT_COLUMNS and value is not None:
            value = int(value)
        elif value is not None and not isinstance(value, str):
            value = str(value)
        row[column] = value
    return row


def record_period_key(record: dict[str, Any]) -> str:
    period = record.get("period_label")
    if not period:
        raise ValueError(f"NSA record missing period key: {record}")
    return str(period).replace("/", "-")


def _flush_record_parquet_batch(
    period: str,
    rows: list[dict[str, Any]],
    writers: dict[str, pq.ParquetWriter],
    source_dir: Path,
) -> None:
    if not rows:
        return
    path = source_dir / f"{period}.parquet"
    schema = record_parquet_schema()
    writer = writers.get(period)
    if writer is None:
        writer = pq.ParquetWriter(path, schema=schema, compression="snappy")
        writers[period] = writer
    table = pa.Table.from_pylist(rows, schema=schema)
    writer.write_table(table)
    rows.clear()


@dataclass
class _NsaDedupState:
    signature: str
    total: int
    winner_index: int
    winner_recency: tuple[int, int, int, int, int]
    conflict: bool = False


def _iter_record_parquet_rows(path: Path, batch_size: int) -> Iterator[dict[str, Any]]:
    parquet = pq.ParquetFile(path)
    columns = list(RECORD_PARQUET_COLUMNS)
    for batch in parquet.iter_batches(batch_size=batch_size, columns=columns):
        table = pa.Table.from_batches([batch], schema=record_parquet_schema())
        yield from table.to_pylist()


def _deduplicate_record_parquet(
    path: Path,
    period: str,
    *,
    batch_size: int,
) -> NsaDedupReport:
    states: dict[tuple[Any, ...], _NsaDedupState] = {}
    rows_before = 0
    for index, record in enumerate(_iter_record_parquet_rows(path, batch_size)):
        rows_before += 1
        natural_key, signature = _nsa_identity(record)
        if natural_key is None:
            continue
        recency = _nsa_source_recency(record, index)
        state = states.get(natural_key)
        if state is None:
            states[natural_key] = _NsaDedupState(
                signature=signature,
                total=1,
                winner_index=index,
                winner_recency=recency,
            )
            continue
        state.total += 1
        if signature != state.signature:
            state.conflict = True
        if recency > state.winner_recency:
            state.winner_index = index
            state.winner_recency = recency

    duplicate_groups = sum(
        state.total > 1 and not state.conflict for state in states.values()
    )
    duplicate_rows_removed = sum(
        state.total - 1
        for state in states.values()
        if state.total > 1 and not state.conflict
    )
    conflict_groups = sum(state.conflict for state in states.values())
    conflict_rows = sum(state.total for state in states.values() if state.conflict)
    report = NsaDedupReport(
        period=period,
        rows_before=rows_before,
        rows_after=rows_before - duplicate_rows_removed,
        duplicate_groups=duplicate_groups,
        duplicate_rows_removed=duplicate_rows_removed,
        conflict_groups=conflict_groups,
        conflict_rows=conflict_rows,
    )

    temp_path = path.with_suffix(".dedup.tmp")
    writer = pq.ParquetWriter(temp_path, schema=record_parquet_schema(), compression="snappy")
    output: list[dict[str, Any]] = []
    try:
        for index, record in enumerate(_iter_record_parquet_rows(path, batch_size)):
            natural_key, _ = _nsa_identity(record)
            state = states.get(natural_key) if natural_key is not None else None
            if state is None or state.conflict or index == state.winner_index:
                output.append(record)
            if len(output) >= batch_size:
                writer.write_table(pa.Table.from_pylist(output, schema=record_parquet_schema()))
                output.clear()
        if output:
            writer.write_table(pa.Table.from_pylist(output, schema=record_parquet_schema()))
    finally:
        writer.close()
    temp_path.replace(path)

    if report.conflict_groups:
        LOGGER.warning(
            "IQVIA NSA natural-key conflicts preserved period=%s groups=%s rows=%s",
            period,
            report.conflict_groups,
            report.conflict_rows,
        )
    if report.duplicate_rows_removed:
        LOGGER.info(
            "IQVIA NSA exact-row dedup period=%s groups=%s removed=%s rows_before=%s rows_after=%s",
            period,
            report.duplicate_groups,
            report.duplicate_rows_removed,
            report.rows_before,
            report.rows_after,
        )
    return report


def materialize_record_parquet(
    files: list[Path],
    output_dir: Path,
    *,
    batch_size: int = 10_000,
    overwrite: bool = False,
    repeat_factor: int = 1,
) -> dict[str, int]:
    """Materialize DB insert records and deduplicate exact facts per period."""
    if repeat_factor < 1:
        raise ValueError(f"repeat_factor must be >= 1, got {repeat_factor}")
    source_dir = output_dir / "nsa"
    source_dir.mkdir(parents=True, exist_ok=True)
    existing = sorted(source_dir.glob("*.parquet"))
    if existing and not overwrite:
        raise FileExistsError(
            f"{source_dir} already has parquet files. "
            "Pass --overwrite-record-parquet to replace generated cache files."
        )
    if overwrite:
        for path in existing:
            path.unlink()

    counts: dict[str, int] = defaultdict(int)
    buffers: dict[str, list[dict[str, Any]]] = defaultdict(list)
    writers: dict[str, pq.ParquetWriter] = {}
    buffered_rows = 0
    try:
        for path in files:
            LOGGER.info("record-parquet materialize NSA %s", path)
            for record in iter_records(path):
                period = record_period_key(record)
                row = normalize_record_for_parquet(record)
                for _ in range(repeat_factor):
                    buffers[period].append(dict(row))
                    counts[period] += 1
                    buffered_rows += 1
                    if buffered_rows >= batch_size:
                        for buffered_period, buffered in buffers.items():
                            _flush_record_parquet_batch(
                                buffered_period,
                                buffered,
                                writers,
                                source_dir,
                            )
                        buffered_rows = 0
        for period, rows in list(buffers.items()):
            _flush_record_parquet_batch(period, rows, writers, source_dir)
    finally:
        for writer in writers.values():
            writer.close()
        writers.clear()
        buffers.clear()

    for period in sorted(counts):
        path = source_dir / f"{period}.parquet"
        report = _deduplicate_record_parquet(path, period, batch_size=batch_size)
        counts[period] = report.rows_after

    for period, count in sorted(counts.items()):
        LOGGER.info("record-parquet wrote nsa/%s.parquet: %s rows", period, f"{count:,}")
    return dict(counts)


def iter_record_parquet_records(output_dir: Path, *, batch_size: int = 10_000) -> Iterator[dict[str, Any]]:
    source_dir = output_dir / "nsa"
    files = sorted(source_dir.glob("*.parquet"))
    if not files:
        raise FileNotFoundError(f"no record parquet files found under {source_dir}")
    columns = list(RECORD_PARQUET_COLUMNS)
    for path in files:
        parquet = pq.ParquetFile(path)
        for batch in parquet.iter_batches(batch_size=batch_size, columns=columns):
            table = pa.Table.from_batches([batch], schema=record_parquet_schema())
            for row in table.to_pylist():
                yield row


def init_target_schema(target_database: str, source_database: str = "jw_mart") -> None:
    """Create a temp target schema and empty raw tables cloned from source DB."""
    conn = connect(source_database)
    try:
        with conn.cursor() as cursor:
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{target_database}`")
            cursor.execute(
                f"CREATE TABLE IF NOT EXISTS `{target_database}`.`{NSA_TABLE}` "
                f"LIKE `{source_database}`.`{NSA_TABLE}`"
            )
            cursor.execute(f"SELECT COUNT(*) FROM `{target_database}`.`{NSA_TABLE}`")
            count = cursor.fetchone()[0]
            if count:
                raise RuntimeError(
                    f"target table {target_database}.{NSA_TABLE} is not empty ({count:,} rows); "
                    "refusing to append reproduction data"
                )
        conn.commit()
    finally:
        conn.close()


def load_record_parquet_source(output_dir: Path, *, target_database: str, batch_size: int) -> int:
    table = table_for_source()
    conn = connect(target_database)
    try:
        return batch_insert(conn, table, iter_record_parquet_records(output_dir, batch_size=batch_size), batch_size)
    finally:
        conn.close()


def dry_run(files: list[Path], out_path: Path | None) -> None:
    lines: list[str] = ["# IQVIA NSA Dry Run", ""]
    lines.append(f"- files: {len(files)}")
    sample_records: list[dict[str, Any]] = []
    total_preview = 0
    errors: list[str] = []

    for path in files[:1] if len(files) > 1 else files:
        lines.append(f"## File: `{path}`")
        try:
            periods = set()
            metrics = set()
            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path, nrows=3, encoding="utf-8-sig", dtype=object)
                headers = list(df.columns)
            else:
                df = pd.read_excel(path, nrows=3, dtype=object)
                headers = list(df.columns)
            for period, metric_cols in nsa_period_columns(headers).items():
                periods.add(period)
                metrics.update(metric_cols)
            lines.append(f"- period count: {len(periods)}")
            lines.append(f"- period range: {min(periods) if periods else None} .. {max(periods) if periods else None}")
            lines.append(f"- metrics: {sorted(metrics)}")
            lines.append(f"- first columns: {headers[:20]}")

            for record in iter_records(path):
                sample_records.append(record)
                total_preview += 1
                if len(sample_records) >= 3:
                    break
            lines.append(f"- preview records generated before stop: {total_preview}")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{path}: {exc}")
            lines.append(f"- ERROR: {exc}")

    lines.append("")
    lines.append("## Sample records")
    for record in sample_records[:3]:
        preview = dict(record)
        payload = preview.pop("payload", "")
        preview["payload_preview"] = payload[:800]
        lines.append("```json")
        lines.append(json.dumps(preview, ensure_ascii=False, indent=2, default=str))
        lines.append("```")

    if errors:
        lines.extend(["", "## Errors", *[f"- {e}" for e in errors]])

    text = "\n".join(lines) + "\n"
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text)
    print(text)
    if errors:
        raise HeaderContractError(f"IQVIA NSA dry-run failed: {'; '.join(errors)}")


def load_source(
    files: list[Path],
    batch_size: int,
    dry: bool = False,
    target_database: str | None = None,
) -> LoadStats:
    stats = LoadStats()
    table = table_for_source()
    if dry:
        dry_run(files, None)
        return stats

    conn = connect(target_database)
    loaded = loaded_sheet_keys(conn, table)
    try:
        pending: list[Path] = []
        for path in files:
            stats.files += 1
            LOGGER.info("reading NSA %s", path)
            if path.suffix.lower() == ".csv":
                sheet_keys = [("CSV",)]
            else:
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_keys = [(s,) for s in wb.sheetnames]
            if all((path.name, sheet_name) in loaded for (sheet_name,) in sheet_keys):
                LOGGER.info("skip already loaded file: %s", path.name)
                continue
            pending.append(path)
            stats.sheets += len(sheet_keys)
        if pending:
            try:
                with tempfile.TemporaryDirectory(prefix="iqvia-nsa-records-") as temp_dir:
                    record_dir = Path(temp_dir)
                    materialize_record_parquet(
                        pending,
                        record_dir,
                        batch_size=batch_size,
                    )
                    count = batch_insert(
                        conn,
                        table,
                        iter_record_parquet_records(record_dir, batch_size=batch_size),
                        batch_size,
                    )
                stats.rows += count
                LOGGER.info("loaded IQVIA NSA files=%s rows=%s", len(pending), f"{count:,}")
            except Exception as exc:  # noqa: BLE001
                conn.rollback()
                message = f"IQVIA NSA batch: {exc}"
                stats.errors.append(message)
                LOGGER.error("ERROR %s", message)
    finally:
        conn.close()
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", type=Path, help="load or dry-run one file")
    parser.add_argument("--dry-run", action="store_true", help="inspect schema without writing")
    parser.add_argument("--batch-size", type=int, default=10000)
    parser.add_argument(
        "--materialize-parquet",
        action="store_true",
        help="write IQVIA NSA period parquet files for Layer0 prototype consumers",
    )
    parser.add_argument(
        "--parquet-output-dir",
        type=Path,
        default=DEFAULT_NSA_PARQUET_DIR,
        help="IQVIA NSA parquet output directory",
    )
    parser.add_argument(
        "--skip-db",
        action="store_true",
        help="skip MariaDB inserts; useful with --materialize-parquet",
    )
    parser.add_argument(
        "--record-parquet-first",
        action="store_true",
        help="materialize DB insert records to output/iqvia/nsa/{period}.parquet, then load from parquet",
    )
    parser.add_argument(
        "--record-parquet-output-dir",
        type=Path,
        default=DEFAULT_RECORD_PARQUET_DIR,
        help="record parquet cache root for --record-parquet-first",
    )
    parser.add_argument(
        "--overwrite-record-parquet",
        action="store_true",
        help="replace generated record parquet files under --record-parquet-output-dir",
    )
    parser.add_argument(
        "--record-repeat-factor",
        type=int,
        default=1,
        help="repeat each parsed record in the record parquet cache to reproduce an existing raw-table baseline",
    )
    parser.add_argument(
        "--target-db",
        help="target schema for --record-parquet-first loads; required unless --skip-db is set",
    )
    parser.add_argument(
        "--source-db",
        default="jw_mart",
        help="source schema to clone raw table definitions from when --init-target-db is set",
    )
    parser.add_argument(
        "--init-target-db",
        action="store_true",
        help="create --target-db and empty raw tables cloned from --source-db",
    )
    return parser.parse_args()


def main() -> int:
    try:
        args = parse_args()
        if args.record_parquet_first and not args.skip_db and not args.target_db:
            LOGGER.error("ERROR: --record-parquet-first needs --target-db unless --skip-db is set")
            return 2
        if args.init_target_db:
            if not args.target_db:
                LOGGER.error("ERROR: --init-target-db needs --target-db")
                return 2
            init_target_schema(args.target_db, args.source_db)

        AUDIT_DIR.mkdir(parents=True, exist_ok=True)
        all_errors: list[str] = []
        files = [args.file] if args.file else discover_files()
        files = [p for p in files if p is not None]
        if args.dry_run:
            out_path = AUDIT_DIR / "dry_run_nsa.md"
            dry_run(files, out_path)
            return 0
        if args.materialize_parquet:
            written = materialize_iqvia_nsa_parquet(files, args.parquet_output_dir)
            LOGGER.info("NSA parquet materialized partitions=%s rows=%s", len(written), f"{sum(written.values()):,}")
        if args.record_parquet_first:
            written = materialize_record_parquet(
                files,
                args.record_parquet_output_dir,
                batch_size=args.batch_size,
                overwrite=args.overwrite_record_parquet,
                repeat_factor=args.record_repeat_factor,
            )
            LOGGER.info(
                "record parquet materialized NSA partitions=%s rows=%s",
                len(written),
                f"{sum(written.values()):,}",
            )
        if args.skip_db:
            return 0
        if args.record_parquet_first:
            count = load_record_parquet_source(
                args.record_parquet_output_dir,
                target_database=args.target_db,
                batch_size=args.batch_size,
            )
            stats = LoadStats(rows=count, files=len(files), sheets=0)
        else:
            stats = load_source(files, args.batch_size)
        LOGGER.info(
            "SUMMARY NSA: files=%s, sheets=%s, rows=%s, errors=%s",
            stats.files,
            stats.sheets,
            f"{stats.rows:,}",
            len(stats.errors),
        )
        all_errors.extend(stats.errors)

        if all_errors:
            LOGGER.error("ERRORS:")
            for error in all_errors:
                LOGGER.error("- %s", error)
            return 1
        return 0
    except Exception:
        LOGGER.exception("IQVIA loader failed")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
