"""Create a deterministic UBIST next-month rehearsal workbook."""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
import re

import openpyxl

from pipeline.etl.io.ubist_loader import CANONICAL_DIMENSIONS


@dataclass(frozen=True, slots=True)
class RehearsalWorkbook:
    period: str
    rows: int


_ROWS: tuple[dict[str, str | float], ...] = (
    {
        "제조사": "JW중외제약",
        "국내/외자": "국내",
        "판매사": "JW중외제약",
        "판매사2": "JW중외제약",
        "제품": "리바로정",
        "ATC": "C10AA",
        "브랜드": "리바로",
        "약가": "551",
        "성분": "pitavastatin calcium",
        "성분용량": "2mg",
        "일반/전문": "전문",
        "약품코드": "REHEARSAL-001",
        "제형": "정제",
        "투여경로": "경구",
        "급여구분": "급여",
        "종별": "의원",
        "진료과": "내과",
        "연령": "전체",
        "성별": "전체",
        "처방조제액(원)": 1000.0,
        "처방건수_P": 10.0,
        "처방량_P": 20.0,
    },
    {
        "제조사": "JW중외제약",
        "국내/외자": "국내",
        "판매사": "JW중외제약",
        "판매사2": "JW중외제약",
        "제품": "리바로젯정",
        "ATC": "C10BA",
        "브랜드": "리바로젯",
        "약가": "784",
        "성분": "pitavastatin, ezetimibe",
        "성분용량": "2mg/10mg",
        "일반/전문": "전문",
        "약품코드": "REHEARSAL-002",
        "제형": "정제",
        "투여경로": "경구",
        "급여구분": "급여",
        "종별": "종합병원",
        "진료과": "순환기내과",
        "연령": "전체",
        "성별": "전체",
        "처방조제액(원)": 2000.0,
        "처방건수_P": 20.0,
        "처방량_P": 40.0,
    },
    {
        "제조사": "테스트제조사",
        "국내/외자": "국내",
        "판매사": "테스트판매사",
        "판매사2": "테스트판매사",
        "제품": "테스트대조정",
        "ATC": "C10AA",
        "브랜드": "테스트대조",
        "약가": "100",
        "성분": "control ingredient",
        "성분용량": "1mg",
        "일반/전문": "전문",
        "약품코드": "REHEARSAL-003",
        "제형": "정제",
        "투여경로": "경구",
        "급여구분": "급여",
        "종별": "의원",
        "진료과": "내과",
        "연령": "전체",
        "성별": "전체",
        "처방조제액(원)": 300.0,
        "처방건수_P": 3.0,
        "처방량_P": 6.0,
    },
)


def write_rehearsal_workbook(output: Path, period: str) -> RehearsalWorkbook:
    if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", period) is None:
        raise ValueError(f"invalid UBIST month: {period!r}")

    metrics = ("처방조제액(원)", "처방건수_P", "처방량_P")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rehearsal"
    for column, header in enumerate((*CANONICAL_DIMENSIONS, *metrics), start=1):
        if header in metrics:
            year, month = period.split("-", maxsplit=1)
            sheet.cell(1, column).value = header
            sheet.cell(2, column).value = f"{year}년 {int(month)}월"
        else:
            sheet.cell(2, column).value = header
    for row_index, row in enumerate(_ROWS, start=3):
        for column, header in enumerate((*CANONICAL_DIMENSIONS, *metrics), start=1):
            sheet.cell(row_index, column).value = row[header]
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output)
    finally:
        workbook.close()
    return RehearsalWorkbook(period=period, rows=len(_ROWS))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=Path)
    parser.add_argument("--period", default="2026-06")
    args = parser.parse_args(argv)
    report = write_rehearsal_workbook(args.output, args.period)
    print(f"period={report.period} rows={report.rows} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
