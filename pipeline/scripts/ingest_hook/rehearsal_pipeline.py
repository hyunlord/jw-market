from __future__ import annotations

import csv
import hashlib
import json
import os
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterator, Sequence

import openpyxl
import duckdb

from pipeline.scripts.etl.brand_activity.csd_core import EXPECTED_HEADERS
from pipeline.scripts.etl.brand_activity.ingest_keyword import KEYWORD_HEADERS
from pipeline.scripts.ingest_hook import category_activation
from pipeline.scripts.ingest_hook.category_map import resolve_category
from pipeline.scripts.ingest_hook.contract import load_manifest
from pipeline.scripts.ingest_hook.g3 import validate
from pipeline.scripts.ingest_hook.rehearsal_nsa_split import split_workbook
from pipeline.scripts.ingest_hook.semantic_replay import (
    ReplayConfig,
    compare_parquet_roots,
)
from pipeline.scripts.ingest_hook.source_fingerprint import fingerprint_source
from pipeline.scripts.ingest_hook.workbook_contracts import WorkbookSummary, summarize


@dataclass(frozen=True, slots=True)
class DryRunActivation:
    target_tables: tuple[str, ...]
    nsa_quarters: int
    dry_run: bool
    published: bool


@dataclass(frozen=True, slots=True)
class IqviaRehearsal:
    history_quarters: int
    latest_quarters: tuple[str, ...]
    history_validation: WorkbookSummary
    latest_validation: WorkbookSummary
    shuffled_columns_identical: bool
    activation: DryRunActivation


@dataclass(frozen=True, slots=True)
class UbistRehearsal:
    category: str
    epoch: str
    validation_rows: int
    actions: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class CsdRehearsal:
    validation: WorkbookSummary
    activation: DryRunActivation


@dataclass(frozen=True, slots=True)
class FullRehearsalReport:
    no_production_writes: bool
    semantic_replay_matches: bool
    iqvia: IqviaRehearsal
    ubist: UbistRehearsal
    csd_channel: CsdRehearsal
    csd_keyword: CsdRehearsal


def build_full_rehearsal(root: Path, *, baseline_month: str = "2026-07") -> FullRehearsalReport:
    root.mkdir(parents=True, exist_ok=True)
    next_month = _next_month(baseline_month)
    return FullRehearsalReport(
        no_production_writes=True,
        semantic_replay_matches=_verify_semantic_replay(root / "semantic-replay"),
        iqvia=_rehearse_iqvia(root / "iqvia"),
        ubist=_rehearse_ubist(root / "ubist", next_month),
        csd_channel=_rehearse_csd(
            root / "csd-channel",
            next_month,
            category="iqvia_csd_channel",
            filename="csd-channel.xlsx",
            tables=("raw_csd_channel_dynamics", "csd_channel_dynamics_stage"),
            writer=_write_csd_channel,
        ),
        csd_keyword=_rehearse_csd(
            root / "csd-keyword",
            next_month,
            category="iqvia_csd_keyword",
            filename="csd-keyword.xlsx",
            tables=("raw_keyword_events", "km_keyword_event_stage"),
            writer=_write_csd_keyword,
        ),
    )


def _rehearse_iqvia(root: Path) -> IqviaRehearsal:
    root.mkdir(parents=True, exist_ok=True)
    source = root / "nsa-source.xlsx"
    history = root / "nsa-history-19q.xlsx"
    latest = root / "nsa-latest-1q.xlsx"
    _write_nsa(source)
    shuffled = root / "renamed-shuffled-extra-column.xlsx"
    _shuffle_nsa_columns(source, shuffled)
    result = split_workbook(source, history, latest, history_quarters=19)
    epoch = result.latest_quarters[0]
    activation = _dry_activation(
        "iqvia_nsa",
        epoch,
        _table_manifest(
            "iqvia_nsa",
            epoch,
            ("iqvia_nsa_quarterly_raw",),
        ),
    )
    return IqviaRehearsal(
        history_quarters=len(result.history_quarters),
        latest_quarters=result.latest_quarters,
        history_validation=summarize("iqvia_nsa", history, epoch),
        latest_validation=summarize("iqvia_nsa", latest, epoch),
        shuffled_columns_identical=(
            fingerprint_source(source, "iqvia_nsa").identity
            == fingerprint_source(shuffled, "iqvia_nsa").identity
        ),
        activation=activation,
    )


def _rehearse_ubist(root: Path, epoch: str) -> UbistRehearsal:
    root.mkdir(parents=True, exist_ok=True)
    csv_path = root / "ubist-next-month.csv"
    _write_ubist_csv(csv_path, epoch)
    manifest_path = _contract_manifest(root, "ubist", epoch, csv_path, rows=3)
    report = validate(load_manifest(manifest_path), resolve_category("ubist"), root)
    return UbistRehearsal(
        category="ubist",
        epoch=epoch,
        validation_rows=report.total_rows,
        actions=(
            "validate manifest with G3",
            "load next-month source into isolated staging target",
            "defer mart publish until explicit production activation",
        ),
    )


def _rehearse_csd(
    root: Path,
    epoch: str,
    *,
    category: str,
    filename: str,
    tables: tuple[str, ...],
    writer: Callable[[Path, str], None],
) -> CsdRehearsal:
    workbook = root / filename
    root.mkdir(parents=True, exist_ok=True)
    writer(workbook, epoch)
    return CsdRehearsal(
        validation=summarize(category, workbook, epoch),
        activation=_dry_activation(category, epoch, _table_manifest(category, epoch, tables)),
    )


def _dry_activation(category: str, epoch: str, manifest: dict[str, object]) -> DryRunActivation:
    with _shadow_activation_env():
        plan = category_activation.prepare(category, manifest, epoch, "rehearsal")
        result = category_activation.activate(
            category,
            manifest,
            epoch,
            "rehearsal",
            dry_run=True,
    )
    return DryRunActivation(
        target_tables=result.target_tables,
        nsa_quarters=len(plan.nsa_quarters),
        dry_run=result.dry_run,
        published=result.published,
    )


@contextmanager
def _shadow_activation_env() -> Iterator[None]:
    values = {
        category_activation.ENV_BUILD_PREFIX: "jw_ingest_shadow_category_build",
        category_activation.ENV_TARGET_IQVIA_NSA_DB: "jw_ingest_shadow_iqvia",
        category_activation.ENV_TARGET_CSD_RAW_DB: "jw_ingest_shadow_csd_raw",
        category_activation.ENV_TARGET_CSD_STAGE_DB: "jw_ingest_shadow_csd_stage",
    }
    old_values = {key: os.environ.get(key) for key in values}
    os.environ.update(values)
    try:
        yield
    finally:
        for key, value in old_values.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def _table_manifest(category: str, epoch: str, tables: tuple[str, ...]) -> dict[str, object]:
    return {
        "schema_version": "ingest-table-load-v1",
        "category": category,
        "epoch": epoch,
        "tables": [
            {
                "schema": f"jw_ingest_stage_{index}",
                "table": table,
                "kind": "replace",
                "rows_before": 0,
                "rows_after": 1,
                "rows_loaded": 1,
                "source_rows": 1,
                "difference_reasons": [],
            }
            for index, table in enumerate(tables, start=1)
        ],
    }


def _contract_manifest(root: Path, category: str, epoch: str, path: Path, *, rows: int) -> Path:
    payload = {
        "contract_version": "v2",
        "epoch": epoch,
        "category": category,
        "complete": True,
        "files": [
            {
                "path": path.relative_to(root).as_posix(),
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                "rows": rows,
                "period_start": epoch,
                "period_end": epoch,
            }
        ],
    }
    manifest_path = root / "manifest.json"
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return manifest_path


def _write_nsa(path: Path) -> None:
    rows: list[Sequence[object]] = [[
        "AUDIT CODE", "MFR CODE", "PRODUCT NAME", "PACK DESC", "DATA PERIOD",
        "Values LC", "Units", "Counting Units", "Dosage Units", "Price",
    ]]
    for index in range(20):
        year = 2021 + index // 4
        month = (index % 4 + 1) * 3
        rows.append([f"A-{index}", "M", "Brand", "Pack", f"{year}-{month:02d}", 1, 2, 3, 4, 5])
    _save_sheet(path, "NSA", rows)


def _write_ubist_csv(path: Path, epoch: str) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(("period", "level", "brand", "value"))
        writer.writerows(
            (
                (epoch, "Class", "리바로", 10.0),
                (epoch, "Class", "리바로젯", 20.0),
                (epoch, "전체", "-", 30.0),
            )
        )


def _write_csd_channel(path: Path, epoch: str) -> None:
    _save_sheet(
        path,
        "Renamed Sheet",
        [*([()] * 6), EXPECTED_HEADERS, (_month_label(epoch), "Market", "TOTAL", "TOTAL", "Brand", "Maker", "JW", 1)],
    )


def _write_csd_keyword(path: Path, epoch: str) -> None:
    _save_sheet(
        path,
        "Renamed Keyword Sheet",
        [KEYWORD_HEADERS, (_month_label(epoch), "Seoul", "IM", "JW", "Brand", "A10", "keyword", "high", "1", "up", "N", "N", "N", "N", "N", "", "")],
    )


def _save_sheet(path: Path, title: str, rows: Sequence[Sequence[object]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _shuffle_nsa_columns(source: Path, target: Path) -> None:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    order = (4, 1, 9, 0, 6, 3, 2, 8, 5, 7)
    shuffled = [["UNUSED EXTRA", *(rows[0][index] for index in order)]]
    shuffled.extend(
        [["ignored", *(row[index] for index in order)] for row in rows[1:]]
    )
    _save_sheet(target, "Human Renamed Sheet", shuffled)


def _verify_semantic_replay(root: Path) -> bool:
    expected = root / "expected"
    actual = root / "actual"
    expected.mkdir(parents=True, exist_ok=True)
    actual.mkdir(parents=True, exist_ok=True)
    rows = [
        (1, "2026-Q1", "100", "first-build"),
        (2, "2026-Q1", "200", "first-build"),
        (3, "2026-Q2", "300", "first-build"),
    ]
    _write_parquet(expected / "all.parquet", rows)
    _write_parquet(
        actual / "part-b.parquet",
        [(3, "2026-Q2", "300", "replay-build")],
    )
    _write_parquet(
        actual / "part-a.parquet",
        [
            (2, "2026-Q1", "200", "replay-build"),
            (1, "2026-Q1", "100", "replay-build"),
        ],
    )
    comparison = compare_parquet_roots(
        expected,
        actual,
        ReplayConfig(
            business_columns=("product_id", "period", "value"),
            memory_limit="64MB",
            threads=1,
            temp_directory=root / "spill",
        ),
    )
    return comparison.matches


def _write_parquet(path: Path, rows: Sequence[Sequence[object]]) -> None:
    with duckdb.connect() as connection:
        connection.execute(
            "CREATE TABLE source("
            "product_id BIGINT, period VARCHAR, value VARCHAR, build_marker VARCHAR)"
        )
        connection.executemany("INSERT INTO source VALUES (?, ?, ?, ?)", rows)
        connection.execute("COPY source TO ? (FORMAT PARQUET)", [str(path)])


def _month_label(epoch: str) -> str:
    year, month = epoch.split("-")
    labels = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
    return f"{labels[int(month) - 1]}. {year[-2:]}"


def _next_month(epoch: str) -> str:
    year, month = (int(part) for part in epoch.split("-"))
    if month == 12:
        return f"{year + 1}-01"
    return f"{year}-{month + 1:02d}"
