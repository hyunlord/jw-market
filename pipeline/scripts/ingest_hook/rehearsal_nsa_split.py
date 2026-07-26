"""Create reproducible 19+1-quarter IQVIA NSA rehearsal workbooks."""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re

import openpyxl

from pipeline.etl.io.iqvia_loader import canonicalize_nsa_headers, period_label_to_quarter


@dataclass(frozen=True, slots=True)
class SplitResult:
    history_quarters: tuple[str, ...]
    latest_quarters: tuple[str, ...]
    history_rows: int
    latest_rows: int


def _quarter(value: object) -> str:
    if isinstance(value, date):
        return f"{value.year:04d}-Q{((value.month - 1) // 3) + 1}"
    text = str(value or "").strip()
    month_match = re.fullmatch(r"(\d{4})-(\d{2})", text)
    if month_match:
        year, month = int(month_match.group(1)), int(month_match.group(2))
        if not 1 <= month <= 12:
            raise ValueError(f"invalid NSA month: {value!r}")
        return f"{year:04d}-Q{((month - 1) // 3) + 1}"
    return period_label_to_quarter(value)


def _copy_row(target, values: tuple[object, ...]) -> None:
    target.append(list(values))


def split_workbook(
    source: Path,
    history_output: Path,
    latest_output: Path,
    *,
    history_quarters: int = 19,
) -> SplitResult:
    """Split a canonical long-format NSA workbook without using its file name.

    The source contract is identified by canonicalized headers. All source
    columns and their order are retained so the generated files exercise the
    same parser contract as the original workbook.
    """
    if history_quarters < 1:
        raise ValueError("history_quarters must be positive")
    with source.open("rb") as source_stream:
        workbook = openpyxl.load_workbook(
            source_stream,
            read_only=True,
            data_only=False,
        )
        try:
            matches: list[tuple[str, tuple[object, ...], tuple[str, ...]]] = []
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                try:
                    raw_headers = tuple(next(rows))
                    headers = tuple(
                        canonicalize_nsa_headers(
                            raw_headers,
                            source=f"{source}:{sheet.title}",
                        )
                    )
                except (StopIteration, ValueError):
                    continue
                if "DATA PERIOD" not in headers:
                    continue
                matches.append((sheet.title, raw_headers, headers))
            if len(matches) != 1:
                raise ValueError(
                    "NSA split requires exactly one long-format data sheet; "
                    f"matched={len(matches)}"
                )
            sheet_name, raw_header, headers = matches[0]
            period_index = headers.index("DATA PERIOD")
            source_sheet = workbook[sheet_name]
            quarters_found: set[str] = set()
            scan_rows = source_sheet.iter_rows(values_only=True)
            next(scan_rows)
            for row in scan_rows:
                if period_index < len(row) and row[period_index] not in (None, ""):
                    quarters_found.add(_quarter(row[period_index]))
            quarters = sorted(quarters_found)
            if len(quarters) < 2:
                raise ValueError("NSA split requires at least two quarters")
            latest_set = {quarters[-1]}
            history_set = set(quarters[-(history_quarters + 1) : -1])

            history_book = openpyxl.Workbook(write_only=True)
            latest_book = openpyxl.Workbook(write_only=True)
            history_sheet = history_book.create_sheet(sheet_name)
            latest_sheet = latest_book.create_sheet(sheet_name)
            _copy_row(history_sheet, raw_header)
            _copy_row(latest_sheet, raw_header)
            history_rows = 0
            latest_rows = 0
            output_rows = source_sheet.iter_rows(values_only=True)
            next(output_rows)
            for row in output_rows:
                if period_index >= len(row) or row[period_index] in (None, ""):
                    continue
                quarter = _quarter(row[period_index])
                if quarter in history_set:
                    _copy_row(history_sheet, row)
                    history_rows += 1
                elif quarter in latest_set:
                    _copy_row(latest_sheet, row)
                    latest_rows += 1
            if not history_rows or not latest_rows:
                raise ValueError(
                    "NSA split produced an empty output "
                    f"history_rows={history_rows} latest_rows={latest_rows}"
                )
            history_output.parent.mkdir(parents=True, exist_ok=True)
            latest_output.parent.mkdir(parents=True, exist_ok=True)
            try:
                history_book.save(history_output)
                latest_book.save(latest_output)
            finally:
                history_book.close()
                latest_book.close()
        finally:
            workbook.close()
    return SplitResult(
        history_quarters=tuple(sorted(history_set)),
        latest_quarters=tuple(sorted(latest_set)),
        history_rows=history_rows,
        latest_rows=latest_rows,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("history_output", type=Path)
    parser.add_argument("latest_output", type=Path)
    parser.add_argument("--history-quarters", type=int, default=19)
    args = parser.parse_args(argv)
    result = split_workbook(
        args.source,
        args.history_output,
        args.latest_output,
        history_quarters=args.history_quarters,
    )
    print(
        f"history={result.history_quarters} rows={result.history_rows} "
        f"latest={result.latest_quarters} rows={result.latest_rows}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
