"""
prototype_11_master_drug_to_parquet.py
======================================
MI Master drug rows -> Parquet.

Phase 09e policy:
- Canonical row-generation logic:
  /Users/rexxa/github/jw-market/etl/master_loader.py::load_market_sheet
- Canonical transform helpers:
  /Users/rexxa/github/jw-market/etl/master_transform.py
- Canonical schema:
  /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_drug
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.
- Master drug parquet is a single file, not partitioned.

Usage, in Step D after user review:
    python3 scripts/prototype_11_master_drug_to_parquet.py
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


ETL_DIR = Path(__file__).resolve().parent / "etl"
sys.path.insert(0, str(ETL_DIR))
from storage import get_mi_master_path  # noqa: E402
try:  # noqa: E402
    from strategic_exclusion_policy import classify_exclusion_cells as classify_exclusion_cells_by_policy
except ModuleNotFoundError:  # noqa: E402
    from pipeline.scripts.strategic_exclusion_policy import classify_exclusion_cells as classify_exclusion_cells_by_policy


DEFAULT_INPUT_FILE = get_mi_master_path()
MASTER_ROOT = DEFAULT_INPUT_FILE.parent
DEFAULT_CATALOG_PATH = Path("docs/reference/master_column_mapping_catalog.md")
DEFAULT_OUTPUT_FILE = Path("parquet/master_drug/master_drug.parquet")

STANDARD_PREFIX = "drug_extra_json."
EXPECTED_ROW_COUNT = 3952
EXPECTED_EXCLUDED_ROWS = 543
EXPECTED_SOURCE_TYPE_DISTRIBUTION = {"IQVIA": 650, "UBIST": 3302}

MASTER_DRUG_COLUMNS = (
    "strategic_market_id",
    "market_name",
    "source_type",
    "drug_index",
    "atc4_code",
    "atc4_desc",
    "molecule",
    "product_name",
    "manufacturer",
    "seller",
    "pack_desc",
    "nhi_type",
    "class",
    "class_2",
    "dosage_form",
    "administration_route",
    "strength",
    "strength_raw",
    "strength_raw_2",
    "formulation",
    "funnel",
    "ox_gx",
    "molecule_disease_definition",
    "composition_type",
    "drug_extra_json",
    "raw_row_json",
    "column_metadata_json",
    "source_sheet",
    "source_file_version",
    "source_row_id",
    "ingested_at",
)

JSON_COLUMNS = ("drug_extra_json", "raw_row_json", "column_metadata_json")


@dataclass(frozen=True)
class MarketSheetConfig:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    source_type: str


@dataclass
class MarketDrugStats:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    max_row: int
    max_col: int
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    excluded_rows: int = 0
    staging_rows: int = 0


MARKET_SHEETS: tuple[MarketSheetConfig, ...] = (
    MarketSheetConfig("strategy_001", "라베칸 라베칸듀오", 5, "UBIST"),
    MarketSheetConfig("strategy_002", "제이클", 5, "IQVIA"),
    MarketSheetConfig("strategy_003", "가드렛 가드메트", 5, "IQVIA"),
    MarketSheetConfig("strategy_004", "타발리스", 5, "IQVIA"),
    MarketSheetConfig("strategy_005", "시그마트", 5, "UBIST"),
    MarketSheetConfig("strategy_006", "리바로 리바로젯", 4, "UBIST"),
    MarketSheetConfig("strategy_007", "리바로페노", 4, "UBIST"),
    MarketSheetConfig("strategy_008", "리바로하이 리바로브이", 5, "UBIST"),
    MarketSheetConfig("strategy_009", "트루패스 피나스타 제이다트", 5, "UBIST"),
    MarketSheetConfig("strategy_010", "뉴트로진 모빌리아", 5, "IQVIA"),
    MarketSheetConfig("strategy_011", "악템라", 5, "IQVIA"),
    MarketSheetConfig("strategy_012", "페린젝트 베노훼럼", 5, "IQVIA"),
    MarketSheetConfig("strategy_013", "헴리브라", 5, "IQVIA"),
    MarketSheetConfig("strategy_014", "위너프 위너프A+", 5, "IQVIA"),
    MarketSheetConfig("strategy_015", "엔커버", 7, "IQVIA"),
    MarketSheetConfig("strategy_016", "플라주오피", 5, "IQVIA"),
)

EXPECTED_MARKET_STATS = {
    # 260518 시트들은 일부 시장에서 실제 데이터 아래에 formatting tail이 남아
    # raw scan은 995행까지 이어진다. 이 표는 raw_rows_scanned/empty_rows를
    # 그대로 기록하되, staging_rows가 실제 약품 행수라는 불변량을 잡는다.
    # raw scan을 과거 짧은 row count에 맞추는 대안은 260518 Excel 형식을
    # 정상 입력으로 처리하지 못해 기각했다.
    "strategy_001": {"sheet_name": "라베칸 라베칸듀오", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 637, "excluded_rows": 0, "staging_rows": 358},
    "strategy_002": {"sheet_name": "제이클", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 950, "excluded_rows": 0, "staging_rows": 45},
    "strategy_003": {"sheet_name": "가드렛 가드메트", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 913, "excluded_rows": 0, "staging_rows": 82},
    "strategy_004": {"sheet_name": "타발리스", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 985, "excluded_rows": 0, "staging_rows": 10},
    "strategy_005": {"sheet_name": "시그마트", "header_row": 5, "raw_rows_scanned": 294, "empty_rows": 0, "excluded_rows": 0, "staging_rows": 294},
    "strategy_006": {"sheet_name": "리바로 리바로젯", "header_row": 4, "raw_rows_scanned": 1295, "empty_rows": 200, "excluded_rows": 48, "staging_rows": 1047},
    "strategy_007": {"sheet_name": "리바로페노", "header_row": 4, "raw_rows_scanned": 996, "empty_rows": 385, "excluded_rows": 494, "staging_rows": 117},
    "strategy_008": {"sheet_name": "리바로하이 리바로브이", "header_row": 5, "raw_rows_scanned": 1096, "empty_rows": 15, "excluded_rows": 0, "staging_rows": 1081},
    "strategy_009": {"sheet_name": "트루패스 피나스타 제이다트", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 589, "excluded_rows": 1, "staging_rows": 405},
    "strategy_010": {"sheet_name": "뉴트로진 모빌리아", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 985, "excluded_rows": 0, "staging_rows": 10},
    "strategy_011": {"sheet_name": "악템라", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 969, "excluded_rows": 0, "staging_rows": 26},
    "strategy_012": {"sheet_name": "페린젝트 베노훼럼", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 919, "excluded_rows": 0, "staging_rows": 76},
    "strategy_013": {"sheet_name": "헴리브라", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 981, "excluded_rows": 0, "staging_rows": 14},
    "strategy_014": {"sheet_name": "위너프 위너프A+", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 664, "excluded_rows": 0, "staging_rows": 331},
    "strategy_015": {"sheet_name": "엔커버", "header_row": 7, "raw_rows_scanned": 993, "empty_rows": 989, "excluded_rows": 0, "staging_rows": 4},
    "strategy_016": {"sheet_name": "플라주오피", "header_row": 5, "raw_rows_scanned": 995, "empty_rows": 943, "excluded_rows": 0, "staging_rows": 52},
}


def utc_now_text() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def resolve_input_file(path: Path) -> Path:
    if path.exists():
        # Match the current project loader path semantics: scripts/load_master.py
        # normally obtains this path from find_master_xlsx(), which returns the
        # actual filesystem directory entry. On macOS that preserves the
        # decomposed Korean filename in xlsx_path.name.
        for candidate in sorted(file for file in path.parent.glob("*.xlsx") if not file.name.startswith("~$")):
            try:
                if candidate.samefile(path):
                    return candidate
            except FileNotFoundError:
                continue
        return path
    candidates = sorted(file for file in MASTER_ROOT.glob("*.xlsx") if not file.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError(f"No Master xlsx found under {MASTER_ROOT}")
    return candidates[-1]


def normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=to_jsonable, sort_keys=True)


def is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _is_class_header(header: Any) -> bool:
    text = normalize_header(header)
    if not text:
        return False
    normalized = re.sub(r"[\s_-]+", "", text).lower()
    return normalized in {"class", "class1", "class2"} or normalized.startswith("class")


def _class_source_indexes(headers: list[Any] | tuple[Any, ...], metadata: dict[str, dict[str, Any]]) -> set[int]:
    indexes: set[int] = {idx for idx, header in enumerate(headers) if _is_class_header(header)}
    for target in ("class", "class_1", "class_2"):
        spec = metadata.get(target) or {}
        if spec.get("position") is not None:
            try:
                indexes.add(int(spec["position"]))
            except (TypeError, ValueError):
                pass
            continue
        source_column = normalize_header(spec.get("source_column"))
        if not source_column:
            continue
        for idx, header in enumerate(headers):
            text = normalize_header(header)
            if text and (text == source_column or text.startswith(source_column)):
                indexes.add(idx)
    return indexes


def is_excluded_row(
    values: list[Any] | tuple[Any, ...],
    class_indexes: set[int] | None = None,
    *,
    strategic_market_id: str | None = None,
    sheet_name: str | None = None,
) -> bool:
    class_indexes = set(class_indexes or ())
    row_excluded, _class_excluded = classify_exclusion_cells_by_policy(
        values,
        class_indexes=class_indexes,
        strategic_market_id=strategic_market_id,
        sheet_name=sheet_name,
    )
    return row_excluded


def make_header_keys(headers: list[Any] | tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    keys: list[str] = []
    for index, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        base = normalized if normalized is not None else f"__blank_col_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        keys.append(base if count == 1 else f"{base}__{count}")
    return keys


def build_raw_row_payload(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    source_row_id: int,
) -> dict[str, Any]:
    width = max(len(headers), len(values))
    padded_headers = list(headers) + [None] * (width - len(headers))
    padded_values = list(values) + [None] * (width - len(values))
    keys = make_header_keys(padded_headers)
    cells = []
    values_by_header: dict[str, Any] = {}
    for index, (header, key, value) in enumerate(zip(padded_headers, keys, padded_values), start=1):
        json_value = to_jsonable(value)
        cell = {
            "column_index": index,
            "header": normalize_header(header),
            "header_key": key,
            "value": json_value,
        }
        cells.append(cell)
        values_by_header[key] = json_value
    return {
        "source_row_id": source_row_id,
        "cells": cells,
        "values_by_header": values_by_header,
    }


def _header_lookup(headers: list[Any] | tuple[Any, ...], values: list[Any] | tuple[Any, ...]) -> dict[str, Any]:
    lookup: dict[str, Any] = {}
    for header, value in zip(headers, values):
        normalized = normalize_header(header)
        if normalized is not None and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def _lookup_source_value(lookup: dict[str, Any], source_column: str | None) -> Any:
    if source_column is None:
        return None
    if source_column in lookup:
        return lookup[source_column]
    for header, value in lookup.items():
        if header.startswith(source_column):
            return value
    return None


def _lookup_position_value(values: list[Any] | tuple[Any, ...], position: Any) -> Any:
    """Return a 0-based physical column value for blank-header source columns."""
    if position is None:
        return None
    try:
        column_index = int(position)
    except (TypeError, ValueError):
        raise ValueError(f"invalid catalog position: {position!r}") from None
    if column_index < 0:
        raise ValueError(f"catalog position must be >= 0: {position!r}")
    if column_index >= len(values):
        return None
    return values[column_index]


def _extra_key(metadata_key: str) -> str:
    return metadata_key[len(STANDARD_PREFIX) :]


def apply_column_mapping(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    metadata: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    lookup = _header_lookup(headers, values)
    standard: dict[str, Any] = {}
    extras: dict[str, Any] = {}
    for target_column, spec in metadata.items():
        source_column = normalize_header(spec.get("source_column"))
        if "position" in spec and spec.get("position") is not None:
            value = _lookup_position_value(values, spec.get("position"))
        else:
            value = _lookup_source_value(lookup, source_column)
        if target_column.startswith(STANDARD_PREFIX):
            extras[_extra_key(target_column)] = to_jsonable(value)
        else:
            standard[target_column] = value
    return standard, extras


def _position_value(values: list[Any] | tuple[Any, ...], column_index: int) -> Any:
    if column_index <= 0 or column_index > len(values):
        return None
    return values[column_index - 1]


def _lookup_key(*values: Any) -> tuple[str, ...] | None:
    key: list[str] = []
    for value in values:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        key.append(text)
    return tuple(key)


def _single_lookup_key(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def explicit_lookup_join(data_rows: list[tuple[int, tuple[Any, ...]]]) -> dict[int, dict[str, Any]]:
    """Return strategy_008 lookup-derived standard column overwrites."""
    lookup1: dict[tuple[str, str], dict[str, Any]] = {}
    for _, values in data_rows:
        key = _lookup_key(_position_value(values, 17), _position_value(values, 18))
        if key and key not in lookup1:
            lookup1[key] = {
                "molecule": _position_value(values, 19),
                "molecule_disease_definition": _position_value(values, 20),
                "composition_type": _position_value(values, 21),
                "class": _position_value(values, 22),
            }

    lookup2: dict[str, Any] = {}
    for _, values in data_rows:
        key = _single_lookup_key(_position_value(values, 25))
        if key and key not in lookup2:
            lookup2[key] = _position_value(values, 26)

    overrides: dict[int, dict[str, Any]] = {}
    for source_row_id, values in data_rows:
        left_key = _lookup_key(_position_value(values, 2), _position_value(values, 3))
        if not left_key or left_key not in lookup1:
            continue
        row_override = dict(lookup1[left_key])
        molecule_key = _single_lookup_key(row_override.get("molecule"))
        row_override["class_2"] = lookup2.get(molecule_key) if molecule_key else None
        overrides[source_row_id] = row_override
    return overrides


def load_column_metadata_catalog(path: Path) -> dict[str, dict[str, dict[str, Any]]]:
    text = path.read_text(encoding="utf-8")
    catalog: dict[str, dict[str, dict[str, Any]]] = {}
    current_market: str | None = None
    lines = iter(text.splitlines())
    for line in lines:
        heading = re.match(r"^###\s+(strategy_\d{3})\s+—", line)
        if heading:
            current_market = heading.group(1)
            continue
        if current_market and line.strip() == "```json":
            block: list[str] = []
            for json_line in lines:
                if json_line.strip() == "```":
                    break
                block.append(json_line)
            catalog[current_market] = json.loads("\n".join(block))
            current_market = None
    missing = sorted({config.strategic_market_id for config in MARKET_SHEETS} - set(catalog))
    if missing:
        raise ValueError(f"column metadata catalog missing markets: {missing}")
    return catalog


def _headers_from_sheet(ws: Any, header_row: int) -> list[Any]:
    return list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))


def _blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_DRUG_COLUMNS}


def load_drug_records(
    xlsx_path: Path,
    catalog_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], list[MarketDrugStats]]:
    metadata_catalog = load_column_metadata_catalog(catalog_path)
    timestamp = ingested_at or utc_now_text()
    records: list[dict[str, Any]] = []
    stats: list[MarketDrugStats] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for config in MARKET_SHEETS:
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}; sheets={wb.sheetnames}")
            ws = wb[config.sheet_name]
            headers = _headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            market_stats = MarketDrugStats(
                strategic_market_id=config.strategic_market_id,
                sheet_name=config.sheet_name,
                header_row=config.header_row,
                max_row=ws.max_row,
                max_col=ws.max_column,
            )
            drug_index = 0

            row_items = list(enumerate(
                ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                start=config.header_row + 1,
            ))
            explicit_overrides = (
                explicit_lookup_join(row_items) if config.strategic_market_id == "strategy_008" else {}
            )
            class_indexes = _class_source_indexes(headers, metadata)

            for source_row_id, values in row_items:
                market_stats.raw_rows_scanned += 1
                if is_empty_row(values):
                    market_stats.empty_rows += 1
                    continue
                raw_payload = build_raw_row_payload(headers, values, source_row_id)
                if is_excluded_row(
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                ):
                    market_stats.excluded_rows += 1
                    continue

                standard_values, extras = apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                drug_index += 1
                market_stats.staging_rows += 1

                record = _blank_record()
                record.update(
                    {
                        "strategic_market_id": config.strategic_market_id,
                        "market_name": config.sheet_name,
                        "source_type": config.source_type,
                        "drug_index": drug_index,
                        "drug_extra_json": dumps_json(extras),
                        "raw_row_json": dumps_json(raw_payload),
                        "column_metadata_json": dumps_json(metadata),
                        "source_sheet": config.sheet_name,
                        "source_file_version": xlsx_path.name,
                        "source_row_id": source_row_id,
                        "ingested_at": timestamp,
                    }
                )
                for key, value in standard_values.items():
                    if key in record:
                        record[key] = value
                records.append(record)
            stats.append(market_stats)
    finally:
        wb.close()

    return records, stats


def _count_by(records: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        value = str(record.get(key) or "")
        counts[value] = counts.get(value, 0) + 1
    return dict(sorted(counts.items()))


def _expected_extra_keys(metadata: dict[str, dict[str, Any]]) -> list[str]:
    return sorted(key.split(".", 1)[1] for key in metadata if key.startswith(STANDARD_PREFIX))


def validate_records(
    records: list[dict[str, Any]],
    stats: list[MarketDrugStats],
    catalog_path: Path,
) -> None:
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"master_drug row count must be {EXPECTED_ROW_COUNT}, found {len(records)}")

    expected_columns = set(MASTER_DRUG_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}")

    stats_by_market = {item.strategic_market_id: item for item in stats}
    expected_market_ids = {config.strategic_market_id for config in MARKET_SHEETS}
    if set(stats_by_market) != expected_market_ids:
        raise ValueError(
            f"market stats mismatch: expected={sorted(expected_market_ids)}, actual={sorted(stats_by_market)}"
        )

    total_excluded = 0
    for market_id, expected in EXPECTED_MARKET_STATS.items():
        actual = stats_by_market[market_id]
        for field in ("sheet_name", "header_row", "raw_rows_scanned", "empty_rows", "excluded_rows", "staging_rows"):
            actual_value = getattr(actual, field)
            expected_value = expected[field]
            if actual_value != expected_value:
                raise ValueError(f"{market_id} {field} mismatch: expected={expected_value}, actual={actual_value}")
        total_excluded += actual.excluded_rows
    if total_excluded != EXPECTED_EXCLUDED_ROWS:
        raise ValueError(f"excluded rows must be {EXPECTED_EXCLUDED_ROWS}, found {total_excluded}")

    pk_values = [(record["strategic_market_id"], str(record["drug_index"])) for record in records]
    if len(set(pk_values)) != EXPECTED_ROW_COUNT:
        duplicate_keys = sorted({value for value in pk_values if pk_values.count(value) > 1})
        raise ValueError(f"compound PK must be unique, duplicate examples={duplicate_keys[:10]}")

    market_distribution = _count_by(records, "strategic_market_id")
    expected_market_distribution = {
        market_id: expected["staging_rows"] for market_id, expected in EXPECTED_MARKET_STATS.items()
    }
    if market_distribution != expected_market_distribution:
        raise ValueError(
            f"market distribution mismatch: expected={expected_market_distribution}, actual={market_distribution}"
        )

    source_type_distribution = _count_by(records, "source_type")
    if source_type_distribution != EXPECTED_SOURCE_TYPE_DISTRIBUTION:
        raise ValueError(
            f"source_type distribution mismatch: expected={EXPECTED_SOURCE_TYPE_DISTRIBUTION}, "
            f"actual={source_type_distribution}"
        )

    metadata_catalog = load_column_metadata_catalog(catalog_path)
    records_by_market: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        records_by_market[record["strategic_market_id"]].append(record)

    for config in MARKET_SHEETS:
        market_records = records_by_market[config.strategic_market_id]
        expected_count = EXPECTED_MARKET_STATS[config.strategic_market_id]["staging_rows"]
        drug_indexes = [int(record["drug_index"]) for record in market_records]
        if drug_indexes != list(range(1, expected_count + 1)):
            raise ValueError(
                f"{config.strategic_market_id} drug_index sequence mismatch: "
                f"expected=1..{expected_count}, actual_first_last={drug_indexes[:3]}...{drug_indexes[-3:]}"
            )

        metadata = metadata_catalog[config.strategic_market_id]
        expected_metadata_json = dumps_json(metadata)
        expected_extra_keys = _expected_extra_keys(metadata)
        for record in market_records:
            for column in JSON_COLUMNS:
                try:
                    json.loads(record[column])
                except Exception as exc:
                    raise ValueError(f"{config.strategic_market_id} {column} is not valid JSON") from exc

            extra = json.loads(record["drug_extra_json"])
            raw_payload = json.loads(record["raw_row_json"])
            metadata_json = json.loads(record["column_metadata_json"])

            if sorted(extra.keys()) != expected_extra_keys:
                raise ValueError(
                    f"{config.strategic_market_id} drug_extra_json keys mismatch: "
                    f"expected={expected_extra_keys}, actual={sorted(extra.keys())}"
                )
            if record["column_metadata_json"] != expected_metadata_json:
                raise ValueError(f"{config.strategic_market_id} column_metadata_json string mismatch")
            if metadata_json != metadata:
                raise ValueError(f"{config.strategic_market_id} column_metadata_json structure mismatch")
            # 시트별 컬럼 수가 달라져도 raw_row_json의 cells와 values_by_header가
            # 같은 폭을 유지하는지만 검증한다. 26칸 고정 검사는 260518의 컬럼
            # 배치 변화를 버그로 오판하므로 기각했다.
            expected_cell_count = len(raw_payload.get("values_by_header", {}))
            if len(raw_payload.get("cells", [])) != expected_cell_count:
                raise ValueError(
                    f"{config.strategic_market_id} raw_row_json cells length mismatch: "
                    f"expected={expected_cell_count}, actual={len(raw_payload.get('cells', []))}"
                )
            if int(record["source_row_id"]) != int(raw_payload.get("source_row_id")):
                raise ValueError(f"{config.strategic_market_id} source_row_id mismatch in raw_row_json")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    schema = pa.schema([pa.field(column, pa.string()) for column in MASTER_DRUG_COLUMNS])
    rows = [
        {column: None if record[column] is None else str(record[column]) for column in MASTER_DRUG_COLUMNS}
        for record in records
    ]
    table = pa.Table.from_pylist(rows, schema=schema)
    pq.write_table(table, output_file, compression="zstd")


def print_summary(records: list[dict[str, Any]], stats: list[MarketDrugStats], output_file: Path) -> None:
    print("Phase 09e master_drug load")
    print(f"master_drug_rows: {len(records)}")
    print(f"compound_pk_unique: {len({(record['strategic_market_id'], str(record['drug_index'])) for record in records})}")
    print(f"source_type_distribution: {_count_by(records, 'source_type')}")
    print("market_distribution:")
    for item in stats:
        print(
            f"  {item.strategic_market_id} {item.sheet_name}: "
            f"raw={item.raw_rows_scanned}, empty={item.empty_rows}, "
            f"excluded={item.excluded_rows}, staging={item.staging_rows}"
        )
    if records:
        print(f"ingested_at: {records[0]['ingested_at']}")
    if output_file.exists():
        print(f"output_file: {output_file} ({output_file.stat().st_size:,} bytes)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load MI Master drug parquet.")
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--catalog-path", type=Path, default=DEFAULT_CATALOG_PATH)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = resolve_input_file(args.input_file)
    records, stats = load_drug_records(xlsx_path, args.catalog_path)
    validate_records(records, stats, args.catalog_path)
    write_parquet(records, args.output_file)
    print_summary(records, stats, args.output_file)
    print("validate_records: PASS")


if __name__ == "__main__":
    main()
