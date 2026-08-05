from __future__ import annotations

from collections import Counter, defaultdict
import json
from pathlib import Path
from typing import Any, Sequence

import openpyxl

from pipeline.scripts.analysis.brand_activity.recheck.inventory import FileRecord
from pipeline.scripts.etl.brand_activity.csd_core import discover_market_sheets, normalize_text


JsonObject = dict[str, Any]


def read_json(path: Path) -> JsonObject:
    """Read one JSON object from a UTF-8 audit artifact."""
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    """Write deterministic JSON audit evidence."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def csd_header_summary(records: Sequence[FileRecord]) -> JsonObject:
    """Verify CSD market sheet naming, header identity, and first data row shape."""
    headers: dict[str, list[str]] = {}
    sheet_counts: dict[str, int] = {}
    market2: dict[str, list[str]] = {}
    first_data_empty: list[dict[str, str]] = []
    for record in records:
        if record.kind != "csd":
            continue
        workbook = openpyxl.load_workbook(record.path, read_only=True, data_only=True)
        try:
            market_sheets = discover_market_sheets(record.path)
            sheet_counts[record.file_name] = len(market_sheets)
            market2[record.file_name] = [name for name in workbook.sheetnames if name.endswith("Market2")]
            for sheet_name in market_sheets:
                sheet = workbook[sheet_name]
                header = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))
                headers[f"{record.file_name}::{sheet_name}"] = [normalize_text(value) for value in header[:8]]
                first_data = next(sheet.iter_rows(min_row=8, max_row=8, values_only=True))
                if not any(normalize_text(value) for value in first_data):
                    first_data_empty.append({"file": record.file_name, "sheet": sheet_name})
        finally:
            workbook.close()
    distinct_headers = {tuple(value) for value in headers.values()}
    return {
        "market_sheet_counts": sheet_counts,
        "files_with_market2": {file: names for file, names in market2.items() if names},
        "checked_market_sheets": len(headers),
        "distinct_header_rows": len(distinct_headers),
        "header_rows_byte_identical": len(distinct_headers) == 1,
        "first_data_row_empty": first_data_empty,
        "header_sample": next(iter(headers.values()), []),
    }


def selected_root_by_kind(records: Sequence[FileRecord]) -> dict[str, str]:
    """Pick the parent folder with the most files for each source kind."""
    counts: dict[str, Counter[Path]] = defaultdict(Counter)
    for record in records:
        counts[record.kind][record.path.parent] += 1
    return {
        kind: str(counter.most_common(1)[0][0])
        for kind, counter in counts.items()
        if kind in {"csd", "keyword", "meeting"} and counter
    }


def duplicate_file_names(records: Sequence[FileRecord]) -> list[str]:
    """Find same-named files across scanned roots that could make SHA comparison ambiguous."""
    counts = Counter(record.file_name for record in records)
    return sorted(file_name for file_name, count in counts.items() if count > 1)


def one_month_file_violations(file_period_distribution: dict[str, dict[str, int]]) -> dict[str, list[str]]:
    """Detect Keyword/Meeting source files whose core sheet contains multiple months."""
    return {
        file_name: sorted(period_counts)
        for file_name, period_counts in file_period_distribution.items()
        if len(period_counts) != 1
    }


def new_enum_values(previous: JsonObject, current: JsonObject) -> JsonObject:
    """Compare current enum distributions to the previous validation artifact."""
    old = previous.get("enum_distribution", {})
    new = current.get("enum_distribution", {})
    result: JsonObject = {}
    for event_kind, fields in new.items():
        result[event_kind] = {}
        for field, current_values in fields.items():
            previous_values = old.get(event_kind, {}).get(field, {})
            additions = sorted(set(current_values) - set(previous_values))
            if additions:
                result[event_kind][field] = additions
    return result


def top_class_counts(class_month_summary: JsonObject, event_kind: str, limit: int = 10) -> list[JsonObject]:
    """Aggregate ATC4 row counts across months for topic follow-up sizing."""
    counts: Counter[str] = Counter()
    for row in class_month_summary.get(event_kind, []):
        counts[str(row["therapeutic_class"])] += int(row["rows"])
    return [
        {"therapeutic_class": class_name, "rows": rows}
        for class_name, rows in counts.most_common(limit)
    ]


def product_variant_hits(products: Sequence[str]) -> list[str]:
    """Surface the known LOWOSMOPERI versus LOW OSMO PERI normalization decision item."""
    variants = []
    for product in sorted(set(products)):
        compact = product.replace(" ", "").upper()
        if "LOWOSMOPERI" in compact:
            variants.append(product)
    return variants


def broken_csd_assumptions(header: JsonObject, validation: JsonObject) -> list[str]:
    """List CSD assumptions that failed under the current source set."""
    broken: list[str] = []
    if not header["header_rows_byte_identical"]:
        broken.append("CSD Market sheet header rows are not byte-identical.")
    if len(set(header["market_sheet_counts"].values())) > 1:
        broken.append("CSD Market sheet count differs by source file.")
    if header["first_data_row_empty"]:
        broken.append("CSD first data row check found an empty row after the header.")
    if validation["dedup_report"]["conflict_groups"] != 0:
        broken.append("CSD rolling-file dedup found conflicting product_details values.")
    if validation["enum_and_measure_checks"]["invalid_channels"]:
        broken.append("CSD JW Channel enum contains unexpected values.")
    if validation["enum_and_measure_checks"]["bad_measure_rows"] != 0:
        broken.append("CSD Product Details contains blank or non-integer values.")
    checks = validation["aggregate_cross_checks"]
    if checks["sample_count"] != checks["match_count"]:
        broken.append("CSD aggregate cross-check samples did not all match TOTAL-region raw rows.")
    return broken


def broken_km_assumptions(validation: JsonObject, enum_additions: JsonObject) -> list[str]:
    """List Keyword/Meeting assumptions that failed under the current source set."""
    broken: list[str] = []
    keyword_months = one_month_file_violations(validation["file_period_distribution"]["keyword"])
    meeting_months = one_month_file_violations(validation["file_period_distribution"]["meeting"])
    if keyword_months:
        broken.append(f"Keyword source file contains multiple core months: {keyword_months}")
    if meeting_months:
        broken.append(f"Meeting source file contains multiple core months: {meeting_months}")
    if validation["message_count_overlap"]["keyword"]["mismatch_cells"] != 0:
        broken.append("Keyword Message Count overlap has value conflicts.")
    if validation["message_count_overlap"]["meeting"]["mismatch_cells"] != 0:
        broken.append("Meeting Message Count overlap has value conflicts.")
    non_empty = {kind: fields for kind, fields in enum_additions.items() if fields}
    if non_empty:
        broken.append(f"Keyword/Meeting enum distribution has new values: {non_empty}")
    if validation.get("db_load") == "skipped":
        broken.append("Keyword/Meeting isolated DB reload was skipped.")
    return broken


def input_completeness_failures(missing_roots: Sequence[str]) -> list[str]:
    """Promote missing required scan roots into the final risk ledger."""
    return [f"Required source scan root missing: {root}" for root in missing_roots]


def table_delta_rows(after: JsonObject, previous_rows: dict[str, int]) -> list[list[object]]:
    """Build previous-vs-current row-count table rows."""
    return [
        [
            kind,
            previous_rows[kind],
            after[kind]["rows"],
            after[kind]["rows"] - previous_rows[kind],
            f"{after[kind].get('period_min', '')}~{after[kind].get('period_max', '')}",
        ]
        for kind in ("csd", "keyword", "meeting")
    ]


def baseline_from_previous_artifacts(csd_path: Path, km_path: Path) -> JsonObject:
    """Build the previous-load baseline from immutable prior audit artifacts."""
    csd = read_json(csd_path)
    km = read_json(km_path)
    csd_periods = [row["period_min"] for row in csd["market_summary"]] + [row["period_max"] for row in csd["market_summary"]]
    keyword_periods = sorted(km["core_period_distribution"]["keyword"])
    meeting_periods = sorted(km["core_period_distribution"]["meeting"])
    return {
        "csd": {
            "rows": csd["stage_rows_after_dedup"],
            "period_min": min(csd_periods),
            "period_max": max(csd_periods),
            "source": str(csd_path),
        },
        "keyword": {
            "rows": km["core_rows"]["keyword"],
            "period_min": keyword_periods[0],
            "period_max": keyword_periods[-1],
            "source": str(km_path),
        },
        "meeting": {
            "rows": km["core_rows"]["meeting"],
            "period_min": meeting_periods[0],
            "period_max": meeting_periods[-1],
            "source": str(km_path),
        },
    }
