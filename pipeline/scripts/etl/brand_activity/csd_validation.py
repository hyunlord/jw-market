from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

import openpyxl

from pipeline.scripts.etl.brand_activity.csd_core import (
    CsdRow,
    MarketSheetScan,
    is_total_region,
    normalize_text,
    parse_period_ym,
    parse_product_details,
    discover_market_sheets,
)


def market_summary(rows: list[CsdRow]) -> list[dict[str, str | int]]:
    grouped: defaultdict[str, list[CsdRow]] = defaultdict(list)
    for row in rows:
        grouped[row.market].append(row)
    summary = []
    for market, market_rows in sorted(grouped.items()):
        periods = sorted({row.period_ym for row in market_rows})
        summary.append(
            {
                "market": market,
                "rows": len(market_rows),
                "period_min": periods[0],
                "period_max": periods[-1],
                "product_details": sum(row.product_details for row in market_rows),
                "unique_master_products": len({row.master_product for row in market_rows}),
                "unique_representing_companies": len({row.representing_company for row in market_rows}),
            }
        )
    return summary


def aggregate_cross_checks(workbooks: list[Path], rows: list[CsdRow], limit: int = 24) -> list[dict[str, str | int | bool | None]]:
    by_source: defaultdict[tuple[str, str, str, str, str, str], int] = defaultdict(int)
    for row in rows:
        channel = "GH+SHPPI+CPPI" if row.jw_channel == "TOTAL" else row.jw_channel
        by_source[(row.source_file, row.period_ym, channel, row.representing_company, row.master_product, row.market)] += row.product_details
    checks: list[dict[str, str | int | bool | None]] = []
    sheet_names = ["TOTAL-Product Detail Monthly", "GH-Product Detail Monthly", "SHPPI-Product Detail Monthly", "CPPI-Product Detail Monthly"]
    for workbook_path in workbooks:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet_name in sheet_names:
                if sheet_name not in workbook.sheetnames:
                    continue
                checks.extend(_aggregate_sheet_checks(workbook_path, workbook[sheet_name], sheet_name, by_source, limit - len(checks)))
                if len(checks) >= limit:
                    return checks
        finally:
            workbook.close()
    return checks


def _aggregate_sheet_checks(workbook_path: Path, sheet: object, sheet_name: str, by_source: dict[tuple[str, str, str, str, str, str], int], limit: int) -> list[dict[str, str | int | bool | None]]:
    header = [normalize_text(value) for value in next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))]
    month_indexes = [(index, parse_period_ym(value)) for index, value in enumerate(header) if value and index >= 5]
    current_channel = ""
    current_company = ""
    current_region = "TOTAL"
    checks: list[dict[str, str | int | bool | None]] = []
    for values in sheet.iter_rows(min_row=8, values_only=True):
        labels = [normalize_text(value) for value in values[:5]]
        current_channel = labels[0] or current_channel
        current_company = labels[1] or current_company
        if "Region" in header:
            region_index = header.index("Region")
            current_region = normalize_text(values[region_index]) or current_region
        product = labels[4] or labels[3]
        if not product or not is_total_region(current_region):
            continue
        for month_index, period in month_indexes:
            aggregate_value = values[month_index]
            if aggregate_value is None:
                continue
            matches = [key for key in by_source if key[:5] == (workbook_path.name, period, current_channel, current_company, product)]
            raw_value = sum(by_source[key] for key in matches)
            if raw_value == 0:
                continue
            checks.append(
                {
                    "source_file": workbook_path.name,
                    "aggregate_sheet": sheet_name,
                    "period_ym": period,
                    "jw_channel": current_channel,
                    "representing_company": current_company,
                    "master_product": product,
                    "aggregate_value": int(aggregate_value),
                    "raw_total_region_value": raw_value,
                    "matches": int(aggregate_value) == raw_value,
                }
            )
            if len(checks) >= limit:
                return checks
    return checks


def hierarchy_samples(workbooks: list[Path], limit: int = 10) -> list[dict[str, str | int]]:
    samples: list[dict[str, str | int]] = []
    for workbook_path in workbooks:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet_name in discover_market_sheets(workbook_path):
                samples.extend(_hierarchy_sheet_samples(workbook_path, workbook[sheet_name], limit - len(samples)))
                if len(samples) >= limit:
                    return samples
        finally:
            workbook.close()
    return samples


def _hierarchy_sheet_samples(workbook_path: Path, sheet: object, limit: int) -> list[dict[str, str | int]]:
    header = [normalize_text(value) for value in next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))]
    indexes = {name: header.index(name) for name in header if name}
    grouped: defaultdict[tuple[str, str, str, str, str], dict[str, int]] = defaultdict(dict)
    for values in sheet.iter_rows(min_row=8, values_only=True):
        if not any(normalize_text(value) for value in values):
            continue
        region = normalize_text(values[indexes["Region"]])
        try:
            amount = parse_product_details(values[indexes["Product Details"]])
        except ValueError:
            continue
        key = (
            parse_period_ym(values[indexes["Related date"]]),
            normalize_text(values[indexes["Market"]]),
            normalize_text(values[indexes["JW Channel"]]),
            normalize_text(values[indexes["Master product"]]),
            normalize_text(values[indexes["Representing Company"]]),
        )
        grouped[key][region] = amount
    return _overcount_samples(workbook_path.name, grouped, limit)


def _overcount_samples(source_file: str, grouped: dict[tuple[str, str, str, str, str], dict[str, int]], limit: int) -> list[dict[str, str | int]]:
    samples: list[dict[str, str | int]] = []
    for key, region_values in grouped.items():
        total = region_values.get("TOTAL")
        if total is None:
            continue
        non_total_sum = sum(value for region, value in region_values.items() if region != "TOTAL")
        if non_total_sum <= total:
            continue
        period, market, channel, product, company = key
        samples.append(
            {
                "source_file": source_file,
                "period_ym": period,
                "market": market,
                "jw_channel": channel,
                "master_product": product,
                "representing_company": company,
                "total_region_value": total,
                "non_total_region_sum": non_total_sum,
                "overcount_multiple_x100": round(non_total_sum / total * 100) if total else 0,
            }
        )
        if len(samples) >= limit:
            return samples
    return samples


def build_validation(workbooks: list[Path], raw_rows: list[CsdRow], scans: list[MarketSheetScan], deduped: list[CsdRow], dedup_report: dict[str, int], missing_months: list[str], ignored_files: list[str], sheet_map: dict[str, list[str]]) -> dict[str, object]:
    aggregate_checks = aggregate_cross_checks(workbooks, raw_rows)
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_files": [path.name for path in workbooks],
        "missing_expected_months": missing_months,
        "ignored_files": ignored_files,
        "sheet_map": sheet_map,
        "raw_total_region_rows_before_dedup": len(raw_rows),
        "dedup_report": dedup_report,
        "stage_rows_after_dedup": len(deduped),
        "stage_product_details_after_dedup": sum(row.product_details for row in deduped),
        "market_summary": market_summary(deduped),
        "region_filter": {
            "raw_market_rows": sum(scan.rows_raw for scan in scans),
            "total_region_rows": sum(scan.rows_total_region for scan in scans),
            "excluded_non_total_rows": sum(scan.rows_raw - scan.rows_total_region for scan in scans),
            "total_region_product_details_sum": sum(scan.product_details_total_region for scan in scans),
            "duplicate_grains_after_total_filter": sum(scan.duplicate_grains_after_total_filter for scan in scans),
        },
        "enum_and_measure_checks": {
            "invalid_channels": {key: value for scan in scans for key, value in scan.invalid_channels.items()},
            "bad_measure_rows": sum(scan.null_or_bad_measure_rows for scan in scans),
        },
        "sheet_scans": [asdict(scan) for scan in scans],
        "hierarchy_overcount_samples": hierarchy_samples(workbooks),
        "aggregate_cross_checks": {
            "sample_count": len(aggregate_checks),
            "match_count": sum(1 for check in aggregate_checks if check["matches"]),
            "checks": aggregate_checks,
        },
    }
