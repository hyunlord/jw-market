"""Message Count sheet parsing for Keyword workbooks."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from pipeline.scripts.etl.brand_activity.km_core import (
    JsonValue,
    KmParseError,
    MESSAGE_MONTH_PATTERN,
    normalize_key,
    normalize_spaces,
    normalize_text,
    parse_count_value,
    parse_period_ym,
    source_period_from_name,
)

if TYPE_CHECKING:
    from pipeline.scripts.etl.brand_activity.km_core import WorksheetLike


@dataclass(frozen=True, slots=True)
class MessageCountCell:
    """Single product-month value from a cumulative `2025 Message Count` sheet."""

    kind: str
    source_file: str
    source_period_ym: str
    product_name: str
    product_key: str
    month_ym: str
    value: int

    def comparison_key(self) -> tuple[str, str]:
        """Return the stable product/month key used across source files."""
        return (self.product_key, self.month_ym)

    def to_dict(self) -> dict[str, JsonValue]:
        """Serialize the non-sensitive count cell for JSON audit output."""
        return asdict(self)


def message_count_header(sheet: "WorksheetLike") -> tuple[int, int, dict[int, str]]:
    """Locate the product column and normalized month columns in Message Count."""
    for row_no in range(1, 10):
        row = next(sheet.iter_rows(min_row=row_no, max_row=row_no, max_col=80, values_only=True))
        headers = [normalize_spaces(normalize_text(value)) for value in row]
        normalized = [normalize_key(header) for header in headers]
        product_index = _product_index(normalized)
        if product_index < 0:
            continue
        month_columns = _month_columns(headers)
        if not month_columns:
            raise KmParseError("Message Count has a product header but no month columns")
        return row_no, product_index, month_columns
    raise KmParseError("Message Count product header not found")


def read_message_count_cells(workbook_path: Path, kind: str) -> list[MessageCountCell]:
    """Read product-month count cells from a workbook's `2025 Message Count` sheet."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "2025 Message Count" not in workbook.sheetnames:
            return []
        sheet = workbook["2025 Message Count"]
        header_row_no, product_index, month_columns = message_count_header(sheet)
        source_period = source_period_from_name(workbook_path)
        cells: list[MessageCountCell] = []
        max_col = max([product_index, *month_columns.keys()]) + 1
        for row in sheet.iter_rows(min_row=header_row_no + 1, max_col=max_col, values_only=True):
            product_name = normalize_spaces(normalize_text(row[product_index]))
            if product_name == "":
                continue
            product_key = normalize_key(product_name)
            for column_index, month_ym in month_columns.items():
                cells.append(
                    MessageCountCell(
                        kind=kind,
                        source_file=workbook_path.name,
                        source_period_ym=source_period,
                        product_name=product_name,
                        product_key=product_key,
                        month_ym=month_ym,
                        value=parse_count_value(row[column_index]),
                    )
                )
        return cells
    finally:
        workbook.close()


def _product_index(normalized_headers: list[str]) -> int:
    """Return the first supported product header index or -1."""
    product_name = normalize_key("PRODUCT NAME")
    product = normalize_key("PRODUCT")
    if product_name in normalized_headers:
        return normalized_headers.index(product_name)
    if product in normalized_headers:
        return normalized_headers.index(product)
    return -1


def _month_columns(headers: list[str]) -> dict[int, str]:
    """Return normalized month columns from a Message Count header row."""
    return {
        index: parse_period_ym(header)
        for index, header in enumerate(headers)
        if MESSAGE_MONTH_PATTERN.search(header) is not None
    }
