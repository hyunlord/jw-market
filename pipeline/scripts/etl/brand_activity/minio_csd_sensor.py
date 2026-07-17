#!/usr/bin/env python3
"""MinIO polling sensor for CSD arrivals (event-driven trigger, mode B).

Watches ``jw-market-raw-iqvia/CSD/ChannelDynamics*`` for new ``.xlsx`` objects,
runs the structure-validation gate, and only then creates the brand-activity
run Job. Design contract (event-driven round, 2026-07-17):

* The sensor only WAKES the pipeline; ingest/topic/row-topic own their own
  gates. A spurious wake is a no-op.
* Fail-closed: a workbook that fails structure validation NEVER triggers a
  run; the failure reason is recorded in the processed marker.
* Idempotent: objects are remembered by (key, etag) in a marker file on the
  shared state volume; re-listing the same object is a no-op.
* Mode A (bucket notification webhook) can later replace the polling loop:
  everything from validation down is unchanged (see stage0_survey.md - MinIO
  supports notify_webhook but enabling it is a platform-team change).

The S3 client is stdlib-only (SigV4) so the pipeline image needs no boto3.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import hmac
import json
import os
import subprocess
import sys
import tempfile
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Final

ROOT = Path(__file__).resolve().parents[4]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pipeline.scripts.etl.brand_activity.csd_core import (  # noqa: E402
    EXPECTED_HEADERS,
    normalize_text,
    select_market_sheets,
)

DEFAULT_BUCKET: Final = "jw-market-raw-iqvia"
DEFAULT_PREFIX: Final = "CSD/ChannelDynamics"
DEFAULT_MARKER_FILE: Final = "/var/lib/jw-pipeline/csd-sensor-marker.json"
DEFAULT_RUN_CRONJOB: Final = "jw-brand-activity-run"
DEFAULT_NAMESPACE: Final = "llmops"
MIN_OBJECT_BYTES: Final = 1024 * 1024  # real reports are ~5MiB; mac metadata is ~212B
EXCLUDED_BASENAME_PREFIXES: Final = ("._", "~$", ".DS_Store")

CommandRunner = Callable[[list[str]], tuple[int, str]]


# --------------------------------------------------------------------------
# Minimal SigV4 S3 client (list + get), stdlib only.
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class S3Config:
    endpoint: str  # e.g. http://minio.llmops.svc.cluster.local:9000
    access_key: str
    secret_key: str
    region: str = "us-east-1"

    @classmethod
    def from_env(cls) -> "S3Config":
        endpoint = os.environ.get("MINIO_ENDPOINT", "http://minio.llmops.svc.cluster.local:9000")
        access_key = os.environ.get("MINIO_ACCESS_KEY", "")
        secret_key = os.environ.get("MINIO_SECRET_KEY", "")
        if not access_key or not secret_key:
            raise SystemExit("MINIO_ACCESS_KEY/MINIO_SECRET_KEY are required")
        return cls(endpoint=endpoint.rstrip("/"), access_key=access_key, secret_key=secret_key)


def _sign(key: bytes, message: str) -> bytes:
    return hmac.new(key, message.encode(), hashlib.sha256).digest()


def _sigv4_headers(config: S3Config, method: str, canonical_uri: str, query: dict[str, str]) -> dict[str, str]:
    host = urllib.parse.urlparse(config.endpoint).netloc
    now = dt.datetime.now(dt.timezone.utc)
    amz_date = now.strftime("%Y%m%dT%H%M%SZ")
    date_stamp = now.strftime("%Y%m%d")
    payload_hash = hashlib.sha256(b"").hexdigest()
    canonical_query = "&".join(
        f"{urllib.parse.quote(k, safe='')}={urllib.parse.quote(v, safe='')}" for k, v in sorted(query.items())
    )
    canonical_headers = f"host:{host}\nx-amz-content-sha256:{payload_hash}\nx-amz-date:{amz_date}\n"
    signed_headers = "host;x-amz-content-sha256;x-amz-date"
    canonical_request = "\n".join(
        [method, canonical_uri, canonical_query, canonical_headers, signed_headers, payload_hash]
    )
    scope = f"{date_stamp}/{config.region}/s3/aws4_request"
    string_to_sign = "\n".join(
        ["AWS4-HMAC-SHA256", amz_date, scope, hashlib.sha256(canonical_request.encode()).hexdigest()]
    )
    signing_key = _sign(_sign(_sign(_sign(b"AWS4" + config.secret_key.encode(), date_stamp), config.region), "s3"), "aws4_request")
    signature = hmac.new(signing_key, string_to_sign.encode(), hashlib.sha256).hexdigest()
    return {
        "x-amz-date": amz_date,
        "x-amz-content-sha256": payload_hash,
        "Authorization": (
            f"AWS4-HMAC-SHA256 Credential={config.access_key}/{scope}, "
            f"SignedHeaders={signed_headers}, Signature={signature}"
        ),
    }


def _s3_request(config: S3Config, method: str, path: str, query: dict[str, str]) -> bytes:
    canonical_uri = urllib.parse.quote(path, safe="/")
    headers = _sigv4_headers(config, method, canonical_uri, query)
    query_string = urllib.parse.urlencode(sorted(query.items()))
    url = f"{config.endpoint}{canonical_uri}" + (f"?{query_string}" if query_string else "")
    request = urllib.request.Request(url, method=method, headers=headers)
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


@dataclass(frozen=True)
class ObjectInfo:
    key: str
    etag: str
    size: int


def list_objects(config: S3Config, bucket: str, prefix: str) -> list[ObjectInfo]:
    objects: list[ObjectInfo] = []
    token: str | None = None
    while True:
        query = {"list-type": "2", "prefix": prefix, "max-keys": "1000"}
        if token:
            query["continuation-token"] = token
        body = _s3_request(config, "GET", f"/{bucket}", query)
        namespace = {"s3": "http://s3.amazonaws.com/doc/2006-03-01/"}
        tree = ET.fromstring(body)
        for content in tree.findall("s3:Contents", namespace):
            objects.append(
                ObjectInfo(
                    key=content.findtext("s3:Key", "", namespace),
                    etag=content.findtext("s3:ETag", "", namespace).strip('"'),
                    size=int(content.findtext("s3:Size", "0", namespace)),
                )
            )
        if tree.findtext("s3:IsTruncated", "false", namespace) == "true":
            token = tree.findtext("s3:NextContinuationToken", None, namespace)
        else:
            return objects


def download_object(config: S3Config, bucket: str, key: str, target: Path) -> None:
    body = _s3_request(config, "GET", f"/{bucket}/{key}", {})
    target.write_bytes(body)


# --------------------------------------------------------------------------
# Candidate filter, validation gate, processed marker.
# --------------------------------------------------------------------------


def is_candidate(obj: ObjectInfo, *, min_bytes: int = MIN_OBJECT_BYTES) -> bool:
    basename = obj.key.rsplit("/", 1)[-1]
    if not basename.lower().endswith(".xlsx"):
        return False
    if any(basename.startswith(prefix) for prefix in EXCLUDED_BASENAME_PREFIXES):
        return False
    return obj.size >= min_bytes


def validate_csd_workbook(path: Path) -> tuple[bool, str]:
    """Structure gate: market sheets exist, expected headers on row 7, data rows exist.

    Reuses the csd_core sheet/header contract (jw_brand_activity_stage lineage)
    without paying for a full scan.
    """

    import openpyxl

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any unreadable workbook fails the gate
        return False, f"unreadable workbook: {exc}"
    try:
        market_sheets = select_market_sheets(tuple(workbook.sheetnames))
        if not market_sheets:
            return False, f"no market sheets among {list(workbook.sheetnames)[:8]}"
        for sheet_name in market_sheets:
            sheet = workbook[sheet_name]
            header = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True), ())
            normalized = {normalize_text(value) for value in header}
            missing = [column for column in EXPECTED_HEADERS if column not in normalized]
            if missing:
                return False, f"sheet {sheet_name!r} missing headers {missing}"
        first = workbook[market_sheets[0]]
        data_rows = sum(
            1 for values in first.iter_rows(min_row=8, max_row=200, values_only=True) if any(normalize_text(v) for v in values)
        )
        if data_rows == 0:
            return False, f"sheet {market_sheets[0]!r} has no data rows"
        return True, f"{len(market_sheets)} market sheets, headers ok, data rows present"
    finally:
        workbook.close()


def load_marker(path: Path) -> dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8") or "{}")
    return {}


def save_marker(path: Path, marker: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(marker, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    tmp.replace(path)


def _default_runner(argv: list[str]) -> tuple[int, str]:
    completed = subprocess.run(argv, capture_output=True, text=True, check=False)
    return completed.returncode, (completed.stdout or "") + (completed.stderr or "")


def trigger_run_job(key: str, etag: str, *, namespace: str, cronjob: str, runner: CommandRunner) -> tuple[str, str]:
    job_name = f"jw-ba-run-{hashlib.sha256(f'{key}:{etag}'.encode()).hexdigest()[:10]}"
    argv = ["kubectl", "-n", namespace, "create", "job", job_name, f"--from=cronjob/{cronjob}"]
    returncode, output = runner(argv)
    if returncode == 0:
        return job_name, "created"
    if "AlreadyExists" in output or "already exists" in output:
        return job_name, "noop_already_exists"
    return job_name, f"error: {output.strip()[:300]}"


def process_once(
    *,
    config: S3Config,
    bucket: str,
    prefix: str,
    marker_path: Path,
    namespace: str,
    cronjob: str,
    dry_run: bool,
    runner: CommandRunner | None = None,
    objects: list[ObjectInfo] | None = None,
    fetch: Callable[[ObjectInfo, Path], None] | None = None,
) -> list[dict]:
    """One polling pass. Returns per-object decisions (also printed as JSON lines)."""

    runner = runner or _default_runner
    fetch = fetch or (lambda obj, target: download_object(config, bucket, obj.key, target))
    marker = load_marker(marker_path)
    decisions: list[dict] = []
    listed = objects if objects is not None else list_objects(config, bucket, prefix)

    for obj in listed:
        decision: dict = {"key": obj.key, "etag": obj.etag, "size": obj.size}
        if not is_candidate(obj):
            continue  # metadata debris etc. - not even worth a decision line
        previous = marker.get(obj.key)
        if previous and previous.get("etag") == obj.etag:
            decision["action"] = "noop_already_processed"
            decisions.append(decision)
            continue

        with tempfile.TemporaryDirectory(prefix="csd_sensor_") as tmp:
            local = Path(tmp) / obj.key.rsplit("/", 1)[-1]
            fetch(obj, local)
            ok, detail = validate_csd_workbook(local)
        decision["validation"] = detail

        if not ok:
            decision["action"] = "blocked_validation_failed"
            if not dry_run:
                marker[obj.key] = {"etag": obj.etag, "status": "validation_failed", "detail": detail,
                                   "at": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")}
        elif dry_run:
            decision["action"] = "would_trigger"
        else:
            job_name, status = trigger_run_job(obj.key, obj.etag, namespace=namespace, cronjob=cronjob, runner=runner)
            decision["action"] = f"triggered:{status}"
            decision["job"] = job_name
            if status in ("created", "noop_already_exists"):
                marker[obj.key] = {"etag": obj.etag, "status": "triggered", "job": job_name,
                                   "at": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")}
        decisions.append(decision)

    if not dry_run:
        save_marker(marker_path, marker)
    for decision in decisions:
        print(json.dumps(decision, ensure_ascii=False), flush=True)
    return decisions


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bucket", default=os.environ.get("CSD_SENSOR_BUCKET", DEFAULT_BUCKET))
    parser.add_argument("--prefix", default=os.environ.get("CSD_SENSOR_PREFIX", DEFAULT_PREFIX))
    parser.add_argument("--marker-file", type=Path,
                        default=Path(os.environ.get("CSD_SENSOR_MARKER_FILE", DEFAULT_MARKER_FILE)))
    parser.add_argument("--namespace", default=os.environ.get("CSD_SENSOR_NAMESPACE", DEFAULT_NAMESPACE))
    parser.add_argument("--run-cronjob", default=os.environ.get("CSD_SENSOR_RUN_CRONJOB", DEFAULT_RUN_CRONJOB))
    parser.add_argument("--dry-run", action="store_true", help="Detect and validate only; no job, no marker write.")
    parser.add_argument(
        "--download-new-to",
        type=Path,
        help="Run-Job helper: download every candidate object under the prefix to this directory and exit "
        "(no validation, no trigger, no marker write; ingest_csd owns dedup/validation downstream).",
    )
    args = parser.parse_args(argv)

    config = S3Config.from_env()
    if args.download_new_to:
        args.download_new_to.mkdir(parents=True, exist_ok=True)
        for obj in list_objects(config, args.bucket, args.prefix):
            if is_candidate(obj):
                target = args.download_new_to / obj.key.rsplit("/", 1)[-1]
                download_object(config, args.bucket, obj.key, target)
                print(json.dumps({"downloaded": obj.key, "to": str(target)}, ensure_ascii=False), flush=True)
        return 0
    process_once(
        config=config,
        bucket=args.bucket,
        prefix=args.prefix,
        marker_path=args.marker_file,
        namespace=args.namespace,
        cronjob=args.run_cronjob,
        dry_run=args.dry_run,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
