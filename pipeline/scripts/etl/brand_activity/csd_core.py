from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Final

import openpyxl

from pipeline.etl.io.source_headers import normalize_source_header


EXPECTED_HEADERS: Final[tuple[str, ...]] = (
    "Related date",
    "Market",
    "JW Channel",
    "Region",
    "Master product",
    "Manufacturer",
    "Representing Company",
    "Product Details",
)
JW_CHANNELS: Final[frozenset[str]] = frozenset({"TOTAL", "GH", "SHPPI", "GH+SHPPI", "CPPI"})
MONTHS: Final[dict[str, int]] = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass(frozen=True, slots=True)
class CsdRow:
    source_file: str
    source_sheet: str
    source_row_no: int
    period_ym: str
    market: str
    jw_channel: str
    master_product: str
    representing_company: str
    product_details: int

    def grain_key(self) -> tuple[str, str, str, str, str]:
        return (
            self.period_ym,
            self.market,
            self.jw_channel,
            self.master_product,
            self.representing_company,
        )

    def to_dict(self) -> dict[str, str | int]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class MarketSheetScan:
    source_file: str
    source_sheet: str
    rows_raw: int
    rows_total_region: int
    product_details_total_region: int
    total_region_raw_sum: int
    duplicate_grains_after_total_filter: int
    regions: dict[str, int]
    invalid_channels: dict[str, int]
    null_or_bad_measure_rows: int


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def parse_period_ym(value: object) -> str:
    text = normalize_text(value)
    match = re.fullmatch(r"([A-Za-z]+)\.?\s*(\d{2}|\d{4})", text)
    if match is None:
        raise ValueError(f"unparseable CSD period: {value!r}")
    month_name = match.group(1).lower()
    month = MONTHS.get(month_name)
    if month is None:
        raise ValueError(f"unparseable CSD month: {value!r}")
    year_raw = int(match.group(2))
    year = 2000 + year_raw if year_raw < 100 else year_raw
    return f"{year}-{month:02d}"


def parse_product_details(value: object) -> int:
    text = normalize_text(value).replace(",", "")
    if not text:
        raise ValueError("missing Product Details")
    number = float(text)
    if not number.is_integer():
        raise ValueError(f"non-integer Product Details: {value!r}")
    return int(number)


def is_total_region(value: object) -> bool:
    return normalize_text(value).upper() == "TOTAL"


def select_market_sheets(sheet_names: list[str] | tuple[str, ...]) -> tuple[str, ...]:
    selected = []
    for sheet_name in sheet_names:
        if re.fullmatch(r"[A-Z].* Market", sheet_name) and not sheet_name.endswith("Market2"):
            selected.append(sheet_name)
    return tuple(selected)


def source_month_key(source_file: str) -> tuple[int, int, str]:
    match = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|June|Jul|July|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s*(\d{2}|\d{4})", source_file, re.IGNORECASE)
    if match is None:
        return (0, 0, source_file)
    month = MONTHS[match.group(1).lower()]
    year_raw = int(match.group(2))
    year = 2000 + year_raw if year_raw < 100 else year_raw
    return (year, month, source_file)


def _header_index(header_row: tuple[object, ...]) -> dict[str, int]:
    normalized = {normalize_source_header(value): index for index, value in enumerate(header_row) if value is not None}
    missing = [header for header in EXPECTED_HEADERS if normalize_source_header(header) not in normalized]
    if missing:
        raise ValueError(f"CSD market sheet header mismatch: missing {missing}")
    return {header: normalized[normalize_source_header(header)] for header in EXPECTED_HEADERS}


def iter_market_rows(workbook_path: Path, sheet_name: str) -> list[CsdRow]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))
        indexes = _header_index(header)
        rows: list[CsdRow] = []
        for source_row_no, values in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
            if not any(normalize_text(value) for value in values):
                continue
            if not is_total_region(values[indexes["Region"]]):
                continue
            rows.append(
                CsdRow(
                    source_file=workbook_path.name,
                    source_sheet=sheet_name,
                    source_row_no=source_row_no,
                    period_ym=parse_period_ym(values[indexes["Related date"]]),
                    market=normalize_text(values[indexes["Market"]]),
                    jw_channel=normalize_text(values[indexes["JW Channel"]]),
                    master_product=normalize_text(values[indexes["Master product"]]),
                    representing_company=normalize_text(values[indexes["Representing Company"]]),
                    product_details=parse_product_details(values[indexes["Product Details"]]),
                )
            )
        return rows
    finally:
        workbook.close()


def scan_market_sheet(workbook_path: Path, sheet_name: str) -> tuple[list[CsdRow], MarketSheetScan]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))
        indexes = _header_index(header)
        rows: list[CsdRow] = []
        regions: defaultdict[str, int] = defaultdict(int)
        invalid_channels: defaultdict[str, int] = defaultdict(int)
        raw_rows = 0
        total_sum = 0
        bad_measure_rows = 0
        for source_row_no, values in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
            if not any(normalize_text(value) for value in values):
                continue
            raw_rows += 1
            region = normalize_text(values[indexes["Region"]])
            regions[region] += 1
            channel = normalize_text(values[indexes["JW Channel"]])
            if channel not in JW_CHANNELS:
                invalid_channels[channel] += 1
            if not is_total_region(region):
                continue
            try:
                product_details = parse_product_details(values[indexes["Product Details"]])
            except ValueError:
                bad_measure_rows += 1
                continue
            total_sum += product_details
            rows.append(
                CsdRow(
                    source_file=workbook_path.name,
                    source_sheet=sheet_name,
                    source_row_no=source_row_no,
                    period_ym=parse_period_ym(values[indexes["Related date"]]),
                    market=normalize_text(values[indexes["Market"]]),
                    jw_channel=channel,
                    master_product=normalize_text(values[indexes["Master product"]]),
                    representing_company=normalize_text(values[indexes["Representing Company"]]),
                    product_details=product_details,
                )
            )
        grain_counts: defaultdict[tuple[str, str, str, str, str], int] = defaultdict(int)
        for row in rows:
            grain_counts[row.grain_key()] += 1
        duplicate_grains = sum(1 for count in grain_counts.values() if count > 1)
        return rows, MarketSheetScan(
            source_file=workbook_path.name,
            source_sheet=sheet_name,
            rows_raw=raw_rows,
            rows_total_region=len(rows),
            product_details_total_region=total_sum,
            total_region_raw_sum=total_sum,
            duplicate_grains_after_total_filter=duplicate_grains,
            regions=dict(regions),
            invalid_channels=dict(invalid_channels),
            null_or_bad_measure_rows=bad_measure_rows,
        )
    finally:
        workbook.close()


def deduplicate_rows(rows: list[CsdRow]) -> tuple[list[CsdRow], dict[str, int]]:
    grouped: defaultdict[tuple[str, str, str, str, str], list[CsdRow]] = defaultdict(list)
    for row in rows:
        grouped[row.grain_key()].append(row)
    deduped: list[CsdRow] = []
    identical_duplicates = 0
    conflict_groups = 0
    conflict_rows = 0
    for group in grouped.values():
        measures = {row.product_details for row in group}
        canonical = max(group, key=lambda row: source_month_key(row.source_file))
        deduped.append(canonical)
        if len(group) > 1:
            identical_duplicates += sum(1 for row in group if row is not canonical and row.product_details == canonical.product_details)
        if len(measures) > 1:
            conflict_groups += 1
            conflict_rows += sum(1 for row in group if row.product_details != canonical.product_details)
    return sorted(deduped, key=lambda row: row.grain_key()), {
        "rows_before": len(rows),
        "rows_after": len(deduped),
        "duplicate_groups": sum(1 for group in grouped.values() if len(group) > 1),
        "identical_duplicate_rows": identical_duplicates,
        "conflict_groups": conflict_groups,
        "conflict_rows": conflict_rows,
    }
