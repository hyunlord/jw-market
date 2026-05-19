"""
verify_ubist_parquet.py
========================
parquet/ubist/*.parquet 의 raw 충실성 검증.

Phase 1: 통계 (partition 별 row, K7 unique, channel 분포)
Phase 2: 랜덤 K7 sample (n=10) 의 메타 + 시계열을 raw xlsx 와 직접 비교
Phase 3: drop 컬럼 검증 (raw 위치 1=10, 6=16, 8=3, 12 vs 17 case)
Phase 4: 국내/외자 분리 검증 (raw 위치 2 ≠ 4 → manufacturer_origin / distributor_origin)
Phase 5: 0 값 vs NULL 구분 + source_files 일관성

성능:
- Phase 2/4 의 raw scan 은 sample K7 의 source_files 에 나오는 파일만
- Phase 3/4 의 drop/origin 검증은 일부 sample 파일 (3-5개) 만
- 전체 5-10분 추정

사용법:
    python3 scripts/verify_ubist_parquet.py \\
        --parquet-dir parquet/ubist \\
        --data-root "data/UBIST/Sales (2021-2026.02)"
"""

import argparse
import json
import random
import re
import sys
import unicodedata
from pathlib import Path

try:
    import duckdb
    import pyarrow.parquet as pq
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install duckdb pyarrow --break-system-packages")

USE_CALAMINE = False
try:
    from python_calamine import CalamineWorkbook
    USE_CALAMINE = True
except ImportError:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: xlsx reader 필요. pip3 install python-calamine --break-system-packages")


# ============================================================================
# 정책 상수 (prototype_02 와 동일)
# ============================================================================
K7_POS = {
    "drug_code":      15,
    "channel":        21,
    "specialty":      22,
    "age_group":      23,
    "gender":         24,
    "insurance_type": 20,
}
K7_COLS = ["drug_code", "channel", "specialty", "age_group", "gender", "insurance_type"]

META_POS_TO_COL = {
    2:  "manufacturer_origin",
    3:  "distributor",
    4:  "distributor_origin",
    5:  "product_name",
    7:  "brand",
    9:  "distributor2",
    10: "manufacturer",
    11: "drug_price",
    13: "molecule_strength",
    14: "drug_class",
    16: "atc",
    17: "molecule",
    18: "formulation",
    19: "route",
}

METRIC_MAP = {
    "처방조제액(원)": "val",
    "처방건수_P":     "count",
    "처방량_P":       "volume",
}

PAT_PERIOD = re.compile(r"^(\d{4})년\s*(\d{1,2})월$")


# ============================================================================
# 헬퍼
# ============================================================================
def clean_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return str(v).strip() or None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.replace(",", "").strip()
        if not v or v.lower() == "nan":
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_period(h2_val):
    if h2_val is None:
        return None
    m = PAT_PERIOD.match(str(h2_val).strip())
    if not m:
        return None
    year, month = m.groups()
    return f"{year}-{int(month):02d}"


def nfc(s):
    if s is None:
        return s
    return unicodedata.normalize("NFC", s)


# ============================================================================
# xlsx streaming reader
# ============================================================================
def stream_xlsx(xlsx_path):
    if USE_CALAMINE:
        wb = CalamineWorkbook.from_path(str(xlsx_path))
        sheet_names = wb.sheet_names
        sheet_name = "Sheet1" if "Sheet1" in sheet_names else sheet_names[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        return iter(sheet.to_python())
    else:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        sheet_name = "Sheet1" if "Sheet1" in sheet_names else sheet_names[0]
        ws = wb[sheet_name]
        return ws.iter_rows(values_only=True)


# ============================================================================
# Raw 파일 인덱스 (NFC 정규화)
# ============================================================================
def build_file_index(data_root):
    """파일 시스템 탐색 → label → path mapping (NFC 정규화)."""
    index = {}
    for sub in data_root.iterdir():
        if sub.is_dir():
            for f in sub.glob("*.xlsx"):
                if f.name.startswith(("~", ".")):
                    continue
                label = nfc(f"{sub.name}/{f.stem}")
                index[label] = f
        elif sub.is_file() and sub.suffix == ".xlsx":
            if sub.name.startswith(("~", ".")):
                continue
            label = nfc(f.stem if False else sub.stem)
            index[label] = sub
    return index


# ============================================================================
# Raw xlsx 한 번 scan → target K7 row 들 + drop/origin sample
# ============================================================================
def scan_file_for_verification(xlsx_path, target_k7_set, drop_sample_n=5000):
    """한 파일을 한 번만 scan 하면서:
    - target_k7_set 에 해당하는 row 의 (meta_24, period_obs) 추출
    - 첫 drop_sample_n row 에서 drop 컬럼 위반 카운트
    - 첫 drop_sample_n row 에서 국내/외자 분리 사례 추출
    """
    rows_iter = stream_xlsx(xlsx_path)
    header_1 = next(rows_iter, None)
    header_2 = next(rows_iter, None)
    
    if header_1 is None or header_2 is None:
        return None
    
    n_cols = len(header_1)
    
    # 시계열 컬럼 분류
    time_cols = []
    for i in range(24, n_cols):
        if i >= len(header_1) or i >= len(header_2):
            continue
        h1 = header_1[i]
        h2 = header_2[i]
        if h1 is None or h2 is None:
            continue
        metric = METRIC_MAP.get(str(h1).strip())
        if metric is None:
            continue
        period = parse_period(h2)
        if period is None:
            continue
        time_cols.append((i, period, metric))
    
    found = {}  # K7 → (meta_24_list, period_obs_dict)
    drop_violations = {
        "1=10 (제조사)":              0,
        "6=16 (ATC)":                 0,
        "8=3 (판매사)":               0,
        "12 vs 17 (성분 case 무시)":  0,
    }
    origin_diff_examples = []
    n_total = 0
    n_drop_checked = 0
    
    for row in rows_iter:
        if row is None or len(row) < 24:
            continue
        n_total += 1
        
        # K7 추출
        k7 = tuple(
            clean_str(row[K7_POS[c] - 1]) if K7_POS[c] - 1 < len(row) else None
            for c in K7_COLS
        )
        
        if k7 in target_k7_set:
            meta_24 = [clean_str(row[i]) if i < len(row) else None for i in range(24)]
            period_obs = {}
            for col_idx, period, metric in time_cols:
                if col_idx < len(row):
                    v = to_float(row[col_idx])
                    if v is not None:
                        period_obs.setdefault(period, {})[metric] = v
            # 같은 K7 가 여러 row 에 있으면 머지
            if k7 in found:
                _, existing_obs = found[k7]
                for p, ms in period_obs.items():
                    existing_obs.setdefault(p, {}).update(ms)
            else:
                found[k7] = (meta_24, period_obs)
        
        # drop / origin 검증 (sample 만)
        if n_total <= drop_sample_n:
            n_drop_checked += 1
            v1  = clean_str(row[0]) if len(row) > 0 else None
            v10 = clean_str(row[9]) if len(row) > 9 else None
            v6  = clean_str(row[5]) if len(row) > 5 else None
            v16 = clean_str(row[15]) if len(row) > 15 else None
            v8  = clean_str(row[7]) if len(row) > 7 else None
            v3  = clean_str(row[2]) if len(row) > 2 else None
            v12 = clean_str(row[11]) if len(row) > 11 else None
            v17 = clean_str(row[16]) if len(row) > 16 else None
            
            if v1 != v10:
                drop_violations["1=10 (제조사)"] += 1
            if v6 != v16:
                drop_violations["6=16 (ATC)"] += 1
            if v8 != v3:
                drop_violations["8=3 (판매사)"] += 1
            if (v12 or "").lower() != (v17 or "").lower():
                drop_violations["12 vs 17 (성분 case 무시)"] += 1
            
            v2 = clean_str(row[1]) if len(row) > 1 else None
            v4 = clean_str(row[3]) if len(row) > 3 else None
            if v2 is not None and v4 is not None and v2 != v4:
                if len(origin_diff_examples) < 3:
                    pn = clean_str(row[4]) if len(row) > 4 else None
                    origin_diff_examples.append((k7, v2, v4, pn))
    
    return {
        "found":                found,
        "drop_violations":      drop_violations,
        "drop_n_checked":       n_drop_checked,
        "origin_diff_examples": origin_diff_examples,
        "total_rows":           n_total,
    }


# ============================================================================
# Phase 0: 전수 row count 검증 (각 파일 단위)
# ============================================================================
def count_sparse_long_rows(xlsx_path):
    """xlsx 한 파일 → sparse long row 수 (적재와 동일 로직)."""
    rows_iter = stream_xlsx(xlsx_path)
    header_1 = next(rows_iter, None)
    header_2 = next(rows_iter, None)
    
    if header_1 is None or header_2 is None:
        return 0
    
    n_cols = len(header_1)
    time_cols = []
    for i in range(24, n_cols):
        if i >= len(header_1) or i >= len(header_2):
            continue
        h1 = header_1[i]
        h2 = header_2[i]
        if h1 is None or h2 is None:
            continue
        metric = METRIC_MAP.get(str(h1).strip())
        if metric is None:
            continue
        period = parse_period(h2)
        if period is None:
            continue
        time_cols.append((i, period))
    
    n_long = 0
    for row in rows_iter:
        if row is None or len(row) < 24:
            continue
        
        drug_code = clean_str(row[K7_POS["drug_code"] - 1]) if K7_POS["drug_code"] - 1 < len(row) else None
        channel = clean_str(row[K7_POS["channel"] - 1]) if K7_POS["channel"] - 1 < len(row) else None
        if not drug_code or not channel:
            continue
        
        periods_with_data = set()
        for col_idx, period in time_cols:
            if col_idx < len(row):
                v = to_float(row[col_idx])
                if v is not None:
                    periods_with_data.add(period)
        
        n_long += len(periods_with_data)
    
    return n_long


def phase0_full_count(con, file_index):
    print()
    print("=" * 72)
    print("[Phase 0] 전수 row count 검증 (file 단위)")
    print("=" * 72)
    
    # multi-source 충돌 여부 확인
    n_multi = con.execute(
        "SELECT COUNT(*) FROM parquet WHERE source_files LIKE '%,%'"
    ).fetchone()[0]
    
    if n_multi > 0:
        print(f"  ⚠ multi-source (충돌) row: {n_multi:,} — file 단위 정확 비교 불가")
        print(f"    합계만 비교 진행")
        single_file_compare = False
    else:
        print(f"  ✓ multi-source row: 0 — file 단위 정확 1:1 비교 가능")
        single_file_compare = True
    
    # parquet 의 source_files 별 count
    pq_counts = {}
    for sf, n in con.execute(
        "SELECT source_files, COUNT(*) FROM parquet GROUP BY source_files"
    ).fetchall():
        pq_counts[nfc(sf)] = n
    
    # raw 파일 scan
    import time
    print(f"\n  raw 파일 scan ({len(file_index)} 파일):")
    print(f"  {'#':>3}  {'label':<55s}  {'raw':>12}  {'parquet':>12}  {'':>2}  {'sec':>6}")
    print(f"  {'-'*3}  {'-'*55}  {'-'*12}  {'-'*12}  {'-'*2}  {'-'*6}")
    
    raw_counts = {}
    total_start = time.time()
    n_pass, n_fail = 0, 0
    
    for i, (label, path) in enumerate(sorted(file_index.items()), 1):
        t0 = time.time()
        raw_n = count_sparse_long_rows(path)
        elapsed = time.time() - t0
        raw_counts[label] = raw_n
        
        pq_n = pq_counts.get(label, 0)
        if single_file_compare:
            sym = "✓" if raw_n == pq_n else "✗"
            if raw_n == pq_n:
                n_pass += 1
            else:
                n_fail += 1
        else:
            sym = "?"
        
        label_short = (label[:55]) if len(label) > 55 else label
        print(f"  {i:>3}  {label_short:<55s}  {raw_n:>12,}  {pq_n:>12,}  "
              f"{sym:>2}  {elapsed:>6.1f}")
    
    total_elapsed = time.time() - total_start
    
    # 합계
    total_raw = sum(raw_counts.values())
    total_pq = sum(pq_counts.values())
    
    print(f"  {'-'*3}  {'-'*55}  {'-'*12}  {'-'*12}  {'-'*2}  {'-'*6}")
    sym = "✓" if total_raw == total_pq else "✗"
    print(f"       {'TOTAL':<55s}  {total_raw:>12,}  {total_pq:>12,}  {sym:>2}")
    
    print(f"\n  scan time: {total_elapsed:.0f}s ({total_elapsed/60:.1f}min)")
    
    if single_file_compare:
        print(f"\n  [Phase 0 결과] pass={n_pass} / fail={n_fail}")
        return n_pass, n_fail
    else:
        # 합계 만으로 판정
        if total_raw == total_pq:
            print(f"\n  [Phase 0 결과] ✓ 합계 일치 (file 별 비교는 multi-source 로 skip)")
            return 1, 0
        else:
            print(f"\n  [Phase 0 결과] ✗ 합계 불일치 (raw - parquet = {total_raw - total_pq:+,})")
            return 0, 1


# ============================================================================
# Phase 1: 통계
# ============================================================================
def phase1_stats(con):
    print()
    print("=" * 72)
    print("[Phase 1] 통계")
    print("=" * 72)
    
    total = con.execute("SELECT COUNT(*) FROM parquet").fetchone()[0]
    k7_unique = con.execute("SELECT COUNT(DISTINCT product_key) FROM parquet").fetchone()[0]
    period_count = con.execute("SELECT COUNT(DISTINCT period) FROM parquet").fetchone()[0]
    
    print(f"  total rows:  {total:,}")
    print(f"  K7 unique:   {k7_unique:,}")
    print(f"  periods:     {period_count}")
    
    # channel 분포
    print(f"\n  channel 분포:")
    channels = con.execute("""
        SELECT channel, COUNT(DISTINCT product_key) AS n_k7, COUNT(*) AS n_rows
        FROM parquet
        GROUP BY channel
        ORDER BY n_rows DESC
    """).fetchall()
    print(f"    {'channel':<28}  {'unique K7':>12}  {'total rows':>14}")
    print(f"    {'-'*28}  {'-'*12}  {'-'*14}")
    for ch, n_k7, n_rows in channels:
        ch_str = (ch or "(NULL)")[:28]
        print(f"    {ch_str:<28}  {n_k7:>12,}  {n_rows:>14,}")


# ============================================================================
# Phase 2: 랜덤 K7 sample 검증
# ============================================================================
def phase2_random_samples(con, file_index, n_samples=10):
    print()
    print("=" * 72)
    print(f"[Phase 2] 랜덤 K7 sample 메타+시계열 검증 (n={n_samples})")
    print("=" * 72)
    
    # K7 sample 추출 (single-source + multi-source)
    n_single = n_samples // 2
    n_multi = n_samples - n_single
    
    single_samples = con.execute(f"""
        SELECT product_key, drug_code, channel, specialty, age_group, gender,
               insurance_type, source_files,
               product_name, brand, manufacturer, manufacturer_origin,
               distributor, distributor2, distributor_origin,
               drug_price, molecule, molecule_strength, drug_class,
               atc, formulation, route
        FROM parquet
        WHERE source_files NOT LIKE '%,%'
        ORDER BY RANDOM()
        LIMIT {n_single}
    """).fetchall()
    
    multi_samples = con.execute(f"""
        SELECT product_key, drug_code, channel, specialty, age_group, gender,
               insurance_type, source_files,
               product_name, brand, manufacturer, manufacturer_origin,
               distributor, distributor2, distributor_origin,
               drug_price, molecule, molecule_strength, drug_class,
               atc, formulation, route
        FROM parquet
        WHERE source_files LIKE '%,%'
        ORDER BY RANDOM()
        LIMIT {n_multi}
    """).fetchall()
    
    cols = ["product_key", "drug_code", "channel", "specialty", "age_group",
            "gender", "insurance_type", "source_files",
            "product_name", "brand", "manufacturer", "manufacturer_origin",
            "distributor", "distributor2", "distributor_origin",
            "drug_price", "molecule", "molecule_strength", "drug_class",
            "atc", "formulation", "route"]
    
    samples = []
    for r in list(single_samples) + list(multi_samples):
        d = dict(zip(cols, r))
        d["k7"] = (d["drug_code"], d["channel"], d["specialty"],
                   d["age_group"], d["gender"], d["insurance_type"])
        d["labels"] = [nfc(s.strip()) for s in d["source_files"].split(",")]
        samples.append(d)
    
    # 검증해야 할 raw 파일 + K7 모으기
    file_to_k7s = {}
    missing_labels = set()
    for s in samples:
        for label in s["labels"]:
            if label not in file_index:
                missing_labels.add(label)
                continue
            path = file_index[label]
            file_to_k7s.setdefault(path, set()).add(s["k7"])
    
    if missing_labels:
        print(f"\n  ⚠ file_index 에 없는 라벨: {sorted(missing_labels)[:5]}")
    
    # Phase 3/4 sample 파일 추가 (drop/origin 검증)
    extra_files_for_drop = sorted(file_index.values())[:5]
    for f in extra_files_for_drop:
        if f not in file_to_k7s:
            file_to_k7s[f] = set()
    
    # 각 파일 한 번씩 scan
    print(f"\n  scan 대상: {len(file_to_k7s)} 파일")
    file_results = {}
    for i, (path, k7_set) in enumerate(file_to_k7s.items(), 1):
        print(f"    [{i:>2}/{len(file_to_k7s)}] {path.name} (K7 {len(k7_set)})")
        result = scan_file_for_verification(path, k7_set)
        if result:
            file_results[path] = result
    
    # Phase 2 검증
    print()
    pass_count, fail_count = 0, 0
    
    for idx, s in enumerate(samples, 1):
        labels = s["labels"]
        k7 = s["k7"]
        pn = s["product_name"]
        
        print(f"\n  [{idx}] {pn} | {s['channel']} | {s['age_group']} | {s['gender']}")
        print(f"      sources: {labels}")
        
        # 모든 source 의 raw 데이터 모음
        sources_data = []
        for label in labels:
            if label not in file_index:
                print(f"      ✗ '{label}' file_index 없음")
                continue
            path = file_index[label]
            result = file_results.get(path)
            if result is None or k7 not in result["found"]:
                continue
            meta_24, period_obs = result["found"][k7]
            sources_data.append((label, meta_24, period_obs))
        
        if not sources_data:
            print(f"      ✗ raw 에서 K7 못 찾음")
            fail_count += 1
            continue
        
        # 메타 검증
        ok_meta = True
        meta_pairs = [
            ("manufacturer",        10, s["manufacturer"]),
            ("manufacturer_origin",  2, s["manufacturer_origin"]),
            ("distributor",          3, s["distributor"]),
            ("distributor_origin",   4, s["distributor_origin"]),
            ("product_name",         5, s["product_name"]),
            ("brand",                7, s["brand"]),
            ("distributor2",         9, s["distributor2"]),
            ("drug_price",          11, s["drug_price"]),
            ("molecule_strength",   13, s["molecule_strength"]),
            ("drug_class",          14, s["drug_class"]),
            ("atc",                 16, s["atc"]),
            ("molecule",            17, s["molecule"]),
            ("formulation",         18, s["formulation"]),
            ("route",               19, s["route"]),
        ]
        meta_msgs = []
        for name, raw_pos, parquet_val in meta_pairs:
            matched = False
            for _, meta_24, _ in sources_data:
                if meta_24[raw_pos - 1] == parquet_val:
                    matched = True
                    break
            if not matched:
                meta_msgs.append(f"      ✗ {name}: parquet={parquet_val!r} ≠ any raw")
                ok_meta = False
        
        # 시계열 검증
        raw_union = {}
        for _, _, period_obs in sources_data:
            for p, ms in period_obs.items():
                for metric, v in ms.items():
                    raw_union[(p, metric)] = v
        
        # parquet 의 K7 시계열 (모든 period rows)
        parquet_rows = con.execute("""
            SELECT period, val, count, volume
            FROM parquet
            WHERE drug_code = ? AND channel = ?
              AND COALESCE(specialty, '') = COALESCE(?, '')
              AND COALESCE(age_group, '') = COALESCE(?, '')
              AND COALESCE(gender, '') = COALESCE(?, '')
              AND COALESCE(insurance_type, '') = COALESCE(?, '')
        """, list(k7)).fetchall()
        
        parquet_time = {}
        for period, val, count, volume in parquet_rows:
            for metric, v in [("val", val), ("count", count), ("volume", volume)]:
                if v is not None:
                    parquet_time[(period, metric)] = v
        
        n_compared, n_mismatch = 0, 0
        mismatches = []
        for k, raw_v in raw_union.items():
            parquet_v = parquet_time.get(k)
            n_compared += 1
            if parquet_v is None or abs(parquet_v - raw_v) > 0.01:
                n_mismatch += 1
                if len(mismatches) < 3:
                    mismatches.append((k, raw_v, parquet_v))
        
        ok_time = (n_mismatch == 0)
        
        if ok_meta and ok_time:
            print(f"      ✓ 메타 14개 + 시계열 {n_compared} values 모두 일치")
            pass_count += 1
        else:
            for m in meta_msgs:
                print(m)
            if mismatches:
                print(f"      ✗ 시계열 {n_mismatch}/{n_compared} 불일치:")
                for (p, met), rv, pv in mismatches:
                    print(f"          [{p}][{met}] raw={rv} parquet={pv}")
            fail_count += 1
    
    print(f"\n  [Phase 2 결과] pass={pass_count} / fail={fail_count}")
    return pass_count, fail_count, file_results


# ============================================================================
# Phase 3: drop 컬럼 검증
# ============================================================================
def phase3_drop_columns(file_results):
    print()
    print("=" * 72)
    print("[Phase 3] drop 컬럼 검증 (raw 위치 1=10, 6=16, 8=3, 12 vs 17 case)")
    print("=" * 72)
    
    total_violations = {k: 0 for k in [
        "1=10 (제조사)", "6=16 (ATC)", "8=3 (판매사)", "12 vs 17 (성분 case 무시)"
    ]}
    total_checked = 0
    
    for path, result in sorted(file_results.items()):
        n = result["drop_n_checked"]
        if n == 0:
            continue
        total_checked += n
        v = result["drop_violations"]
        all_zero = all(x == 0 for x in v.values())
        sym = "✓" if all_zero else "✗"
        print(f"\n  {path.name} ({n} rows) {sym}")
        for k, count in v.items():
            total_violations[k] += count
            if count > 0:
                print(f"      ✗ {k}: {count}")
    
    print(f"\n  [Phase 3 종합] {total_checked:,} sample rows 검증")
    for k, count in total_violations.items():
        sym = "✓" if count == 0 else "✗"
        print(f"    {sym} {k}: {count}")


# ============================================================================
# Phase 4: 국내/외자 분리
# ============================================================================
def phase4_origin_split(con, file_results):
    print()
    print("=" * 72)
    print("[Phase 4] 국내/외자 분리 (raw 위치 2 ≠ 4 → parquet 분리 보존)")
    print("=" * 72)
    
    pass_count, fail_count = 0, 0
    n_examples = 0
    
    for path, result in sorted(file_results.items()):
        examples = result["origin_diff_examples"]
        if not examples:
            continue
        print(f"\n  {path.name}:")
        for k7, v2, v4, raw_pn in examples:
            n_examples += 1
            res = con.execute("""
                SELECT manufacturer_origin, distributor_origin, product_name
                FROM parquet
                WHERE drug_code = ? AND channel = ?
                  AND COALESCE(specialty, '') = COALESCE(?, '')
                  AND COALESCE(age_group, '') = COALESCE(?, '')
                  AND COALESCE(gender, '') = COALESCE(?, '')
                  AND COALESCE(insurance_type, '') = COALESCE(?, '')
                LIMIT 1
            """, list(k7)).fetchone()
            
            if res is None:
                print(f"      ✗ K7 parquet 에서 못 찾음")
                fail_count += 1
                continue
            
            p_mfr, p_dst, p_pn = res
            mfr_ok = (p_mfr == v2)
            dst_ok = (p_dst == v4)
            sym_mfr = "✓" if mfr_ok else "✗"
            sym_dst = "✓" if dst_ok else "✗"
            print(f"      [{n_examples}] {p_pn}")
            print(f"          raw 위치2={v2!r} → parquet mfr_origin={p_mfr!r}  {sym_mfr}")
            print(f"          raw 위치4={v4!r} → parquet dst_origin={p_dst!r}  {sym_dst}")
            if mfr_ok and dst_ok:
                pass_count += 1
            else:
                fail_count += 1
    
    if n_examples == 0:
        print("\n  (검증 sample 에서 위치 2 ≠ 4 인 row 없음)")
    print(f"\n  [Phase 4 결과] pass={pass_count} / fail={fail_count} / examples={n_examples}")
    return pass_count, fail_count


# ============================================================================
# Phase 5: 0 값 vs NULL + source_files
# ============================================================================
def phase5_zero_null_and_source(con):
    print()
    print("=" * 72)
    print("[Phase 5] 0 값 vs NULL + source_files")
    print("=" * 72)
    
    print("\n  metric 별 NULL/0/value 분포:")
    print(f"  {'metric':<10}  {'NULL':>14}  {'0.0':>14}  {'value':>14}")
    print(f"  {'-'*10}  {'-'*14}  {'-'*14}  {'-'*14}")
    for metric in ["val", "count", "volume"]:
        r = con.execute(f"""
            SELECT 
                SUM(CASE WHEN {metric} IS NULL THEN 1 ELSE 0 END) as n_null,
                SUM(CASE WHEN {metric} = 0 THEN 1 ELSE 0 END) as n_zero,
                SUM(CASE WHEN {metric} != 0 AND {metric} IS NOT NULL THEN 1 ELSE 0 END) as n_value
            FROM parquet
        """).fetchone()
        n_null, n_zero, n_value = r
        print(f"  {metric:<10}  {n_null:>14,}  {n_zero:>14,}  {n_value:>14,}")
    
    print(f"\n  source_files 분포 (top 5):")
    rows = con.execute("""
        SELECT source_files, COUNT(*) AS n
        FROM parquet
        GROUP BY source_files
        ORDER BY n DESC
        LIMIT 5
    """).fetchall()
    for sf, n in rows:
        sf_short = sf[:60] + ("..." if len(sf) > 60 else "")
        print(f"    {n:>10,}  {sf_short}")


# ============================================================================
# main
# ============================================================================
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--parquet-dir", default="parquet/ubist")
    p.add_argument("--data-root", required=True)
    p.add_argument("--n-samples", type=int, default=10)
    args = p.parse_args()
    
    parquet_dir = Path(args.parquet_dir)
    data_root = Path(args.data_root)
    
    if not parquet_dir.is_dir():
        sys.exit(f"ERROR: {parquet_dir} 없음")
    if not data_root.is_dir():
        sys.exit(f"ERROR: {data_root} 없음")
    
    print("=" * 72)
    print("UBIST Parquet 무결성 검증")
    print("=" * 72)
    print(f"  parquet:   {parquet_dir}")
    print(f"  data root: {data_root}")
    
    # DuckDB 연결 + parquet view
    con = duckdb.connect(":memory:")
    con.execute(f"CREATE VIEW parquet AS SELECT * FROM '{parquet_dir}/*.parquet'")
    
    n_files = len(list(parquet_dir.glob("*.parquet")))
    print(f"  partitions: {n_files}")
    
    # 파일 인덱스
    file_index = build_file_index(data_root)
    print(f"  raw 파일 인덱스: {len(file_index)} 라벨")
    
    # Phase 0 — 전수 row count (가장 오래 걸림)
    p0_pass, p0_fail = phase0_full_count(con, file_index)
    
    # Phase 1
    phase1_stats(con)
    
    # Phase 2 (Phase 3, 4 의 file scan 도 동시)
    p2_pass, p2_fail, file_results = phase2_random_samples(
        con, file_index, n_samples=args.n_samples
    )
    
    # Phase 3
    phase3_drop_columns(file_results)
    
    # Phase 4
    p4_pass, p4_fail = phase4_origin_split(con, file_results)
    
    # Phase 5
    phase5_zero_null_and_source(con)
    
    # 종합
    print()
    print("=" * 72)
    print("종합")
    print("=" * 72)
    print(f"  Phase 0 (전수 row count): pass={p0_pass} / fail={p0_fail}")
    print(f"  Phase 2 (메타+시계열):    pass={p2_pass} / fail={p2_fail} / total={args.n_samples}")
    print(f"  Phase 4 (국내/외자):      pass={p4_pass} / fail={p4_fail}")
    
    con.close()


if __name__ == "__main__":
    main()
