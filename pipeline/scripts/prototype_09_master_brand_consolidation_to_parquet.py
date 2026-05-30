"""
prototype_09_master_brand_consolidation_to_parquet.py
=====================================================
MI Master 악템라 brand consolidation -> Parquet.

Phase 09c policy:
- Canonical row-generation logic:
  /Users/rexxa/github/jw-market/etl/master_brand_consolidation.py
- Canonical market-sheet scan logic:
  /Users/rexxa/github/jw-market/etl/master_loader.py, limited to strategy_011
- Canonical schema:
  /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_brand_consolidation
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.

Usage, in Step D after user review:
    python3 scripts/prototype_09_master_brand_consolidation_to_parquet.py
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


ETL_DIR = Path(__file__).resolve().parent / "etl"
sys.path.insert(0, str(ETL_DIR))
from storage import get_mi_master_path  # noqa: E402


DEFAULT_INPUT_FILE = get_mi_master_path()
DEFAULT_OUTPUT_FILE = Path(
    "parquet/master_brand_consolidation/master_brand_consolidation.parquet"
)

STRATEGIC_MARKET_ID = "strategy_011"
SOURCE_SHEET = "악템라"
HEADER_ROW = 5
PRODUCT_NAME_SOURCE_COLUMN = "PRODUCT NAME KOR"
EXPECTED_DRUG_ROWS = 26
EXPECTED_ROW_COUNT = 6
EXPECTED_MEMBER_DRUG_INDEXES = {5, 6, 18, 19, 22, 23}
SOURCE_REMARK = "Master Remark indicates one-brand consolidation"

MASTER_BRAND_CONSOLIDATION_COLUMNS = (
    "strategic_market_id",
    "brand_group",
    "member_drug_index",
    "member_drug_name",
    "source_remark",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)

BRAND_GROUP_MEMBERS = {
    "strategy_011": {
        "엔브렐": {"엔브렐", "엔브렐마이클릭"},
        "오렌시아": {"오렌시아", "오렌시아서브큐"},
        "젤잔즈": {"젤잔즈", "젤잔즈엑스알"},
    }
}

EXPECTED_BRAND_GROUP_COUNTS = {
    "엔브렐": 2,
    "오렌시아": 2,
    "젤잔즈": 2,
}


@dataclass
class BrandConsolidationStats:
    strategic_market_id: str = STRATEGIC_MARKET_ID
    sheet_name: str = SOURCE_SHEET
    header_row: int = HEADER_ROW
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    excluded_rows: int = 0
    staging_drug_rows: int = 0
    brand_consolidation_rows: int = 0


def utc_now_text() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def is_excluded_row(values: list[Any] | tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if "제외" in text and not text.startswith("비제외"):
            return True
    return False


def header_index(headers: list[Any], source_column: str) -> int:
    present_headers = {str(header).strip(): index for index, header in enumerate(headers) if header}
    if source_column not in present_headers:
        raise ValueError(f"required source column not found: {source_column}")
    return present_headers[source_column]


def build_brand_consolidation_rows(
    strategic_market_id: str,
    drug_record: dict[str, Any],
    ingested_at: str,
) -> list[dict[str, Any]]:
    groups = BRAND_GROUP_MEMBERS.get(strategic_market_id, {})
    product_name = drug_record.get("product_name")
    if product_name is None:
        return []
    product_text = str(product_name).strip()
    rows: list[dict[str, Any]] = []
    for brand_group, members in groups.items():
        if product_text in members:
            rows.append(
                {
                    "strategic_market_id": strategic_market_id,
                    "brand_group": brand_group,
                    "member_drug_index": drug_record["drug_index"],
                    "member_drug_name": product_text,
                    "source_remark": SOURCE_REMARK,
                    "source_sheet": drug_record["source_sheet"],
                    "source_file_version": drug_record["source_file_version"],
                    "ingested_at": ingested_at,
                }
            )
    return rows


def load_brand_consolidation_records(
    xlsx_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], BrandConsolidationStats]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            raise ValueError(f"required sheet not found: {SOURCE_SHEET!r}; sheets={wb.sheetnames}")

        ws = wb[SOURCE_SHEET]
        headers = list(
            next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
        )
        product_idx = header_index(headers, PRODUCT_NAME_SOURCE_COLUMN)
        timestamp = ingested_at or utc_now_text()
        stats = BrandConsolidationStats()
        records: list[dict[str, Any]] = []
        drug_index = 0

        for source_row_id, values in enumerate(
            ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
            start=HEADER_ROW + 1,
        ):
            stats.raw_rows_scanned += 1
            if is_empty_row(values):
                stats.empty_rows += 1
                continue
            if is_excluded_row(values):
                stats.excluded_rows += 1
                continue

            drug_index += 1
            stats.staging_drug_rows += 1
            drug_record = {
                "strategic_market_id": STRATEGIC_MARKET_ID,
                "drug_index": drug_index,
                "product_name": values[product_idx] if len(values) > product_idx else None,
                "source_sheet": SOURCE_SHEET,
                "source_file_version": xlsx_path.name,
                "source_row_id": source_row_id,
                "ingested_at": timestamp,
            }
            brand_rows = build_brand_consolidation_rows(
                STRATEGIC_MARKET_ID,
                drug_record,
                timestamp,
            )
            records.extend(brand_rows)
            stats.brand_consolidation_rows += len(brand_rows)

        return records, stats
    finally:
        wb.close()


def validate_records(records: list[dict[str, Any]], stats: BrandConsolidationStats) -> None:
    if stats.raw_rows_scanned != EXPECTED_DRUG_ROWS:
        raise ValueError(
            f"raw scanned rows must be {EXPECTED_DRUG_ROWS}, found {stats.raw_rows_scanned}"
        )
    if stats.staging_drug_rows != EXPECTED_DRUG_ROWS:
        raise ValueError(
            f"staging drug rows must be {EXPECTED_DRUG_ROWS}, found {stats.staging_drug_rows}"
        )
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(
            f"brand consolidation row count must be {EXPECTED_ROW_COUNT}, found {len(records)}"
        )

    pk_values = [
        (record["strategic_market_id"], record["brand_group"], record["member_drug_index"])
        for record in records
    ]
    if len(set(pk_values)) != EXPECTED_ROW_COUNT:
        raise ValueError(f"compound PK must be unique, found duplicates in {pk_values}")

    member_indexes = {int(record["member_drug_index"]) for record in records}
    if member_indexes != EXPECTED_MEMBER_DRUG_INDEXES:
        raise ValueError(
            f"member_drug_index mismatch: "
            f"expected={sorted(EXPECTED_MEMBER_DRUG_INDEXES)}, actual={sorted(member_indexes)}"
        )

    group_counts: dict[str, int] = {}
    for record in records:
        group_counts[record["brand_group"]] = group_counts.get(record["brand_group"], 0) + 1
    if group_counts != EXPECTED_BRAND_GROUP_COUNTS:
        raise ValueError(
            f"brand_group distribution mismatch: "
            f"expected={EXPECTED_BRAND_GROUP_COUNTS}, actual={group_counts}"
        )

    expected_columns = set(MASTER_BRAND_CONSOLIDATION_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )
        if record["strategic_market_id"] != STRATEGIC_MARKET_ID:
            raise ValueError(f"row {index} strategic_market_id mismatch: {record}")
        if record["source_sheet"] != SOURCE_SHEET:
            raise ValueError(f"row {index} source_sheet mismatch: {record}")
        if record["source_remark"] != SOURCE_REMARK:
            raise ValueError(f"row {index} source_remark mismatch: {record}")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    schema = pa.schema([pa.field(column, pa.string()) for column in MASTER_BRAND_CONSOLIDATION_COLUMNS])
    rows = [
        {column: None if record[column] is None else str(record[column]) for column in MASTER_BRAND_CONSOLIDATION_COLUMNS}
        for record in records
    ]
    table = pa.Table.from_pylist(rows, schema=schema)
    pq.write_table(table, output_file, compression="zstd", compression_level=3)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("--output-file", default=str(DEFAULT_OUTPUT_FILE))
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_file = Path(args.output_file)
    if not input_file.exists():
        sys.exit(f"ERROR: input file not found: {input_file}")

    print("=" * 72)
    print("MI Master brand_consolidation -> Parquet")
    print("=" * 72)
    print(f"  input file:   {input_file}")
    print(f"  source sheet: {SOURCE_SHEET}")
    print(f"  output file:  {output_file}")
    print(f"  columns:      {len(MASTER_BRAND_CONSOLIDATION_COLUMNS)} DDL columns")
    print("  helpers:      none")

    records, stats = load_brand_consolidation_records(input_file)
    validate_records(records, stats)
    write_parquet(records, output_file)

    print("\nResult")
    print(f"  raw rows scanned:             {stats.raw_rows_scanned}")
    print(f"  empty rows:                   {stats.empty_rows}")
    print(f"  excluded rows:                {stats.excluded_rows}")
    print(f"  staging drug rows:            {stats.staging_drug_rows}")
    print(f"  brand consolidation rows:     {stats.brand_consolidation_rows}")
    print(f"  compound PK unique:           {len(records)}")
    print(f"  output size:                  {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  ingested_at:                  {records[0]['ingested_at'] if records else None}")
    print("\nDone")


if __name__ == "__main__":
    main()
