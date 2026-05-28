"""
prototype_21_strategic_product_to_parquet.py
===========================================
Phase 14 Step 14-6 strategic_product -> Parquet.

Policy:
- Product rows are generated under strategic_brand.
- If source sales parquet has product candidates for a strategic_brand row,
  create one strategic_product row per unique source product candidate.
- Q-52: if no source product candidate exists, keep a fallback row with
  strategic_product.name = strategic_brand.name.
- D-32: UBIST matching uses NFC/whitespace normalization and manufacturer
  aliasing (제이더블유중외제약 = JW중외제약).
- D-33: IQVIA matching uses pack-grain where possible, then
  manufacturer+ATC+molecule or ATC+molecule filter-grain.
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STRATEGIC_BRAND_FILE = Path("output/catalog/strategic_brand/strategic_brand.parquet")
DEFAULT_ML_MARKET_FILE = Path("output/catalog/ml_market/ml_market.parquet")
DEFAULT_CD_MARKET_FILE = Path("output/catalog/cd_market/cd_market.parquet")
DEFAULT_UBIST_BASE_DIR = Path("output/ubist")
DEFAULT_IQVIA_DIR = Path("output/iqvia_nsa")
DEFAULT_OUTPUT_FILE = Path("output/catalog/strategic_product/strategic_product.parquet")
DEFAULT_COVERAGE_CACHE = Path("data/cache/prototype_14_step6_product_match_coverage.csv")
MASTER_DRUG_SCRIPT = Path("scripts/prototype_11_master_drug_to_parquet.py")
STRATEGIC_BRAND_SCRIPT = Path("scripts/prototype_20_strategic_brand_to_parquet.py")

EXPECTED_COLUMNS = (
    "product_id",
    "name",
    "merge_name",
    "brand_id",
    "ml_id",
    "cd_id",
    "class",
    "molecule",
    "dosage_form",
    "strength_pack",
    "nhi_type",
    "ox_gx",
    "fish_oil",
    "판매사",
    "제조사",
    "source_file_version",
    "ingested_at",
)

STRATEGIC_PRODUCT_SCHEMA = pa.schema(
    [
        pa.field("product_id", pa.string(), nullable=False),
        pa.field("name", pa.string(), nullable=False),
        pa.field("merge_name", pa.string(), nullable=False),
        pa.field("brand_id", pa.string(), nullable=False),
        pa.field("ml_id", pa.string(), nullable=False),
        pa.field("cd_id", pa.string(), nullable=True),
        pa.field("class", pa.string(), nullable=True),
        pa.field("molecule", pa.string(), nullable=True),
        pa.field("dosage_form", pa.string(), nullable=True),
        pa.field("strength_pack", pa.string(), nullable=True),
        pa.field("nhi_type", pa.string(), nullable=True),
        pa.field("ox_gx", pa.string(), nullable=True),
        pa.field("fish_oil", pa.string(), nullable=True),
        pa.field("판매사", pa.string(), nullable=True),
        pa.field("제조사", pa.string(), nullable=True),
        pa.field("source_file_version", pa.string(), nullable=False),
        pa.field("ingested_at", pa.timestamp("us"), nullable=False),
    ]
)

UBIST_JOIN_KEY_BY_SMID = {
    "strategy_001": "ubist_brand_manufacturer",
    "strategy_005": "ubist_brand_manufacturer",
    "strategy_006": "ubist_product_manufacturer",
    "strategy_007": "ubist_product_manufacturer",
    "strategy_008": "ubist_brand_manufacturer",
    "strategy_009": "ubist_brand_manufacturer",
    "strategy_015": "ubist_brand_manufacturer",
}

IQVIA_JOIN_KEY_BY_SMID = {
    "strategy_002": "iqvia_atc4_molecule",
    "strategy_003": "iqvia_atc4_molecule",
    "strategy_004": "iqvia_manufacturer_atc4_molecule",
    "strategy_010": "iqvia_manufacturer_atc4_molecule",
    "strategy_011": "iqvia_manufacturer_atc4_molecule",
    "strategy_012": "iqvia_atc4_molecule",
    "strategy_013": "iqvia_atc4_molecule",
    "strategy_014": "iqvia_atc4_molecule",
    "strategy_015": "iqvia_pack_manufacturer_atc4",
    "strategy_016": "iqvia_manufacturer_atc4_molecule",
}


def utc_now_datetime() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = unicodedata.normalize("NFC", str(value)).strip()
    if not text or text.lower() == "nan":
        return None
    if text in {"#N/A", "N/A", "NA"}:
        return None
    return text.replace("위너프A+", "위너프에이플러스")


def normalize_key(value: Any) -> str:
    text = clean_text(value) or ""
    text = re.sub(r"\s+", "", text)
    return text.upper().replace("_", "-")


def manufacturer_key(value: Any) -> str:
    key = normalize_key(value)
    aliases = {
        normalize_key("제이더블유중외제약"): normalize_key("JW중외제약"),
        normalize_key("제이더블유생명과학"): normalize_key("JW생명과학"),
    }
    return aliases.get(key, key)


def source_row_id_from_brand_id(brand_id: str) -> int:
    return int(brand_id.rsplit("_", 1)[1])


def ml_index_from_brand_id(brand_id: str) -> int:
    return int(brand_id.split("_")[1])


def import_module(path: Path, name: str) -> Any:
    spec = importlib.util.spec_from_file_location(name, REPO_ROOT / path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot import module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def read_parquet_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"required parquet not found: {path}")
    return pq.read_table(path).to_pylist()


def resolve_ubist_latest(base_dir: Path = DEFAULT_UBIST_BASE_DIR) -> Path:
    parts = sorted(base_dir.glob("year=*/month=*/data.parquet"))
    if not parts:
        raise FileNotFoundError(f"no UBIST parquet partitions under {base_dir}")
    return parts[-1]


def resolve_iqvia_latest(base_dir: Path = DEFAULT_IQVIA_DIR) -> Path:
    parts = sorted(base_dir.glob("*.parquet"))
    if not parts:
        raise FileNotFoundError(f"no IQVIA NSA parquet partitions under {base_dir}")
    return parts[-1]


def read_compat_rows(path: Path, columns: list[str], aliases: dict[str, str]) -> list[dict[str, Any]]:
    schema_names = set(pq.read_schema(path).names)
    if set(columns).issubset(schema_names):
        return pq.read_table(path, columns=columns).to_pylist()
    source_columns = [aliases.get(column, column) for column in columns]
    missing = [column for column in source_columns if column not in schema_names]
    if missing:
        raise ValueError(f"{path} missing columns for compatibility read: {missing}")
    rows = pq.read_table(path, columns=source_columns).to_pylist()
    source_to_target = {source: target for target, source in aliases.items()}
    return [
        {source_to_target.get(column, column): value for column, value in row.items()}
        for row in rows
    ]


def extract_atc_code(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    bracket = re.search(r"\[([A-Z0-9]+)\]", text.upper())
    if bracket:
        return bracket.group(1)
    plain = re.search(r"\b([A-Z][0-9][A-Z0-9]{2,3})\b", text.upper())
    return plain.group(1) if plain else text


def make_product_name(product_name: Any, pack_or_strength: Any) -> str | None:
    product = clean_text(product_name)
    pack = clean_text(pack_or_strength)
    if product is None:
        return None
    if pack is None:
        return product
    if normalize_key(pack) in normalize_key(product):
        return product
    return f"{product} {pack}"


def sheet_product_name(brand_row: dict[str, Any]) -> str:
    """Use the sheet row as product grain when it already carries pack/strength."""
    name = str(brand_row["name"])
    strength = clean_text(brand_row.get("strength_pack"))
    if strength is None:
        return name
    if normalize_key(strength) in normalize_key(name):
        return name
    return f"{name} {strength}"


def is_sheet_product_grain(brand_row: dict[str, Any], context: dict[str, Any] | None = None) -> bool:
    """Rows with an explicit strength/pack are already product-grain in MI Master."""
    if context and context.get("strategic_market_id") == "strategy_007":
        # strategy_007 strength_pack is Phase 14 serving materialization from 성분용량;
        # keep UBIST product expansion semantics from Step 14-6.
        return False
    return clean_text(brand_row.get("strength_pack")) is not None


def load_context_by_brand_id() -> dict[str, dict[str, Any]]:
    helpers = import_module(MASTER_DRUG_SCRIPT, "prototype_11_master_drug_helpers_for_product")
    brand_helpers = import_module(STRATEGIC_BRAND_SCRIPT, "prototype_20_strategic_brand_helpers_for_product")
    xlsx_path = helpers.resolve_input_file(helpers.DEFAULT_INPUT_FILE)
    metadata_catalog = helpers.load_column_metadata_catalog(helpers.DEFAULT_CATALOG_PATH)

    contexts: dict[str, dict[str, Any]] = {}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for ml_index, config in enumerate(helpers.MARKET_SHEETS, start=1):
            ws = wb[config.sheet_name]
            headers = helpers._headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                helpers.explicit_lookup_join(row_items)
                if config.strategic_market_id == "strategy_008"
                else {}
            )
            for source_row_id, values in row_items:
                if helpers.is_empty_row(values):
                    continue
                standard_values, extras = helpers.apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                fields = brand_helpers.strategic_fields(standard_values, extras)
                brand_id = f"sb_{ml_index:03d}_{source_row_id:05d}"
                contexts[brand_id] = {
                    "strategic_market_id": config.strategic_market_id,
                    "source_row_id": source_row_id,
                    "atc4_code": extract_atc_code(fields.get("atc4_code")),
                    "product_name": brand_helpers.make_name(
                        standard_values,
                        config.strategic_market_id,
                        source_row_id,
                    ),
                    "manufacturer": fields.get("제조사"),
                    "seller": fields.get("판매사"),
                    "pack_desc": clean_text(standard_values.get("pack_desc")) or clean_text(extras.get("product_pack")),
                    "strength": fields.get("strength_pack"),
                    "molecule": fields.get("molecule"),
                    "class": fields.get("class"),
                    "dosage_form": fields.get("dosage_form"),
                    "nhi_type": fields.get("nhi_type"),
                    "ox_gx": fields.get("ox_gx"),
                    "fish_oil": fields.get("fish_oil"),
                }
    finally:
        wb.close()
    return contexts


def unique_candidates(rows: list[dict[str, Any]], identity_columns: tuple[str, ...]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    out: list[dict[str, Any]] = []
    for row in rows:
        key = tuple(row.get(column) for column in identity_columns)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def load_ubist_indexes(path: Path) -> dict[str, dict[tuple[str, str], list[dict[str, Any]]]]:
    columns = [
        "product_key",
        "product_name",
        "brand",
        "manufacturer",
        "molecule_strength",
        "molecule",
        "formulation",
        "insurance_type",
    ]
    rows = read_compat_rows(
        path,
        columns,
        {
            "product_key": "약품코드",
            "product_name": "제품",
            "brand": "브랜드",
            "manufacturer": "제조사",
            "molecule_strength": "성분용량",
            "molecule": "성분",
            "formulation": "제형",
            "insurance_type": "급여구분",
        },
    )
    brand_index: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    product_index: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        candidate = {
            "source_view": "UBIST",
            "source_product_key": clean_text(row.get("product_key")),
            "product_name": clean_text(row.get("product_name")),
            "brand": clean_text(row.get("brand")),
            "manufacturer": clean_text(row.get("manufacturer")),
            "molecule": clean_text(row.get("molecule")),
            "strength_pack": clean_text(row.get("molecule_strength")),
            "dosage_form": clean_text(row.get("formulation")),
            "nhi_type": clean_text(row.get("insurance_type")),
        }
        mfr = manufacturer_key(candidate["manufacturer"])
        brand_key = normalize_key(candidate["brand"])
        product_key = normalize_key(candidate["product_name"])
        if brand_key and mfr:
            brand_index[(brand_key, mfr)].append(candidate)
        if product_key and mfr:
            product_index[(product_key, mfr)].append(candidate)

    identity = ("product_name", "manufacturer", "strength_pack", "dosage_form", "nhi_type")
    return {
        "ubist_brand_manufacturer": {
            key: unique_candidates(value, identity) for key, value in brand_index.items()
        },
        "ubist_product_manufacturer": {
            key: unique_candidates(value, identity) for key, value in product_index.items()
        },
    }


def load_iqvia_indexes(path: Path) -> dict[str, dict[tuple[str, ...], list[dict[str, Any]]]]:
    columns = [
        "product_key",
        "product_name",
        "pack_desc",
        "mfr_name_kor",
        "atc4_code",
        "strength",
        "molecule_desc",
        "nhi_type",
        "nfc3_desc",
    ]
    table = pq.read_table(path, columns=columns)
    pack_index: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    mfr_atc_mol_index: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    atc_mol_index: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in table.to_pylist():
        candidate = {
            "source_view": "IQVIA",
            "source_product_key": clean_text(row.get("product_key")),
            "product_name": clean_text(row.get("product_name")),
            "pack_desc": clean_text(row.get("pack_desc")),
            "manufacturer": clean_text(row.get("mfr_name_kor")),
            "atc4_code": clean_text(row.get("atc4_code")),
            "molecule": clean_text(row.get("molecule_desc")),
            "strength_pack": clean_text(row.get("pack_desc")) or clean_text(row.get("strength")),
            "dosage_form": clean_text(row.get("nfc3_desc")),
            "nhi_type": clean_text(row.get("nhi_type")),
        }
        mfr = manufacturer_key(candidate["manufacturer"])
        atc = normalize_key(candidate["atc4_code"])
        molecule = normalize_key(candidate["molecule"])
        pack = normalize_key(candidate["pack_desc"])
        if pack and mfr and atc:
            pack_index[(pack, mfr, atc)].append(candidate)
        if mfr and atc and molecule:
            mfr_atc_mol_index[(mfr, atc, molecule)].append(candidate)
        if atc and molecule:
            atc_mol_index[(atc, molecule)].append(candidate)

    identity = ("product_name", "pack_desc", "manufacturer", "atc4_code", "molecule", "nhi_type")
    return {
        "iqvia_pack_manufacturer_atc4": {
            key: unique_candidates(value, identity) for key, value in pack_index.items()
        },
        "iqvia_manufacturer_atc4_molecule": {
            key: unique_candidates(value, identity) for key, value in mfr_atc_mol_index.items()
        },
        "iqvia_atc4_molecule": {
            key: unique_candidates(value, identity) for key, value in atc_mol_index.items()
        },
    }


def ubist_candidates(
    brand_row: dict[str, Any],
    context: dict[str, Any],
    indexes: dict[str, dict[tuple[str, str], list[dict[str, Any]]]],
) -> tuple[str | None, list[dict[str, Any]]]:
    smid = str(context["strategic_market_id"])
    join_key = UBIST_JOIN_KEY_BY_SMID.get(smid)
    if join_key is None:
        return None, []
    name_key = normalize_key(context.get("product_name") or brand_row.get("name"))
    mfr_key = manufacturer_key(context.get("manufacturer") or brand_row.get("제조사"))
    if not name_key or not mfr_key:
        return join_key, []
    return join_key, indexes[join_key].get((name_key, mfr_key), [])


def iqvia_candidates(
    context: dict[str, Any],
    indexes: dict[str, dict[tuple[str, ...], list[dict[str, Any]]]],
) -> tuple[str | None, list[dict[str, Any]]]:
    smid = str(context["strategic_market_id"])
    join_key = IQVIA_JOIN_KEY_BY_SMID.get(smid)
    if join_key is None:
        return None, []
    atc = normalize_key(context.get("atc4_code"))
    molecule = normalize_key(context.get("molecule"))
    mfr = manufacturer_key(context.get("manufacturer"))
    pack = normalize_key(context.get("pack_desc") or context.get("strength"))

    if join_key == "iqvia_pack_manufacturer_atc4":
        candidates = indexes[join_key].get((pack, mfr, atc), []) if pack and mfr and atc else []
        if candidates:
            return join_key, candidates
        # Fall back to manufacturer+ATC+molecule for rows where pack text
        # differs between the sheet and the latest IQVIA partition.
        fallback_key = "iqvia_manufacturer_atc4_molecule"
        return f"{join_key}->fallback_mfr_atc_molecule", (
            indexes[fallback_key].get((mfr, atc, molecule), []) if mfr and atc and molecule else []
        )
    if join_key == "iqvia_manufacturer_atc4_molecule":
        return join_key, indexes[join_key].get((mfr, atc, molecule), []) if mfr and atc and molecule else []
    if join_key == "iqvia_atc4_molecule":
        return join_key, indexes[join_key].get((atc, molecule), []) if atc and molecule else []
    raise ValueError(f"unknown IQVIA join key: {join_key}")


def source_order_for_data_source(data_source: str) -> tuple[str, ...]:
    if data_source == "ubist":
        return ("UBIST",)
    if data_source == "iqvia":
        return ("IQVIA",)
    if data_source == "both":
        return ("UBIST", "IQVIA")
    raise ValueError(f"unknown data_source: {data_source}")


def product_record_from_candidate(
    brand_row: dict[str, Any],
    context: dict[str, Any],
    candidate: dict[str, Any] | None,
    product_id: str,
    ingested_at: datetime,
) -> dict[str, Any]:
    if candidate is None:
        return {
            "product_id": product_id,
            "name": str(brand_row["name"]),
            "merge_name": str(brand_row["merge_name"]),
            "brand_id": str(brand_row["brand_id"]),
            "ml_id": str(brand_row["ml_id"]),
            "cd_id": brand_row.get("cd_id"),
            "class": brand_row.get("class"),
            "molecule": brand_row.get("molecule"),
            "dosage_form": brand_row.get("dosage_form"),
            "strength_pack": brand_row.get("strength_pack"),
            "nhi_type": brand_row.get("nhi_type"),
            "ox_gx": brand_row.get("ox_gx"),
            "fish_oil": brand_row.get("fish_oil"),
            "판매사": brand_row.get("판매사"),
            "제조사": brand_row.get("제조사"),
            "source_file_version": str(brand_row["source_file_version"]),
            "ingested_at": ingested_at,
        }

    source_name = make_product_name(
        candidate.get("product_name"),
        candidate.get("pack_desc") if candidate.get("source_view") == "IQVIA" else None,
    ) or str(brand_row["name"])
    return {
        "product_id": product_id,
        "name": source_name,
        "merge_name": str(brand_row["merge_name"]),
        "brand_id": str(brand_row["brand_id"]),
        "ml_id": str(brand_row["ml_id"]),
        "cd_id": brand_row.get("cd_id"),
        "class": brand_row.get("class"),
        "molecule": clean_text(candidate.get("molecule")) or brand_row.get("molecule"),
        "dosage_form": clean_text(candidate.get("dosage_form")) or brand_row.get("dosage_form"),
        "strength_pack": clean_text(candidate.get("strength_pack")) or brand_row.get("strength_pack"),
        "nhi_type": clean_text(candidate.get("nhi_type")) or brand_row.get("nhi_type"),
        "ox_gx": brand_row.get("ox_gx"),
        "fish_oil": brand_row.get("fish_oil"),
        "판매사": brand_row.get("판매사"),
        "제조사": clean_text(candidate.get("manufacturer")) or brand_row.get("제조사"),
        "source_file_version": str(brand_row["source_file_version"]),
        "ingested_at": ingested_at,
    }


def product_record_from_sheet_product(
    brand_row: dict[str, Any],
    product_id: str,
    ingested_at: datetime,
) -> dict[str, Any]:
    return {
        "product_id": product_id,
        "name": sheet_product_name(brand_row),
        "merge_name": str(brand_row["merge_name"]),
        "brand_id": str(brand_row["brand_id"]),
        "ml_id": str(brand_row["ml_id"]),
        "cd_id": brand_row.get("cd_id"),
        "class": brand_row.get("class"),
        "molecule": brand_row.get("molecule"),
        "dosage_form": brand_row.get("dosage_form"),
        "strength_pack": brand_row.get("strength_pack"),
        "nhi_type": brand_row.get("nhi_type"),
        "ox_gx": brand_row.get("ox_gx"),
        "fish_oil": brand_row.get("fish_oil"),
        "판매사": brand_row.get("판매사"),
        "제조사": brand_row.get("제조사"),
        "source_file_version": str(brand_row["source_file_version"]),
        "ingested_at": ingested_at,
    }


def load_strategic_product_records(
    strategic_brand_path: Path,
    ml_market_path: Path,
    cd_market_path: Path,
    ubist_path: Path,
    iqvia_path: Path,
    ingested_at: datetime | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    brand_rows = read_parquet_rows(strategic_brand_path)
    ml_rows = read_parquet_rows(ml_market_path)
    cd_rows = read_parquet_rows(cd_market_path)
    ml_by_id = {str(row["ml_id"]): row for row in ml_rows}
    contexts = load_context_by_brand_id()
    ubist_indexes = load_ubist_indexes(ubist_path)
    iqvia_indexes = load_iqvia_indexes(iqvia_path)
    timestamp = ingested_at or utc_now_datetime()

    records: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []

    for brand_row in brand_rows:
        brand_id = str(brand_row["brand_id"])
        context = contexts.get(brand_id)
        if context is None:
            raise ValueError(f"missing source context for brand_id={brand_id}")
        data_source = str(ml_by_id[str(brand_row["ml_id"])]["data_source"])
        matched_candidates: list[dict[str, Any]] = []
        join_keys: list[str] = []
        source_views: list[str] = []

        if is_sheet_product_grain(brand_row, context):
            product_id = f"sp_{ml_index_from_brand_id(brand_id):03d}_{source_row_id_from_brand_id(brand_id):05d}_001"
            records.append(product_record_from_sheet_product(brand_row, product_id, timestamp))
            match_status = "sheet_product"
            matched_count = 1
            sample_names = sheet_product_name(brand_row)
            join_keys.append("sheet_product_1_to_1")
        else:
            for source_view in source_order_for_data_source(data_source):
                if source_view == "UBIST":
                    join_key, candidates = ubist_candidates(brand_row, context, ubist_indexes)
                else:
                    join_key, candidates = iqvia_candidates(context, iqvia_indexes)
                if join_key:
                    join_keys.append(join_key)
                if candidates:
                    source_views.append(source_view)
                    matched_candidates.extend(candidates)

            # De-duplicate across BOTH branches by source view and final product
            # identity so the same branch cannot inflate product rows.
            matched_candidates = unique_candidates(
                matched_candidates,
                ("source_view", "product_name", "pack_desc", "manufacturer", "strength_pack", "dosage_form", "nhi_type"),
            )

            if not matched_candidates:
                product_id = f"sp_{ml_index_from_brand_id(brand_id):03d}_{source_row_id_from_brand_id(brand_id):05d}_001"
                records.append(
                    product_record_from_candidate(brand_row, context, None, product_id, timestamp)
                )
                match_status = "fallback"
                matched_count = 0
                sample_names = ""
            else:
                for seq, candidate in enumerate(matched_candidates, start=1):
                    product_id = (
                        f"sp_{ml_index_from_brand_id(brand_id):03d}_"
                        f"{source_row_id_from_brand_id(brand_id):05d}_{seq:03d}"
                    )
                    records.append(
                        product_record_from_candidate(brand_row, context, candidate, product_id, timestamp)
                    )
                match_status = "matched"
                matched_count = len(matched_candidates)
                sample_names = " | ".join(
                    clean_text(candidate.get("product_name")) or ""
                    for candidate in matched_candidates[:5]
                )

        coverage_rows.append(
            {
                "brand_id": brand_id,
                "strategic_market_id": context["strategic_market_id"],
                "ml_id": brand_row["ml_id"],
                "cd_id": brand_row.get("cd_id") or "",
                "brand_name": brand_row["name"],
                "data_source": data_source,
                "join_keys_attempted": ";".join(join_keys),
                "source_views_matched": ";".join(source_views),
                "match_status": match_status,
                "matched_product_count": matched_count,
                "sample_product_names": sample_names,
            }
        )

    validate_records(records, coverage_rows, brand_rows, ml_rows, cd_rows)
    return records, coverage_rows


def validate_records(
    records: list[dict[str, Any]],
    coverage_rows: list[dict[str, Any]],
    brand_rows: list[dict[str, Any]],
    ml_rows: list[dict[str, Any]],
    cd_rows: list[dict[str, Any]],
) -> None:
    if not records:
        raise ValueError("strategic_product must not be empty")
    for index, record in enumerate(records, start=1):
        if tuple(record.keys()) != EXPECTED_COLUMNS:
            raise ValueError(
                f"row {index} columns mismatch: expected={EXPECTED_COLUMNS}, actual={tuple(record.keys())}"
            )
        if not isinstance(record["ingested_at"], datetime):
            raise ValueError(f"row {index} ingested_at must be datetime")
    product_ids = [record["product_id"] for record in records]
    if len(set(product_ids)) != len(product_ids):
        raise ValueError("product_id must be unique")

    brand_by_id = {str(row["brand_id"]): row for row in brand_rows}
    ml_ids = {str(row["ml_id"]) for row in ml_rows}
    cd_ids = {str(row["cd_id"]) for row in cd_rows}
    for record in records:
        brand = brand_by_id.get(str(record["brand_id"]))
        if brand is None:
            raise ValueError(f"{record['product_id']} missing brand FK: {record['brand_id']}")
        if record["ml_id"] not in ml_ids:
            raise ValueError(f"{record['product_id']} missing ml FK: {record['ml_id']}")
        if record["cd_id"] is not None and record["cd_id"] not in cd_ids:
            raise ValueError(f"{record['product_id']} missing cd FK: {record['cd_id']}")
        for column in ("merge_name", "ml_id", "cd_id"):
            if record[column] != brand[column]:
                raise ValueError(
                    f"{record['product_id']} {column} inheritance mismatch: "
                    f"product={record[column]!r}, brand={brand[column]!r}"
                )

    coverage_by_brand = {row["brand_id"]: row for row in coverage_rows}
    if set(coverage_by_brand) != set(brand_by_id):
        raise ValueError("coverage cache must have exactly one row per strategic_brand")
    fallback_count = sum(1 for row in coverage_rows if row["match_status"] == "fallback")
    if fallback_count == 0:
        raise ValueError("Q-52 fallback path was not exercised; review matching logic")

    expected_merge_names = {"엔브렐", "오렌시아", "젤잔즈"}
    for merge_name in expected_merge_names:
        product_names = {
            row["brand_id"]
            for row in records
            if row["merge_name"] == merge_name
        }
        brand_names = {
            row["brand_id"]
            for row in brand_rows
            if row["merge_name"] == merge_name
        }
        if not brand_names.issubset(product_names):
            raise ValueError(f"merge_name inheritance missing product rows for {merge_name}")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    table = pa.Table.from_pylist(records, schema=STRATEGIC_PRODUCT_SCHEMA)
    pq.write_table(table, output_file, compression="zstd", compression_level=3)


def validate_written_parquet(output_file: Path) -> None:
    table = pq.read_table(output_file)
    if table.schema != STRATEGIC_PRODUCT_SCHEMA:
        raise ValueError(f"written schema mismatch:\nexpected={STRATEGIC_PRODUCT_SCHEMA}\nactual={table.schema}")


def write_coverage_cache(rows: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    columns = (
        "brand_id",
        "strategic_market_id",
        "ml_id",
        "cd_id",
        "brand_name",
        "data_source",
        "join_keys_attempted",
        "source_views_matched",
        "match_status",
        "matched_product_count",
        "sample_product_names",
    )
    with output_file.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})


def print_summary(records: list[dict[str, Any]], coverage_rows: list[dict[str, Any]], output_file: Path) -> None:
    print("prototype Phase 14 Step 14-6 strategic_product -> Parquet")
    print(f"rows={len(records)}")
    print(f"columns={len(EXPECTED_COLUMNS)}")
    print(f"output={output_file}")
    print(f"source_file_version={records[0]['source_file_version']}")
    print(f"ingested_at={records[0]['ingested_at'].isoformat(sep=' ', timespec='seconds')}")
    print("product_rows_by_ml:")
    for ml_id, count in sorted(Counter(record["ml_id"] for record in records).items()):
        print(f"  {ml_id}: {count}")
    print("coverage_by_status:")
    for status, count in sorted(Counter(row["match_status"] for row in coverage_rows).items()):
        print(f"  {status}: {count}")
    print("coverage_by_data_source_and_status:")
    for (data_source, status), count in sorted(Counter((row["data_source"], row["match_status"]) for row in coverage_rows).items()):
        print(f"  {data_source}/{status}: {count}")
    print("expanded_product_rows_by_match_status:")
    brand_status = {row["brand_id"]: row["match_status"] for row in coverage_rows}
    for status, count in sorted(Counter(brand_status[record["brand_id"]] for record in records).items()):
        print(f"  {status}: {count}")
    print("validate_records: PASS")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load Phase 14 strategic_product parquet.")
    parser.add_argument("--strategic-brand", type=Path, default=DEFAULT_STRATEGIC_BRAND_FILE)
    parser.add_argument("--ml-market", type=Path, default=DEFAULT_ML_MARKET_FILE)
    parser.add_argument("--cd-market", type=Path, default=DEFAULT_CD_MARKET_FILE)
    parser.add_argument("--ubist", "--ubist-path", dest="ubist", type=Path, default=None)
    parser.add_argument("--iqvia", "--iqvia-path", dest="iqvia", type=Path, default=None)
    parser.add_argument("--ubist-base-dir", type=Path, default=DEFAULT_UBIST_BASE_DIR)
    parser.add_argument("--iqvia-dir", type=Path, default=DEFAULT_IQVIA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--coverage-cache", type=Path, default=DEFAULT_COVERAGE_CACHE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ubist_path = args.ubist or resolve_ubist_latest(args.ubist_base_dir)
    iqvia_path = args.iqvia or resolve_iqvia_latest(args.iqvia_dir)
    records, coverage_rows = load_strategic_product_records(
        args.strategic_brand,
        args.ml_market,
        args.cd_market,
        ubist_path,
        iqvia_path,
    )
    write_parquet(records, args.output)
    validate_written_parquet(args.output)
    write_coverage_cache(coverage_rows, args.coverage_cache)
    print_summary(records, coverage_rows, args.output)


if __name__ == "__main__":
    main()
