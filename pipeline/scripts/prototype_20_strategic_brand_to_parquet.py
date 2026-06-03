"""
prototype_20_strategic_brand_to_parquet.py
=========================================
Phase 14 Step 14-5 strategic_brand -> Parquet.

Policy:
- Q-42 / D-39: include every non-empty MI Master detail row, including rows
  that were previously excluded by the master_drug staging loader.
- A non-Class cell that contains "제외" marks the row as strict-excluded for
  downstream strategic marts. A Class-only "제외" keeps the row in market
  totals and sets ``is_class_excluded`` so Class analysis can skip it.
- Q-50: brand_id is readable and stable: sb_{ml_index:03d}_{source_row_id:05d}.
- Q-51: CD assignment is strict. 0 matches -> NULL, 1 match -> cd_id,
  2+ matches -> stop condition.
- D-30: recode/redefine columns overwrite the analysis column. strategy_008
  and strategy_011 promote class_2 into class when present, while retaining
  class_1/class_2 as explicit audit and downstream analysis axes.
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")

try:
    from strategic_exclusion_policy import (
        classify_exclusion_cells as classify_exclusion_cells_by_policy,
        contains_exclusion_marker,
    )
except ModuleNotFoundError:
    from pipeline.scripts.strategic_exclusion_policy import (
        classify_exclusion_cells as classify_exclusion_cells_by_policy,
        contains_exclusion_marker,
    )


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_FILE = Path("output/catalog/strategic_brand/strategic_brand.parquet")
DEFAULT_GADRELET_CACHE = Path("data/cache/prototype_14_step5_gadrelet_brand_mapping.csv")
DEFAULT_ML_MARKET_FILE = Path("output/catalog/ml_market/ml_market.parquet")
DEFAULT_CD_FILTER_FILE = Path("output/catalog/cd_filter/cd_filter.parquet")
DEFAULT_CD_MARKET_FILE = Path("output/catalog/cd_market/cd_market.parquet")
MASTER_DRUG_SCRIPT = Path("scripts/prototype_11_master_drug_to_parquet.py")

EXPECTED_ROW_COUNT = 4495
EXPECTED_STAGING_ROWS = 3952
EXPECTED_EXCLUDED_ROWS = 543
EXPECTED_COLUMNS = (
    "brand_id",
    "name",
    "merge_name",
    "ml_id",
    "cd_id",
    "is_excluded",
    "is_class_excluded",
    "allowed_atc4_codes_json",
    "class",
    "class_1",
    "class_2",
    "molecule",
    "dosage_form",
    "strength_pack",
    "nhi_type",
    "ox_gx",
    "fish_oil",
    "판매사",
    "제조사",
    "source_file_version",
    "ingested_at",
)
EXPECTED_ML_COUNTS = {
    "ml_001": 358,
    "ml_002": 45,
    "ml_003": 82,
    "ml_004": 10,
    "ml_005": 294,
    "ml_006": 1095,
    "ml_007": 611,
    "ml_008": 1081,
    "ml_009": 406,
    "ml_010": 10,
    "ml_011": 26,
    "ml_012": 76,
    "ml_013": 14,
    "ml_014": 331,
    "ml_015": 4,
    "ml_016": 52,
}
PHASE12_CD_BASELINE = {
    "cd_001": 116,
    "cd_002": 24,
    "cd_003": 18,
    "cd_004": 10,
    "cd_005": 11,
    "cd_006": 1047,
    "cd_007": 117,
    "cd_008": 20,
    "cd_009": 26,
    "cd_010": 160,
    "cd_011": 140,
    "cd_012": 8,
    "cd_013": 2,
    "cd_014": 26,
    "cd_015": 16,
    "cd_016": 13,
    "cd_017": 4,
    "cd_018": 64,
    "cd_019": 8,
}
SHEET_TOTAL_FILTER_IDS = {"cdf_004", "cdf_006", "cdf_007", "cdf_014", "cdf_016", "cdf_017"}
MERGE_NAME_BY_NAME = {
    "엔브렐마이클릭": "엔브렐",
    "엔브렐": "엔브렐",
    "오렌시아": "오렌시아",
    "오렌시아서브큐": "오렌시아",
    "젤잔즈": "젤잔즈",
    "젤잔즈엑스알": "젤잔즈",
}

STRATEGIC_BRAND_SCHEMA = pa.schema(
    [
        pa.field("brand_id", pa.string(), nullable=False),
        pa.field("name", pa.string(), nullable=False),
        pa.field("merge_name", pa.string(), nullable=False),
        pa.field("ml_id", pa.string(), nullable=False),
        pa.field("cd_id", pa.string(), nullable=True),
        pa.field("is_excluded", pa.bool_(), nullable=False),
        pa.field("is_class_excluded", pa.bool_(), nullable=False),
        pa.field("allowed_atc4_codes_json", pa.string(), nullable=True),
        pa.field("class", pa.string(), nullable=True),
        pa.field("class_1", pa.string(), nullable=True),
        pa.field("class_2", pa.string(), nullable=True),
        pa.field("molecule", pa.string(), nullable=True),
        pa.field("dosage_form", pa.string(), nullable=True),
        pa.field("strength_pack", pa.string(), nullable=True),
        pa.field("nhi_type", pa.string(), nullable=True),
        pa.field("ox_gx", pa.string(), nullable=True),
        pa.field("fish_oil", pa.string(), nullable=True),
        pa.field("판매사", pa.string(), nullable=True),
        pa.field("제조사", pa.string(), nullable=True),
        pa.field("source_file_version", pa.string(), nullable=False),
        pa.field("ingested_at", pa.timestamp("us"), nullable=False),
    ]
)


def utc_now_datetime() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = unicodedata.normalize("NFC", str(value)).strip()
    if not text or text.lower() == "nan":
        return None
    if text in {"#N/A", "N/A", "NA"}:
        return None
    return text.replace("위너프A+", "위너프에이플러스")


def normalize_for_match(value: Any) -> str:
    text = clean_text(value) or ""
    return re.sub(r"\s+", "", text).upper().replace("_", "-")


def contains_excluded(value: Any) -> bool:
    return contains_exclusion_marker(value)


def _is_class_header(header: Any) -> bool:
    text = clean_text(header)
    if not text:
        return False
    normalized = re.sub(r"[\s_-]+", "", text).lower()
    return normalized in {"class", "class1", "class2"} or normalized.startswith("class")


def _class_source_indexes(headers: list[Any] | tuple[Any, ...], metadata: dict[str, dict[str, Any]]) -> set[int]:
    indexes: set[int] = {idx for idx, header in enumerate(headers) if _is_class_header(header)}
    for target in ("class", "class_1", "class_2"):
        spec = metadata.get(target) or {}
        if spec.get("position") is not None:
            try:
                indexes.add(int(spec["position"]))
            except (TypeError, ValueError):
                pass
            continue
        source_column = clean_text(spec.get("source_column"))
        if not source_column:
            continue
        for idx, header in enumerate(headers):
            text = clean_text(header)
            if text and (text == source_column or text.startswith(source_column)):
                indexes.add(idx)
    return indexes


def classify_exclusion_cells(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    class_indexes: set[int] | None = None,
    *,
    strategic_market_id: str | None = None,
    sheet_name: str | None = None,
) -> tuple[bool, bool]:
    class_indexes = set(class_indexes or ())
    if not class_indexes:
        class_indexes = {idx for idx, header in enumerate(headers) if _is_class_header(header)}
    return classify_exclusion_cells_by_policy(
        values,
        class_indexes=class_indexes,
        strategic_market_id=strategic_market_id,
        sheet_name=sheet_name,
    )


def null_if_excluded(value: Any) -> str | None:
    return None if contains_excluded(value) else clean_text(value)


def load_master_drug_helpers() -> Any:
    script_path = REPO_ROOT / MASTER_DRUG_SCRIPT
    spec = importlib.util.spec_from_file_location("prototype_11_master_drug_helpers", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot import helper script: {script_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def read_parquet_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"required parquet not found: {path}")
    return pq.read_table(path).to_pylist()


def parse_json_array(value: Any) -> list[str]:
    text = clean_text(value)
    if text is None:
        return []
    parsed = json.loads(text)
    if not isinstance(parsed, list):
        raise ValueError(f"expected JSON array string, found={text!r}")
    return [str(item) for item in parsed]


def extract_atc_code(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    bracket = re.search(r"\[([A-Z0-9]+)\]", text.upper())
    if bracket:
        return bracket.group(1)
    plain = re.search(r"\b([A-Z][0-9][A-Z0-9]{2,3})\b", text.upper())
    return plain.group(1) if plain else text


def dumps_json_array(values: list[str]) -> str | None:
    cleaned = sorted({str(value).strip().upper() for value in values if str(value).strip()})
    return json.dumps(cleaned, ensure_ascii=False) if cleaned else None


def canonical_nhi(value: Any) -> str | None:
    text = normalize_for_match(value)
    if not text:
        return None
    if text in {"급여", "NHI"}:
        return "NHI"
    if text in {"비급여", "NON-NHI", "NONNHI"}:
        return "NON-NHI"
    return text


def match_text(actual: Any, expected: str, *, mode: str = "exact") -> bool:
    actual_text = normalize_for_match(actual)
    expected_text = normalize_for_match(expected)
    if not actual_text or not expected_text:
        return False
    if mode == "contains":
        return expected_text in actual_text
    if mode == "prefix":
        return actual_text.startswith(expected_text)
    return actual_text == expected_text


def field_matches(row: dict[str, Any], field: str, values: list[str] | str | None) -> bool:
    if values is None or values == []:
        return True
    expected_values = values if isinstance(values, list) else [values]
    if not expected_values:
        return True

    if field == "atc3":
        actual_atc = extract_atc_code(row.get("atc4_code"))
        return any(match_text(actual_atc, expected, mode="prefix") for expected in expected_values)
    if field == "atc4":
        actual_atc = extract_atc_code(row.get("atc4_code"))
        return any(match_text(actual_atc, expected, mode="exact") for expected in expected_values)
    if field == "molecule":
        return any(match_text(row.get("molecule"), expected, mode="contains") for expected in expected_values)
    if field == "class":
        actual = row.get("class")
        for expected in expected_values:
            if "/" in expected:
                if match_text(actual, expected, mode="exact"):
                    return True
            elif match_text(actual, expected, mode="exact") or match_text(actual, expected, mode="prefix"):
                return True
        return False
    if field == "nhi":
        actual_nhi = canonical_nhi(row.get("nhi_type"))
        return any(actual_nhi == canonical_nhi(expected) for expected in expected_values)
    if field == "dosage_form":
        return any(match_text(row.get("dosage_form"), expected, mode="exact") for expected in expected_values)
    raise ValueError(f"unknown filter field: {field}")


def cd_filter_conditions(filter_row: dict[str, Any]) -> dict[str, list[str] | str | None]:
    return {
        "atc3": parse_json_array(filter_row.get("atc3")),
        "atc4": parse_json_array(filter_row.get("atc4")),
        "molecule": parse_json_array(filter_row.get("molecule")),
        "class": parse_json_array(filter_row.get("class")),
        "nhi": clean_text(filter_row.get("nhi")),
        "dosage_form": clean_text(filter_row.get("dosage_form")),
    }


def is_sheet_total_filter(cd_filter_id: str, conditions: dict[str, Any]) -> bool:
    return cd_filter_id in SHEET_TOTAL_FILTER_IDS and all(not value for value in conditions.values())


def assign_cd_id(
    row: dict[str, Any],
    cd_markets_for_ml: dict[str, list[dict[str, Any]]],
    filter_by_id: dict[str, dict[str, Any]],
) -> tuple[str | None, list[str]]:
    candidates: list[str] = []
    for cd_market in cd_markets_for_ml.get(str(row["ml_id"]), []):
        cd_filter_id = str(cd_market["cd_filter_id"])
        conditions = cd_filter_conditions(filter_by_id[cd_filter_id])
        if is_sheet_total_filter(cd_filter_id, conditions):
            candidates.append(str(cd_market["cd_id"]))
            continue
        if all(field_matches(row, field, value) for field, value in conditions.items() if value):
            candidates.append(str(cd_market["cd_id"]))
    if len(candidates) == 1:
        return candidates[0], candidates
    if len(candidates) == 0:
        return None, candidates
    return None, candidates


def source_version_from_ml_market(ml_rows: list[dict[str, Any]]) -> str:
    versions = {clean_text(row.get("source_file_version")) for row in ml_rows}
    versions.discard(None)
    if len(versions) != 1:
        raise ValueError(f"ml_market source_file_version must be single-valued: {sorted(versions)}")
    return str(next(iter(versions)))


def first_present(*values: Any) -> str | None:
    for value in values:
        text = null_if_excluded(value)
        if text is not None:
            return text
    return None


def make_name(
    standard_values: dict[str, Any],
    strategic_market_id: str,
    source_row_id: int,
) -> str:
    if strategic_market_id == "strategy_003":
        name = first_present(standard_values.get("molecule"), standard_values.get("atc4_code"))
    else:
        name = first_present(standard_values.get("product_name"), standard_values.get("molecule"), standard_values.get("atc4_code"))
    if name is None:
        name = f"unknown_row_{source_row_id}"
    return name


def strategic_fields(
    standard_values: dict[str, Any],
    extras: dict[str, Any],
) -> dict[str, str | None]:
    class_2_value = first_present(standard_values.get("class_2"), standard_values.get("class"), extras.get("class_raw"))
    class_1_value = first_present(standard_values.get("class_1"))
    if class_1_value is None and first_present(standard_values.get("class_2")) is not None:
        class_1_value = first_present(standard_values.get("class"))
    return {
        "class": class_2_value,
        "class_1": class_1_value,
        "class_2": class_2_value if first_present(standard_values.get("class_2")) is not None else None,
        "molecule": first_present(standard_values.get("molecule")),
        "dosage_form": first_present(standard_values.get("dosage_form"), extras.get("administration_route")),
        "strength_pack": first_present(standard_values.get("strength"), standard_values.get("pack_desc"), extras.get("product_pack")),
        "nhi_type": first_present(standard_values.get("nhi_type")),
        "ox_gx": first_present(standard_values.get("ox_gx"), extras.get("ox_gx"), extras.get("ox_gx_biosimilar")),
        "fish_oil": first_present(extras.get("fish_oil_yn")),
        "판매사": first_present(standard_values.get("seller")),
        "제조사": first_present(standard_values.get("manufacturer")),
        "atc4_code": first_present(standard_values.get("atc4_code")),
    }


def load_strategic_brand_records(
    ml_market_path: Path,
    cd_filter_path: Path,
    cd_market_path: Path,
    ingested_at: datetime | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    helpers = load_master_drug_helpers()
    xlsx_path = helpers.resolve_input_file(helpers.DEFAULT_INPUT_FILE)
    metadata_catalog = helpers.load_column_metadata_catalog(helpers.DEFAULT_CATALOG_PATH)

    ml_rows = read_parquet_rows(ml_market_path)
    cd_filter_rows = read_parquet_rows(cd_filter_path)
    cd_market_rows = read_parquet_rows(cd_market_path)
    ml_ids = {str(row["ml_id"]) for row in ml_rows}
    filter_by_id = {str(row["cd_filter_id"]): row for row in cd_filter_rows}
    cd_markets_for_ml: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cd_market_rows:
        cd_markets_for_ml[str(row["ml_id"])].append(row)

    timestamp = ingested_at or utc_now_datetime()
    source_file_version = source_version_from_ml_market(ml_rows)
    records: list[dict[str, Any]] = []
    gadrelet_rows: list[dict[str, Any]] = []
    stats = {
        "raw_rows_scanned": Counter(),
        "empty_rows": Counter(),
        "excluded_rows": Counter(),
        "included_rows": Counter(),
        "nullified_cells": Counter(),
        "overlap_rows": [],
        "unknown_name_rows": [],
    }

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for ml_index, config in enumerate(helpers.MARKET_SHEETS, start=1):
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}")
            ws = wb[config.sheet_name]
            headers = helpers._headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                helpers.explicit_lookup_join(row_items)
                if config.strategic_market_id == "strategy_008"
                else {}
            )
            class_indexes = _class_source_indexes(headers, metadata)
            ml_id = f"ml_{ml_index:03d}"
            if ml_id not in ml_ids:
                raise ValueError(f"{config.strategic_market_id} missing ml_market FK: {ml_id}")

            allowed_atc4_by_name: dict[str, set[str]] = defaultdict(set)
            for source_row_id, values in row_items:
                if helpers.is_empty_row(values):
                    continue
                row_excluded, _class_excluded = classify_exclusion_cells(
                    headers,
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                )
                if row_excluded:
                    continue
                standard_values, extras = helpers.apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])
                name = make_name(standard_values, config.strategic_market_id, source_row_id)
                fields = strategic_fields(standard_values, extras)
                atc4_code = extract_atc_code(fields.get("atc4_code"))
                if atc4_code:
                    allowed_atc4_by_name[normalize_for_match(name)].add(atc4_code)

            for source_row_id, values in row_items:
                stats["raw_rows_scanned"][config.strategic_market_id] += 1
                if helpers.is_empty_row(values):
                    stats["empty_rows"][config.strategic_market_id] += 1
                    continue

                excluded, class_excluded = classify_exclusion_cells(
                    headers,
                    values,
                    class_indexes,
                    strategic_market_id=config.strategic_market_id,
                    sheet_name=config.sheet_name,
                )
                if excluded:
                    stats["excluded_rows"][config.strategic_market_id] += 1

                standard_values, extras = helpers.apply_column_mapping(headers, values, metadata)
                if source_row_id in explicit_overrides:
                    standard_values.update(explicit_overrides[source_row_id])

                name = make_name(standard_values, config.strategic_market_id, source_row_id)
                if name.startswith("unknown_row_"):
                    stats["unknown_name_rows"].append(
                        {
                            "strategic_market_id": config.strategic_market_id,
                            "source_row_id": source_row_id,
                        }
                    )
                fields = strategic_fields(standard_values, extras)
                match_context = dict(fields)
                match_context["ml_id"] = ml_id
                cd_id, candidates = assign_cd_id(match_context, cd_markets_for_ml, filter_by_id)
                if len(candidates) > 1:
                    stats["overlap_rows"].append(
                        {
                            "strategic_market_id": config.strategic_market_id,
                            "source_row_id": source_row_id,
                            "name": name,
                            "candidates": ",".join(candidates),
                        }
                    )

                record = {
                    "brand_id": f"sb_{ml_index:03d}_{source_row_id:05d}",
                    "name": name,
                    "merge_name": MERGE_NAME_BY_NAME.get(name, name),
                    "ml_id": ml_id,
                    "cd_id": cd_id,
                    "is_excluded": bool(excluded),
                    "is_class_excluded": bool(class_excluded),
                    "allowed_atc4_codes_json": dumps_json_array(list(allowed_atc4_by_name.get(normalize_for_match(name), set()))),
                    "class": fields["class"],
                    "class_1": fields["class_1"],
                    "class_2": fields["class_2"],
                    "molecule": fields["molecule"],
                    "dosage_form": fields["dosage_form"],
                    "strength_pack": fields["strength_pack"],
                    "nhi_type": fields["nhi_type"],
                    "ox_gx": fields["ox_gx"],
                    "fish_oil": fields["fish_oil"],
                    "판매사": fields["판매사"],
                    "제조사": fields["제조사"],
                    "source_file_version": source_file_version,
                    "ingested_at": timestamp,
                }

                for column, value in {
                    "class": standard_values.get("class_2") or standard_values.get("class"),
                    "class_1": standard_values.get("class"),
                    "class_2": standard_values.get("class_2"),
                    "molecule": standard_values.get("molecule"),
                    "dosage_form": standard_values.get("dosage_form"),
                    "strength_pack": standard_values.get("strength") or standard_values.get("pack_desc") or extras.get("product_pack"),
                    "nhi_type": standard_values.get("nhi_type"),
                    "ox_gx": standard_values.get("ox_gx") or extras.get("ox_gx") or extras.get("ox_gx_biosimilar"),
                    "fish_oil": extras.get("fish_oil_yn"),
                    "판매사": standard_values.get("seller"),
                    "제조사": standard_values.get("manufacturer"),
                }.items():
                    if contains_excluded(value):
                        stats["nullified_cells"][column] += 1
                records.append({column: record.get(column) for column in EXPECTED_COLUMNS})
                stats["included_rows"][config.strategic_market_id] += 1

                if config.strategic_market_id == "strategy_003":
                    gadrelet_rows.append(
                        {
                            "brand_id": record["brand_id"],
                            "source_row_id": source_row_id,
                            "atc4_code": fields["atc4_code"],
                            "molecule": fields["molecule"],
                            "class": fields["class"],
                            "dosage_form": fields["dosage_form"],
                            "strategic_brand_name": name,
                            "cd_id": cd_id,
                        }
                    )
    finally:
        wb.close()

    summary = {
        "stats": stats,
        "gadrelet_rows": gadrelet_rows,
        "source_file_version": source_file_version,
    }
    validate_records(records, summary, ml_rows, cd_market_rows)
    return records, summary


def validate_records(
    records: list[dict[str, Any]],
    summary: dict[str, Any],
    ml_rows: list[dict[str, Any]],
    cd_market_rows: list[dict[str, Any]],
) -> None:
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"strategic_brand row count must be {EXPECTED_ROW_COUNT}, found={len(records)}")
    for index, record in enumerate(records, start=1):
        if tuple(record.keys()) != EXPECTED_COLUMNS:
            raise ValueError(
                f"row {index} columns mismatch: expected={EXPECTED_COLUMNS}, actual={tuple(record.keys())}"
            )
        if not isinstance(record["ingested_at"], datetime):
            raise ValueError(f"row {index} ingested_at must be datetime")
    brand_ids = [record["brand_id"] for record in records]
    if len(set(brand_ids)) != len(brand_ids):
        raise ValueError("brand_id must be unique")

    ml_ids = {str(row["ml_id"]) for row in ml_rows}
    cd_ids = {str(row["cd_id"]) for row in cd_market_rows}
    for record in records:
        if record["ml_id"] not in ml_ids:
            raise ValueError(f"{record['brand_id']} missing ml FK: {record['ml_id']}")
        if record["cd_id"] is not None and record["cd_id"] not in cd_ids:
            raise ValueError(f"{record['brand_id']} missing cd FK: {record['cd_id']}")

    stats = summary["stats"]
    included_counts = dict(sorted(stats["included_rows"].items()))
    expected_by_smid = {
        f"strategy_{index:03d}": EXPECTED_ML_COUNTS[f"ml_{index:03d}"]
        for index in range(1, 17)
    }
    if included_counts != expected_by_smid:
        raise ValueError(f"market row distribution mismatch: expected={expected_by_smid}, actual={included_counts}")
    if sum(stats["excluded_rows"].values()) != EXPECTED_EXCLUDED_ROWS:
        raise ValueError(
            f"excluded rows must be {EXPECTED_EXCLUDED_ROWS}, found={sum(stats['excluded_rows'].values())}"
        )
    strict_excluded = sum(1 for record in records if record.get("is_excluded") is True)
    if strict_excluded != EXPECTED_EXCLUDED_ROWS:
        raise ValueError(f"is_excluded rows must be {EXPECTED_EXCLUDED_ROWS}, found={strict_excluded}")
    if sum(stats["included_rows"].values()) - sum(stats["excluded_rows"].values()) != EXPECTED_STAGING_ROWS:
        raise ValueError(f"included - strict_excluded must equal Phase 12 master_drug {EXPECTED_STAGING_ROWS} rows")
    if stats["overlap_rows"]:
        raise ValueError(f"Q-51 overlap rows found: {stats['overlap_rows'][:5]}")
    if stats["unknown_name_rows"]:
        raise ValueError(f"unknown brand name fallback rows found: {stats['unknown_name_rows'][:5]}")

    ml_counts = dict(sorted(Counter(record["ml_id"] for record in records).items()))
    if ml_counts != EXPECTED_ML_COUNTS:
        raise ValueError(f"ml distribution mismatch: expected={EXPECTED_ML_COUNTS}, actual={ml_counts}")

    merge_groups = defaultdict(set)
    for record in records:
        merge_groups[record["merge_name"]].add(record["name"])
    expected_merge_sets = {
        "엔브렐": {"엔브렐마이클릭", "엔브렐"},
        "오렌시아": {"오렌시아", "오렌시아서브큐"},
        "젤잔즈": {"젤잔즈", "젤잔즈엑스알"},
    }
    for merge_name, expected_names in expected_merge_sets.items():
        if merge_groups[merge_name] != expected_names:
            raise ValueError(
                f"merge_name {merge_name} mismatch: expected={expected_names}, actual={merge_groups[merge_name]}"
            )
    for merge_name, names in merge_groups.items():
        if merge_name not in expected_merge_sets and names != {merge_name}:
            raise ValueError(f"unexpected many-to-one merge_name mapping: {merge_name} -> {sorted(names)}")

    key_rows = {record["brand_id"]: record for record in records}
    for brand_id, expected_cd in {
        "sb_005_00017": "cd_005",
        "sb_008_00958": "cd_008",
        "sb_008_00978": "cd_008",
        "sb_008_01015": "cd_009",
    }.items():
        if brand_id in key_rows and key_rows[brand_id]["cd_id"] != expected_cd:
            raise ValueError(f"{brand_id} expected cd_id={expected_cd}, actual={key_rows[brand_id]['cd_id']}")


def write_gadrelet_cache(rows: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    columns = (
        "brand_id",
        "source_row_id",
        "atc4_code",
        "molecule",
        "class",
        "dosage_form",
        "strategic_brand_name",
        "cd_id",
    )
    with output_file.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    table = pa.Table.from_pylist(records, schema=STRATEGIC_BRAND_SCHEMA)
    pq.write_table(table, output_file, compression="zstd", compression_level=3)


def validate_written_parquet(output_file: Path) -> None:
    table = pq.read_table(output_file)
    if table.schema != STRATEGIC_BRAND_SCHEMA:
        raise ValueError(f"written schema mismatch:\nexpected={STRATEGIC_BRAND_SCHEMA}\nactual={table.schema}")
    records = table.to_pylist()
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"written row count mismatch: {len(records)}")


def print_summary(records: list[dict[str, Any]], summary: dict[str, Any], output_file: Path) -> None:
    stats = summary["stats"]
    print("prototype Phase 14 Step 14-5 strategic_brand -> Parquet")
    print(f"rows={len(records)}")
    print(f"columns={len(EXPECTED_COLUMNS)}")
    print(f"output={output_file}")
    print(f"source_file_version={summary['source_file_version']}")
    print(f"ingested_at={records[0]['ingested_at'].isoformat(sep=' ', timespec='seconds')}")
    print(f"phase12_staging_rows={len(records) - sum(stats['excluded_rows'].values())}")
    print(f"formerly_excluded_rows_included={sum(stats['excluded_rows'].values())}")
    print(f"strict_excluded_flagged={sum(1 for record in records if record.get('is_excluded') is True)}")
    print("ml_distribution:")
    for ml_id, count in sorted(Counter(record["ml_id"] for record in records).items()):
        print(f"  {ml_id}: {count}")
    print("cd_distribution:")
    cd_counts = Counter(record["cd_id"] or "NULL" for record in records)
    for cd_id, count in sorted(cd_counts.items()):
        baseline = PHASE12_CD_BASELINE.get(cd_id)
        suffix = f" (phase12={baseline}, delta={count - baseline:+d})" if baseline is not None else ""
        print(f"  {cd_id}: {count}{suffix}")
    print("excluded_distribution:")
    for smid, count in sorted(stats["excluded_rows"].items()):
        print(f"  {smid}: {count}")
    print("nullified_cells:")
    for column, count in sorted(stats["nullified_cells"].items()):
        print(f"  {column}: {count}")
    print(f"q51_overlap_rows={len(stats['overlap_rows'])}")
    print(f"gadrelet_mapping_rows={len(summary['gadrelet_rows'])}")
    print("validate_records: PASS")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load Phase 14 strategic_brand parquet.")
    parser.add_argument("--ml-market", type=Path, default=DEFAULT_ML_MARKET_FILE)
    parser.add_argument("--cd-filter", type=Path, default=DEFAULT_CD_FILTER_FILE)
    parser.add_argument("--cd-market", type=Path, default=DEFAULT_CD_MARKET_FILE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--gadrelet-cache", type=Path, default=DEFAULT_GADRELET_CACHE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    records, summary = load_strategic_brand_records(args.ml_market, args.cd_filter, args.cd_market)
    write_parquet(records, args.output)
    validate_written_parquet(args.output)
    write_gadrelet_cache(summary["gadrelet_rows"], args.gadrelet_cache)
    print_summary(records, summary, args.output)


if __name__ == "__main__":
    main()
