"""
prototype_08_master_qa_to_parquet.py
====================================
MI Master "Q&A" sheet -> Parquet.

Phase 09b policy:
- Canonical logic: /Users/rexxa/github/jw-market/etl/master_qa.py
- Canonical schema: /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_qa
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.

Usage, in Step D after user review:
    python3 scripts/prototype_08_master_qa_to_parquet.py
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


ETL_DIR = Path(__file__).resolve().parent / "etl"
sys.path.insert(0, str(ETL_DIR))
from storage import get_mi_master_path  # noqa: E402


DEFAULT_INPUT_FILE = get_mi_master_path()
DEFAULT_OUTPUT_FILE = Path("parquet/master_qa/master_qa.parquet")
SOURCE_SHEET = "Q&A"
HEADER_ROW = 2
EXPECTED_ROW_COUNT = 13

MASTER_QA_COLUMNS = (
    "qa_id",
    "strategic_market_id",
    "question_text",
    "answer_text",
    "application_actions_json",
    "source_remark",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)

EXPECTED_QA_IDS = tuple(f"qa_{index:04d}" for index in range(1, EXPECTED_ROW_COUNT + 1))

MARKET_NAME_ALIASES = {
    "공통": None,
    "라베칸": "strategy_001",
    "라베칸 라베칸듀오": "strategy_001",
    "라베칸/라베칸듀오": "strategy_001",
    "제이클": "strategy_002",
    "가드렛 가드메트": "strategy_003",
    "가드렛/가드메트": "strategy_003",
    "타발리스": "strategy_004",
    "시그마트": "strategy_005",
    "리바로 리바로젯": "strategy_006",
    "리바로/리바로젯": "strategy_006",
    "리바로페노": "strategy_007",
    "리바로하이": "strategy_008",
    "리바로브이": "strategy_008",
    "리바로하이 리바로브이": "strategy_008",
    "트루패스": "strategy_009",
    "피나스타/제이다트": "strategy_009",
    "트루패스 피나스타 제이다트": "strategy_009",
    "뉴트로진": "strategy_010",
    "모빌리아": "strategy_010",
    "뉴트로진 모빌리아": "strategy_010",
    "악템라": "strategy_011",
    "페린젝트": "strategy_012",
    "베노훼럼": "strategy_012",
    "페린젝트 베노훼럼": "strategy_012",
    "헴리브라": "strategy_013",
    "위너프/A+": "strategy_014",
    "위너프 위너프A+": "strategy_014",
    "엔커버": "strategy_015",
    "플라주오피": "strategy_016",
}


@dataclass
class QALoadStats:
    sheet_name: str = SOURCE_SHEET
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    staging_rows: int = 0


def utc_now_text() -> str:
    return datetime.now(timezone.utc).replace(tzinfo=None).isoformat(sep=" ", timespec="seconds")


def normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=to_jsonable, sort_keys=True)


def is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def make_header_keys(headers: list[Any] | tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    keys: list[str] = []
    for index, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        base = normalized if normalized is not None else f"__blank_col_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        keys.append(base if count == 1 else f"{base}__{count}")
    return keys


def build_raw_row_payload(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    source_row_id: int,
) -> dict[str, Any]:
    width = max(len(headers), len(values))
    padded_headers = list(headers) + [None] * (width - len(headers))
    padded_values = list(values) + [None] * (width - len(values))
    keys = make_header_keys(padded_headers)
    cells = []
    values_by_header: dict[str, Any] = {}
    for index, (header, key, value) in enumerate(zip(padded_headers, keys, padded_values), start=1):
        json_value = to_jsonable(value)
        cell = {
            "column_index": index,
            "header": normalize_header(header),
            "header_key": key,
            "value": json_value,
        }
        cells.append(cell)
        values_by_header[key] = json_value
    return {
        "source_row_id": source_row_id,
        "cells": cells,
        "values_by_header": values_by_header,
    }


def market_id_for_name(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return MARKET_NAME_ALIASES.get(text)


def blank_record() -> dict[str, Any]:
    return {column: None for column in MASTER_QA_COLUMNS}


def load_qa_records(
    xlsx_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], QALoadStats]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            # 260518 MI Master에는 Q&A 시트가 빠졌다. 이 단계의 역할은
            # 원본 Q&A sheet를 parquet로 옮기는 것뿐이고, 수프렙/TIRZEPATIDE
            # 같은 실제 override는 별도 override/config 경로에서 검증한다.
            # 시트 부재를 fatal로 두면 run_layer0_postfix가 원본 변경 하나로
            # 멈추므로, 빈 산출로 skip한다. 더미 Q&A를 생성하는 대안은 원본에
            # 없는 데이터를 만든다는 점에서 기각했다.
            return [], QALoadStats()

        ws = wb[SOURCE_SHEET]
        headers = list(
            next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
        )
        timestamp = ingested_at or utc_now_text()
        stats = QALoadStats()
        records: list[dict[str, Any]] = []
        qa_index = 0

        for source_row_id, values in enumerate(
            ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
            start=HEADER_ROW + 1,
        ):
            stats.raw_rows_scanned += 1
            if is_empty_row(values):
                stats.empty_rows += 1
                continue

            qa_index += 1
            raw_payload = build_raw_row_payload(headers, values, source_row_id=source_row_id)
            question_type = values[1] if len(values) > 1 else None
            market_name = values[2] if len(values) > 2 else None
            question_text = values[3] if len(values) > 3 else None
            answer_text = values[4] if len(values) > 4 else None
            source_remark = values[5] if len(values) > 5 else None
            actions = {
                "question_type": question_type,
                "market_name": market_name,
                "raw_answer": answer_text,
                "raw_marketing_note": source_remark,
                "auto_apply_in_phase_2": False,
                "raw_row": raw_payload,
            }

            record = blank_record()
            record.update(
                {
                    "qa_id": f"qa_{qa_index:04d}",
                    "strategic_market_id": market_id_for_name(market_name),
                    "question_text": question_text,
                    "answer_text": answer_text,
                    "application_actions_json": dumps_json(actions),
                    "source_remark": source_remark,
                    "source_sheet": SOURCE_SHEET,
                    "source_file_version": xlsx_path.name,
                    "ingested_at": timestamp,
                }
            )
            records.append(record)
            stats.staging_rows += 1

        return records, stats
    finally:
        wb.close()


def validate_records(records: list[dict[str, Any]]) -> None:
    if not records:
        return
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"master_qa row count must be {EXPECTED_ROW_COUNT}, found {len(records)}")

    qa_ids = [record["qa_id"] for record in records]
    if tuple(qa_ids) != EXPECTED_QA_IDS:
        raise ValueError(f"qa_id sequence mismatch: expected={EXPECTED_QA_IDS}, actual={tuple(qa_ids)}")
    if len(set(qa_ids)) != EXPECTED_ROW_COUNT:
        raise ValueError(f"qa_id must be unique, found duplicates in {qa_ids}")

    expected_columns = set(MASTER_QA_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )

        actions = json.loads(record["application_actions_json"])
        expected_action_keys = {
            "question_type",
            "market_name",
            "raw_answer",
            "raw_marketing_note",
            "auto_apply_in_phase_2",
            "raw_row",
        }
        if set(actions) != expected_action_keys:
            raise ValueError(
                f"row {index} application_actions_json key mismatch: "
                f"expected={sorted(expected_action_keys)}, actual={sorted(actions)}"
            )
        if actions["auto_apply_in_phase_2"] is not False:
            raise ValueError(f"row {index} auto_apply_in_phase_2 must be false")
        raw_row = actions["raw_row"]
        if not isinstance(raw_row, dict) or "source_row_id" not in raw_row:
            raise ValueError(f"row {index} raw_row payload is invalid")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    schema = pa.schema([pa.field(column, pa.string()) for column in MASTER_QA_COLUMNS])
    table = pa.Table.from_pylist(records, schema=schema)
    pq.write_table(table, output_file, compression="zstd", compression_level=3)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("--output-file", default=str(DEFAULT_OUTPUT_FILE))
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_file = Path(args.output_file)
    if not input_file.exists():
        sys.exit(f"ERROR: input file not found: {input_file}")

    print("=" * 72)
    print("MI Master Q&A -> Parquet")
    print("=" * 72)
    print(f"  input file:   {input_file}")
    print(f"  source sheet: {SOURCE_SHEET}")
    print(f"  output file:  {output_file}")
    print(f"  columns:      {len(MASTER_QA_COLUMNS)} DDL columns")
    print("  helpers:      none")

    records, stats = load_qa_records(input_file)
    validate_records(records)
    write_parquet(records, output_file)

    print("\nResult")
    print(f"  raw rows scanned: {stats.raw_rows_scanned}")
    print(f"  empty rows:       {stats.empty_rows}")
    print(f"  staging rows:     {stats.staging_rows}")
    print(f"  unique qa_ids:    {len({row['qa_id'] for row in records})}")
    print(f"  output size:      {output_file.stat().st_size / 1024:.1f} KB")
    print(f"  ingested_at:      {records[0]['ingested_at'] if records else None}")
    print("\nDone")


if __name__ == "__main__":
    main()
