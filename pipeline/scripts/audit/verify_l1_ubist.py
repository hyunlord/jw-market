#!/usr/bin/env python3
"""Verify Layer 1 UBIST raw parquet load without mutating data.

The verifier intentionally keeps the expensive checks bounded:
- row counts come from Parquet metadata
- distribution and null/zero/negative checks use representative partitions
- JW brand counts scan only the brand column across partitions
"""

from __future__ import annotations

import json
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook

from ops_utils import find_project_root


PROJECT_ROOT = find_project_root(Path(__file__).resolve())
AUDIT_DIR = PROJECT_ROOT / "docs" / "audit" / "phase_16g4_side_verify_l1"
UBIST_PARQUET_GLOB = "output/ubist/year=*/month=*/data.parquet"
REQUESTED_UBIST_SOURCE_ROOT = PROJECT_ROOT / "data new" / "UBIST"
ACTUAL_UBIST_SOURCE_ROOT = PROJECT_ROOT / "data" / "UBIST"
EXPECTED_TOTAL_ROWS = 145_000_000
EXPECTED_TOTAL_TOLERANCE = 5_000_000
JW_BRANDS = [
    "리바로",
    "리바로젯",
    "리바로브이",
    "리바로페노",
    "리바로하이",
    "페린젝트",
    "시그마트",
    "가드메트",
    "타발리스",
]


def nfc(value: object) -> str:
    return unicodedata.normalize("NFC", str(value))


def expected_months(start: str, end: str) -> list[str]:
    """Return inclusive YYYY-MM month labels."""
    start_year, start_month = [int(part) for part in start.split("-")]
    end_year, end_month = [int(part) for part in end.split("-")]
    months: list[str] = []
    year, month = start_year, start_month
    while (year, month) <= (end_year, end_month):
        months.append(f"{year:04d}-{month:02d}")
        month += 1
        if month == 13:
            year += 1
            month = 1
    return months


def partition_period_from_path(path: Path) -> str:
    year = None
    month = None
    for part in path.parts:
        if part.startswith("year="):
            year = part.split("=", 1)[1]
        elif part.startswith("month="):
            month = part.split("=", 1)[1]
    if not year or not month:
        raise ValueError(f"not a UBIST hive partition path: {path}")
    return f"{int(year):04d}-{int(month):02d}"


def discover_partitions() -> list[Path]:
    return sorted(PROJECT_ROOT.glob(UBIST_PARQUET_GLOB), key=partition_period_from_path)


def parquet_row_count(path: Path) -> int:
    return int(pq.ParquetFile(path).metadata.num_rows)


def file_size_mb(path: Path) -> float:
    return round(path.stat().st_size / 1024 / 1024, 2)


def choose_sample_files(files: list[Path], max_count: int = 5) -> list[Path]:
    if len(files) <= max_count:
        return files
    indices = {0, 1, len(files) // 2, len(files) - 2, len(files) - 1}
    return [files[i] for i in sorted(i for i in indices if 0 <= i < len(files))]


def status_for_total(total_rows: int) -> str:
    return "PASS" if abs(total_rows - EXPECTED_TOTAL_ROWS) <= EXPECTED_TOTAL_TOLERANCE else "WARN"


def discover_sample_external_xlsx() -> Path | None:
    if not ACTUAL_UBIST_SOURCE_ROOT.exists():
        return None
    preferred = [path for path in ACTUAL_UBIST_SOURCE_ROOT.rglob("*.xlsx") if nfc(path.name) == "병원 2021.xlsx"]
    if preferred:
        return sorted(preferred)[0]
    files = sorted(ACTUAL_UBIST_SOURCE_ROOT.rglob("*.xlsx"))
    return files[0] if files else None


def count_xlsx_data_rows(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return max(int(ws.max_row or 0) - 2, 0)
    finally:
        wb.close()


def count_layer1_rows_for_source_file(files: list[Path], source_file: Path) -> dict[str, Any]:
    source_stem = nfc(source_file.stem)
    counts_by_period: dict[str, int] = {}
    total = 0
    for path in files:
        df = pd.read_parquet(path, columns=["source_file"])
        series = df["source_file"].dropna().astype(str).map(nfc)
        matched = int(series.str.contains(source_stem, regex=False).sum())
        if matched:
            period = partition_period_from_path(path)
            counts_by_period[period] = matched
            total += matched
    return {
        "source_file": nfc(source_file.name),
        "matched_stem": source_stem,
        "layer1_rows": total,
        "months_observed": sorted(counts_by_period),
        "rows_by_period": counts_by_period,
    }


def sample_numeric_quality(files: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in files:
        df = pd.read_parquet(path, columns=["rx_amt", "rx_cnt", "rx_qty"])
        total = len(df)
        for col in ["rx_amt", "rx_cnt", "rx_qty"]:
            null_count = int(df[col].isnull().sum())
            zero_count = int((df[col] == 0).sum())
            negative_count = int((df[col] < 0).sum())
            rows.append(
                {
                    "partition": partition_period_from_path(path),
                    "column": col,
                    "sample_rows": total,
                    "null_count": null_count,
                    "null_pct": round(null_count / total * 100, 4) if total else 0,
                    "zero_count": zero_count,
                    "zero_pct": round(zero_count / total * 100, 4) if total else 0,
                    "negative_count": negative_count,
                    "negative_pct": round(negative_count / total * 100, 4) if total else 0,
                    "status": "WARN" if negative_count else "PASS",
                }
            )
    return rows


def sample_value_counts(files: list[Path], column: str, limit: int = 20) -> dict[str, int]:
    counter: Counter[str] = Counter()
    for path in files:
        df = pd.read_parquet(path, columns=[column])
        counter.update(str(value) for value in df[column].dropna())
    return {key: int(value) for key, value in counter.most_common(limit)}


def sample_unique_counts(files: list[Path]) -> dict[str, int]:
    values: dict[str, set[str]] = {"브랜드": set(), "제품": set(), "ATC": set()}
    for path in files:
        df = pd.read_parquet(path, columns=list(values))
        for col in values:
            values[col].update(str(value) for value in df[col].dropna().unique())
    return {col: len(col_values) for col, col_values in values.items()}


def jw_brand_counts(files: list[Path]) -> dict[str, int]:
    brands = set(JW_BRANDS)
    counts = {brand: 0 for brand in JW_BRANDS}
    for path in files:
        df = pd.read_parquet(path, columns=["브랜드"])
        series = df["브랜드"].dropna()
        matched = series[series.isin(brands)].value_counts()
        for brand, count in matched.items():
            counts[str(brand)] += int(count)
    return counts


def source_inventory() -> dict[str, Any]:
    actual_files = sorted(ACTUAL_UBIST_SOURCE_ROOT.rglob("*.xlsx")) if ACTUAL_UBIST_SOURCE_ROOT.exists() else []
    return {
        "requested_source_root": str(REQUESTED_UBIST_SOURCE_ROOT.relative_to(PROJECT_ROOT)),
        "requested_source_root_exists": REQUESTED_UBIST_SOURCE_ROOT.exists(),
        "actual_source_root": str(ACTUAL_UBIST_SOURCE_ROOT.relative_to(PROJECT_ROOT)),
        "actual_source_root_exists": ACTUAL_UBIST_SOURCE_ROOT.exists(),
        "actual_xlsx_file_count": len(actual_files),
        "note": "The request named data new/UBIST, but this repository's UBIST loader uses data/UBIST.",
    }


def verify_ubist() -> dict[str, Any]:
    generated_at = datetime.now().isoformat(timespec="seconds")
    files = discover_partitions()
    if not files:
        raise FileNotFoundError(f"No UBIST parquet partitions found under {UBIST_PARQUET_GLOB}")

    partition_breakdown = [
        {
            "period": partition_period_from_path(path),
            "year": partition_period_from_path(path).split("-")[0],
            "month": partition_period_from_path(path).split("-")[1],
            "rows": parquet_row_count(path),
            "file_size_mb": file_size_mb(path),
            "path": str(path.relative_to(PROJECT_ROOT)),
        }
        for path in files
    ]
    parquet_total = sum(row["rows"] for row in partition_breakdown)
    actual_periods = [row["period"] for row in partition_breakdown]
    requested_periods = expected_months("2021-01", "2026-02")
    missing_periods = [period for period in requested_periods if period not in actual_periods]
    extra_periods = [period for period in actual_periods if period not in requested_periods]

    sample_files = choose_sample_files(files)
    numeric_quality = sample_numeric_quality(sample_files)
    channel_distribution = sample_value_counts(sample_files, "종별")
    specialty_distribution = sample_value_counts(sample_files, "진료과")
    unique_counts = sample_unique_counts(sample_files)
    jw_counts = jw_brand_counts(files)

    checks: list[dict[str, Any]] = [
        {
            "name": "parquet total row count",
            "value": parquet_total,
            "expected_approx": EXPECTED_TOTAL_ROWS,
            "tolerance": EXPECTED_TOTAL_TOLERANCE,
            "status": status_for_total(parquet_total),
            "partition_count": len(files),
        },
        {
            "name": "period continuity (requested 2021-01 to 2026-02)",
            "value": len(actual_periods),
            "expected": len(requested_periods),
            "status": "PASS" if not missing_periods and not extra_periods else ("WARN" if not missing_periods else "FAIL"),
            "missing_periods": missing_periods,
            "extra_periods": extra_periods,
            "actual_first_period": actual_periods[0],
            "actual_last_period": actual_periods[-1],
        },
        {
            "name": "numeric null/zero/negative stats",
            "status": "PASS" if all(row["negative_count"] == 0 for row in numeric_quality) else "WARN",
            "sample_partitions": [partition_period_from_path(path) for path in sample_files],
        },
        {
            "name": "JW brand row counts",
            "status": "PASS" if any(count > 0 for count in jw_counts.values()) else "WARN",
            "brands_checked": len(jw_counts),
        },
        {
            "name": "known issue check: 페린젝트 UBIST expected zero",
            "status": "PASS" if jw_counts.get("페린젝트", 0) == 0 else "WARN",
            "value": jw_counts.get("페린젝트", 0),
            "expected": 0,
            "note": "DATA_ANALYSIS_COMPREHENSIVE expected FERINJECT to be absent from UBIST; current L1 contains rows.",
        },
    ]

    sample_external_file = discover_sample_external_xlsx()
    external_cross_check: dict[str, Any] | None = None
    if sample_external_file:
        external_rows = count_xlsx_data_rows(sample_external_file)
        match = count_layer1_rows_for_source_file(files, sample_external_file)
        months_observed = len(match["months_observed"])
        expected_upper = external_rows * max(months_observed, 1)
        status = "PASS" if match["layer1_rows"] > 0 and match["layer1_rows"] <= expected_upper else "WARN"
        external_cross_check = {
            "name": f"external cross-check: {nfc(sample_external_file.name)}",
            "status": status,
            "external_rows": external_rows,
            "layer1_rows": match["layer1_rows"],
            "months_observed": months_observed,
            "expected_upper_rows": expected_upper,
            "layer1_to_external_ratio": round(match["layer1_rows"] / external_rows, 4) if external_rows else None,
            "source_path": nfc(str(sample_external_file.relative_to(PROJECT_ROOT))),
            "note": "UBIST source is wide by month; Layer 1 is one row per populated source row and period.",
            "rows_by_period": match["rows_by_period"],
        }
        checks.append(external_cross_check)
    else:
        checks.append(
            {
                "name": "external cross-check: UBIST xlsx discovery",
                "status": "WARN",
                "note": "No UBIST xlsx source file found under data/UBIST.",
            }
        )

    result: dict[str, Any] = {
        "phase": "16-G-4-Side-Verify-L1",
        "layer": "L1 UBIST raw",
        "generated_at": generated_at,
        "source_inventory": source_inventory(),
        "checks": checks,
        "partition_breakdown": partition_breakdown,
        "numeric_quality_sample": numeric_quality,
        "channel_distribution_sample": channel_distribution,
        "top_specialty_distribution_sample": specialty_distribution,
        "sample_unique_counts": unique_counts,
        "jw_brand_row_counts": jw_counts,
        "external_cross_check": external_cross_check,
        "notes": [
            f"페린젝트(FERINJECT, B03A2)는 UBIST에서 0 row가 예상됐으나 현재 {jw_counts.get('페린젝트', 0):,} row가 관측됐다.",
            "요청 범위는 2021-01~2026-02였지만 현재 적재물은 2026-04까지 64개 partition을 포함한다.",
            "본 검증은 raw/mart/cache/ETL을 변경하지 않고 Parquet와 원본 xlsx를 읽기만 한다.",
        ],
    }
    return result


def write_result(result: dict[str, Any]) -> Path:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = AUDIT_DIR / "01_ubist_verification.json"
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    return out_path


def main() -> int:
    result = verify_ubist()
    out_path = write_result(result)
    total_check = next(check for check in result["checks"] if check["name"] == "parquet total row count")
    period_check = next(check for check in result["checks"] if check["name"].startswith("period continuity"))
    print(f"UBIST rows: {total_check['value']:,} ({total_check['status']})")
    print(
        "UBIST periods: "
        f"{period_check['value']} actual / {period_check['expected']} requested "
        f"({period_check['status']})"
    )
    print(f"Wrote {out_path.relative_to(PROJECT_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
