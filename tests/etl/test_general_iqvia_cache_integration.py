from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from pipeline.etl.io.iqvia_cache_builder import build_iqvia_parquet_cache
from pipeline.etl.io.iqvia_cache_contract import CacheUnavailableError, sha256_file
from pipeline.etl.io.iqvia_cache_storage import LocalCacheStorage
from pipeline.etl.io.mart import general_iqvia


HEADERS = [
    "DATA PERIOD",
    "AUDIT CODE",
    "AUDIT DESC",
    "MFR CODE",
    "MFR NAME",
    "PRODUCT NAME",
    "PRODUCT NAME KOR",
    "PACK DESC",
    "ATC 4 CODE",
    "ATC 4 DESC",
    "Values LC",
    "Units",
    "Counting Units",
    "Dosage Units",
    "Price",
]


def _write_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "NSA"
    sheet.append(HEADERS)
    sheet.append(
        [
            "2026-03-01",
            "KCPA",
            "Clinic",
            "M1",
            "MAKER",
            "ALPHA",
            "알파",
            "10MG",
            "A01A",
            "A01A DESC",
            100,
            50,
            20,
            25,
            2,
        ]
    )
    sheet.append(
        [
            "2026-06-01",
            "KCPA",
            "Clinic",
            "M2",
            "MAKER",
            "BETA",
            "베타",
            "20MG",
            "C10A",
            "C10A DESC",
            200,
            100,
            40,
            50,
            2,
        ]
    )
    workbook.save(path)
    return path


def _configure_cache(
    monkeypatch: pytest.MonkeyPatch,
    source: Path,
    storage: LocalCacheStorage,
) -> None:
    monkeypatch.setenv("S4_INPUT_MODE", "enriched")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_ENDPOINT", "http://cache.invalid:9000")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_ACCESS_KEY", "scoped-access")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_SECRET_KEY", "scoped-secret")
    monkeypatch.setenv("IQVIA_CACHE_SOURCE_SHA256", sha256_file(source))
    monkeypatch.setattr(
        general_iqvia,
        "build_iqvia_minio_cache_storage",
        lambda: storage,
    )


def test_general_iqvia_uses_verified_cache_without_db_or_workbook_access(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "nsa.xlsx")
    storage = LocalCacheStorage(tmp_path / "cache")
    build_iqvia_parquet_cache(source, storage)
    _configure_cache(monkeypatch, source, storage)
    monkeypatch.setattr(
        general_iqvia,
        "mariadb_connect",
        lambda: pytest.fail("MariaDB fallback must not run"),
    )
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("XLSX fallback must not run"),
    )

    frame = general_iqvia.load_iqvia_base_frame(
        quarters=("2026-Q2",),
        atc4_codes=("C10A",),
    )

    assert len(frame) == 1
    assert frame.iloc[0]["period_yyyymm"] == "2026-Q2"
    assert frame.iloc[0]["atc4_code"] == "C10A"
    assert frame.iloc[0]["brand_name"] == "베타"
    assert frame.iloc[0]["raw_sales"] == 200.0


def test_general_iqvia_cache_miss_fails_closed_without_db_fallback(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "nsa.xlsx")
    storage = LocalCacheStorage(tmp_path / "empty-cache")
    _configure_cache(monkeypatch, source, storage)
    monkeypatch.setattr(
        general_iqvia,
        "mariadb_connect",
        lambda: pytest.fail("MariaDB fallback must not run"),
    )

    with pytest.raises(CacheUnavailableError, match="_SUCCESS"):
        general_iqvia.load_iqvia_base_frame()


def test_general_iqvia_cache_requires_approved_source_sha(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("S4_INPUT_MODE", "raw")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_ENDPOINT", "http://cache.invalid:9000")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_ACCESS_KEY", "scoped-access")
    monkeypatch.setenv("IQVIA_CACHE_MINIO_SECRET_KEY", "scoped-secret")
    monkeypatch.delenv("IQVIA_CACHE_SOURCE_SHA256", raising=False)
    monkeypatch.setattr(
        general_iqvia,
        "mariadb_connect",
        lambda: pytest.fail("MariaDB fallback must not run"),
    )

    with pytest.raises(RuntimeError, match="IQVIA_CACHE_SOURCE_SHA256"):
        general_iqvia.load_iqvia_base_frame()


def test_cache_backed_bounded_iterator_discovers_atc4_without_db(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "nsa.xlsx")
    storage = LocalCacheStorage(tmp_path / "cache")
    build_iqvia_parquet_cache(source, storage)
    _configure_cache(monkeypatch, source, storage)
    monkeypatch.setattr(
        general_iqvia,
        "mariadb_connect",
        lambda: pytest.fail("MariaDB discovery must not run in cache mode"),
    )
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("XLSX discovery must not run"),
    )

    frames = list(
        general_iqvia.iter_iqvia_base_frames(quarters=("2026-Q2",))
    )

    assert [(atc4, len(frame)) for atc4, frame in frames] == [("C10A", 1)]


class _CountingStorage(LocalCacheStorage):
    def __init__(self, root: Path) -> None:
        super().__init__(root)
        self.partition_reads: list[str] = []

    def get_file(self, key: str, destination: Path) -> None:
        self.partition_reads.append(key)
        super().get_file(key, destination)


def test_cache_max_rows_bounds_partition_downloads(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "nsa.xlsx")
    storage = _CountingStorage(tmp_path / "cache")
    build_iqvia_parquet_cache(source, storage, batch_size=1)
    _configure_cache(monkeypatch, source, storage)

    frame = general_iqvia.load_iqvia_base_frame(max_rows=1)

    assert len(frame) == 1
    assert len(storage.partition_reads) == 1


def test_partial_cache_configuration_fails_closed(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("IQVIA_CACHE_MINIO_ENDPOINT", "http://cache.invalid:9000")
    monkeypatch.delenv("IQVIA_CACHE_MINIO_ACCESS_KEY", raising=False)
    monkeypatch.delenv("IQVIA_CACHE_MINIO_SECRET_KEY", raising=False)
    monkeypatch.delenv("IQVIA_CACHE_SOURCE_SHA256", raising=False)
    monkeypatch.setattr(
        general_iqvia,
        "mariadb_connect",
        lambda: pytest.fail("partial cache config must not fall back to MariaDB"),
    )

    with pytest.raises(RuntimeError, match="IQVIA_CACHE_SOURCE_SHA256"):
        general_iqvia.load_iqvia_base_frame()
