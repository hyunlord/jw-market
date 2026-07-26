from __future__ import annotations

import hashlib
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook
import pandas as pd
import yaml

from pipeline.etl.mi_master_registry import (
    MiMasterRegistry,
    apply_record_rules,
    discover_mi_master_registry,
)
from pipeline.etl.io.catalog.brand.strategic_product_context import (
    load_context_by_brand_id,
)
from pipeline.etl.io.catalog.postfix.canonical import build_canonical_brands
from pipeline.etl.io.catalog.dim.jw_products import build_jw_product_specs
from pipeline.etl.io.catalog.dim.market_competitive_dynamics_schema import (
    competitive_dynamics_contract,
)
from pipeline.etl.io.catalog.dim.market_landscape_schema import (
    market_landscape_contract,
)
from pipeline.etl.io.catalog.postfix.rebuild_strategic import (
    clean_strategic_brand,
    strategic_brand_contract,
)
from pipeline.etl.io.enrich import catalog as enrich_catalog
from pipeline.etl.lib.storage import get_mi_master_path
from pipeline.scripts.api.catalog import build_display_brands
from pipeline.scripts.api.metadata.ml_market_meta import build_brand_metadata


EXPECTED_ANALYZE_MATRIX_SHA256 = (
    "81b597b8dc6792122c9ea185a859f2fa862f6fa6e7d55539c2c234807deb2e6c"
)
EXPECTED_CD_SPECS_SHA256 = (
    "6b5e674a9e53c1b0d5150ad27f19018f0633f88cf39a181097374ed2b4b5818e"
)


def _canonical_sha256(value: object) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=list,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _registry_with_seventeenth_market(tmp_path: Path) -> MiMasterRegistry:
    source = get_mi_master_path()
    candidate = tmp_path / source.name
    shutil.copy2(source, candidate)

    workbook = load_workbook(candidate)
    try:
        source_sheet = workbook["타발리스"]
        new_sheet = workbook.copy_worksheet(source_sheet)
        new_sheet.title = "신규시장"
        definition = workbook["시장정의 & Target"]
        definition.cell(5, 23, "테스트팀")
        definition.cell(6, 23, "신규시장")
        definition.cell(7, 23, "Z99A1")
        definition.cell(10, 23, "IQVIA")
        definition.cell(14, 23, "Class")
        definition.cell(15, 23, "Molecule")
        definition.cell(48, 23, "Z99A1")
        workbook.save(candidate)
    finally:
        workbook.close()

    return discover_mi_master_registry(candidate)


def test_canonical_workbook_discovers_existing_topology_without_static_counts() -> None:
    registry = discover_mi_master_registry(get_mi_master_path())

    assert tuple(registry.market_by_id) == tuple(
        f"strategy_{index:03d}" for index in range(1, 17)
    )
    assert tuple(registry.analyze_matrix) == tuple(
        f"ml_{index:03d}" for index in range(1, 17)
    )
    assert tuple(spec["cd_id"] for spec in registry.cd_specs) == tuple(
        f"cd_{index:03d}" for index in range(1, 20)
    )
    assert _canonical_sha256(registry.analyze_matrix) == EXPECTED_ANALYZE_MATRIX_SHA256
    assert _canonical_sha256(registry.cd_specs) == EXPECTED_CD_SPECS_SHA256


def test_expected_row_counts_does_not_freeze_mi_master_topology() -> None:
    config_path = (
        Path(__file__).resolve().parents[2]
        / "pipeline"
        / "etl"
        / "config"
        / "expected_row_counts.yaml"
    )
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))

    assert "postfix_rebuild_strategic" not in config


def test_temporary_seventeenth_market_is_discovered_without_code_change(
    tmp_path: Path,
) -> None:
    registry = _registry_with_seventeenth_market(tmp_path)

    assert registry.market_by_id["strategy_017"]["sheet_name"] == "신규시장"
    assert registry.market_definition_columns["strategy_017"] == (23,)
    assert registry.analyze_matrix["ml_017"]["class"] is True
    assert registry.analyze_matrix["ml_017"]["molecule"] is True
    assert registry.cd_specs[-1] == {
        "cd_id": "cd_020",
        "name": "신규시장",
        "ml_id": "ml_017",
        "cd_filter_id": "cdf_020",
        "strategic_market_id": "strategy_017",
        "column_ids": (23,),
    }
    assert registry.target_brands[-1].brand_name == "신규시장"
    assert registry.target_brands[-1].strategic_market_id == "strategy_017"
    assert registry.target_brands[-1].cd_id == "cd_020"


def test_temporary_seventeenth_market_reaches_catalog_and_api_registries(
    tmp_path: Path,
) -> None:
    registry = _registry_with_seventeenth_market(tmp_path)

    display = build_display_brands(registry)
    metadata = build_brand_metadata(registry)
    canonical = build_canonical_brands(registry)
    jw_products = build_jw_product_specs(registry)
    market_ids, market_count = market_landscape_contract(registry)
    cd_ids, cd_count = competitive_dynamics_contract(registry)
    strategic_market_count, canonical_brand_count = strategic_brand_contract(
        registry
    )

    assert display[-1].brand_name == "신규시장"
    assert display[-1].market_id == "strategy_017"
    assert display[-1].sources == ["IQVIA"]
    assert metadata[-1].brand == "신규시장"
    assert metadata[-1].market_id == "strategy_017"
    assert metadata[-1].market_label_kor == "신규시장"
    assert canonical[-1].name == "신규시장"
    assert canonical[-1].ml_id == "ml_017"
    assert canonical[-1].cd_id == "cd_020"
    assert jw_products[-1] == (
        "strategy_017",
        "신규시장",
        "신규시장",
        "sheet split",
    )
    assert market_ids[-1] == "strategy_017"
    assert market_count == 17
    assert cd_ids[-1] == "cd_020"
    assert cd_count == 20
    assert strategic_market_count == 17
    assert canonical_brand_count == 26


def test_market_metadata_uses_discovered_topology(
    tmp_path: Path,
    monkeypatch,
) -> None:
    registry = _registry_with_seventeenth_market(tmp_path)
    monkeypatch.setattr(
        enrich_catalog,
        "default_mi_master_registry",
        lambda: registry,
    )

    metadata = enrich_catalog.load_market_metadata()

    assert metadata["counts"]["ml_market"] == 17
    assert metadata["counts"]["cd_market"] == 20
    assert metadata["counts"]["cd_filter"] == 20
    assert metadata["counts"]["detail_sheets"] == 17
    assert metadata["markets"]["ml_017"] == {
        "name": "신규시장",
        "data_source": "iqvia",
        "atc_codes": ["Z99A1"],
        "analyze_axes": registry.analyze_matrix["ml_017"],
        "detail_sheet": "신규시장",
    }
    assert metadata["ml_cd_mapping"]["ml_017"] == ["cd_020"]
    assert metadata["detail_sheets"][-1] == "신규시장"


def test_jacle_molecule_source_is_declarative() -> None:
    record = {"molecule": "Trisulfate", "class": "Trisulfate"}
    raw = {"MOLECULE DESC": "SODIUM PICOSULFATE"}

    updated = apply_record_rules(
        record,
        stage="strategic_brand_source",
        context={"sheet_name": "제이클", "raw": raw},
    )

    assert updated["molecule"] == "SODIUM PICOSULFATE"


def test_jacle_equal_class_and_molecule_is_nullified_declaratively() -> None:
    updated = apply_record_rules(
        {"molecule": "Trisulfate", "class": "Trisulfate"},
        stage="strategic_brand_fields",
        context={"sheet_name": "제이클"},
    )

    assert updated["molecule"] is None


def test_jacle_product_context_uses_the_same_source_rule_as_brand_catalog() -> None:
    contexts = load_context_by_brand_id()
    trisulfate_rows = [
        row
        for row in contexts.values()
        if row["strategic_market_id"] == "strategy_002"
        and row["class"] == "Trisulfate"
    ]

    assert trisulfate_rows
    assert all(row["molecule"] != "Trisulfate" for row in trisulfate_rows)
    assert all(row["molecule"] for row in trisulfate_rows)


def test_gardlet_tirzepatide_is_present_and_reclassified_declaratively() -> None:
    workbook = load_workbook(get_mi_master_path(), read_only=True, data_only=True)
    try:
        sheet = workbook["가드렛 가드메트"]
        source_rows = [
            {
                "molecule": sheet.cell(row_id, 4).value,
                "class": sheet.cell(row_id, 5).value,
            }
            for row_id in range(6, 40)
            if str(sheet.cell(row_id, 4).value or "").strip().upper()
            == "TIRZEPATIDE"
        ]
    finally:
        workbook.close()

    assert source_rows, "TIRZEPATIDE must remain a materialized MI Master source row"
    updated = apply_record_rules(
        source_rows[0],
        stage="ml_postfix",
        context={"sheet_name": "가드렛 가드메트"},
    )
    assert updated["class"] == "GLP-1RA"


def test_gardlet_tirzepatide_placeholder_does_not_replace_korean_brand() -> None:
    catalog = pd.DataFrame(
        [
            {
                "ml_id": "ml_003",
                "brand_id": "sb_003_00018",
                "name": "TIRZEPATIDE",
                "is_jw": False,
                "is_target": False,
            },
            {
                "ml_id": "ml_003",
                "brand_id": "sb_003_atc4_00427",
                "name": "마운자로",
                "is_jw": False,
                "is_target": False,
            },
        ]
    )

    cleaned, removed = clean_strategic_brand(catalog)

    assert cleaned["name"].tolist() == ["마운자로"]
    assert removed["name"].tolist() == ["TIRZEPATIDE"]
