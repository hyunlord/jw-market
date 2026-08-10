from __future__ import annotations

from pathlib import Path
import sys

import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from pipeline.scripts.etl.brand_activity.ingest_keyword import (  # noqa: E402
    KEYWORD_HEADERS,
    _keyword_sheet,
    read_keyword_events,
    read_keyword_message_counts,
)
from pipeline.scripts.etl.brand_activity.ingest_keyword_stage import (  # noqa: E402
    quote_schema_name,
    stage_ddl,
)
from pipeline.scripts.etl.brand_activity.km_validation import (  # noqa: E402
    compare_core_to_message_count,
    compare_message_count_overlaps,
)


def _save_keyword_fixture(path: Path, message_july: int = 2) -> None:
    """Create a minimal Keyword workbook with duplicate event rows preserved."""
    workbook = openpyxl.Workbook()
    workbook.active.title = "Main"
    count_sheet = workbook.create_sheet("2025 Message Count")
    count_sheet.append(["PRODUCT NAME", "June. 25", "July. 25"])
    count_sheet.append(["ATOZET", 95, message_july])
    sheet = workbook.create_sheet("Keywords")
    sheet.append(
        [
            "Related date",
            "VISIT LOCATION",
            "SPECIALTY NAME",
            "REP# CO",
            "PRODUCT NAME",
            "THERAPEUTIC CLASS",
            "KEYWORDS",
            "INTEREST",
            "Prescription frequency",
            "Prescription evolution",
            "Abstract and clinical literature / data",
            "Patient educational literature",
            "Promotional product literature",
            "SAMPLES LEFT",
            "OTHER MATERIALS LEFT",
            "WHAT OTHER MATERIALS",
            "OTHER COMMENTS",
        ]
    )
    duplicate_row = [
        "July 2025",
        "Clinic",
        "Cardiology",
        "JW",
        "ATOZET",
        "C10C0",
        "KW_FIXTURE_TEXT",
        "VERY USEFUL",
        "frequently",
        "increase",
        "YES",
        "NO",
        "NO",
        "NO",
        "NO",
        "",
        "",
    ]
    sheet.append(duplicate_row)
    sheet.append(duplicate_row)
    workbook.save(path)
    workbook.close()


def test_keyword_reader_preserves_full_duplicate_rows_and_redacts_text(tmp_path: Path) -> None:
    # Given: a Keyword workbook has duplicate-looking source rows.
    workbook_path = tmp_path / "Keywords for JW July. 25.xlsx"
    _save_keyword_fixture(workbook_path)

    # When: the keyword reader parses the workbook.
    events = read_keyword_events(workbook_path)

    # Then: duplicate source rows are preserved and sensitive text is redacted in audits.
    assert len(events) == 2
    assert [event.source_row_no for event in events] == [2, 3]
    assert events[0].keyword_text == events[1].keyword_text
    redacted = events[0].to_redacted_row()
    assert "keyword_text" not in redacted
    assert redacted["keyword_text_len"] == 15
    assert len(str(redacted["keyword_text_sha256"])) == 64


def test_keyword_reader_discovers_sheet_by_headers_not_sheet_name(tmp_path: Path) -> None:
    workbook_path = tmp_path / "renamed.xlsx"
    _save_keyword_fixture(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    workbook["Keywords"].title = "데이터"
    workbook.save(workbook_path)
    workbook.close()

    events = read_keyword_events(workbook_path)

    assert len(events) == 2
    assert {event.source_sheet for event in events} == {"데이터"}


def _append_keyword_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    sheet.append(list(KEYWORD_HEADERS))


def test_keyword_sheet_selects_exact_normalized_name_from_multiple_candidates() -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = "May 24 NTENSE"
    _append_keyword_headers(workbook.active)
    exact = workbook.create_sheet(" Keywords ")
    _append_keyword_headers(exact)

    selected, _, _ = _keyword_sheet(workbook)

    assert selected.title == " Keywords "
    workbook.close()


@pytest.mark.parametrize(
    ("titles", "exact_name_matches"),
    [
        (("May 24 NTENSE", "Archive"), 0),
        (("Keywords", " Keywords "), 2),
    ],
)
def test_keyword_sheet_rejects_ambiguous_multiple_candidates(
    titles: tuple[str, str],
    exact_name_matches: int,
) -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = titles[0]
    _append_keyword_headers(workbook.active)
    other = workbook.create_sheet(titles[1])
    _append_keyword_headers(other)

    with pytest.raises(
        ValueError,
        match=rf"found 2; exact sheet name matches {exact_name_matches}",
    ):
        _keyword_sheet(workbook)
    workbook.close()


def test_keyword_sheet_rejects_workbook_without_canonical_headers() -> None:
    workbook = openpyxl.Workbook()
    workbook.active.append(["not", "canonical"])

    with pytest.raises(ValueError, match="found 0; exact sheet name matches 0"):
        _keyword_sheet(workbook)
    workbook.close()


def test_keyword_reader_rejects_period_only_row(tmp_path: Path) -> None:
    workbook_path = tmp_path / "period-only.xlsx"
    _save_keyword_fixture(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook["Keywords"]
    header_width = sheet.max_column
    for column in range(2, header_width + 1):
        sheet.cell(2, column).value = None
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(ValueError, match="missing required Keyword values"):
        read_keyword_events(workbook_path)


def test_keyword_stage_ddl_excludes_meeting_table() -> None:
    # Given / When: keyword stage DDL is generated for the isolated schema.
    ddl = stage_ddl("jw_brand_activity_stage").lower()

    # Then: the reproducible stage builder creates Keyword only.
    assert "km_keyword_event_stage" in ddl
    assert "km_meeting_event_stage" not in ddl
    assert "meeting_topic" not in ddl


def test_message_count_overlap_reports_mismatches_and_product_differences(tmp_path: Path) -> None:
    # Given: two Keyword message-count sheets disagree for one overlapping month.
    july_path = tmp_path / "Keywords for JW July. 25.xlsx"
    aug_path = tmp_path / "Keywords for JW Aug. 25.xlsx"
    _save_keyword_fixture(july_path, message_july=2)
    _save_keyword_fixture(aug_path, message_july=3)

    # When: message-count overlap evidence is built.
    july_cells = read_keyword_message_counts(july_path)
    aug_cells = read_keyword_message_counts(aug_path)
    report = compare_message_count_overlaps([july_cells, aug_cells])

    # Then: the overlap report names the exact product-month mismatch.
    assert report["compared_cells"] == 2
    assert report["matched_cells"] == 1
    assert report["mismatch_cells"] == 1
    assert report["mismatches"][0]["product_name"] == "ATOZET"
    assert report["mismatches"][0]["month_ym"] == "2025-07"


def test_stage_schema_name_must_be_exact_isolated_schema() -> None:
    # Given / When / Then: only the production isolated schema and scratch schemas pass.
    assert quote_schema_name("jw_brand_activity_stage") == "jw_brand_activity_stage"
    assert quote_schema_name("jw_brand_activity_repro") == "jw_brand_activity_repro"
    with pytest.raises(ValueError, match="brand-activity scratch"):
        quote_schema_name("jw_brand_activity")
    with pytest.raises(ValueError, match="brand-activity scratch"):
        quote_schema_name("other_stage")


def test_core_to_message_count_alignment_uses_core_event_counts(tmp_path: Path) -> None:
    # Given: a Keyword workbook whose core rows match the pivot count for July.
    workbook_path = tmp_path / "Keywords for JW July. 25.xlsx"
    _save_keyword_fixture(workbook_path, message_july=2)
    events = read_keyword_events(workbook_path)
    message_cells = read_keyword_message_counts(workbook_path)

    # When: core rows are reconciled to Message Count cells.
    report = compare_core_to_message_count("keyword", events, message_cells)

    # Then: the product-month aligns exactly.
    assert report["checked_product_months"] == 1
    assert report["matched_product_months"] == 1
    assert report["mismatch_product_months"] == 0
