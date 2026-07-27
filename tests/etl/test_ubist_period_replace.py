"""기간 단위 교체 + Generic 보존 회귀 테스트.

기존 test_ubist_incremental_loader.py 의 13건은 건드리지 않는다.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pyarrow.parquet as pq
import pytest

from pipeline.etl.io import ubist_loader as ul


def _missing(value: object) -> bool:
    """pandas 결측은 None 이 아니라 nan 이다. 둘 다 본다."""
    return value is None or (isinstance(value, float) and value != value)


def _workbook(path: Path, rows: list[tuple[str, str, float]]) -> Path:
    """rows = [(product, period, amount)] — Generic 열 없는 판매 워크북."""
    path.parent.mkdir(parents=True, exist_ok=True)
    periods = sorted({period for _, period, _ in rows})
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(2, 1).value = "제품"
    for idx, period in enumerate(periods, start=2):
        year, month = period.split("-")
        sheet.cell(1, idx).value = "처방조제액(원)"
        sheet.cell(2, idx).value = f"{year}년 {int(month)}월"
    for row_no, product in enumerate(sorted({product for product, _, _ in rows}), start=3):
        sheet.cell(row_no, 1).value = product
        for idx, period in enumerate(periods, start=2):
            match = [amount for prod, per, amount in rows if prod == product and per == period]
            if match:
                sheet.cell(row_no, idx).value = match[0]
    workbook.save(path)
    workbook.close()
    return path


def _ingredient_workbook(path: Path, product: str, generic: str, amount: float) -> Path:
    """Generic 열을 담는 성분 워크북."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(2, 1).value = "제품"
    sheet.cell(2, 2).value = "Generic"
    sheet.cell(1, 3).value = "처방조제액(원)"
    sheet.cell(2, 3).value = "2026년 3월"
    sheet.cell(3, 1).value = product
    sheet.cell(3, 2).value = generic
    sheet.cell(3, 3).value = amount
    workbook.save(path)
    workbook.close()
    return path


def _frame(target: Path, period: str):
    year, month = period.split("-")
    path = target / f"year={year}" / f"month={month}" / "data.parquet"
    if not path.exists():
        return []
    return pq.read_table(path).to_pandas().to_dict("records")


def _amounts(target: Path, period: str) -> dict[str, float]:
    return {
        str(record["제품"]): float(record["rx_amt"])
        for record in _frame(target, period)
        if record["제품"] is not None
    }


def _generic_of(target: Path, source_file: str, period: str = "2026-03") -> list:
    return [
        record["Generic"]
        for record in _frame(target, period)
        if str(record["source_file"]) == source_file
    ]


@pytest.fixture
def loaded_target(tmp_path, monkeypatch):
    """clinic 2행 + hospital 2행이 2026-03 에 적재된 상태."""
    source_root = tmp_path / "src"
    source_root.mkdir()
    monkeypatch.setattr(ul, "UBIST_ROOT", source_root)
    target = tmp_path / "target"
    clinic = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 100.0), ("A2", "2026-03", 200.0)])
    hospital = _workbook(source_root / "hospital.xlsx", [("B1", "2026-03", 300.0), ("B2", "2026-03", 400.0)])
    ul.run_ubist_load(target=target, mode="replace", truncate=True, paths=[clinic, hospital], all_sources=False)
    return target, source_root


def test_same_period_new_content_replaces_only_that_uploads_share(loaded_target):
    target, source_root = loaded_target
    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 111.0), ("A2", "2026-03", 222.0)])

    plan = ul.incremental_plan([corrected], target)
    assert [summary.source_file for summary in plan.replace] == ["clinic.xlsx"]
    assert plan.add == []

    ul.run_incremental_ubist_load(target=target, paths=[corrected], all_sources=False)

    # clinic 몫만 새 값으로 바뀌고 hospital 몫은 그대로여야 한다.
    assert _amounts(target, "2026-03") == {"A1": 111.0, "A2": 222.0, "B1": 300.0, "B2": 400.0}


def test_same_period_same_content_is_idempotent(loaded_target):
    target, source_root = loaded_target
    before = _amounts(target, "2026-03")
    same = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 100.0), ("A2", "2026-03", 200.0)])

    plan = ul.incremental_plan([same], target)
    assert plan.replace == []
    assert [summary.source_file for summary in plan.skip] == ["clinic.xlsx"]

    ul.run_incremental_ubist_load(target=target, paths=[same], all_sources=False)
    assert _amounts(target, "2026-03") == before


def test_other_period_upload_leaves_existing_period_untouched(loaded_target):
    target, source_root = loaded_target
    before = _amounts(target, "2026-03")
    april = _workbook(source_root / "clinic_april.xlsx", [("A1", "2026-04", 555.0)])

    ul.run_incremental_ubist_load(target=target, paths=[april], all_sources=False)

    assert _amounts(target, "2026-03") == before
    assert _amounts(target, "2026-04") == {"A1": 555.0}


def test_failure_during_replace_preserves_existing_rows(loaded_target, monkeypatch):
    target, source_root = loaded_target
    before = _amounts(target, "2026-03")
    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 999.0)])

    def boom(*args, **kwargs):
        raise RuntimeError("injected failure after purge")

    monkeypatch.setattr(ul, "deduplicate_written_partitions", boom)

    with pytest.raises(RuntimeError, match="injected failure after purge"):
        ul.run_incremental_ubist_load(target=target, paths=[corrected], all_sources=False)

    # 원자 rename 전이라 target 은 손대지 않은 상태여야 한다.
    assert _amounts(target, "2026-03") == before


def test_replace_disabled_falls_back_to_name_based_skip(loaded_target, monkeypatch):
    target, source_root = loaded_target
    before = _amounts(target, "2026-03")
    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 111.0), ("A2", "2026-03", 222.0)])
    monkeypatch.setenv("INGEST_PERIOD_REPLACE", "0")

    plan = ul.incremental_plan([corrected], target)
    assert plan.replace == []

    ul.run_incremental_ubist_load(target=target, paths=[corrected], all_sources=False)
    # 플래그를 끄면 예전 동작 그대로 정정본이 반영되지 않는다.
    assert _amounts(target, "2026-03") == before


def test_scope_period_drops_sibling_files_in_that_period(loaded_target, monkeypatch):
    target, source_root = loaded_target
    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 111.0), ("A2", "2026-03", 222.0)])
    monkeypatch.setenv("INGEST_PERIOD_REPLACE_SCOPE", "period")

    ul.run_incremental_ubist_load(target=target, paths=[corrected], all_sources=False)

    # scope=period 는 그 달 전체를 갈아끼우므로 hospital 행이 사라진다.
    # 기본값으로 두면 안 되는 이유를 계약으로 고정해 둔다.
    assert _amounts(target, "2026-03") == {"A1": 111.0, "A2": 222.0}


def test_missing_digest_is_reported_as_undetermined_not_silently_skipped(loaded_target):
    target, source_root = loaded_target
    manifest_path = target / "_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.pop("source_file_digests", None)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")

    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 111.0)])
    plan = ul.incremental_plan([corrected], target)

    # digest 가 없으면 정정 여부를 알 수 없다. 모른다는 사실이 목록으로 남아야 한다.
    assert [summary.source_file for summary in plan.undetermined] == ["clinic.xlsx"]
    assert plan.replace == []


def test_invalid_scope_is_rejected(monkeypatch):
    monkeypatch.setenv("INGEST_PERIOD_REPLACE_SCOPE", "nonsense")
    with pytest.raises(ValueError, match="INGEST_PERIOD_REPLACE_SCOPE"):
        ul.period_replace_scope()


# --- pinned month (b1f284e6) 와의 상호작용 -----------------------------------


def test_replacing_a_pinned_period_fails_closed(loaded_target):
    target, source_root = loaded_target
    before = _amounts(target, "2026-03")
    corrected = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 777.0)])

    # pinned 기간은 s1 이 다시 쓰지 않는다. 걷어내기만 하면 영구 결손이므로 멈춰야 한다.
    with pytest.raises(RuntimeError, match="pinned periods"):
        ul.run_incremental_ubist_load(
            target=target,
            paths=[corrected],
            all_sources=False,
            exclude_periods=frozenset({"2026-03"}),
        )

    assert _amounts(target, "2026-03") == before


def test_pinned_period_outside_replace_scope_is_untouched(tmp_path, monkeypatch):
    source_root = tmp_path / "src"
    source_root.mkdir()
    monkeypatch.setattr(ul, "UBIST_ROOT", source_root)
    target = tmp_path / "target"
    clinic = _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 100.0)])
    may = _workbook(source_root / "may.xlsx", [("B1", "2026-05", 500.0)])
    ul.run_ubist_load(target=target, mode="replace", truncate=True, paths=[clinic, may], all_sources=False)

    _workbook(source_root / "clinic.xlsx", [("A1", "2026-03", 777.0)])
    ul.run_incremental_ubist_load(
        target=target,
        paths=[source_root / "clinic.xlsx"],
        all_sources=False,
        exclude_periods=frozenset({"2026-05"}),
    )

    assert _amounts(target, "2026-03") == {"A1": 777.0}
    assert _amounts(target, "2026-05") == {"B1": 500.0}


# --- Generic 보존 (안0 · 전량 코퍼스 룩업) ------------------------------------


@pytest.fixture
def generic_target(tmp_path, monkeypatch):
    """Generic 을 담은 성분 워크북 + 담지 않은 판매 워크북이 함께 적재된 상태."""
    source_root = tmp_path / "src"
    source_root.mkdir()
    monkeypatch.setattr(ul, "UBIST_ROOT", source_root)
    target = tmp_path / "target"
    sales = _workbook(source_root / "sales.xlsx", [("테스트제품", "2026-03", 100.0)])
    ingredient = _ingredient_workbook(source_root / "ingredient.xlsx", "테스트제품", "Original", 50.0)
    ul.run_ubist_load(target=target, mode="replace", truncate=True, paths=[sales, ingredient], all_sources=False)
    return target, source_root


def test_period_replace_preserves_generic_backfill(generic_target):
    target, source_root = generic_target
    assert _generic_of(target, "sales.xlsx") == ["Original"]

    _workbook(source_root / "sales.xlsx", [("테스트제품", "2026-03", 777.0)])
    ul.run_incremental_ubist_load(target=target, paths=[source_root / "sales.xlsx"], all_sources=False)

    # 금액은 반영되고 Generic 은 살아 있어야 한다. 판매 워크북에는 Generic 열이 없으므로
    # 이미 적재된 코퍼스에서 채워져야 한다.
    assert _amounts(target, "2026-03")["테스트제품"] == 777.0
    assert _generic_of(target, "sales.xlsx") == ["Original"]


def test_incremental_add_preserves_generic_backfill(generic_target):
    target, source_root = generic_target
    new = _workbook(source_root / "sales_new.xlsx", [("테스트제품", "2026-03", 555.0)])

    ul.run_incremental_ubist_load(
        target=target, paths=[new], all_sources=False, allow_overlap_dedup=True
    )

    generics = _generic_of(target, "sales_new.xlsx")
    assert generics and not any(_missing(value) for value in generics)
    assert generics == ["Original"]


def test_uploaded_workbook_generic_wins_over_corpus(generic_target):
    target, source_root = generic_target
    _ingredient_workbook(source_root / "ingredient.xlsx", "테스트제품", "Generic", 50.0)

    ul.run_incremental_ubist_load(
        target=target, paths=[source_root / "ingredient.xlsx"], all_sources=False
    )

    # 원천이 실제로 바뀌면 코퍼스의 옛 값이 이기면 안 된다.
    assert _generic_of(target, "ingredient.xlsx") == ["Generic"]
