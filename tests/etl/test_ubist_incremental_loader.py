from __future__ import annotations

import json
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

import pytest

from pipeline.etl.io.ubist_loader import (
    BUSINESS_GRAIN_COLUMNS,
    COLUMNS,
    SCHEMA,
    build_generic_lookup,
    deduplicate_business_grain,
    deduplicate_partition_file,
    iter_xlsx_rows,
    prune_ubist_partitions,
    run_incremental_ubist_load,
)


def _write_manifest(target: Path, source_files: list[str]) -> None:
    target.mkdir(parents=True)
    (target / "_manifest.json").write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "partitions": [
                    {
                        "period_yyyymm": "2025-07",
                        "path": "year=2025/month=07/data.parquet",
                        "row_count": 1,
                        "source_files": source_files,
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _workbook(path: Path, periods: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1).value = None
    sheet.cell(2, 1).value = "제품"
    for idx, period in enumerate(periods, start=2):
        year, month = period.split("-")
        sheet.cell(1, idx).value = "처방조제액(원)"
        sheet.cell(2, idx).value = f"{year}년 {int(month)}월"
        sheet.cell(3, idx).value = 1
    sheet.cell(3, 1).value = "테스트"
    workbook.save(path)
    workbook.close()
    return path


def _ubist_row(**overrides):
    row = dict.fromkeys(COLUMNS)
    row.update(
        {
            "제조사": "테스트제조사",
            "국내/외자": "국내",
            "판매사": "테스트판매사",
            "판매사2": "테스트판매사2",
            "제품": "테스트제품",
            "ATC": "A10BA",
            "브랜드": "테스트브랜드",
            "약가": "100",
            "성분": "테스트성분",
            "성분용량": "10mg",
            "일반/전문": "전문",
            "약품코드": "P001",
            "제형": "정제",
            "투여경로": "경구",
            "급여구분": "급여",
            "종별": "의원",
            "진료과": "내과",
            "연령": "전체",
            "성별": "전체",
            "period_yyyymm": "2026-02",
            "rx_amt": 100.0,
            "rx_cnt": 10.0,
            "rx_qty": 20.0,
            "source_file": "a.xlsx",
            "source_folder": "A",
            "source_sheet": "Sheet1",
            "source_row_no": 3,
            "ingested_at": "2026-06-17T00:00:00",
        }
    )
    row.update(overrides)
    return row


class _ReadOnlyWorksheet:
    def iter_rows(self, *, values_only: bool):
        assert values_only is True
        return iter(
            [
                (None, None, "처방조제액(원)"),
                ("제품", "Generic", "2025년 7월"),
                ("테스트제품", "Original", 1),
            ]
        )


class _ReadOnlyWorkbook:
    sheetnames = ["Sheet1"]

    def __init__(self) -> None:
        self.closed = False

    def __getitem__(self, sheet_name: str) -> _ReadOnlyWorksheet:
        assert sheet_name == "Sheet1"
        return _ReadOnlyWorksheet()

    def close(self) -> None:
        self.closed = True


def test_generic_lookup_closes_each_read_only_workbook(monkeypatch, tmp_path):
    workbook = _ReadOnlyWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: workbook)

    lookup = build_generic_lookup([tmp_path / "source.xlsx"])

    assert lookup["product:테스트제품"] == "Original"
    assert workbook.closed is True


def test_row_iterator_closes_read_only_workbook_after_full_consumption(monkeypatch, tmp_path):
    workbook = _ReadOnlyWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: workbook)

    rows = list(iter_xlsx_rows(tmp_path / "source.xlsx"))

    assert len(rows) == 1
    assert workbook.closed is True


def test_row_iterator_closes_read_only_workbook_when_consumer_stops_early(monkeypatch, tmp_path):
    workbook = _ReadOnlyWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: workbook)
    rows = iter_xlsx_rows(tmp_path / "source.xlsx")

    next(rows)
    rows.close()

    assert workbook.closed is True


def test_row_iterator_rejects_non_numeric_metric_instead_of_converting_to_null(tmp_path):
    workbook_path = _workbook(tmp_path / "invalid-metric.xlsx", ["2025-07"])
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.active.cell(3, 2).value = "not-a-number"
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(ValueError, match="non-numeric UBIST metric"):
        list(iter_xlsx_rows(workbook_path))


def _write_partition(target: Path, period: str, rows: list[dict[str, object]]) -> None:
    year, month = period.split("-")
    path = target / f"year={year}" / f"month={month}" / "data.parquet"
    path.parent.mkdir(parents=True, exist_ok=True)
    table = pa.Table.from_pandas(pd.DataFrame(rows).reindex(columns=COLUMNS), schema=SCHEMA, preserve_index=False)
    pq.write_table(table, path)


def test_business_grain_merge_treats_patent_metadata_as_values():
    rows = [
        _ubist_row(source_file="a.xlsx", source_row_no=4, ingested_at="2026-06-17T00:00:00"),
        _ubist_row(
            source_file="b.xlsx",
            source_row_no=3,
            PMS만료일="2030-01-01",
            Generic="Original",
            ingested_at="2026-06-18T00:00:00",
        ),
    ]

    deduped, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    assert len(deduped) == 1
    assert report.duplicate_groups == 0
    assert report.duplicate_rows_removed == 0
    assert report.conflict_groups == 1
    assert deduped.iloc[0]["source_file"] == "b.xlsx"
    assert deduped.iloc[0]["PMS만료일"] == "2030-01-01"
    assert deduped.iloc[0]["Generic"] == "Original"


def test_business_grain_merge_resolves_metric_conflicts_by_ingest_time():
    rows = [
        _ubist_row(source_file="a.xlsx", rx_amt=100.0, ingested_at="2026-06-17T00:00:00"),
        _ubist_row(source_file="b.xlsx", rx_amt=101.0, ingested_at="2026-06-18T00:00:00"),
    ]

    deduped, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    assert len(deduped) == 1
    assert deduped.iloc[0]["rx_amt"] == 101.0
    assert report.duplicate_rows_removed == 0
    assert report.conflict_groups == 1
    assert report.conflict_rows == 2


def test_business_grain_dedup_preserves_non_patent_canonical_differences():
    rows = [
        _ubist_row(source_file="a.xlsx", 제조사="제조사A"),
        _ubist_row(source_file="b.xlsx", 제조사="제조사B"),
    ]

    deduped, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    assert len(deduped) == 2
    assert report.duplicate_rows_removed == 0
    assert set(deduped["제조사"]) == {"제조사A", "제조사B"}


def test_partition_dedup_spills_without_materializing_full_table(monkeypatch, tmp_path):
    _write_partition(
        tmp_path,
        "2026-02",
        [
            _ubist_row(source_file="a.xlsx", source_row_no=4, ingested_at="2026-06-17T00:00:00"),
            _ubist_row(
                source_file="b.xlsx",
                source_row_no=3,
                PMS만료일="2030-01-01",
                Generic="Original",
                ingested_at="2026-06-18T00:00:00",
            ),
            _ubist_row(
                source_file="c.xlsx",
                source_row_no=5,
                제품="충돌제품",
                약품코드="P002",
                ingested_at="2026-06-17T00:00:00",
            ),
            _ubist_row(
                source_file="d.xlsx",
                source_row_no=6,
                제품="충돌제품",
                약품코드="P002",
                rx_amt=101.0,
                ingested_at="2026-06-18T00:00:00",
            ),
        ],
    )
    partition = tmp_path / "year=2026" / "month=02" / "data.parquet"

    def fail_full_table_read(*_args, **_kwargs):
        raise AssertionError("partition dedup must not materialize the full Parquet table")

    monkeypatch.setattr(pq, "read_table", fail_full_table_read)

    report = deduplicate_partition_file(partition, "2026-02")
    assert pq.ParquetFile(partition).schema_arrow == SCHEMA
    with duckdb.connect() as connection:
        rows = connection.execute(
            "SELECT source_file, rx_amt, PMS만료일, Generic FROM read_parquet(?) ORDER BY source_file",
            [str(partition)],
        ).fetchall()

    assert report.rows_before == 4
    assert report.rows_after == 2
    assert report.duplicate_rows_removed == 0
    assert report.conflict_groups == 2
    assert report.conflict_rows == 4
    assert rows == [
        ("b.xlsx", 100.0, "2030-01-01", "Original"),
        ("d.xlsx", 101.0, None, None),
    ]


def test_partition_dedup_matches_in_memory_contract(tmp_path):
    rows = [
        _ubist_row(source_file="a.xlsx", source_row_no=4, ingested_at="2026-06-17T00:00:00"),
        _ubist_row(
            source_file="b.xlsx",
            source_row_no=3,
            PMS만료일="2030-01-01",
            Generic="Original",
            ingested_at="2026-06-18T00:00:00",
        ),
        _ubist_row(
            source_file="c.xlsx",
            source_row_no=5,
            제품="충돌제품",
            약품코드="P002",
            ingested_at="2026-06-17T00:00:00",
        ),
        _ubist_row(
            source_file="d.xlsx",
            source_row_no=6,
            제품="충돌제품",
            약품코드="P002",
            rx_amt=101.0,
            ingested_at="2026-06-18T00:00:00",
        ),
        _ubist_row(source_file="e.xlsx", source_row_no=7, 제품="고유제품", 약품코드="P003", rx_amt=None),
    ]
    expected_frame, expected_report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")
    _write_partition(tmp_path, "2026-02", rows)
    partition = tmp_path / "year=2026" / "month=02" / "data.parquet"

    actual_report = deduplicate_partition_file(partition, "2026-02")
    with duckdb.connect() as connection:
        actual_frame = connection.execute(
            f"SELECT {', '.join(f'\"{column}\"' for column in COLUMNS)} FROM read_parquet(?)",
            [str(partition)],
        ).df()

    assert actual_report == expected_report
    pd.testing.assert_frame_equal(
        actual_frame.reset_index(drop=True),
        expected_frame.reset_index(drop=True),
        check_dtype=False,
    )


def test_row_merge_is_idempotent_for_same_file_reupload() -> None:
    # Given: the same business row is uploaded twice at different times.
    rows = [
        _ubist_row(ingested_at="2026-07-27T10:00:00"),
        _ubist_row(ingested_at="2026-07-27T11:00:00"),
    ]

    # When: rows are merged by business identity and values.
    merged, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    # Then: one latest lineage row remains without a value conflict.
    assert len(merged) == 1
    assert merged.iloc[0]["ingested_at"] == "2026-07-27T11:00:00"
    assert report.conflict_groups == 0


def test_row_merge_deduplicates_same_content_with_different_file_names() -> None:
    # Given: two differently named files contain the same business row and values.
    rows = [
        _ubist_row(source_file="z-old.xlsx", ingested_at="2026-07-27T10:00:00"),
        _ubist_row(source_file="a-new.xlsx", ingested_at="2026-07-27T11:00:00"),
    ]

    # When: rows are merged.
    merged, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    # Then: the row is not double-counted and filename ordering is irrelevant.
    assert len(merged) == 1
    assert merged.iloc[0]["source_file"] == "a-new.xlsx"
    assert report.duplicate_rows_removed == 1


def test_row_merge_uses_later_ingested_value_and_records_metric_conflict() -> None:
    # Given: the same identity has different metrics at different ingest times.
    rows = [
        _ubist_row(source_file="a-old.xlsx", rx_amt=100.0, ingested_at="2026-07-27T10:00:00"),
        _ubist_row(source_file="z-new.xlsx", rx_amt=125.0, ingested_at="2026-07-27T11:00:00"),
    ]

    # When: rows are merged.
    merged, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    # Then: the later value wins and the discarded value remains auditable.
    assert len(merged) == 1
    assert merged.iloc[0]["rx_amt"] == 125.0
    assert report.conflict_groups == 1
    assert report.conflicts[0]["winner"]["rx_amt"] == 125.0
    assert report.conflicts[0]["discarded"][0]["rx_amt"] == 100.0


def test_row_merge_stops_when_different_values_have_same_ingest_time() -> None:
    # Given: one ingest batch contains two different values for the same identity.
    rows = [
        _ubist_row(source_file="a.xlsx", rx_amt=100.0, ingested_at="2026-07-27T10:00:00"),
        _ubist_row(source_file="b.xlsx", rx_amt=125.0, ingested_at="2026-07-27T10:00:00"),
    ]

    # When/Then: the loader refuses to invent a winner.
    with pytest.raises(RuntimeError, match="same ingest time"):
        deduplicate_business_grain(pd.DataFrame(rows), "2026-02")


def test_row_merge_uses_later_ingested_patent_value_and_records_conflict() -> None:
    # Given: only a patent value differs and filename order favors the older row.
    rows = [
        _ubist_row(
            source_file="a-old.xlsx",
            PMS만료일="2030-01-01",
            ingested_at="2026-07-27T10:00:00",
        ),
        _ubist_row(
            source_file="z-new.xlsx",
            PMS만료일="2031-01-01",
            ingested_at="2026-07-27T11:00:00",
        ),
    ]

    # When: rows are merged.
    merged, report = deduplicate_business_grain(pd.DataFrame(rows), "2026-02")

    # Then: patent fields participate in value comparison and later ingest wins.
    assert len(merged) == 1
    assert merged.iloc[0]["PMS만료일"] == "2031-01-01"
    assert report.conflict_groups == 1


def test_incremental_load_reads_all_three_files_without_manifest_skip(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Given: three files are submitted and one filename already appears in the manifest.
    target = tmp_path / "target"
    _write_manifest(target, ["first.xlsx"])
    paths = [
        _workbook(tmp_path / "first.xlsx", ["2026-06"]),
        _workbook(tmp_path / "second.xlsx", ["2026-06"]),
        _workbook(tmp_path / "third.xlsx", ["2026-06"]),
    ]
    loaded_paths: list[Path] = []

    def fake_load_to_parquet(paths, target, *, mode, truncate, previous_manifest):
        loaded_paths.extend(paths)
        return {}

    monkeypatch.setattr("pipeline.etl.io.ubist_loader.load_to_parquet", fake_load_to_parquet)

    # When: the incremental loader runs.
    run_incremental_ubist_load(target=target, paths=paths)

    # Then: every submitted workbook reaches row-level merge regardless of filename history.
    assert loaded_paths == [path.resolve() for path in paths]


def test_generic_lookup_uses_existing_corpus_when_increment_has_no_generic(
    tmp_path: Path,
) -> None:
    # Given: the durable corpus knows Generic but the increment contains no patent block.
    target = tmp_path / "target"
    _write_partition(
        target,
        "2026-05",
        [_ubist_row(제품="전량제품", 약품코드="FULL001", Generic="Original")],
    )

    # When: the deterministic lookup is built for an empty incremental file set.
    lookup = build_generic_lookup([], parquet_root=target)

    # Then: Generic is derived from the full durable corpus.
    assert lookup["code:full001"] == "Original"
    assert lookup["product:전량제품"] == "Original"


def test_incremental_load_stops_when_no_workbooks_are_selected(tmp_path: Path) -> None:
    # Given: a valid target manifest exists but the upload contains no xlsx files.
    target = tmp_path / "target"
    _write_manifest(target, [])

    # When/Then: absence is explicit rather than a successful no-op.
    with pytest.raises(RuntimeError, match="No xlsx files selected"):
        run_incremental_ubist_load(target=target, paths=[])


def _create_partition_markers(target: Path, periods: list[str]) -> None:
    for period in periods:
        year, month = period.split("-")
        partition = target / f"year={year}" / f"month={month}"
        partition.mkdir(parents=True)
        (partition / "data.parquet").touch()


def test_load_retention_keeps_all_66_months(tmp_path: Path) -> None:
    periods = pd.period_range("2021-01", periods=66, freq="M").strftime("%Y-%m").tolist()
    _create_partition_markers(tmp_path, periods)

    removed = prune_ubist_partitions(tmp_path)

    assert removed == ()
    assert len(list(tmp_path.glob("year=*/month=*/data.parquet"))) == 66


def test_load_retention_removes_oldest_partition_after_72_months(tmp_path: Path) -> None:
    periods = pd.period_range("2021-01", periods=73, freq="M").strftime("%Y-%m").tolist()
    _create_partition_markers(tmp_path, periods)

    removed = prune_ubist_partitions(tmp_path)

    assert removed == ("2021-01",)
    remaining = sorted(
        f"{path.parent.parent.name.removeprefix('year=')}-{path.parent.name.removeprefix('month=')}"
        for path in tmp_path.glob("year=*/month=*/data.parquet")
    )
    assert len(remaining) == 72
    assert remaining[0] == "2021-02"
