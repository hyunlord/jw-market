from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from pipeline.scripts.etl.brand_activity.ingest_keyword import (  # noqa: E402
    read_keyword_events,
    read_keyword_message_counts,
)
from pipeline.scripts.etl.brand_activity.ingest_keyword_meeting import quote_schema_name  # noqa: E402
from pipeline.scripts.etl.brand_activity.ingest_meeting import read_meeting_events  # noqa: E402
from pipeline.scripts.etl.brand_activity.km_validation import (  # noqa: E402
    compare_message_count_overlaps,
    compare_core_to_message_count,
)


def _save_keyword_fixture(path: Path, message_july: int = 2) -> None:
    """Create a minimal Keyword workbook with duplicate event rows preserved."""
    workbook = openpyxl.Workbook()
    workbook.active.title = "Main"
    count_sheet = workbook.create_sheet("2025 Message Count")
    count_sheet.append(["PRODUCT NAME", "June. 25", "July. 25"])
    count_sheet.append(["ATOZET", 95, message_july])
    sheet = workbook.create_sheet("Keywords")
    sheet.append(
        [
            "Related date",
            "VISIT LOCATION",
            "SPECIALTY NAME",
            "REP# CO",
            "PRODUCT NAME",
            "THERAPEUTIC CLASS",
            "KEYWORDS",
            "INTEREST",
            "Prescription frequency",
            "Prescription evolution",
            "Abstract and clinical literature / data",
            "Patient educational literature",
            "Promotional product literature",
            "SAMPLES LEFT",
            "OTHER MATERIALS LEFT",
            "WHAT OTHER MATERIALS",
            "OTHER COMMENTS",
        ]
    )
    duplicate_row = [
        "July 2025",
        "Clinic",
        "Cardiology",
        "JW",
        "ATOZET",
        "C10C0",
        "KW_FIXTURE_TEXT",
        "VERY USEFUL",
        "frequently",
        "increase",
        "YES",
        "NO",
        "NO",
        "NO",
        "NO",
        "",
        "",
    ]
    sheet.append(duplicate_row)
    sheet.append(duplicate_row)
    workbook.save(path)
    workbook.close()


def _save_meeting_fixture(path: Path) -> None:
    """Create a minimal Meeting workbook with the observed source aliases."""
    workbook = openpyxl.Workbook()
    workbook.active.title = "Main"
    count_sheet = workbook.create_sheet("2025 Message Count")
    count_sheet.append(["Product name", "July 25"])
    count_sheet.append(["ATOZET", 1])
    sheet = workbook.create_sheet("Meetings")
    sheet.append(
        [
            "Meeting date",
            "Meeting Topic",
            "Meeting Format",
            "Pharmaceutical Sponsor",
            "Non-Pharmaceutical Sponsor",
            "No at Meeting",
            "Product name",
            "Therapeutic Class",
            "Prescription frequency text",
            "Prescription change text",
            "Interest (Information Presented)",
            "Verbatim Message (Information Caught Attention)",
            "Other Comments",
            "",
        ]
    )
    sheet.append(
        [
            datetime(2025, 7, 15),
            "TOPIC_FIXTURE",
            "Internet live broadcast",
            "JW",
            "",
            20,
            "ATOZET",
            "C10C0",
            "occasionally",
            "remain unchanged",
            "SOMEWHAT USEFUL",
            "VERBATIM_FIXTURE_TEXT",
            "",
            "",
        ]
    )
    workbook.save(path)
    workbook.close()


def test_keyword_reader_preserves_full_duplicate_rows_and_redacts_text(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Keywords for JW July. 25.xlsx"
    _save_keyword_fixture(workbook_path)

    events = read_keyword_events(workbook_path)

    assert len(events) == 2
    assert [event.source_row_no for event in events] == [2, 3]
    assert events[0].keyword_text == events[1].keyword_text
    redacted = events[0].to_redacted_row()
    assert "keyword_text" not in redacted
    assert redacted["keyword_text_len"] == 15
    assert len(str(redacted["keyword_text_sha256"])) == 64


def test_meeting_reader_uses_alias_headers_and_normalizes_dates(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Meetings for JW July. 25.xlsx"
    _save_meeting_fixture(workbook_path)

    events = read_meeting_events(workbook_path)

    assert len(events) == 1
    assert events[0].meeting_date == "2025-07-15"
    assert events[0].period_ym == "2025-07"
    assert events[0].prescription_frequency == "occasionally"
    assert events[0].verbatim_message == "VERBATIM_FIXTURE_TEXT"
    redacted = events[0].to_redacted_row()
    assert "verbatim_message" not in redacted
    assert redacted["verbatim_message_len"] == 21


def test_message_count_overlap_reports_mismatches_and_product_differences(tmp_path: Path) -> None:
    july_path = tmp_path / "Keywords for JW July. 25.xlsx"
    aug_path = tmp_path / "Keywords for JW Aug. 25.xlsx"
    _save_keyword_fixture(july_path, message_july=2)
    _save_keyword_fixture(aug_path, message_july=3)

    july_cells = read_keyword_message_counts(july_path)
    aug_cells = read_keyword_message_counts(aug_path)
    report = compare_message_count_overlaps([july_cells, aug_cells])

    assert report["compared_cells"] == 2
    assert report["matched_cells"] == 1
    assert report["mismatch_cells"] == 1
    assert report["mismatches"][0]["product_name"] == "ATOZET"
    assert report["mismatches"][0]["month_ym"] == "2025-07"


def test_stage_schema_name_must_be_exact_isolated_schema() -> None:
    assert quote_schema_name("jw_brand_activity_stage") == "jw_brand_activity_stage"
    assert quote_schema_name("jw_brand_activity_repro") == "jw_brand_activity_repro"
    with pytest.raises(ValueError, match="brand-activity scratch"):
        quote_schema_name("jw_brand_activity")
    with pytest.raises(ValueError, match="brand-activity scratch"):
        quote_schema_name("other_stage")


def test_core_to_message_count_alignment_uses_core_event_counts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Keywords for JW July. 25.xlsx"
    _save_keyword_fixture(workbook_path, message_july=2)
    events = read_keyword_events(workbook_path)
    message_cells = read_keyword_message_counts(workbook_path)

    report = compare_core_to_message_count("keyword", events, message_cells)

    assert report["checked_product_months"] == 1
    assert report["matched_product_months"] == 1
    assert report["mismatch_product_months"] == 0
