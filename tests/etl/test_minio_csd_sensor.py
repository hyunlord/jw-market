"""CSD MinIO sensor contract: filter, validation gate, idempotency, dry-run."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from pipeline.scripts.etl.brand_activity.csd_core import EXPECTED_HEADERS
from pipeline.scripts.etl.brand_activity.minio_csd_sensor import (
    ObjectInfo,
    is_candidate,
    process_once,
    validate_csd_workbook,
)


def _make_workbook(path: Path, *, sheets=("LIVALO Market",), headers=EXPECTED_HEADERS, data_rows=2) -> Path:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_name in sheets:
        sheet = workbook.create_sheet(sheet_name)
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=7, column=column, value=header)
        for row_index in range(data_rows):
            sheet.cell(row=8 + row_index, column=1, value="2025-08")
            sheet.cell(row=8 + row_index, column=2, value="LIVALO Market")
            sheet.cell(row=8 + row_index, column=3, value="TOTAL")
            sheet.cell(row=8 + row_index, column=4, value="TOTAL")
            sheet.cell(row=8 + row_index, column=5, value="리바로")
            sheet.cell(row=8 + row_index, column=6, value="JW중외제약")
            sheet.cell(row=8 + row_index, column=7, value="JW중외제약")
            sheet.cell(row=8 + row_index, column=8, value=3)
    workbook.save(path)
    return path


GOOD_KEY = "CSD/ChannelDynamics (test)/ChannelDynamics_JW Pharma Regional Report_Aug.25.xlsx"


def _obj(key=GOOD_KEY, etag="etag-1", size=5 * 1024 * 1024) -> ObjectInfo:
    return ObjectInfo(key=key, etag=etag, size=size)


def _run(tmp_path, objects, workbook_builder, *, dry_run=False, marker=None, runner_calls=None):
    marker_path = tmp_path / "marker.json"
    if marker is not None:
        marker_path.write_text(json.dumps(marker))
    calls = runner_calls if runner_calls is not None else []

    def fetch(obj, target):
        workbook_builder(target)

    decisions = process_once(
        config=None,  # fetch/objects injected; no network
        bucket="jw-market-raw-iqvia",
        prefix="CSD/ChannelDynamics",
        marker_path=marker_path,
        namespace="llmops",
        cronjob="jw-brand-activity-run",
        dry_run=dry_run,
        runner=lambda argv: calls.append(argv) or (0, "created"),
        objects=objects,
        fetch=fetch,
    )
    return decisions, marker_path, calls


def test_candidate_filter_excludes_mac_debris_and_small_files():
    assert is_candidate(_obj())
    assert not is_candidate(_obj(key="CSD/x/._ChannelDynamics_...xlsx", size=212))
    assert not is_candidate(_obj(key="CSD/x/.DS_Store", size=8192))
    assert not is_candidate(_obj(key="CSD/x/~$report.xlsx"))
    assert not is_candidate(_obj(key="CSD/x/report.csv"))
    assert not is_candidate(_obj(size=4096))  # too small to be a real report


def test_valid_workbook_triggers_run_job(tmp_path):
    decisions, marker_path, calls = _run(tmp_path, [_obj()], lambda p: _make_workbook(p))

    assert decisions[0]["action"] == "triggered:created"
    assert len(calls) == 1
    assert calls[0][:5] == ["kubectl", "-n", "llmops", "create", "job"]
    marker = json.loads(marker_path.read_text())
    assert marker[GOOD_KEY]["status"] == "triggered"


def test_structurally_broken_workbook_is_blocked(tmp_path):
    broken_headers = tuple(h for h in EXPECTED_HEADERS if h != "Master product")
    decisions, marker_path, calls = _run(
        tmp_path, [_obj()], lambda p: _make_workbook(p, headers=broken_headers)
    )

    assert decisions[0]["action"] == "blocked_validation_failed"
    assert "Master product" in decisions[0]["validation"]
    assert calls == []  # fail-closed: no job
    marker = json.loads(marker_path.read_text())
    assert marker[GOOD_KEY]["status"] == "validation_failed"


def test_workbook_without_market_sheets_is_blocked(tmp_path):
    decisions, _, calls = _run(tmp_path, [_obj()], lambda p: _make_workbook(p, sheets=("Summary",)))

    assert decisions[0]["action"] == "blocked_validation_failed"
    assert calls == []


def test_duplicate_object_is_noop(tmp_path):
    marker = {GOOD_KEY: {"etag": "etag-1", "status": "triggered"}}
    decisions, _, calls = _run(tmp_path, [_obj()], lambda p: _make_workbook(p), marker=marker)

    assert decisions[0]["action"] == "noop_already_processed"
    assert calls == []


def test_changed_etag_reprocesses(tmp_path):
    marker = {GOOD_KEY: {"etag": "old-etag", "status": "triggered"}}
    decisions, _, calls = _run(tmp_path, [_obj(etag="etag-2")], lambda p: _make_workbook(p), marker=marker)

    assert decisions[0]["action"] == "triggered:created"
    assert len(calls) == 1


def test_dry_run_writes_nothing_and_triggers_nothing(tmp_path):
    decisions, marker_path, calls = _run(tmp_path, [_obj()], lambda p: _make_workbook(p), dry_run=True)

    assert decisions[0]["action"] == "would_trigger"
    assert calls == []
    assert not marker_path.exists()


def test_validate_rejects_empty_data(tmp_path):
    path = _make_workbook(tmp_path / "empty.xlsx", data_rows=0)
    ok, detail = validate_csd_workbook(path)

    assert not ok
    assert "no data rows" in detail
