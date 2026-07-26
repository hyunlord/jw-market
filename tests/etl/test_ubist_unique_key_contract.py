from __future__ import annotations

from pathlib import Path

import duckdb
import openpyxl
import pytest

from pipeline.etl.io.ubist_loader import (
    BUSINESS_GRAIN_COLUMNS,
    BUSINESS_METRIC_COLUMNS,
    CANONICAL_DIMENSIONS,
    canonical_header,
    iter_xlsx_rows,
    load_to_parquet,
)


def _values_by_header(**overrides: object) -> dict[str, object]:
    values: dict[str, object] = {
        "제조사": "테스트제조사",
        "국내/외자": "국내",
        "판매사": "테스트판매사",
        "판매사2": "테스트판매사2",
        "제품": "테스트제품",
        "ATC": "A10BA",
        "브랜드": "테스트브랜드",
        "약가": "100",
        "성분": "테스트성분",
        "성분용량": "10mg",
        "일반/전문": "전문",
        "약품코드": "P001",
        "제형": "정제",
        "투여경로": "경구",
        "급여구분": "급여",
        "종별": "의원",
        "진료과": "내과",
        "연령": "전체",
        "성별": "전체",
        "처방조제액(원)": 100.0,
        "처방건수_P": 10.0,
        "처방량_P": 20.0,
    }
    values.update(overrides)
    return values


def _write_ubist_workbook(
    path: Path,
    *,
    rows: list[dict[str, object]] | None = None,
    dimension_headers: list[str] | None = None,
    extra_header: str | None = None,
) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "UBIST"
    headers = list(dimension_headers or CANONICAL_DIMENSIONS)
    if extra_header:
        headers.append(extra_header)
    metric_headers = ["처방조제액(원)", "처방건수_P", "처방량_P"]
    for column_index, header in enumerate(headers + metric_headers, start=1):
        if header in metric_headers:
            sheet.cell(1, column_index).value = header
            sheet.cell(2, column_index).value = "2026년 2월"
        else:
            sheet.cell(2, column_index).value = header
    for row_index, row_values in enumerate(rows or [_values_by_header()], start=3):
        for column_index, header in enumerate(headers + metric_headers, start=1):
            canonical = canonical_header(header) or header
            sheet.cell(row_index, column_index).value = row_values.get(canonical)
    workbook.save(path)
    workbook.close()
    return path


def _business_metric_values(path: Path) -> tuple[object, ...]:
    rows = list(iter_xlsx_rows(path))
    assert len(rows) == 1
    return tuple(rows[0][1][column] for column in BUSINESS_METRIC_COLUMNS)


def test_iter_xlsx_rows_uses_normalized_header_names_when_columns_are_shuffled(tmp_path: Path) -> None:
    # Given: the same UBIST fact in fixed-order and shuffled workbooks.
    baseline = _write_ubist_workbook(tmp_path / "baseline.xlsx")
    shuffled_headers = [
        "제품",
        " 약품코드 ",
        "성별",
        "연령",
        "진료과",
        "종별",
        "급여구분",
        "투여경로",
        "제형",
        "일반/전문",
        "성분용량",
        "성분",
        "약가",
        "브랜드",
        "ATC",
        "판매사2",
        "판매사",
        "국내/외자",
        "제조사",
    ]
    shuffled = _write_ubist_workbook(
        tmp_path / "shuffled.xlsx",
        dimension_headers=shuffled_headers,
        extra_header="원천 메모",
    )

    # When: rows are streamed through the workbook loader.
    baseline_values = _business_metric_values(baseline)
    shuffled_values = _business_metric_values(shuffled)

    # Then: business identity and metrics match by normalized column name, not position.
    assert shuffled_values == baseline_values


def test_iter_xlsx_rows_rejects_null_natural_key_before_staging(tmp_path: Path) -> None:
    # Given: a workbook row with a missing normalized natural-key value.
    workbook = _write_ubist_workbook(
        tmp_path / "missing-key.xlsx",
        rows=[_values_by_header(약품코드=None)],
    )

    # When / Then: the row is rejected before partition staging.
    with pytest.raises(RuntimeError, match="missing UBIST natural key column"):
        list(iter_xlsx_rows(workbook))


def test_iter_xlsx_rows_rejects_duplicate_natural_key_inside_same_file(tmp_path: Path) -> None:
    # Given: one source workbook repeats the same natural key.
    workbook = _write_ubist_workbook(
        tmp_path / "duplicate-key.xlsx",
        rows=[_values_by_header(), _values_by_header(**{"처방조제액(원)": 999.0})],
    )

    # When / Then: the duplicate is rejected before partition staging.
    with pytest.raises(RuntimeError, match="duplicate UBIST natural key"):
        list(iter_xlsx_rows(workbook))


def test_load_to_parquet_preserves_cross_source_overlap_for_partition_dedup(tmp_path: Path) -> None:
    # Given: two different source workbooks contain the same legal overlap.
    first = _write_ubist_workbook(tmp_path / "first.xlsx")
    second = _write_ubist_workbook(tmp_path / "second.xlsx")
    target = tmp_path / "target"

    # When: the loader stages both sources.
    stats = load_to_parquet([first, second], target, mode="replace", truncate=False)

    # Then: source-level validation allowed the overlap, and partition dedup collapsed it.
    assert stats["2026-02"].row_count == 1
    with duckdb.connect() as connection:
        rows = connection.execute(
            "SELECT source_file, rx_amt FROM read_parquet(?)",
            [str(target / "year=2026" / "month=02" / "data.parquet")],
        ).fetchall()
    assert rows == [("first.xlsx", 100.0)]
