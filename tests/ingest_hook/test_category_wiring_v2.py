from __future__ import annotations

import hashlib
import json
import shutil
import unicodedata
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from pipeline.scripts.ingest_hook import config, job_runner
from pipeline.scripts.ingest_hook.app import IngestService, create_app
from pipeline.scripts.ingest_hook.category_map import resolve_category
from pipeline.scripts.ingest_hook.contract import load_manifest
from pipeline.scripts.ingest_hook.g3 import G3Error, validate


def _manifest(root: Path, category: str, workbook: Path, epoch: str = "2026-03", rows: int | None = 1) -> Path:
    payload = {
        "contract_version": "v2",
        "epoch": epoch,
        "category": category,
        "complete": True,
        "files": [{
            "path": workbook.name,
            "sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
            "rows": rows,
            "period_start": epoch,
            "period_end": epoch,
        }],
    }
    path = root / f"{category}.manifest.json"
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def _nsa(path: Path, headers: list[str] | None = None) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "NSA"
    sheet.append(headers or [
        "AUDIT CODE", "MFR CODE", "PRODUCT NAME", "PACK DESC", "DATA PERIOD",
        "Values LC", "Units", "Counting Units", "Dosage Units", "Price",
    ])
    sheet.append(["A", "M", "Brand", "Pack", "2026-03", 1, 1, 1, 1, 1])
    workbook.save(path)


def _keyword(path: Path, headers: list[str] | None = None) -> None:
    from pipeline.scripts.etl.brand_activity.ingest_keyword import KEYWORD_HEADERS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Keywords"
    sheet.append(headers or list(KEYWORD_HEADERS))
    sheet.append(["Mar. 26", "Seoul", "IM", "JW", "Brand", "A10B0", "keyword", "high", "1", "up", "N", "N", "N", "N", "N", "", ""])
    workbook.save(path)


def _csd(path: Path, headers: list[str] | None = None) -> None:
    from pipeline.scripts.etl.brand_activity.csd_core import EXPECTED_HEADERS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "A Market"
    for _ in range(6):
        sheet.append([])
    sheet.append(headers or list(EXPECTED_HEADERS))
    sheet.append(["Mar. 26", "Market", "TOTAL", "TOTAL", "Brand", "Maker", "JW", 1])
    workbook.save(path)


def _mi_master(path: Path) -> None:
    source = Path(
        "data/JW 주요 약품 수동 매핑/"
        "MI팀_시장분석 AI_시장 분석 Master Version (원본파일 점검용 재공유 2026.05.18).xlsx"
    )
    shutil.copyfile(source, path)


@pytest.mark.parametrize(
    ("category", "builder", "epoch"),
    [
        ("iqvia_nsa", _nsa, "2026-Q1"),
        ("iqvia_csd_channel", _csd, "2026-03"),
        ("iqvia_csd_keyword", _keyword, "2026-03"),
        ("mi_master", _mi_master, "2026-03"),
    ],
)
def test_v1_each_new_category_has_a_real_workbook_contract(tmp_path, category, builder, epoch):
    workbook = tmp_path / f"{category}.xlsx"
    builder(workbook)
    declared_rows = None if category == "mi_master" else 1
    report = validate(load_manifest(_manifest(tmp_path, category, workbook, epoch, declared_rows)), resolve_category(category), tmp_path)
    assert report.total_rows >= 1
    assert report.observed_periods == {epoch}


def test_v2_wrong_category_workbook_is_rejected(tmp_path):
    workbook = tmp_path / "nsa.xlsx"
    _nsa(workbook)
    manifest = load_manifest(_manifest(tmp_path, "iqvia_csd_keyword", workbook, "2026-Q1"))
    with pytest.raises(G3Error, match="Keyword|Keywords|structure"):
        validate(manifest, resolve_category("iqvia_csd_keyword"), tmp_path)


def test_v2_missing_required_header_is_rejected(tmp_path):
    workbook = tmp_path / "keyword.xlsx"
    _keyword(workbook, headers=["Related date", "PRODUCT NAME"])
    manifest = load_manifest(_manifest(tmp_path, "iqvia_csd_keyword", workbook))
    with pytest.raises(G3Error, match="missing"):
        validate(manifest, resolve_category("iqvia_csd_keyword"), tmp_path)


def test_v2_header_nfkc_case_and_space_variants_are_accepted(tmp_path):
    workbook = tmp_path / "keyword.xlsx"
    from pipeline.scripts.etl.brand_activity.ingest_keyword import KEYWORD_HEADERS

    variants = [unicodedata.normalize("NFD", f"  {header.swapcase()}  ") for header in KEYWORD_HEADERS]
    _keyword(workbook, headers=variants)
    report = validate(
        load_manifest(_manifest(tmp_path, "iqvia_csd_keyword", workbook)),
        resolve_category("iqvia_csd_keyword"),
        tmp_path,
    )
    assert report.total_rows == 1


def test_legacy_collapsed_categories_fail_closed():
    for category in ("iqvia", "mimaster", "ubist_weekly", "iqvia_chso", "iqvia_csd_meeting"):
        with pytest.raises(ValueError, match="unknown ingest category"):
            resolve_category(category)


def test_legacy_category_new_upload_is_rejected_but_history_remains_queryable(
    sqlite_ledger, bucket, fake_transport
):
    client = TestClient(create_app(IngestService(sqlite_ledger, bucket, transport=fake_transport)))
    workbook = bucket / "legacy.csv"
    workbook.write_text("period,brand,value\n2026-Q1,A,1\n", encoding="utf-8")
    manifest = _manifest(bucket, "iqvia", workbook, "2026-Q1")

    response = client.post(
        "/ingest/webhook",
        json={"manifest_path": str(manifest.relative_to(bucket))},
    )

    assert response.status_code == 422
    assert sqlite_ledger.queued_categories() == []


@pytest.mark.parametrize(
    ("category", "builder", "epoch"),
    [
        ("iqvia_nsa", _nsa, "2026-Q1"),
        ("iqvia_csd_channel", _csd, "2026-03"),
        ("iqvia_csd_keyword", _keyword, "2026-03"),
        ("mi_master", _mi_master, "2026-03"),
    ],
)
def test_v1_new_categories_use_the_table_loader_contract(
    tmp_path, monkeypatch, category, builder, epoch
):
    spec = resolve_category(category)

    assert "pipeline.scripts.ingest_hook.category_table_load" in spec.load_argv
    assert spec.load_verify == "table_manifest"
    assert spec.load_batch_files is True
    assert spec.production_load_supported is (category == "iqvia_nsa")


@pytest.mark.parametrize(
    "category", ["iqvia_csd_channel", "iqvia_csd_keyword", "mi_master"]
)
def test_staging_artifact_loader_fails_closed_in_production(
    tmp_path, monkeypatch, category
):
    workbook = tmp_path / f"{category}.xlsx"
    {
        "iqvia_nsa": _nsa,
        "iqvia_csd_channel": _csd,
        "iqvia_csd_keyword": _keyword,
        "mi_master": _mi_master,
    }[category](workbook)
    epoch = "2026-Q1" if category == "iqvia_nsa" else "2026-03"
    manifest = load_manifest(_manifest(tmp_path, category, workbook, epoch, None))
    monkeypatch.delenv(config.ENV_LOAD_STAGING_ROOT, raising=False)
    monkeypatch.setenv(config.ENV_LOAD_TARGET_ROOT, str(tmp_path / "production-root"))

    with pytest.raises(RuntimeError, match="refusing production completion"):
        job_runner._real_load(manifest, resolve_category(category), tmp_path)
