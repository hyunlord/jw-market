"""Opt-in MariaDB proof for the isolated category table adapters.

Set ``INGEST_LOADER_INTEGRATION_DB`` to a disposable ``jw_ingest_*`` schema.
The fixture creates and drops only that schema and its derived ``_raw``/``_stage``
schemas. It must never point at an operational database.
"""
from __future__ import annotations

import os
import shutil
from pathlib import Path

import openpyxl
import pytest

from pipeline.scripts.ingest_hook import category_table_load, config


TARGET_DB = os.environ.get("INGEST_LOADER_INTEGRATION_DB", "")
pytestmark = pytest.mark.skipif(
    not TARGET_DB.startswith("jw_ingest_"),
    reason="requires disposable INGEST_LOADER_INTEGRATION_DB=jw_ingest_*",
)


def _connect(database: str | None = None):
    import pymysql

    return pymysql.connect(
        host=os.environ.get("MARIADB_HOST", "127.0.0.1"),
        port=int(os.environ.get("MARIADB_PORT", "3306")),
        user=os.environ.get("MARIADB_USER", "root"),
        password=os.environ.get("MARIADB_PASSWORD", ""),
        database=database,
        charset="utf8mb4",
        autocommit=True,
    )


@pytest.fixture(scope="module", autouse=True)
def disposable_schemas():
    if not TARGET_DB.startswith("jw_ingest_"):
        yield
        return
    schemas = (TARGET_DB, f"{TARGET_DB}_raw", f"{TARGET_DB}_stage")
    connection = _connect()
    try:
        with connection.cursor() as cursor:
            for schema in schemas:
                cursor.execute(f"DROP SCHEMA IF EXISTS `{schema}`")
            cursor.execute(f"CREATE SCHEMA `{TARGET_DB}`")
            cursor.execute(
                f"""
                CREATE TABLE `{TARGET_DB}`.`iqvia_nsa_quarterly_raw` (
                  `id` bigint unsigned NOT NULL AUTO_INCREMENT,
                  `source_file` varchar(255) NOT NULL,
                  `sheet_name` varchar(128),
                  `source_row_no` int NOT NULL,
                  `audit_code` varchar(64),
                  `audit_desc` varchar(255),
                  `mfr_code` varchar(64),
                  `mfr_name` varchar(255),
                  `period_yyyy` int,
                  `period_quarter` int,
                  `period_label` varchar(16),
                  `payload` longtext NOT NULL,
                  `source_master_version` varchar(255),
                  PRIMARY KEY (`id`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """
            )
        yield
    finally:
        with connection.cursor() as cursor:
            for schema in reversed(schemas):
                cursor.execute(f"DROP SCHEMA IF EXISTS `{schema}`")
        connection.close()


def _nsa(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "NSA"
    sheet.append([
        "DATA PERIOD", "AUDIT CODE", "AUDIT DESC", "MFR CODE", "MFR NAME",
        "PRODUCT NAME", "PACK DESC", "Values LC", "Units", "Counting Units",
        "Dosage Units", "Price",
    ])
    sheet.append([
        "2026-03-01", "KCPA", "Clinic", "JW", "JW", "BRAND", "PACK",
        100, 2, 2, 2, 50,
    ])
    workbook.save(path)


def _csd(path: Path) -> None:
    from pipeline.scripts.etl.brand_activity.csd_core import EXPECTED_HEADERS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "A Market"
    for _ in range(6):
        sheet.append([])
    sheet.append(list(EXPECTED_HEADERS))
    sheet.append(["Mar. 26", "Market", "TOTAL", "TOTAL", "Brand", "Maker", "JW", 7])
    workbook.save(path)


def _keyword(path: Path) -> None:
    from pipeline.scripts.etl.brand_activity.ingest_keyword import KEYWORD_HEADERS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Keywords"
    sheet.append(list(KEYWORD_HEADERS))
    sheet.append([
        "Mar. 26", "Seoul", "IM", "JW", "Brand", "A10B0", "keyword", "high",
        "1", "up", "N", "N", "N", "N", "N", "", "",
    ])
    workbook.save(path)


def _mi_master(path: Path) -> None:
    source = Path(
        "data/JW 주요 약품 수동 매핑/"
        "MI팀_시장분석 AI_시장 분석 Master Version (원본파일 점검용 재공유 2026.05.18).xlsx"
    )
    shutil.copyfile(source, path)


@pytest.mark.parametrize(
    ("category", "filename", "builder", "schema_suffix", "table"),
    [
        ("iqvia_nsa", "NSA.xlsx", _nsa, "", "iqvia_nsa_quarterly_raw"),
        ("iqvia_csd_channel", "ChannelDynamics Mar. 26.xlsx", _csd, "_raw", "raw_csd_channel_dynamics"),
        ("iqvia_csd_keyword", "Keywords for JW Mar. 26.xlsx", _keyword, "_raw", "raw_keyword_events"),
    ],
)
def test_real_adapter_loads_and_retries_idempotently(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    category: str,
    filename: str,
    builder,
    schema_suffix: str,
    table: str,
) -> None:
    source = tmp_path / filename
    builder(source)
    monkeypatch.setenv(config.ENV_LOAD_STAGING_DB, TARGET_DB)
    first = category_table_load.load(category, source, tmp_path / category / "first", "2026-03")
    second = category_table_load.load(category, source, tmp_path / category / "second", "2026-03")

    assert first["rows_loaded"] == 1
    assert first["rows_after"] == first["rows_before"] + 1
    assert second["rows_loaded"] == 0
    assert second["rows_after"] == second["rows_before"]
    connection = _connect(TARGET_DB)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT COUNT(*) FROM `{TARGET_DB}{schema_suffix}`.`{table}`")
            assert int(cursor.fetchone()[0]) >= 1
    finally:
        connection.close()


def test_one_category_parser_failure_does_not_block_another(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    invalid_csd = tmp_path / "ChannelDynamics Apr. 26.xlsx"
    invalid_csd.write_bytes(b"not-an-xlsx")
    keyword = tmp_path / "Keywords for JW Apr. 26.xlsx"
    _keyword(keyword)
    monkeypatch.setenv(config.ENV_LOAD_STAGING_DB, TARGET_DB)

    with pytest.raises(Exception):
        category_table_load.load(
            "iqvia_csd_channel", invalid_csd, tmp_path / "bad-csd", "2026-04"
        )
    result = category_table_load.load(
        "iqvia_csd_keyword", keyword, tmp_path / "good-keyword", "2026-04"
    )

    assert result["rows_loaded"] == 1


def test_mi_master_replaces_both_isolated_stage_tables_idempotently(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "MI Master.xlsx"
    _mi_master(source)
    monkeypatch.setenv(config.ENV_LOAD_STAGING_DB, TARGET_DB)

    first = category_table_load.load("mi_master", source, tmp_path / "mi-first", "2026-03")
    second = category_table_load.load("mi_master", source, tmp_path / "mi-second", "2026-03")

    assert first["rows_before"] == 0
    assert first["rows_after"] == first["rows_loaded"] > 0
    assert second["rows_before"] == second["rows_after"] == second["rows_loaded"]
