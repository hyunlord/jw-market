"""
prototype_06_chso_to_parquet.py
================================
IQVIA CHSO Sell-Out xlsx -> Parquet monthly partitions.

Reference policy mirrored from jw-market:
- SQL DDL source: sql/schema_iqvia.sql, stg_iqvia_chso_raw
- Product key: SHA256(audit_desc|mfr_name_kor|product_name_kor|
  pack_description|atc4|chc4)[:32]
- observations_json stores monthly Sell-Out observations parsed from the raw
  month headers.
- raw_extra_json preserves source fields not promoted to dedicated columns.
- Sparse policy: empty raw metric cells are skipped; explicit 0 is preserved.

Prototype storage:
- 1 row = 1 product_key x 1 month, matching the NSA/UBIST long partition
  pattern.
- Output partition key is period (YYYY-MM), written to
  parquet/chso/YYYY-MM.parquet.
- Core DDL columns are preserved; prototype helper columns are period,
  source_files, and is_grand_total.

Usage:
    python3 scripts/prototype_06_chso_to_parquet.py \\
        --input-file "data/IQVIA/CHSO/CHSO_KOR_SellOut_Basic_Feb-19-2026(2026년 1월까지).xlsx" \\
        --output-dir parquet/chso
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

try:
    import duckdb
    import pyarrow as pa
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install duckdb pyarrow openpyxl --break-system-packages")


# SQL DDL columns from /Users/rexxa/github/jw-market/sql/schema_iqvia.sql
DDL_COLS = [
    "product_key",
    "audit_desc",
    "mfr_name_kor",
    "product_name_kor",
    "pack_description",
    "chc1",
    "chc2",
    "chc3",
    "chc4",
    "atc1",
    "atc2",
    "atc3",
    "atc4",
    "observations_json",
    "raw_extra_json",
    "source_file",
    "source_sheet",
    "source_row_id",
    "ingested_at",
]

# Prototype helper columns for partition traceability, source tracking, and the
# user-requested Grand Total preservation flag.
OUTPUT_COLS = [
    "product_key",
    "audit_desc",
    "mfr_name_kor",
    "product_name_kor",
    "pack_description",
    "chc1",
    "chc2",
    "chc3",
    "chc4",
    "atc1",
    "atc2",
    "atc3",
    "atc4",
    "period",
    "observations_json",
    "raw_extra_json",
    "source_file",
    "source_sheet",
    "source_row_id",
    "source_files",
    "is_grand_total",
    "ingested_at",
]

META_HEADER_TO_COL = {
    "AUDIT DESC": "audit_desc",
    "MFR NAME KOR": "mfr_name_kor",
    "PRODUCT NAME KOR": "product_name_kor",
    "PACK DESCRIPTION": "pack_description",
    "CHC 1": "chc1",
    "CHC 2": "chc2",
    "CHC 3": "chc3",
    "CHC 4": "chc4",
    "ATC 1": "atc1",
    "ATC 2": "atc2",
    "ATC 3": "atc3",
    "ATC 4": "atc4",
}

PRODUCT_KEY_COLS = [
    "audit_desc",
    "mfr_name_kor",
    "product_name_kor",
    "pack_description",
    "atc4",
    "chc4",
]

METRIC_TO_JSON_KEY = {
    "VALUES LC SI PRICE": "value_lc_si_price",
    "UNITS": "units",
    "SELL OUT PRICE AVERAGE": "sell_out_price_average",
    "SELL IN PRICE": "sell_in_price",
}

CHSO_MONTH_HEADER_RE = re.compile(r"^(.+?)\s*\n\s*(\d{1,2})/(\d{4})\s*$")


def clean_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text or None


def clean_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=to_jsonable, sort_keys=True)


def normalize_metric_value(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace(",", "").strip()
        if not text:
            return None
        try:
            value = float(text)
        except ValueError:
            return None
    if isinstance(value, Decimal):
        value = float(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    try:
        as_float = float(value)
    except (TypeError, ValueError):
        return None
    if as_float.is_integer():
        return int(as_float)
    return as_float


def is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def compute_product_key(meta: dict[str, Any]) -> str:
    raw = "|".join(clean_str(meta.get(col)) or "" for col in PRODUCT_KEY_COLS)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]


def normalize_meta_header(header: Any) -> str | None:
    text = clean_header(header)
    if text is None:
        return None
    return re.sub(r"\s+", " ", text).strip()


def parse_chso_header(headers: tuple[Any, ...]) -> tuple[dict[str, int], list[dict[str, Any]]]:
    meta_indexes: dict[str, int] = {}
    time_columns: list[dict[str, Any]] = []

    for index, header in enumerate(headers):
        raw_header = clean_header(header)
        if raw_header is None:
            continue
        normalized_meta = normalize_meta_header(raw_header)
        staging_col = META_HEADER_TO_COL.get(normalized_meta or "")
        if staging_col:
            meta_indexes[staging_col] = index
            continue

        match = CHSO_MONTH_HEADER_RE.match(raw_header)
        if not match:
            continue
        raw_metric, month, year = match.groups()
        metric = re.sub(r"\s+", " ", raw_metric).strip()
        metric_key = METRIC_TO_JSON_KEY.get(metric)
        if metric_key is None:
            continue
        period = f"{int(year):04d}-{int(month):02d}"
        time_columns.append(
            {
                "index": index,
                "raw_header": raw_header,
                "raw_metric": metric,
                "metric_key": metric_key,
                "period": period,
            }
        )

    missing_meta = [col for col in META_HEADER_TO_COL.values() if col not in meta_indexes]
    if missing_meta:
        raise ValueError(f"CHSO required meta headers missing: {missing_meta}")
    if not time_columns:
        raise ValueError("CHSO month metric headers not found")
    return meta_indexes, time_columns


def resolve_input_file(input_file: str | None, input_dir: str) -> Path:
    if input_file:
        path = Path(input_file)
        if not path.exists():
            sys.exit(f"ERROR: input-file 없음: {path}")
        return path

    input_path = Path(input_dir)
    files = sorted(
        path for path in input_path.glob("*.xlsx") if not path.name.startswith(("~", "."))
    )
    if len(files) != 1:
        sys.exit(f"ERROR: CHSO input xlsx 는 1개여야 함: found={len(files)} dir={input_path}")
    return files[0]


def build_raw_extra_json(
    headers: tuple[Any, ...],
    values: tuple[Any, ...],
    promoted_indexes: set[int],
) -> str:
    extra: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if index in promoted_indexes:
            continue
        key = normalize_meta_header(header) or f"__blank_col_{index + 1}"
        value = values[index] if index < len(values) else None
        if value is None or str(value).strip() == "":
            continue
        extra[key] = to_jsonable(value)
    return dumps_json(extra)


def iter_chso_rows(
    xlsx_path: Path,
    ingested_at: str | None = None,
) -> Iterable[dict[str, Any]]:
    timestamp = ingested_at or datetime.now(timezone.utc).isoformat()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) != 1:
            raise ValueError(f"CHSO workbook expected 1 sheet, found {len(wb.sheetnames)}: {wb.sheetnames}")
        ws = wb[wb.sheetnames[0]]
        header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = tuple(header_values)
        meta_indexes, time_columns = parse_chso_header(headers)
        promoted_indexes = set(meta_indexes.values()) | {col["index"] for col in time_columns}

        for source_row_id, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values_tuple = tuple(values)
            if is_empty_row(values_tuple):
                continue

            meta = {}
            for col, index in meta_indexes.items():
                value = values_tuple[index] if index < len(values_tuple) else None
                meta[col] = clean_str(value)

            product_key = compute_product_key(meta)
            audit_desc = clean_str(meta.get("audit_desc"))
            is_grand_total = bool(audit_desc and audit_desc.lower() == "grand total")

            observations_by_period: dict[str, dict[str, int | float]] = defaultdict(dict)
            for col in time_columns:
                index = col["index"]
                value = values_tuple[index] if index < len(values_tuple) else None
                normalized = normalize_metric_value(value)
                if normalized is None:
                    continue
                observations_by_period[col["period"]][col["metric_key"]] = normalized

            raw_extra = build_raw_extra_json(headers, values_tuple, promoted_indexes)
            for period in sorted(observations_by_period):
                row = {"product_key": product_key}
                for col in META_HEADER_TO_COL.values():
                    row[col] = meta.get(col)
                row.update(
                    {
                        "period": period,
                        "observations_json": dumps_json({period: observations_by_period[period]}),
                        "raw_extra_json": raw_extra,
                        "source_file": xlsx_path.name,
                        "source_sheet": ws.title,
                        "source_row_id": source_row_id,
                        "source_files": xlsx_path.name,
                        "is_grand_total": is_grand_total,
                        "ingested_at": timestamp,
                    }
                )
                yield row
    finally:
        wb.close()


def analyze_chso_raw(xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        ws = wb[sheet_names[0]]
        headers = tuple(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        meta_indexes, time_columns = parse_chso_header(headers)

        metric_periods: dict[str, list[str]] = defaultdict(list)
        periods = set()
        for col in time_columns:
            metric_periods[col["raw_metric"]].append(col["period"])
            periods.add(col["period"])

        raw_rows = 0
        blank_rows = 0
        missing_key_rows = 0
        grand_total_rows = 0
        product_key_counts: Counter[str] = Counter()
        period_counts: Counter[str] = Counter()
        metric_value_counts: Counter[str] = Counter()
        metric_zero_counts: Counter[str] = Counter()
        metric_empty_counts: Counter[str] = Counter()
        explicit_zero_cells = 0
        non_empty_metric_cells = 0
        empty_metric_cells = 0
        row_period_counts: Counter[int] = Counter()

        for source_row_id, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values_tuple = tuple(values)
            if is_empty_row(values_tuple):
                blank_rows += 1
                continue
            raw_rows += 1

            meta = {}
            for col, index in meta_indexes.items():
                value = values_tuple[index] if index < len(values_tuple) else None
                meta[col] = clean_str(value)

            if any(clean_str(meta.get(col)) is None for col in PRODUCT_KEY_COLS):
                missing_key_rows += 1
            if (clean_str(meta.get("audit_desc")) or "").lower() == "grand total":
                grand_total_rows += 1
            product_key_counts[compute_product_key(meta)] += 1

            periods_with_value = set()
            for col in time_columns:
                index = col["index"]
                value = values_tuple[index] if index < len(values_tuple) else None
                normalized = normalize_metric_value(value)
                metric = col["raw_metric"]
                if normalized is None:
                    metric_empty_counts[metric] += 1
                    empty_metric_cells += 1
                    continue
                periods_with_value.add(col["period"])
                metric_value_counts[metric] += 1
                non_empty_metric_cells += 1
                if normalized == 0:
                    metric_zero_counts[metric] += 1
                    explicit_zero_cells += 1

            for period in periods_with_value:
                period_counts[period] += 1
            row_period_counts[len(periods_with_value)] += 1

        duplicate_groups = {key: count for key, count in product_key_counts.items() if count > 1}
        periods_sorted = sorted(periods)
        metric_summary = {}
        for metric, metric_period_list in sorted(metric_periods.items()):
            metric_summary[metric] = {
                "metric_key": METRIC_TO_JSON_KEY[metric],
                "columns": len(metric_period_list),
                "period_min": min(metric_period_list),
                "period_max": max(metric_period_list),
                "non_empty_cells": metric_value_counts[metric],
                "explicit_zero_cells": metric_zero_counts[metric],
                "empty_cells": metric_empty_counts[metric],
            }

        return {
            "source_file": xlsx_path.name,
            "size_mb": xlsx_path.stat().st_size / 1024 / 1024,
            "sheet_names": sheet_names,
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "header_row": 1,
            "meta_columns": len(meta_indexes),
            "time_columns": len(time_columns),
            "metrics": metric_summary,
            "period_count": len(periods_sorted),
            "period_min": periods_sorted[0],
            "period_max": periods_sorted[-1],
            "period_first5": periods_sorted[:5],
            "period_last5": periods_sorted[-5:],
            "raw_rows": raw_rows,
            "blank_rows": blank_rows,
            "missing_key_rows": missing_key_rows,
            "grand_total_raw_rows": grand_total_rows,
            "unique_product_keys": len(product_key_counts),
            "duplicate_product_key_groups": len(duplicate_groups),
            "duplicate_product_key_extra_rows": sum(count - 1 for count in duplicate_groups.values()),
            "sparse_long_rows": sum(period_counts.values()),
            "period_counts": dict(sorted(period_counts.items())),
            "non_empty_metric_cells": non_empty_metric_cells,
            "explicit_zero_cells": explicit_zero_cells,
            "empty_metric_cells": empty_metric_cells,
            "row_period_count_distribution": dict(sorted(row_period_counts.items())),
        }
    finally:
        wb.close()


def create_staging_table(con: duckdb.DuckDBPyConnection) -> None:
    types = []
    for col in OUTPUT_COLS:
        if col == "source_row_id":
            types.append((col, "INTEGER"))
        elif col == "is_grand_total":
            types.append((col, "BOOLEAN"))
        else:
            types.append((col, "VARCHAR"))
    cols_def = ", ".join(f'"{col}" {typ}' for col, typ in types)
    con.execute(f"CREATE OR REPLACE TABLE stg_raw ({cols_def})")


def bulk_insert(con: duckdb.DuckDBPyConnection, batch_rows: list[dict[str, Any]]) -> None:
    if not batch_rows:
        return
    batch_dict = {col: [] for col in OUTPUT_COLS}
    for row in batch_rows:
        for col in OUTPUT_COLS:
            batch_dict[col].append(row.get(col))
    arrow_table = pa.Table.from_pydict(batch_dict)
    con.register("_batch", arrow_table)
    con.execute("INSERT INTO stg_raw SELECT * FROM _batch")
    con.unregister("_batch")


def write_partition_parquets(con: duckdb.DuckDBPyConnection, output_dir: Path) -> list[tuple[str, int, float]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale in output_dir.glob("*.parquet"):
        stale.unlink()

    periods = [row[0] for row in con.execute("SELECT DISTINCT period FROM stg_raw ORDER BY period").fetchall()]
    select_cols = ", ".join(f'"{col}"' for col in OUTPUT_COLS)
    written = []
    for period in periods:
        out_path = output_dir / f"{period}.parquet"
        con.execute(
            f"""
            COPY (
                SELECT {select_cols}
                FROM stg_raw
                WHERE period = ?
                ORDER BY source_file, source_sheet, source_row_id, product_key
            ) TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)
            """,
            [period],
        )
        n_rows = con.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
        size_mb = out_path.stat().st_size / 1024 / 1024
        written.append((period, n_rows, size_mb))
    return written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", default=None)
    parser.add_argument("--input-dir", default="data/IQVIA/CHSO")
    parser.add_argument("--output-dir", default="parquet/chso")
    parser.add_argument("--db-path", default="staging_chso.duckdb")
    parser.add_argument("--chunk-size", type=int, default=5000)
    parser.add_argument("--keep-db", action="store_true")
    args = parser.parse_args()

    input_file = resolve_input_file(args.input_file, args.input_dir)
    output_dir = Path(args.output_dir)
    db_path = Path(args.db_path)

    print("=" * 72)
    print("IQVIA CHSO Sell-Out -> Parquet (monthly partitions)")
    print("=" * 72)
    print(f"  input file:    {input_file}")
    print(f"  output dir:    {output_dir}")
    print(f"  duckdb path:   {db_path}")
    print(f"  chunk size:    {args.chunk_size:,}")
    print(f"  core columns:  {len(DDL_COLS)} DDL + period/source_files/is_grand_total helpers")

    print("\n[Step 1] raw structure analysis")
    t0 = time.time()
    analysis = analyze_chso_raw(input_file)
    print(f"  sheet names:      {analysis['sheet_names']}")
    print(f"  sheet shape:      {analysis['max_row']:,} rows x {analysis['max_col']:,} cols")
    print(f"  header row:       {analysis['header_row']}")
    print(f"  meta columns:     {analysis['meta_columns']}")
    print(f"  time columns:     {analysis['time_columns']}")
    print(f"  metrics:          {', '.join(analysis['metrics'].keys())}")
    print(
        f"  periods:          {analysis['period_min']} ~ {analysis['period_max']} "
        f"({analysis['period_count']})"
    )
    print(f"  raw rows:         {analysis['raw_rows']:,}")
    print(f"  unique keys:      {analysis['unique_product_keys']:,}")
    print(f"  duplicate keys:   {analysis['duplicate_product_key_groups']:,}")
    print(f"  sparse long rows: {analysis['sparse_long_rows']:,}")
    print(f"  grand total rows: {analysis['grand_total_raw_rows']:,}")
    print(f"  analysis time:    {time.time() - t0:.1f}s")

    if db_path.exists():
        print(f"\n  기존 DuckDB 삭제: {db_path}")
        db_path.unlink()
    con = duckdb.connect(str(db_path))
    create_staging_table(con)

    print("\n[Step 2] DuckDB insert")
    ingested_at = datetime.now(timezone.utc).isoformat()
    total_inserted = 0
    chunk: list[dict[str, Any]] = []
    t0 = time.time()
    for row in iter_chso_rows(input_file, ingested_at=ingested_at):
        chunk.append(row)
        total_inserted += 1
        if len(chunk) >= args.chunk_size:
            bulk_insert(con, chunk)
            chunk = []
            print(f"  inserted {total_inserted:,} rows...")
    if chunk:
        bulk_insert(con, chunk)
    print(f"  total inserted: {total_inserted:,} long rows ({time.time() - t0:.1f}s)")

    print("\n[Step 3] staging stats")
    total_rows = con.execute("SELECT COUNT(*) FROM stg_raw").fetchone()[0]
    distinct_product_keys = con.execute("SELECT COUNT(DISTINCT product_key) FROM stg_raw").fetchone()[0]
    distinct_periods = con.execute("SELECT COUNT(DISTINCT period) FROM stg_raw").fetchone()[0]
    duplicate_product_period = con.execute(
        """
        SELECT COUNT(*) FROM (
            SELECT product_key, period, COUNT(*) AS n
            FROM stg_raw
            GROUP BY product_key, period
            HAVING COUNT(*) > 1
        )
        """
    ).fetchone()[0]
    grand_total_rows = con.execute("SELECT COUNT(*) FROM stg_raw WHERE is_grand_total").fetchone()[0]
    print(f"  total rows:                    {total_rows:,}")
    print(f"  distinct product_key:          {distinct_product_keys:,}")
    print(f"  distinct periods:              {distinct_periods:,}")
    print(f"  duplicate product_key+period:  {duplicate_product_period:,}")
    print(f"  grand total long rows:         {grand_total_rows:,}")

    print("\n  period distribution:")
    for period, n_rows in con.execute(
        "SELECT period, COUNT(*) FROM stg_raw GROUP BY period ORDER BY period"
    ).fetchall():
        print(f"    {period}: {n_rows:,}")

    print("\n[Step 4] write Parquet")
    t0 = time.time()
    written = write_partition_parquets(con, output_dir)
    print(f"  partitions written: {len(written)} ({time.time() - t0:.1f}s)")

    con.close()
    if args.keep_db:
        print(f"\n  DuckDB preserved: {db_path}")
    else:
        db_path.unlink(missing_ok=True)
        print(f"\n  DuckDB deleted:   {db_path}")

    print()
    print("=" * 72)
    print("Result")
    print("=" * 72)
    print(f"  output:      {output_dir}/")
    print(f"  partitions:  {len(written)}")
    print(f"  rows:        {sum(n for _, n, _ in written):,}")
    print()
    print(f"  {'period':<10}  {'rows':>8}  {'size_mb':>9}")
    print(f"  {'-' * 10}  {'-' * 8}  {'-' * 9}")
    for period, n_rows, size_mb in written:
        print(f"  {period:<10}  {n_rows:>8,}  {size_mb:>9.2f}")


if __name__ == "__main__":
    main()
