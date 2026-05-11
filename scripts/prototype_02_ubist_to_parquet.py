"""
prototype_02_ubist_to_parquet.py
=================================
UBIST xlsx (모든 폴더 + 단일 파일) → Parquet 월별 partition

검증된 정책 (prototype_sqlite_v1 의 결과):
- K7 = (drug_code, channel, specialty, age_group, gender, insurance_type)
- 메타: raw 24 컬럼 → staging 14 (drop 1, 6, 8, 12 — 동일/case 차이)
  - canonical: raw 10 (제조사), 3 (판매사), 16 (ATC), 17 (성분)
  - split:     raw 2 (manufacturer_origin), 4 (distributor_origin)
  - 별도:      raw 9 (distributor2)
- 시점: 'YYYY-MM' (2-row header 결합)
- metric: 처방조제액(원) → val, 처방건수_P → count, 처방량_P → volume
- 0값 보존, 빈 셀 skip (sparse)

DuckDB 사용 (메모리 안전, 의원_내과 제외 24M row 처리):
- streaming xlsx read → DuckDB INSERT (chunk)
- 모든 파일 끝 → GROUP BY (K7, period) + ARG_MAX (source_file 우선) + COPY TO Parquet

사용법:
    python3 scripts/prototype_02_ubist_to_parquet.py \\
        --data-dirs \\
            "data/UBIST/Sales (2021-2026.02)/보건소" \\
            "data/UBIST/Sales (2021-2026.02)/병원" \\
            "data/UBIST/Sales (2021-2026.02)/상급종병" \\
            "data/UBIST/Sales (2021-2026.02)/의원_내과" \\
            "data/UBIST/Sales (2021-2026.02)/의원_내과 제외-002" \\
            "data/UBIST/Sales (2021-2026.02)/기타 2021-2026.02.xlsx" \\
            "data/UBIST/Sales (2021-2026.02)/종병 병원 2601-02.xlsx" \\
        --output-dir parquet/ubist
"""

import argparse
import hashlib
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import duckdb
    import pyarrow as pa
except ImportError as e:
    sys.exit(
        f"ERROR: {e}\n"
        "  pip3 install duckdb pyarrow --break-system-packages"
    )

# xlsx reader — calamine 우선 (빠름), 없으면 openpyxl
USE_CALAMINE = False
try:
    from python_calamine import CalamineWorkbook
    USE_CALAMINE = True
except ImportError:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "ERROR: xlsx reader 필요.\n"
            "  pip3 install python-calamine --break-system-packages  (권장, 5-10배 빠름)\n"
            "  pip3 install openpyxl --break-system-packages         (대안)"
        )


# ============================================================================
# 정책 상수 (prototype_sqlite_v1 검증 결과)
# ============================================================================

# K7 분리축 — raw 위치 (1-indexed)
K7_POS = {
    "drug_code":      15,
    "channel":        21,
    "specialty":      22,
    "age_group":      23,
    "gender":         24,
    "insurance_type": 20,
}

K7_COLS = ["drug_code", "channel", "specialty", "age_group", "gender", "insurance_type"]

# 메타 raw 위치 (1-indexed) → staging 컬럼명 (drop 1, 6, 8, 12 제외)
META_POS_TO_COL = {
    2:  "manufacturer_origin",
    3:  "distributor",
    4:  "distributor_origin",
    5:  "product_name",
    7:  "brand",
    9:  "distributor2",
    10: "manufacturer",
    11: "drug_price",
    13: "molecule_strength",
    14: "drug_class",
    16: "atc",
    17: "molecule",
    18: "formulation",
    19: "route",
}

# 메타 컬럼 순서
META_COLS = [
    "manufacturer", "manufacturer_origin",
    "distributor", "distributor_origin", "distributor2",
    "product_name", "brand", "drug_price",
    "molecule_strength", "drug_class",
    "atc", "molecule", "formulation", "route",
]

# 시계열 metric 매핑
METRIC_MAP = {
    "처방조제액(원)": "val",
    "처방건수_P":     "count",
    "처방량_P":       "volume",
}

METRIC_COLS = ["val", "count", "volume"]

# 시점 형식: 'YYYY년 N월'
PAT_PERIOD = re.compile(r"^(\d{4})년\s*(\d{1,2})월$")

# 전체 staging 컬럼 (INSERT 순서)
ALL_COLS = (
    ["product_key"] + K7_COLS + META_COLS +
    ["period"] + METRIC_COLS + ["source_file"]
)


# ============================================================================
# 헬퍼
# ============================================================================
def clean_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return str(v).strip() or None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.replace(",", "").strip()
        if not v or v.lower() == "nan":
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_period(h2_val):
    if h2_val is None:
        return None
    m = PAT_PERIOD.match(str(h2_val).strip())
    if not m:
        return None
    year, month = m.groups()
    return f"{year}-{int(month):02d}"


def compute_k7_hash(k7_values):
    s = "|".join((v or "") for v in k7_values)
    return hashlib.sha256(s.encode("utf-8")).hexdigest()[:32]


def build_file_label(path, base_dir_name):
    if base_dir_name is None:
        return path.stem
    return f"{base_dir_name}/{path.stem}"


# ============================================================================
# xlsx streaming reader (calamine 또는 openpyxl)
# ============================================================================
def stream_xlsx_with_calamine(xlsx_path):
    """calamine 으로 row 단위 streaming. yield list of row values."""
    wb = CalamineWorkbook.from_path(str(xlsx_path))
    sheet_names = wb.sheet_names
    sheet_name = "Sheet1" if "Sheet1" in sheet_names else sheet_names[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    data = sheet.to_python()
    for row in data:
        yield row


def stream_xlsx_with_openpyxl(xlsx_path):
    """openpyxl read_only streaming."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        sheet_name = "Sheet1" if "Sheet1" in sheet_names else sheet_names[0]
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


def stream_xlsx(xlsx_path):
    if USE_CALAMINE:
        return stream_xlsx_with_calamine(xlsx_path)
    else:
        return stream_xlsx_with_openpyxl(xlsx_path)


# ============================================================================
# row generator (K7 + meta + period observation → long rows)
# ============================================================================
def extract_rows(xlsx_path, source_label):
    """xlsx 한 파일 → 여러 (K7, period, val, count, volume) long row 로 yield."""
    rows_iter = iter(stream_xlsx(xlsx_path))
    
    header_1 = next(rows_iter, None)
    header_2 = next(rows_iter, None)
    
    if header_1 is None or header_2 is None:
        return
    
    n_cols = len(header_1)
    
    # 시계열 컬럼 분류
    time_cols = []
    for i in range(24, n_cols):
        if i >= len(header_1) or i >= len(header_2):
            continue
        h1 = header_1[i]
        h2 = header_2[i]
        if h1 is None or h2 is None:
            continue
        metric = METRIC_MAP.get(str(h1).strip())
        if metric is None:
            continue
        period = parse_period(h2)
        if period is None:
            continue
        time_cols.append((i, period, metric))
    
    for raw_row in rows_iter:
        if raw_row is None or len(raw_row) < 24:
            continue
        
        # K7 추출
        k7_vals = []
        for col_name in K7_COLS:
            pos = K7_POS[col_name]
            if pos - 1 < len(raw_row):
                k7_vals.append(clean_str(raw_row[pos - 1]))
            else:
                k7_vals.append(None)
        
        # drug_code 또는 channel 없으면 skip
        if not k7_vals[0] or not k7_vals[1]:
            continue
        
        pk = compute_k7_hash(k7_vals)
        
        # 메타 추출 (위치 기반)
        meta = {}
        for raw_pos, staging_col in META_POS_TO_COL.items():
            if raw_pos - 1 < len(raw_row):
                meta[staging_col] = clean_str(raw_row[raw_pos - 1])
            else:
                meta[staging_col] = None
        
        # period 별 metric 수집 (sparse: None 은 저장 안 함)
        period_obs = {}
        for col_idx, period, metric in time_cols:
            if col_idx < len(raw_row):
                v = to_float(raw_row[col_idx])
                if v is not None:
                    period_obs.setdefault(period, {})[metric] = v
        
        # 각 period 별 long row yield
        for period, ms in period_obs.items():
            row = {"product_key": pk}
            for i, col_name in enumerate(K7_COLS):
                row[col_name] = k7_vals[i]
            for col_name in META_COLS:
                row[col_name] = meta.get(col_name)
            row["period"] = period
            for metric in METRIC_COLS:
                row[metric] = ms.get(metric)
            row["source_file"] = source_label
            yield row


# ============================================================================
# DuckDB 적재
# ============================================================================
def create_staging_table(con):
    cols_sql = [("product_key", "VARCHAR")]
    for c in K7_COLS:
        cols_sql.append((c, "VARCHAR"))
    for c in META_COLS:
        cols_sql.append((c, "VARCHAR"))
    cols_sql.extend([
        ("period", "VARCHAR"),
        ("val", "DOUBLE"),
        ("count", "DOUBLE"),
        ("volume", "DOUBLE"),
        ("source_file", "VARCHAR"),
    ])
    cols_def = ", ".join(f'"{c}" {t}' for c, t in cols_sql)
    con.execute(f"CREATE OR REPLACE TABLE stg_raw ({cols_def})")


def bulk_insert(con, batch_rows):
    if not batch_rows:
        return
    # column-wise dict
    batch_dict = {col: [] for col in ALL_COLS}
    for row in batch_rows:
        for col in ALL_COLS:
            batch_dict[col].append(row.get(col))
    arrow_table = pa.Table.from_pydict(batch_dict)
    con.register("_batch", arrow_table)
    con.execute("INSERT INTO stg_raw SELECT * FROM _batch")
    con.unregister("_batch")


# ============================================================================
# Parquet write per period (GROUP BY K7+period + ARG_MAX)
# ============================================================================
def write_partition_parquets(con, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    
    periods = [r[0] for r in con.execute(
        "SELECT DISTINCT period FROM stg_raw ORDER BY period"
    ).fetchall()]
    
    print(f"\n  periods: {len(periods)} ({periods[0]} ~ {periods[-1]})")
    
    ingested_at = datetime.now(timezone.utc).isoformat()
    
    # GROUP BY 시 메타/metric 의 충돌 해결: ARG_MAX (source_file alphabetical 마지막 우선)
    meta_select = ", ".join(f'ARG_MAX("{c}", source_file) AS "{c}"' for c in META_COLS)
    metric_select = ", ".join(f'ARG_MAX("{c}", source_file) AS "{c}"' for c in METRIC_COLS)
    k7_select = ", ".join(f'"{c}"' for c in K7_COLS)
    k7_group = ", ".join(f'"{c}"' for c in K7_COLS)
    
    written = []
    
    for period in periods:
        out_path = output_dir / f"{period}.parquet"
        con.execute(f"""
            COPY (
                SELECT
                    product_key,
                    {k7_select},
                    {meta_select},
                    period,
                    {metric_select},
                    STRING_AGG(DISTINCT source_file, ',' ORDER BY source_file) AS source_files,
                    '{ingested_at}' AS ingested_at
                FROM stg_raw
                WHERE period = ?
                GROUP BY product_key, {k7_group}, period
            ) TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """, [period])
        
        n_rows = con.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
        size_mb = out_path.stat().st_size / 1024 / 1024
        written.append((period, n_rows, size_mb))
    
    return written


# ============================================================================
# main
# ============================================================================
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--data-dirs", nargs="+", required=True,
                   help="UBIST 폴더 또는 xlsx 파일 경로")
    p.add_argument("--output-dir", default="parquet/ubist")
    p.add_argument("--db-path", default="staging_ubist.duckdb",
                   help="DuckDB 임시 db 경로")
    p.add_argument("--keep-db", action="store_true",
                   help="parquet write 후 db 보존 (기본: 삭제)")
    p.add_argument("--chunk-size", type=int, default=20000,
                   help="DuckDB INSERT chunk 크기 (기본: 20,000)")
    args = p.parse_args()
    
    output_dir = Path(args.output_dir)
    db_path = Path(args.db_path)
    
    # 폴더/파일 수집
    all_files = []
    for path_str in args.data_dirs:
        p_obj = Path(path_str)
        if p_obj.is_dir():
            xlsx_in_d = sorted(
                f for f in p_obj.glob("*.xlsx")
                if not f.name.startswith(("~", "."))
            )
            if not xlsx_in_d:
                sys.exit(f"ERROR: {p_obj} 안에 xlsx 없음")
            for f in xlsx_in_d:
                all_files.append((f, p_obj.name))
        elif p_obj.is_file() and p_obj.suffix == ".xlsx":
            all_files.append((p_obj, None))
        else:
            sys.exit(f"ERROR: {p_obj} 디렉토리도 xlsx 파일도 아님")
    
    print("=" * 72)
    print("UBIST → Parquet (월별 partition, DuckDB streaming)")
    print("=" * 72)
    print(f"  inputs:           {len(args.data_dirs)}")
    print(f"  total xlsx files: {len(all_files)}")
    print(f"  output:           {output_dir}/")
    print(f"  duckdb path:      {db_path}")
    print(f"  xlsx reader:      {'calamine (fast)' if USE_CALAMINE else 'openpyxl'}")
    print(f"  chunk size:       {args.chunk_size:,}")
    
    # DuckDB 연결
    if db_path.exists():
        print(f"\n  ⚠ 기존 db 삭제: {db_path}")
        db_path.unlink()
    
    con = duckdb.connect(str(db_path))
    create_staging_table(con)
    
    # Step 1: streaming + insert
    print(f"\n[Step 1] xlsx streaming + DuckDB insert")
    total_inserted = 0
    start_time = time.time()
    
    for idx, (xlsx_path, base_dir_name) in enumerate(all_files, 1):
        label = build_file_label(xlsx_path, base_dir_name)
        size_mb = xlsx_path.stat().st_size / 1024 / 1024
        t0 = time.time()
        
        chunk = []
        n_rows_for_file = 0
        for row in extract_rows(xlsx_path, label):
            chunk.append(row)
            n_rows_for_file += 1
            if len(chunk) >= args.chunk_size:
                bulk_insert(con, chunk)
                chunk = []
        if chunk:
            bulk_insert(con, chunk)
        
        total_inserted += n_rows_for_file
        elapsed = time.time() - t0
        rate = n_rows_for_file / max(elapsed, 0.01)
        print(f"  [{idx:>2}/{len(all_files)}] {label:<55s} "
              f"({size_mb:>6.1f}MB) → {n_rows_for_file:>10,} rows "
              f"({elapsed:>5.1f}s, {rate:>6.0f} r/s)")
    
    total_elapsed = time.time() - start_time
    print(f"\n  total inserted: {total_inserted:,} long rows ({total_elapsed:.1f}s)")
    
    # 통계
    print(f"\n  통계:")
    k7_count = con.execute(
        "SELECT COUNT(DISTINCT product_key) FROM stg_raw"
    ).fetchone()[0]
    print(f"    K7 unique:        {k7_count:,}")
    
    k7p_count = con.execute("""
        SELECT COUNT(*) FROM (
            SELECT DISTINCT product_key, period FROM stg_raw
        )
    """).fetchone()[0]
    print(f"    K7+period unique: {k7p_count:,}")
    
    if k7p_count < total_inserted:
        n_dup = total_inserted - k7p_count
        print(f"    ⚠ K7+period 충돌 (multi-source): {n_dup:,} "
              f"({n_dup/total_inserted*100:.1f}%)")
    
    # Step 2: parquet write
    print(f"\n[Step 2] period 별 Parquet write")
    t0 = time.time()
    written = write_partition_parquets(con, output_dir)
    print(f"  ({time.time() - t0:.1f}s)")
    
    # 정리
    con.close()
    if not args.keep_db:
        db_path.unlink(missing_ok=True)
        print(f"\n  duckdb 삭제: {db_path}")
    else:
        size_gb = db_path.stat().st_size / 1024 / 1024 / 1024
        print(f"\n  duckdb 보존: {db_path} ({size_gb:.2f}GB)")
    
    # 결과 요약
    print()
    print("=" * 72)
    print("결과")
    print("=" * 72)
    print(f"  partitions: {len(written)}")
    print(f"  K7 unique:  {k7_count:,}")
    print()
    print(f"  {'period':<10}  {'rows':>10}  {'size (MB)':>10}")
    print(f"  {'-'*10}  {'-'*10}  {'-'*10}")
    total_rows = 0
    total_size = 0.0
    for period, n, s in written:
        print(f"  {period:<10}  {n:>10,}  {s:>10.2f}")
        total_rows += n
        total_size += s
    print(f"  {'-'*10}  {'-'*10}  {'-'*10}")
    print(f"  {'TOTAL':<10}  {total_rows:>10,}  {total_size:>10.2f}")
    
    print(f"\n  검증 (DuckDB):")
    print(f"    python3 -c \"import duckdb; "
          f"duckdb.sql(\\\"SELECT COUNT(*), COUNT(DISTINCT product_key) FROM "
          f"'{output_dir}/*.parquet'\\\").show()\"")


if __name__ == "__main__":
    main()
