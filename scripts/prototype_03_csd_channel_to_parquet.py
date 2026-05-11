"""
prototype_03_csd_channel_to_parquet.py
======================================
IQVIA CSD ChannelDynamics xlsx -> Parquet monthly partitions.

Reference policy mirrored from jw-market:
- SQL DDL source: sql/schema_iqvia.sql, stg_iqvia_csd_channel_raw
- CSD row id: SHA256(channel|source_file|source_sheet|source_row_id)[:32]
- Header detection / sheet classification / observation extraction follow
  jw-market/etl/iqvia_csd_loader.py and iqvia_transform.py.
- raw_row_json preserves the full source row because CSD sheet layouts vary.
- Sparse policy: empty monthly cells are skipped; explicit 0 is preserved.

Prototype storage:
- 1 row = 1 raw source row, matching stg_iqvia_csd_channel_raw.
- Output partition key is the report month inferred from the workbook filename.
- Core DDL columns are preserved; prototype helper columns are period and source_files.

Usage:
    python3 scripts/prototype_03_csd_channel_to_parquet.py \\
        --input-dir "data/IQVIA/CSD/ChannelDynamics (콜 수=영업 횟수)" \\
        --output-dir parquet/csd_channel
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

try:
    import duckdb
    import pyarrow as pa
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install duckdb pyarrow --break-system-packages")


USE_CALAMINE = False
try:
    from python_calamine import CalamineWorkbook

    USE_CALAMINE = True
except ImportError:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "ERROR: xlsx reader 필요.\n"
            "  pip3 install python-calamine --break-system-packages  (권장)\n"
            "  pip3 install openpyxl --break-system-packages         (대안)"
        )


# SQL DDL columns from /Users/rexxa/github/jw-market/sql/schema_iqvia.sql
DDL_COLS = [
    "csd_row_id",
    "source_file",
    "source_sheet",
    "source_row_id",
    "report_family",
    "channel",
    "report_section",
    "market_sheet_name",
    "product",
    "manufacturer",
    "product_details",
    "region",
    "metric_name",
    "observations_json",
    "raw_row_json",
    "ingested_at",
]

# Prototype helper columns for monthly Parquet partitions.
OUTPUT_COLS = [
    "csd_row_id",
    "source_file",
    "source_sheet",
    "source_row_id",
    "report_family",
    "channel",
    "report_section",
    "market_sheet_name",
    "product",
    "manufacturer",
    "product_details",
    "region",
    "metric_name",
    "period",
    "observations_json",
    "raw_row_json",
    "source_files",
    "ingested_at",
]

MONTH_MAP = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "june": "06",
    "jul": "07",
    "july": "07",
    "aug": "08",
    "sep": "09",
    "sept": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


def normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=to_jsonable, sort_keys=True)


def is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        numeric = text.replace(",", "")
        if re.fullmatch(r"-?\d+", numeric):
            try:
                return int(numeric)
            except ValueError:
                return text
        if re.fullmatch(r"-?\d+\.\d+", numeric):
            try:
                return float(numeric)
            except ValueError:
                return text
        return text
    return to_jsonable(value)


def to_varchar(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def make_csd_row_id(table_scope: str, source_file: str, source_sheet: str, source_row_id: int) -> str:
    key = "|".join([table_scope, source_file, source_sheet, str(source_row_id)])
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:32]


def make_header_keys(headers: list[Any] | tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    keys: list[str] = []
    for index, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        base = normalized if normalized is not None else f"__blank_col_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        keys.append(base if count == 1 else f"{base}__{count}")
    return keys


def build_raw_row_payload(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    source_row_id: int,
) -> dict[str, Any]:
    width = max(len(headers), len(values))
    padded_headers = list(headers) + [None] * (width - len(headers))
    padded_values = list(values) + [None] * (width - len(values))
    keys = make_header_keys(padded_headers)
    cells = []
    values_by_header: dict[str, Any] = {}
    for index, (header, key, value) in enumerate(zip(padded_headers, keys, padded_values), start=1):
        json_value = to_jsonable(value)
        cells.append(
            {
                "column_index": index,
                "header": normalize_header(header),
                "header_key": key,
                "value": json_value,
            }
        )
        values_by_header[key] = json_value
    return {
        "source_row_id": source_row_id,
        "cells": cells,
        "values_by_header": values_by_header,
    }


def trim_row_width(headers: tuple[Any, ...], values: tuple[Any, ...]) -> tuple[list[Any], list[Any]]:
    width = max(len(headers), len(values))
    padded_headers = list(headers) + [None] * (width - len(headers))
    padded_values = list(values) + [None] * (width - len(values))
    last = 0
    for index, (header, value) in enumerate(zip(padded_headers, padded_values), start=1):
        if normalize_header(header) is not None or not is_empty_row([value]):
            last = index
    return padded_headers[:last], padded_values[:last]


def raw_row_json(headers: tuple[Any, ...], values: tuple[Any, ...], source_row_id: int) -> str:
    trimmed_headers, trimmed_values = trim_row_width(headers, values)
    return dumps_json(build_raw_row_payload(trimmed_headers, trimmed_values, source_row_id))


def row_lookup(headers: tuple[Any, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    trimmed_headers, trimmed_values = trim_row_width(headers, values)
    lookup: dict[str, Any] = {}
    for header, value in zip(trimmed_headers, trimmed_values):
        normalized = normalize_header(header)
        if normalized is not None and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def first_matching_value(lookup: dict[str, Any], names: list[str], contains: list[str] | None = None) -> Any:
    for name in names:
        if name in lookup:
            return lookup[name]
    lowered = {key.lower(): key for key in lookup}
    for name in names:
        key = lowered.get(name.lower())
        if key is not None:
            return lookup[key]
    for token in contains or []:
        for key, value in lookup.items():
            if token.lower() in key.lower():
                return value
    return None


def _header_score(values: tuple[Any, ...]) -> int:
    non_empty = [normalize_header(value) for value in values]
    non_empty = [value for value in non_empty if value is not None]
    if len(non_empty) < 3:
        return 0
    joined = " ".join(non_empty).lower()
    known_tokens = [
        "product",
        "jw channel",
        "representing company",
        "region",
        "related date",
        "meeting",
        "keywords",
        "rank",
        "manufacturer",
        "master product",
    ]
    token_score = sum(1 for token in known_tokens if token in joined)
    return len(non_empty) + token_score * 10


def detect_header_row(rows: list[list[Any] | tuple[Any, ...]], max_scan_rows: int = 12) -> int | None:
    best_row: int | None = None
    best_score = 0
    for row_number, values in enumerate(rows[:max_scan_rows], start=1):
        score = _header_score(tuple(values))
        if score > best_score:
            best_row = row_number
            best_score = score
    return best_row if best_score else None


def classify_csd_sheet(sheet_name: str) -> tuple[str | None, str | None, str | None, str | None]:
    normalized = re.sub(r"\s+", " ", sheet_name).strip()
    if "Market" in normalized:
        market = re.sub(r"\s+Market\d*$", "", normalized).strip()
        return "Market", None, "Market", market or None

    channel_patterns = [
        ("GH+SHPPI", r"^GH\s*\+\s*SHPPI"),
        ("TOTAL", r"^TOTAL"),
        ("SHPPI", r"^SHPPI"),
        ("CPPI", r"^CPPI"),
        ("GH", r"^GH"),
    ]
    for channel, pattern in channel_patterns:
        if re.match(pattern, normalized):
            for section in [
                "Team Trend",
                "Product Detail",
                "Product-Detail",
                "TOP20",
                "TOP10",
                "Call Rank",
                "Detail",
            ]:
                if section in normalized:
                    normalized_section = "Product Detail" if section == "Product-Detail" else section
                    return channel, channel, normalized_section, None
            return channel, channel, "Other", None
    return None, None, None, None


def extract_csd_observations(lookup: dict[str, Any]) -> dict[str, Any]:
    observations: dict[str, Any] = {}
    for key, value in lookup.items():
        normalized = normalize_header(key)
        if normalized is None:
            continue
        if re.match(r"^[A-Za-z]{3,5}\.?\s+\d{2}$", normalized) or re.match(
            r"^(Jan|Feb|Mar|Apr|May|June|Jul|July|Aug|Sep|Oct|Nov|Dec)\s+\d{2}$",
            normalized,
            re.I,
        ):
            observations[normalized] = normalize_scalar(value)
    return observations


def parse_observation_period(raw_label: str) -> str | None:
    match = re.match(r"^([A-Za-z]{3,5})\.?\s+(\d{2})$", raw_label.strip())
    if not match:
        return None
    month_name, year_2digit = match.groups()
    month = MONTH_MAP.get(month_name.lower())
    if month is None:
        return None
    return f"20{year_2digit}-{month}"


def parse_report_month_from_filename(filename: str) -> str | None:
    """Parse report month from names such as Dec.24, June25, July25."""
    match = re.search(r"([A-Za-z]{3,5})\.?\s*(\d{2})", filename)
    if not match:
        return None
    month_name, year_2digit = match.groups()
    month = MONTH_MAP.get(month_name.lower())
    if month is None:
        return None
    return f"20{year_2digit}-{month}"


def extract_channel_record_fields(lookup: dict[str, Any]) -> dict[str, Any]:
    return {
        "product": first_matching_value(
            lookup,
            ["Master product", "Product name", "PRODUCT NAME", "Product Name"],
            contains=["product"],
        ),
        "manufacturer": first_matching_value(
            lookup,
            ["Manufacturer", "Representing Company", "Pharmaceutical Sponsor"],
            contains=["manufacturer", "company"],
        ),
        "product_details": first_matching_value(
            lookup,
            ["Product Details", "Product Detail", "Detail"],
            contains=["details"],
        ),
        "region": first_matching_value(lookup, ["Region", "JW Channel"], contains=["region"]),
        "metric_name": first_matching_value(
            lookup,
            ["Metric", "Value Type", "Usefulness", "Weighted Calls"],
            contains=["usefulness", "weighted"],
        ),
    }


def iter_workbook_sheets(xlsx_path: Path) -> Iterable[tuple[str, list[list[Any] | tuple[Any, ...]]]]:
    if USE_CALAMINE:
        wb = CalamineWorkbook.from_path(str(xlsx_path))
        for sheet_name in wb.sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            yield sheet_name, sheet.to_python()
        return

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            yield sheet_name, list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def iter_csd_raw_rows(xlsx_path: Path, ingested_at: str | None = None) -> Iterable[dict[str, Any]]:
    timestamp = ingested_at or datetime.now(timezone.utc).isoformat()
    report_month = parse_report_month_from_filename(xlsx_path.name)
    if report_month is None:
        raise ValueError(f"report month parse 실패: {xlsx_path.name}")

    for sheet_name, rows in iter_workbook_sheets(xlsx_path):
        if normalize_header(sheet_name) == "Main":
            continue
        header_row = detect_header_row(rows)
        if header_row is None:
            continue

        headers = tuple(rows[header_row - 1])
        report_family, channel, report_section, market_sheet_name = classify_csd_sheet(sheet_name)

        for source_row_id, values in enumerate(rows[header_row:], start=header_row + 1):
            values_tuple = tuple(values)
            if is_empty_row(values_tuple):
                continue

            lookup = row_lookup(headers, values_tuple)
            fields = extract_channel_record_fields(lookup)
            observations = extract_csd_observations(lookup)

            csd_row_id = make_csd_row_id("channel", xlsx_path.name, sheet_name, source_row_id)
            row_json = raw_row_json(headers, values_tuple, source_row_id)
            yield {
                "csd_row_id": csd_row_id,
                "source_file": xlsx_path.name,
                "source_sheet": sheet_name,
                "source_row_id": source_row_id,
                "report_family": to_varchar(report_family),
                "channel": to_varchar(channel),
                "report_section": to_varchar(report_section),
                "market_sheet_name": to_varchar(market_sheet_name),
                "product": to_varchar(fields["product"]),
                "manufacturer": to_varchar(fields["manufacturer"]),
                "product_details": to_varchar(fields["product_details"]),
                "region": to_varchar(fields["region"]),
                "metric_name": to_varchar(fields["metric_name"]),
                "period": report_month,
                "observations_json": dumps_json(observations),
                "raw_row_json": row_json,
                "source_files": xlsx_path.name,
                "ingested_at": timestamp,
            }


# Backward-compatible name used by the verifier import in earlier drafts.
iter_csd_long_rows = iter_csd_raw_rows


def collect_raw_structure(xlsx_path: Path) -> dict[str, Any]:
    sheet_count = 0
    loaded_sheets = 0
    skipped_sheets = 0
    header_rows: dict[int, int] = {}
    data_rows = 0
    periods: set[str] = set()
    report_month = parse_report_month_from_filename(xlsx_path.name)

    for sheet_name, rows in iter_workbook_sheets(xlsx_path):
        sheet_count += 1
        if normalize_header(sheet_name) == "Main":
            skipped_sheets += 1
            continue
        header_row = detect_header_row(rows)
        if header_row is None:
            skipped_sheets += 1
            continue

        loaded_sheets += 1
        header_rows[header_row] = header_rows.get(header_row, 0) + 1
        headers = tuple(rows[header_row - 1])
        for values in rows[header_row:]:
            values_tuple = tuple(values)
            if is_empty_row(values_tuple):
                continue
            data_rows += 1
            if report_month is not None:
                periods.add(report_month)

    return {
        "source_file": xlsx_path.name,
        "sheet_count": sheet_count,
        "loaded_sheets": loaded_sheets,
        "skipped_sheets": skipped_sheets,
        "header_rows": dict(sorted(header_rows.items())),
        "data_rows": data_rows,
        "long_rows": data_rows,
        "periods": sorted(periods),
        "filename_report_month": report_month,
    }


def create_staging_table(con: duckdb.DuckDBPyConnection) -> None:
    col_types = []
    for col in OUTPUT_COLS:
        if col == "source_row_id":
            col_types.append((col, "INTEGER"))
        else:
            col_types.append((col, "VARCHAR"))
    cols_def = ", ".join(f'"{col}" {typ}' for col, typ in col_types)
    con.execute(f"CREATE OR REPLACE TABLE stg_raw ({cols_def})")


def bulk_insert(con: duckdb.DuckDBPyConnection, batch_rows: list[dict[str, Any]]) -> None:
    if not batch_rows:
        return
    batch_dict = {col: [] for col in OUTPUT_COLS}
    for row in batch_rows:
        for col in OUTPUT_COLS:
            batch_dict[col].append(row.get(col))
    arrow_table = pa.Table.from_pydict(batch_dict)
    con.register("_batch", arrow_table)
    con.execute("INSERT INTO stg_raw SELECT * FROM _batch")
    con.unregister("_batch")


def write_partition_parquets(con: duckdb.DuckDBPyConnection, output_dir: Path) -> list[tuple[str, int, float]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale in output_dir.glob("*.parquet"):
        stale.unlink()

    periods = [r[0] for r in con.execute("SELECT DISTINCT period FROM stg_raw ORDER BY period").fetchall()]
    if not periods:
        return []

    written = []
    select_cols = ", ".join(f'"{col}"' for col in OUTPUT_COLS)
    for period in periods:
        out_path = output_dir / f"{period}.parquet"
        con.execute(
            f"""
            COPY (
                SELECT {select_cols}
                FROM stg_raw
                WHERE period = ?
                ORDER BY source_file, source_sheet, source_row_id, csd_row_id
            ) TO '{out_path}' (FORMAT PARQUET, COMPRESSION ZSTD)
            """,
            [period],
        )
        n_rows = con.execute(f"SELECT COUNT(*) FROM '{out_path}'").fetchone()[0]
        size_mb = out_path.stat().st_size / 1024 / 1024
        written.append((period, n_rows, size_mb))
    return written


def resolve_input_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists() or not input_dir.is_dir():
        sys.exit(f"ERROR: input-dir 가 없거나 디렉토리가 아님: {input_dir}")
    files = sorted(
        p for p in input_dir.glob("*.xlsx") if not p.name.startswith(("~", "."))
    )
    if not files:
        sys.exit(f"ERROR: input-dir 안에 xlsx 파일 없음: {input_dir}")
    return files


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-dir",
        default="data/IQVIA/CSD/ChannelDynamics (콜 수=영업 횟수)",
        help="ChannelDynamics xlsx 폴더",
    )
    parser.add_argument("--output-dir", default="parquet/csd_channel")
    parser.add_argument("--db-path", default="staging_csd_channel.duckdb")
    parser.add_argument("--chunk-size", type=int, default=5000)
    parser.add_argument("--keep-db", action="store_true")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    db_path = Path(args.db_path)
    files = resolve_input_files(input_dir)

    print("=" * 72)
    print("IQVIA CSD ChannelDynamics -> Parquet (report-month partitions)")
    print("=" * 72)
    print(f"  input dir:     {input_dir}")
    print(f"  files:         {len(files)}")
    print(f"  output dir:    {output_dir}")
    print(f"  duckdb path:   {db_path}")
    print(f"  xlsx reader:   {'calamine' if USE_CALAMINE else 'openpyxl'}")
    print(f"  chunk size:    {args.chunk_size:,}")
    print(f"  core columns:  {len(DDL_COLS)} DDL + period/source_files helpers")

    if db_path.exists():
        print(f"\n  기존 DuckDB 삭제: {db_path}")
        db_path.unlink()

    con = duckdb.connect(str(db_path))
    create_staging_table(con)

    ingested_at = datetime.now(timezone.utc).isoformat()
    file_stats: list[dict[str, Any]] = []
    total_inserted = 0

    print("\n[Step 1] raw structure check + DuckDB insert")
    start = time.time()
    for index, xlsx_path in enumerate(files, start=1):
        t0 = time.time()
        structure = collect_raw_structure(xlsx_path)
        file_stats.append(structure)

        chunk: list[dict[str, Any]] = []
        inserted_for_file = 0
        for row in iter_csd_raw_rows(xlsx_path, ingested_at=ingested_at):
            chunk.append(row)
            inserted_for_file += 1
            if len(chunk) >= args.chunk_size:
                bulk_insert(con, chunk)
                chunk = []
        if chunk:
            bulk_insert(con, chunk)

        total_inserted += inserted_for_file
        elapsed = time.time() - t0
        periods = structure["periods"]
        period_span = f"{periods[0]}~{periods[-1]}" if periods else "(none)"
        print(
            f"  [{index}/{len(files)}] {xlsx_path.name:<62s} "
            f"sheets {structure['loaded_sheets']:>2}/{structure['sheet_count']:<2} "
            f"rows {inserted_for_file:>8,} report_month {period_span:<17s} "
            f"({elapsed:>5.1f}s)"
        )

    elapsed_total = time.time() - start
    print(f"\n  total inserted: {total_inserted:,} raw rows ({elapsed_total:.1f}s)")

    print("\n[Step 2] staging stats")
    total_rows = con.execute("SELECT COUNT(*) FROM stg_raw").fetchone()[0]
    distinct_csd_rows = con.execute("SELECT COUNT(DISTINCT csd_row_id) FROM stg_raw").fetchone()[0]
    distinct_periods = con.execute("SELECT COUNT(DISTINCT period) FROM stg_raw").fetchone()[0]
    duplicate_csd_row_id = con.execute(
        """
        SELECT COUNT(*) FROM (
            SELECT csd_row_id, COUNT(*) AS n
            FROM stg_raw
            GROUP BY csd_row_id
            HAVING COUNT(*) > 1
        )
        """
    ).fetchone()[0]
    print(f"  total rows:              {total_rows:,}")
    print(f"  distinct csd_row_id:     {distinct_csd_rows:,}")
    print(f"  distinct periods:        {distinct_periods:,}")
    print(f"  duplicate csd_row_id groups: {duplicate_csd_row_id:,}")

    print("\n  file distribution:")
    for source_file, n in con.execute(
        "SELECT source_file, COUNT(*) FROM stg_raw GROUP BY source_file ORDER BY source_file"
    ).fetchall():
        print(f"    {source_file:<70s} {n:>10,}")

    print("\n[Step 3] write Parquet")
    t0 = time.time()
    written = write_partition_parquets(con, output_dir)
    print(f"  partitions written: {len(written)} ({time.time() - t0:.1f}s)")

    con.close()
    if args.keep_db:
        print(f"\n  DuckDB preserved: {db_path}")
    else:
        db_path.unlink(missing_ok=True)
        print(f"\n  DuckDB deleted:   {db_path}")

    print()
    print("=" * 72)
    print("Result")
    print("=" * 72)
    print(f"  output:      {output_dir}/")
    print(f"  partitions:  {len(written)}")
    print(f"  rows:        {sum(n for _, n, _ in written):,}")
    print()
    print(f"  {'period':<10}  {'rows':>10}  {'size_mb':>9}")
    print(f"  {'-' * 10}  {'-' * 10}  {'-' * 9}")
    for period, n_rows, size_mb in written:
        print(f"  {period:<10}  {n_rows:>10,}  {size_mb:>9.2f}")


if __name__ == "__main__":
    main()
