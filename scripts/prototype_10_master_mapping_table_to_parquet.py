"""
prototype_10_master_mapping_table_to_parquet.py
================================================
MI Master manual mapping table -> Parquet.

Phase 09d policy:
- Canonical row-generation logic:
  /Users/rexxa/github/jw-market/etl/master_loader.py::_manual_mapping_records
- Canonical market-sheet scan / mapping helpers:
  /Users/rexxa/github/jw-market/etl/master_loader.py and
  /Users/rexxa/github/jw-market/etl/master_transform.py
- Canonical schema:
  /Users/rexxa/github/jw-market/sql/schema_master.sql,
  stg_master_mapping_table
- Output schema is DDL columns only. No prototype helper columns such as
  source_files or period are added.

Usage, in Step D after user review:
    python3 scripts/prototype_10_master_mapping_table_to_parquet.py
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit(f"ERROR: {e}\n  pip3 install pyarrow openpyxl --break-system-packages")


MASTER_ROOT = Path("data/JW 주요 약품 수동 매핑")
DEFAULT_INPUT_FILE = MASTER_ROOT / "MI팀_시장분석 AI_시장 분석 Master Version (260422).xlsx"
DEFAULT_CATALOG_PATH = Path("docs/reference/master_column_mapping_catalog.md")
DEFAULT_OUTPUT_FILE = Path("parquet/master_mapping_table/master_mapping_table.parquet")

STANDARD_PREFIX = "drug_extra_json."
EXPECTED_ROW_COUNT = 5932
ZERO_MAPPING_MARKETS = {"strategy_006", "strategy_007", "strategy_009"}

MASTER_MAPPING_TABLE_COLUMNS = (
    "mapping_id",
    "strategic_market_id",
    "source_value",
    "target_column",
    "target_value",
    "mapping_type",
    "source_sheet",
    "source_file_version",
    "ingested_at",
)


@dataclass(frozen=True)
class MarketSheetConfig:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    source_type: str


MARKET_SHEETS: tuple[MarketSheetConfig, ...] = (
    MarketSheetConfig("strategy_001", "라베칸 라베칸듀오", 5, "UBIST"),
    MarketSheetConfig("strategy_002", "제이클", 5, "IQVIA"),
    MarketSheetConfig("strategy_003", "가드렛 가드메트", 5, "IQVIA"),
    MarketSheetConfig("strategy_004", "타발리스", 5, "IQVIA"),
    MarketSheetConfig("strategy_005", "시그마트", 5, "UBIST"),
    MarketSheetConfig("strategy_006", "리바로 리바로젯", 4, "UBIST"),
    MarketSheetConfig("strategy_007", "리바로페노", 4, "UBIST"),
    MarketSheetConfig("strategy_008", "리바로하이 리바로브이", 5, "UBIST"),
    MarketSheetConfig("strategy_009", "트루패스 피나스타 제이다트", 5, "UBIST"),
    MarketSheetConfig("strategy_010", "뉴트로진 모빌리아", 5, "IQVIA"),
    MarketSheetConfig("strategy_011", "악템라", 5, "IQVIA"),
    MarketSheetConfig("strategy_012", "페린젝트 베노훼럼", 5, "IQVIA"),
    MarketSheetConfig("strategy_013", "헴리브라", 5, "IQVIA"),
    MarketSheetConfig("strategy_014", "위너프 위너프A+", 5, "IQVIA"),
    MarketSheetConfig("strategy_015", "엔커버", 7, "IQVIA"),
    MarketSheetConfig("strategy_016", "플라주오피", 5, "IQVIA"),
)

EXPECTED_MARKET_STATS = {
    "strategy_001": {
        "sheet_name": "라베칸 라베칸듀오",
        "header_row": 5,
        "raw_rows_scanned": 358,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 358,
        "manual_specs": 3,
        "mapping_rows": 731,
    },
    "strategy_002": {
        "sheet_name": "제이클",
        "header_row": 5,
        "raw_rows_scanned": 45,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 45,
        "manual_specs": 3,
        "mapping_rows": 135,
    },
    "strategy_003": {
        "sheet_name": "가드렛 가드메트",
        "header_row": 5,
        "raw_rows_scanned": 113,
        "empty_rows": 31,
        "excluded_rows": 0,
        "staging_rows": 82,
        "manual_specs": 2,
        "mapping_rows": 164,
    },
    "strategy_004": {
        "sheet_name": "타발리스",
        "header_row": 5,
        "raw_rows_scanned": 10,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 10,
        "manual_specs": 1,
        "mapping_rows": 9,
    },
    "strategy_005": {
        "sheet_name": "시그마트",
        "header_row": 5,
        "raw_rows_scanned": 294,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 294,
        "manual_specs": 4,
        "mapping_rows": 1176,
    },
    "strategy_006": {
        "sheet_name": "리바로 리바로젯",
        "header_row": 4,
        "raw_rows_scanned": 1295,
        "empty_rows": 200,
        "excluded_rows": 48,
        "staging_rows": 1047,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_007": {
        "sheet_name": "리바로페노",
        "header_row": 4,
        "raw_rows_scanned": 996,
        "empty_rows": 385,
        "excluded_rows": 494,
        "staging_rows": 117,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_008": {
        "sheet_name": "리바로하이 리바로브이",
        "header_row": 5,
        "raw_rows_scanned": 1096,
        "empty_rows": 15,
        "excluded_rows": 0,
        "staging_rows": 1081,
        "manual_specs": 5,
        "mapping_rows": 1653,
    },
    "strategy_009": {
        "sheet_name": "트루패스 피나스타 제이다트",
        "header_row": 5,
        "raw_rows_scanned": 418,
        "empty_rows": 12,
        "excluded_rows": 1,
        "staging_rows": 405,
        "manual_specs": 0,
        "mapping_rows": 0,
    },
    "strategy_010": {
        "sheet_name": "뉴트로진 모빌리아",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 985,
        "excluded_rows": 0,
        "staging_rows": 10,
        "manual_specs": 4,
        "mapping_rows": 38,
    },
    "strategy_011": {
        "sheet_name": "악템라",
        "header_row": 5,
        "raw_rows_scanned": 26,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 26,
        "manual_specs": 3,
        "mapping_rows": 78,
    },
    "strategy_012": {
        "sheet_name": "페린젝트 베노훼럼",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 919,
        "excluded_rows": 0,
        "staging_rows": 76,
        "manual_specs": 5,
        "mapping_rows": 182,
    },
    "strategy_013": {
        "sheet_name": "헴리브라",
        "header_row": 5,
        "raw_rows_scanned": 14,
        "empty_rows": 0,
        "excluded_rows": 1,
        "staging_rows": 13,
        "manual_specs": 2,
        "mapping_rows": 26,
    },
    "strategy_014": {
        "sheet_name": "위너프 위너프A+",
        "header_row": 5,
        "raw_rows_scanned": 995,
        "empty_rows": 664,
        "excluded_rows": 8,
        "staging_rows": 323,
        "manual_specs": 6,
        "mapping_rows": 1696,
    },
    "strategy_015": {
        "sheet_name": "엔커버",
        "header_row": 7,
        "raw_rows_scanned": 4,
        "empty_rows": 0,
        "excluded_rows": 0,
        "staging_rows": 4,
        "manual_specs": 1,
        "mapping_rows": 4,
    },
    "strategy_016": {
        "sheet_name": "플라주오피",
        "header_row": 5,
        "raw_rows_scanned": 54,
        "empty_rows": 2,
        "excluded_rows": 31,
        "staging_rows": 21,
        "manual_specs": 2,
        "mapping_rows": 40,
    },
}

EXPECTED_MARKET_DISTRIBUTION = {
    "strategy_001": 731,
    "strategy_002": 135,
    "strategy_003": 164,
    "strategy_004": 9,
    "strategy_005": 1176,
    "strategy_008": 1653,
    "strategy_010": 38,
    "strategy_011": 78,
    "strategy_012": 182,
    "strategy_013": 26,
    "strategy_014": 1696,
    "strategy_015": 4,
    "strategy_016": 40,
}

EXPECTED_MAPPING_TYPE_DISTRIBUTION = {
    "class_recode": 1353,
    "manual_mapping": 2024,
    "molecule_recode": 1984,
    "nhi_overlay": 490,
    "strength_recode": 81,
}


@dataclass
class MarketMappingStats:
    strategic_market_id: str
    sheet_name: str
    header_row: int
    max_row: int
    max_col: int
    raw_rows_scanned: int = 0
    empty_rows: int = 0
    excluded_rows: int = 0
    staging_rows: int = 0
    manual_specs: int = 0
    mapping_rows: int = 0


def utc_now_text() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def resolve_input_file(path: Path) -> Path:
    if path.exists():
        return path
    candidates = sorted(file for file in MASTER_ROOT.glob("*.xlsx") if not file.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError(f"No Master xlsx found under {MASTER_ROOT}")
    return candidates[-1]


def normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def is_excluded_row(values: list[Any] | tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if "제외" in text and not text.startswith("비제외"):
            return True
    return False


def _header_lookup(headers: list[Any] | tuple[Any, ...], values: list[Any] | tuple[Any, ...]) -> dict[str, Any]:
    lookup: dict[str, Any] = {}
    for header, value in zip(headers, values):
        normalized = normalize_header(header)
        if normalized is not None and normalized not in lookup:
            lookup[normalized] = value
    return lookup


def _lookup_source_value(lookup: dict[str, Any], source_column: str | None) -> Any:
    if source_column is None:
        return None
    if source_column in lookup:
        return lookup[source_column]
    for header, value in lookup.items():
        if header.startswith(source_column):
            return value
    return None


def _position_value(values: list[Any] | tuple[Any, ...], column_index: int) -> Any:
    if column_index <= 0 or column_index > len(values):
        return None
    return values[column_index - 1]


def _lookup_key(*values: Any) -> tuple[str, ...] | None:
    key: list[str] = []
    for value in values:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        key.append(text)
    return tuple(key)


def _single_lookup_key(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def explicit_lookup_join(data_rows: list[tuple[int, tuple[Any, ...]]]) -> dict[int, dict[str, Any]]:
    """Return strategy_008 lookup-derived standard column overwrites."""
    lookup1: dict[tuple[str, str], dict[str, Any]] = {}
    for _, values in data_rows:
        key = _lookup_key(_position_value(values, 17), _position_value(values, 18))
        if key and key not in lookup1:
            lookup1[key] = {
                "molecule": _position_value(values, 19),
                "molecule_disease_definition": _position_value(values, 20),
                "composition_type": _position_value(values, 21),
                "class": _position_value(values, 22),
            }

    lookup2: dict[str, Any] = {}
    for _, values in data_rows:
        key = _single_lookup_key(_position_value(values, 25))
        if key and key not in lookup2:
            lookup2[key] = _position_value(values, 26)

    overrides: dict[int, dict[str, Any]] = {}
    for source_row_id, values in data_rows:
        left_key = _lookup_key(_position_value(values, 2), _position_value(values, 3))
        if not left_key or left_key not in lookup1:
            continue
        row_override = dict(lookup1[left_key])
        molecule_key = _single_lookup_key(row_override.get("molecule"))
        row_override["class_2"] = lookup2.get(molecule_key) if molecule_key else None
        overrides[source_row_id] = row_override
    return overrides


def _extra_key(metadata_key: str) -> str:
    return metadata_key[len(STANDARD_PREFIX) :]


def apply_column_mapping(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    metadata: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    lookup = _header_lookup(headers, values)
    standard: dict[str, Any] = {}
    extras: dict[str, Any] = {}
    for target_column, spec in metadata.items():
        source_column = normalize_header(spec.get("source_column"))
        value = _lookup_source_value(lookup, source_column)
        if target_column.startswith(STANDARD_PREFIX):
            extras[_extra_key(target_column)] = to_jsonable(value)
        else:
            standard[target_column] = value
    return standard, extras


def load_column_metadata_catalog(path: Path) -> dict[str, dict[str, dict[str, Any]]]:
    text = path.read_text(encoding="utf-8")
    catalog: dict[str, dict[str, dict[str, Any]]] = {}
    current_market: str | None = None
    lines = iter(text.splitlines())
    for line in lines:
        heading = re.match(r"^###\s+(strategy_\d{3})\s+—", line)
        if heading:
            current_market = heading.group(1)
            continue
        if current_market and line.strip() == "```json":
            block: list[str] = []
            for json_line in lines:
                if json_line.strip() == "```":
                    break
                block.append(json_line)
            catalog[current_market] = json.loads("\n".join(block))
            current_market = None
    missing = sorted({config.strategic_market_id for config in MARKET_SHEETS} - set(catalog))
    if missing:
        raise ValueError(f"column metadata catalog missing markets: {missing}")
    return catalog


def make_mapping_id(
    strategic_market_id: str,
    source_column: object,
    source_value: object,
    target_column: object,
    target_value: object,
    source_row_id: int | None = None,
) -> str:
    parts = [
        strategic_market_id,
        str(source_row_id or ""),
        str(source_column or ""),
        str(source_value or ""),
        str(target_column or ""),
        str(target_value or ""),
    ]
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"{strategic_market_id}:{digest}"


def mapping_type_for(target_column: str, source_column: object) -> str:
    lowered = f"{target_column} {source_column or ''}".lower()
    if "molecule" in lowered or "성분" in lowered:
        return "molecule_recode"
    if "class" in lowered:
        return "class_recode"
    if "strength" in lowered or "규격" in lowered:
        return "strength_recode"
    if "nhi" in lowered or "급여" in lowered:
        return "nhi_overlay"
    return "manual_mapping"


def raw_value_for_mapping(
    headers: list[Any] | tuple[Any, ...],
    values: list[Any] | tuple[Any, ...],
    spec: dict[str, Any],
) -> Any:
    lookup = _header_lookup(headers, values)
    overlay_target = normalize_header(spec.get("overlay_target"))
    source_column = normalize_header(spec.get("source_column"))
    if overlay_target and overlay_target in lookup:
        return lookup.get(overlay_target)
    return _lookup_source_value(lookup, source_column)


def _headers_from_sheet(ws: Any, header_row: int) -> list[Any]:
    return list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))


def _manual_mapping_records(
    strategic_market_id: str,
    source_sheet: str,
    source_file_version: str,
    source_row_id: int,
    headers: list[Any],
    values: tuple[Any, ...],
    metadata: dict[str, dict[str, Any]],
    standard_values: dict[str, Any],
    extras: dict[str, Any],
    ingested_at: str,
    inclusion_standard_values: dict[str, Any] | None = None,
    inclusion_extras: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    inclusion_standard_values = inclusion_standard_values or standard_values
    inclusion_extras = inclusion_extras or extras
    for target_column, spec in metadata.items():
        metadata_type = str(spec.get("type") or "")
        if not metadata_type.startswith("manual"):
            continue
        source_column = spec.get("source_column")
        if target_column.startswith("drug_extra_json."):
            extra_key = target_column.split(".", 1)[1]
            target_value = extras.get(extra_key)
            inclusion_target_value = inclusion_extras.get(extra_key)
        else:
            target_value = standard_values.get(target_column)
            inclusion_target_value = inclusion_standard_values.get(target_column)
        source_value = raw_value_for_mapping(headers, values, spec)
        if source_value is None and inclusion_target_value is None:
            continue
        records.append(
            {
                "mapping_id": make_mapping_id(
                    strategic_market_id,
                    source_column,
                    source_value,
                    target_column,
                    target_value,
                    source_row_id=source_row_id,
                ),
                "strategic_market_id": strategic_market_id,
                "source_value": None if source_value is None else str(source_value),
                "target_column": target_column,
                "target_value": None if target_value is None else str(target_value),
                "mapping_type": mapping_type_for(target_column, source_column),
                "source_sheet": source_sheet,
                "source_file_version": source_file_version,
                "ingested_at": ingested_at,
            }
        )
    return records


def load_mapping_records(
    xlsx_path: Path,
    catalog_path: Path,
    ingested_at: str | None = None,
) -> tuple[list[dict[str, Any]], list[MarketMappingStats]]:
    metadata_catalog = load_column_metadata_catalog(catalog_path)
    timestamp = ingested_at or utc_now_text()
    records: list[dict[str, Any]] = []
    stats: list[MarketMappingStats] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for config in MARKET_SHEETS:
            if config.sheet_name not in wb.sheetnames:
                raise ValueError(f"required sheet not found: {config.sheet_name!r}; sheets={wb.sheetnames}")
            ws = wb[config.sheet_name]
            headers = _headers_from_sheet(ws, config.header_row)
            metadata = metadata_catalog[config.strategic_market_id]
            manual_specs = sum(
                1 for spec in metadata.values() if str(spec.get("type") or "").startswith("manual")
            )
            market_stats = MarketMappingStats(
                strategic_market_id=config.strategic_market_id,
                sheet_name=config.sheet_name,
                header_row=config.header_row,
                max_row=ws.max_row,
                max_col=ws.max_column,
                manual_specs=manual_specs,
            )

            row_items = list(
                enumerate(
                    ws.iter_rows(min_row=config.header_row + 1, values_only=True),
                    start=config.header_row + 1,
                )
            )
            explicit_overrides = (
                explicit_lookup_join(row_items) if config.strategic_market_id == "strategy_008" else {}
            )

            for source_row_id, values in row_items:
                market_stats.raw_rows_scanned += 1
                if is_empty_row(values):
                    market_stats.empty_rows += 1
                    continue
                if is_excluded_row(values):
                    market_stats.excluded_rows += 1
                    continue

                market_stats.staging_rows += 1
                standard_values, extras = apply_column_mapping(headers, values, metadata)
                inclusion_standard_values = standard_values
                if source_row_id in explicit_overrides:
                    standard_values = dict(standard_values)
                    standard_values.update(explicit_overrides[source_row_id])
                mapping_rows = _manual_mapping_records(
                    config.strategic_market_id,
                    config.sheet_name,
                    xlsx_path.name,
                    source_row_id,
                    headers,
                    values,
                    metadata,
                    standard_values,
                    extras,
                    timestamp,
                    inclusion_standard_values=inclusion_standard_values,
                )
                records.extend(mapping_rows)
                market_stats.mapping_rows += len(mapping_rows)

            stats.append(market_stats)
    finally:
        wb.close()

    return records, stats


def _count_by(records: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        value = str(record.get(key) or "")
        counts[value] = counts.get(value, 0) + 1
    return counts


def validate_records(records: list[dict[str, Any]], stats: list[MarketMappingStats]) -> None:
    if len(records) != EXPECTED_ROW_COUNT:
        raise ValueError(f"mapping row count must be {EXPECTED_ROW_COUNT}, found {len(records)}")

    mapping_ids = [record["mapping_id"] for record in records]
    if len(set(mapping_ids)) != EXPECTED_ROW_COUNT:
        duplicates = sorted({value for value in mapping_ids if mapping_ids.count(value) > 1})
        raise ValueError(f"mapping_id must be unique, duplicate examples={duplicates[:10]}")

    stats_by_market = {item.strategic_market_id: item for item in stats}
    expected_market_ids = {config.strategic_market_id for config in MARKET_SHEETS}
    if set(stats_by_market) != expected_market_ids:
        raise ValueError(
            f"market stats mismatch: expected={sorted(expected_market_ids)}, actual={sorted(stats_by_market)}"
        )

    for market_id, expected in EXPECTED_MARKET_STATS.items():
        actual = stats_by_market[market_id]
        for field in (
            "sheet_name",
            "header_row",
            "raw_rows_scanned",
            "empty_rows",
            "excluded_rows",
            "staging_rows",
            "manual_specs",
            "mapping_rows",
        ):
            actual_value = getattr(actual, field)
            expected_value = expected[field]
            if actual_value != expected_value:
                raise ValueError(
                    f"{market_id} {field} mismatch: expected={expected_value}, actual={actual_value}"
                )

    zero_mapping_actual = {item.strategic_market_id for item in stats if item.mapping_rows == 0}
    if zero_mapping_actual != ZERO_MAPPING_MARKETS:
        raise ValueError(
            f"zero mapping market mismatch: expected={sorted(ZERO_MAPPING_MARKETS)}, "
            f"actual={sorted(zero_mapping_actual)}"
        )

    market_distribution = _count_by(records, "strategic_market_id")
    if market_distribution != EXPECTED_MARKET_DISTRIBUTION:
        raise ValueError(
            f"market distribution mismatch: expected={EXPECTED_MARKET_DISTRIBUTION}, "
            f"actual={market_distribution}"
        )

    type_distribution = _count_by(records, "mapping_type")
    if type_distribution != EXPECTED_MAPPING_TYPE_DISTRIBUTION:
        raise ValueError(
            f"mapping_type distribution mismatch: expected={EXPECTED_MAPPING_TYPE_DISTRIBUTION}, "
            f"actual={type_distribution}"
        )

    expected_columns = set(MASTER_MAPPING_TABLE_COLUMNS)
    for index, record in enumerate(records, start=1):
        extra_columns = sorted(set(record) - expected_columns)
        missing_columns = sorted(expected_columns - set(record))
        if extra_columns or missing_columns:
            raise ValueError(
                f"row {index} schema mismatch: extra={extra_columns}, missing={missing_columns}"
            )
        if not record["mapping_id"]:
            raise ValueError(f"row {index} has blank mapping_id")
        if not record["target_column"]:
            raise ValueError(f"row {index} has blank target_column: {record}")
        if record["mapping_type"] not in EXPECTED_MAPPING_TYPE_DISTRIBUTION:
            raise ValueError(f"row {index} unexpected mapping_type: {record['mapping_type']}")


def write_parquet(records: list[dict[str, Any]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    schema = pa.schema([pa.field(column, pa.string()) for column in MASTER_MAPPING_TABLE_COLUMNS])
    rows = [
        {column: None if record[column] is None else str(record[column]) for column in MASTER_MAPPING_TABLE_COLUMNS}
        for record in records
    ]
    table = pa.Table.from_pylist(rows, schema=schema)
    pq.write_table(table, output_file, compression="zstd")


def print_summary(
    records: list[dict[str, Any]],
    stats: list[MarketMappingStats],
    output_file: Path,
) -> None:
    print("Phase 09d master_mapping_table load")
    print(f"mapping_rows: {len(records)}")
    print(f"unique_mapping_id: {len({record['mapping_id'] for record in records})}")
    print(f"mapping_type_distribution: {_count_by(records, 'mapping_type')}")
    print("market_distribution:")
    for item in stats:
        print(
            f"  {item.strategic_market_id} {item.sheet_name}: "
            f"raw={item.raw_rows_scanned}, empty={item.empty_rows}, "
            f"excluded={item.excluded_rows}, staging={item.staging_rows}, "
            f"manual_specs={item.manual_specs}, mapping={item.mapping_rows}"
        )
    if output_file.exists():
        print(f"output_file: {output_file} ({output_file.stat().st_size:,} bytes)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load MI Master mapping_table parquet.")
    parser.add_argument("--input-file", type=Path, default=DEFAULT_INPUT_FILE)
    parser.add_argument("--catalog-path", type=Path, default=DEFAULT_CATALOG_PATH)
    parser.add_argument("--output-file", type=Path, default=DEFAULT_OUTPUT_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = resolve_input_file(args.input_file)
    records, stats = load_mapping_records(xlsx_path, args.catalog_path)
    validate_records(records, stats)
    write_parquet(records, args.output_file)
    print_summary(records, stats, args.output_file)
    print("validate_records: PASS")


if __name__ == "__main__":
    main()
