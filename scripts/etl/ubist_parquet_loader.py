#!/usr/bin/env python3
"""Load UBIST xlsx raw sales into hive-partitioned Parquet.

The source workbook layout is fixed across the 53 UBIST workbooks:
row 1 contains metric names, row 2 contains dimensions or periods, and
data begins at row 3. The loader streams rows into one Parquet file per
period partition so the full load does not need to materialize in memory.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq


ROOT = Path(__file__).resolve().parents[2]
UBIST_ROOT = ROOT / "data" / "UBIST"
TARGET_DIR = ROOT / "parquet" / "ubist"
KST = ZoneInfo("Asia/Seoul")

CANONICAL_DIMENSIONS = [
    "제조사",
    "국내/외자",
    "판매사",
    "판매사2",
    "제품",
    "ATC",
    "브랜드",
    "약가",
    "성분",
    "성분용량",
    "일반/전문",
    "약품코드",
    "제형",
    "투여경로",
    "급여구분",
    "종별",
    "진료과",
    "연령",
    "성별",
]

PATENT_DIMENSIONS = [
    "물질특허만료일",
    "마지막특허만료일",
    "마지막특허특성",
    "PMS만료일",
    "약품허가일",
    "Generic",
]

METRIC_MAP = {
    "처방조제액(원)": "rx_amt",
    "처방건수_P": "rx_cnt",
    "처방량_P": "rx_qty",
}

PATENT_ALIASES = {
    "물질특허 만료일": "물질특허만료일",
    "마지막특허 만료일": "마지막특허만료일",
    "마지막특허 특성": "마지막특허특성",
}

COLUMNS = (
    CANONICAL_DIMENSIONS
    + PATENT_DIMENSIONS
    + list(METRIC_MAP.values())
    + [
        "period_yyyymm",
        "source_file",
        "source_folder",
        "source_sheet",
        "source_row_no",
        "ingested_at",
    ]
)

STRING_COLUMNS = (
    CANONICAL_DIMENSIONS
    + PATENT_DIMENSIONS
    + ["period_yyyymm", "source_file", "source_folder", "source_sheet", "ingested_at"]
)

SCHEMA = pa.schema(
    [(col, pa.string()) for col in CANONICAL_DIMENSIONS + PATENT_DIMENSIONS]
    + [(metric, pa.float64()) for metric in METRIC_MAP.values()]
    + [
        ("period_yyyymm", pa.string()),
        ("source_file", pa.string()),
        ("source_folder", pa.string()),
        ("source_sheet", pa.string()),
        ("source_row_no", pa.int64()),
        ("ingested_at", pa.string()),
    ]
)


@dataclass
class SheetMapping:
    sheet_name: str
    dim_cols: list[tuple[int, str, str]]
    duplicate_cols: list[tuple[int, str, str]]
    metric_cols: list[tuple[int, str, str, str]]


@dataclass
class PartitionStats:
    row_count: int = 0
    source_files: set[str] = field(default_factory=set)


def now_kst() -> str:
    return datetime.now(KST).isoformat(timespec="seconds")


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return unicodedata.normalize("NFC", text)


def to_string_or_none(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    return unicodedata.normalize("NFC", text) if text else None


def to_number_or_none(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_period(value: object) -> str:
    text = normalize_text(value)
    if not text:
        raise ValueError(f"unparseable period: {value!r}")
    match = re.match(r"^(\d{4})\s*년\s*(\d{1,2})\s*월$", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    raise ValueError(f"unparseable period: {value!r}")


def canonical_header(value: object) -> str | None:
    header = normalize_text(value)
    if not header:
        return None
    header = PATENT_ALIASES.get(header, header)
    if header in CANONICAL_DIMENSIONS or header in PATENT_DIMENSIONS:
        return header
    return None


def resolve_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if not path.is_absolute():
        path = ROOT / path
    if path.exists():
        return path

    candidates = [
        Path(unicodedata.normalize("NFC", str(path))),
        Path(unicodedata.normalize("NFD", str(path))),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    normalized_target = unicodedata.normalize("NFC", str(path))
    for candidate in ROOT.rglob(path.name):
        if unicodedata.normalize("NFC", str(candidate)) == normalized_target:
            return candidate

    raise FileNotFoundError(f"Path not found: {raw_path}")


def discover_xlsx(args: argparse.Namespace) -> list[Path]:
    paths: list[Path] = []
    if args.all:
        paths.extend(sorted(UBIST_ROOT.rglob("*.xlsx")))
    if args.folder:
        folder = resolve_path(args.folder)
        paths.extend(sorted(folder.rglob("*.xlsx")))
    if args.file:
        paths.append(resolve_path(args.file))
    unique = sorted({p.resolve() for p in paths})
    if not unique:
        raise RuntimeError("No xlsx files selected. Use --all, --folder, or --file.")
    return unique


def source_folder_for(path: Path) -> str:
    try:
        return str(path.parent.relative_to(UBIST_ROOT))
    except ValueError:
        return str(path.parent)


def classify_sheet(sheet_name: str, header1: tuple[object, ...], header2: tuple[object, ...]) -> SheetMapping:
    dim_cols: list[tuple[int, str, str]] = []
    duplicate_cols: list[tuple[int, str, str]] = []
    metric_cols: list[tuple[int, str, str, str]] = []
    seen_dims: set[str] = set()

    for idx in range(max(len(header1), len(header2))):
        h1 = normalize_text(header1[idx] if idx < len(header1) else None)
        h2 = normalize_text(header2[idx] if idx < len(header2) else None)

        if h1 in METRIC_MAP:
            try:
                period = parse_period(h2)
            except ValueError:
                continue
            metric_cols.append((idx, h1, METRIC_MAP[h1], period))
            continue

        canonical = canonical_header(h2)
        if canonical:
            if canonical in seen_dims:
                duplicate_cols.append((idx, h2 or "", canonical))
            else:
                seen_dims.add(canonical)
                dim_cols.append((idx, h2 or "", canonical))

    if not metric_cols:
        raise RuntimeError(f"No metric columns found in sheet: {sheet_name}")
    return SheetMapping(sheet_name, dim_cols, duplicate_cols, metric_cols)


def row_has_identifier(base: dict[str, object]) -> bool:
    return any(base.get(col) for col in ("약품코드", "제품", "성분", "브랜드"))


def iter_xlsx_rows(xlsx_path: Path):
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header1 = next(rows_iter)
            header2 = next(rows_iter)
        except StopIteration:
            continue

        mapping = classify_sheet(sheet_name, header1, header2)
        loaded_at = now_kst()

        for row_no, row in enumerate(rows_iter, start=3):
            base = {canonical: to_string_or_none(row[idx] if idx < len(row) else None) for idx, _, canonical in mapping.dim_cols}
            if not row_has_identifier(base):
                continue
            base.update(
                {
                    "source_file": xlsx_path.name,
                    "source_folder": source_folder_for(xlsx_path),
                    "source_sheet": sheet_name,
                    "source_row_no": row_no,
                    "ingested_at": loaded_at,
                }
            )

            period_metrics: dict[str, dict[str, float | None]] = defaultdict(dict)
            for idx, _, metric, period in mapping.metric_cols:
                period_metrics[period][metric] = to_number_or_none(row[idx] if idx < len(row) else None)

            for period, metrics in period_metrics.items():
                output = dict(base)
                output.update(metrics)
                output["period_yyyymm"] = period
                yield period, output


def prepare_frame(rows: list[dict[str, object]]) -> pd.DataFrame:
    frame = pd.DataFrame(rows)
    for col in COLUMNS:
        if col not in frame.columns:
            frame[col] = None
    for col in STRING_COLUMNS:
        frame[col] = frame[col].map(to_string_or_none)
    for metric in METRIC_MAP.values():
        frame[metric] = pd.to_numeric(frame[metric], errors="coerce")
    frame["source_row_no"] = pd.to_numeric(frame["source_row_no"], errors="coerce").fillna(0).astype("int64")
    return frame.reindex(columns=COLUMNS)


class PartitionWriter:
    def __init__(self, target_root: Path):
        self.target_root = target_root
        self.writers: dict[str, pq.ParquetWriter] = {}
        self.stats: dict[str, PartitionStats] = defaultdict(PartitionStats)
        self.loaded_existing: set[str] = set()

    def _path_for(self, period: str) -> Path:
        year, month = period.split("-")
        return self.target_root / f"year={year}" / f"month={month}" / "data.parquet"

    def _open_writer(self, period: str) -> pq.ParquetWriter:
        if period in self.writers:
            return self.writers[period]
        path = self._path_for(period)
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists() and period not in self.loaded_existing:
            existing = pq.read_table(path).select(COLUMNS)
            path.unlink()
            writer = pq.ParquetWriter(path, SCHEMA, compression="snappy")
            writer.write_table(existing)
            stats = self.stats[period]
            stats.row_count += existing.num_rows
            if "source_file" in existing.column_names:
                stats.source_files.update(str(value.as_py()) for value in existing["source_file"].unique() if value.as_py())
            self.loaded_existing.add(period)
        else:
            writer = pq.ParquetWriter(path, SCHEMA, compression="snappy")
        self.writers[period] = writer
        return writer

    def write_rows(self, period: str, rows: list[dict[str, object]]) -> None:
        if not rows:
            return
        frame = prepare_frame(rows)
        table = pa.Table.from_pandas(frame, schema=SCHEMA, preserve_index=False)
        self._open_writer(period).write_table(table)
        stats = self.stats[period]
        stats.row_count += len(rows)
        stats.source_files.update(str(row["source_file"]) for row in rows)

    def close(self) -> None:
        for writer in self.writers.values():
            writer.close()
        self.writers.clear()


def flush_buffers(writer: PartitionWriter, buffers: dict[str, list[dict[str, object]]], *, final: bool = False) -> None:
    threshold = 25_000
    for period in list(buffers):
        if final or len(buffers[period]) >= threshold:
            rows = buffers.pop(period)
            writer.write_rows(period, rows)


def load_to_parquet(xlsx_paths: list[Path], target: Path, *, mode: str, truncate: bool) -> dict[str, PartitionStats]:
    timestamp = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    tmp_target = target.parent / f"{target.name}.__tmp_{timestamp}"
    backup_target = target.parent / f"{target.name}.__backup_{timestamp}"

    if tmp_target.exists():
        shutil.rmtree(tmp_target)
    tmp_target.mkdir(parents=True)

    if mode == "append" and target.exists():
        shutil.copytree(target, tmp_target, dirs_exist_ok=True)
    elif mode != "replace":
        raise ValueError(f"Unsupported mode: {mode}")

    buffers: dict[str, list[dict[str, object]]] = defaultdict(list)
    writer = PartitionWriter(tmp_target)
    total_rows = 0
    try:
        for idx, xlsx_path in enumerate(xlsx_paths, start=1):
            print(f"[{idx}/{len(xlsx_paths)}] reading {xlsx_path}")
            for period, row in iter_xlsx_rows(xlsx_path):
                buffers[period].append(row)
                total_rows += 1
                if total_rows % 250_000 == 0:
                    print(f"  rows={total_rows:,} active_partitions={len(buffers)}")
                    flush_buffers(writer, buffers)
            flush_buffers(writer, buffers)
        flush_buffers(writer, buffers, final=True)
    finally:
        writer.close()

    write_manifest(tmp_target, writer.stats, xlsx_paths)

    if target.exists():
        if truncate or mode == "replace":
            target.rename(backup_target)
        else:
            raise RuntimeError(f"Target already exists: {target}")
    tmp_target.rename(target)
    if backup_target.exists():
        shutil.rmtree(backup_target)
    print(f"loaded rows={total_rows:,} partitions={len(writer.stats)} target={target}")
    return writer.stats


def partition_path(period: str) -> str:
    year, month = period.split("-")
    return f"year={year}/month={month}/data.parquet"


def write_manifest(target: Path, stats: dict[str, PartitionStats], xlsx_paths: list[Path]) -> None:
    generated_at = now_kst()
    manifest = {
        "schema_version": "1.0",
        "generated_at": generated_at,
        "storage": "parquet_hive_partition",
        "compression": "snappy",
        "metric_map": METRIC_MAP,
        "canonical_dimensions": CANONICAL_DIMENSIONS,
        "patent_dimensions": PATENT_DIMENSIONS,
        "dimension_duplicate_policy": {
            "decision": "collapse_duplicate_semantic_headers",
            "kept": "first occurrence in the fixed dimension block",
            "dropped_after_sample_match": ["제조사", "국내/외자", "판매사", "ATC", "성분"],
            "retained_distinct": ["판매사2"],
        },
        "source_file_count": len(xlsx_paths),
        "partitions": [
            {
                "period_yyyymm": period,
                "path": partition_path(period),
                "row_count": stats[period].row_count,
                "source_files": sorted(stats[period].source_files),
                "loaded_at": generated_at,
            }
            for period in sorted(stats)
        ],
    }
    target.mkdir(parents=True, exist_ok=True)
    (target / "_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def dry_run(xlsx_paths: list[Path], limit_rows: int = 3) -> None:
    print("# UBIST Parquet Loader Dry Run\n")
    for xlsx_path in xlsx_paths:
        print(f"## Source: {xlsx_path}\n")
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            header1 = next(rows_iter)
            header2 = next(rows_iter)
            mapping = classify_sheet(sheet_name, header1, header2)
            periods = sorted({period for _, _, _, period in mapping.metric_cols})
            metrics = sorted({metric for _, _, metric, _ in mapping.metric_cols})
            print(f"### Sheet: {sheet_name}")
            print(f"- rows: {worksheet.max_row:,}")
            print(f"- cols: {worksheet.max_column:,}")
            print(f"- kept dimensions: {len(mapping.dim_cols)}")
            for idx, raw, canonical in mapping.dim_cols:
                print(f"  - col {idx + 1}: {raw} -> {canonical}")
            print(f"- duplicate dimensions dropped: {len(mapping.duplicate_cols)}")
            for idx, raw, canonical in mapping.duplicate_cols:
                print(f"  - col {idx + 1}: {raw} -> {canonical}")
            print(f"- metrics: {metrics}")
            print(f"- periods: {periods[0]} .. {periods[-1]} ({len(periods)})")
            print("\n#### Sample Output Rows")
            samples = []
            for _, row in iter_xlsx_rows(xlsx_path):
                samples.append(row)
                if len(samples) >= limit_rows:
                    break
            frame = prepare_frame(samples)
            print(frame.head(limit_rows).to_string(index=False))
            print()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--folder", help="Folder containing UBIST xlsx files")
    source.add_argument("--file", help="Single UBIST xlsx file")
    source.add_argument("--all", action="store_true", help="Load all xlsx files below data/UBIST")
    parser.add_argument("--dry-run", action="store_true", help="Analyze schema and sample rows without writing")
    parser.add_argument("--truncate", action="store_true", help="Replace the existing parquet/ubist target")
    parser.add_argument("--mode", choices=["replace", "append"], default="replace")
    parser.add_argument("--target-dir", default=str(TARGET_DIR))
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    try:
        xlsx_paths = discover_xlsx(args)
        if args.dry_run:
            dry_run(xlsx_paths)
            return 0
        stats = load_to_parquet(xlsx_paths, Path(args.target_dir), mode=args.mode, truncate=args.truncate)
        print("partition summary:")
        for period in sorted(stats):
            print(f"  {period}: {stats[period].row_count:,}")
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
