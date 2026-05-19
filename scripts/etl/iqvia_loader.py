#!/usr/bin/env python3
"""Load IQVIA NSA/CSD/CHSO raw sources into MariaDB Layer 1 tables."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Iterator

import openpyxl
import pandas as pd
import pymysql


REPO_ROOT = Path(__file__).resolve().parents[2]
IQVIA_ROOT = REPO_ROOT / "data" / "IQVIA"
AUDIT_DIR = REPO_ROOT / "audit" / "phase16c3_iqvia_mariadb"

NSA_TABLE = "iqvia_nsa_quarterly_raw"
CSD_TABLE = "iqvia_csd_monthly_raw"
CHSO_TABLE = "iqvia_chso_monthly_raw"


MONTH_NAME_TO_NUM = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


INSERT_SQL = {
    NSA_TABLE: f"""
        INSERT INTO {NSA_TABLE}
        (source_file, sheet_name, source_row_no, audit_code, audit_desc, mfr_code,
         mfr_name, period_yyyy, period_quarter, period_label, payload, source_master_version)
        VALUES
        (%(source_file)s, %(sheet_name)s, %(source_row_no)s, %(audit_code)s, %(audit_desc)s,
         %(mfr_code)s, %(mfr_name)s, %(period_yyyy)s, %(period_quarter)s,
         %(period_label)s, %(payload)s, %(source_master_version)s)
    """,
    CSD_TABLE: f"""
        INSERT INTO {CSD_TABLE}
        (source_file, sheet_name, source_row_no, period_yyyymm, channel, region,
         payload, source_master_version)
        VALUES
        (%(source_file)s, %(sheet_name)s, %(source_row_no)s, %(period_yyyymm)s,
         %(channel)s, %(region)s, %(payload)s, %(source_master_version)s)
    """,
    CHSO_TABLE: f"""
        INSERT INTO {CHSO_TABLE}
        (source_file, sheet_name, source_row_no, period_yyyymm, payload, source_master_version)
        VALUES
        (%(source_file)s, %(sheet_name)s, %(source_row_no)s, %(period_yyyymm)s,
         %(payload)s, %(source_master_version)s)
    """,
}


@dataclass(frozen=True)
class SourceSheet:
    path: Path
    sheet_name: str | None = None


@dataclass
class LoadStats:
    rows: int = 0
    files: int = 0
    sheets: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def connect() -> pymysql.connections.Connection:
    env = load_env(REPO_ROOT / "docker" / ".env")
    return pymysql.connect(
        host="127.0.0.1",
        port=int(env.get("HOST_PORT", "3307")),
        user=env.get("MARIADB_USER", "jwapp"),
        password=env["MARIADB_PASSWORD"],
        database=env.get("MARIADB_DATABASE", "jw_mart"),
        charset="utf8mb4",
        autocommit=False,
    )


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if pd.isna(value) if not isinstance(value, (list, dict, tuple, set)) else False:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def clean_key(value: Any, idx: int) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return f"__blank_{idx}"
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text if text else f"__blank_{idx}"


def dedupe_keys(keys: list[str]) -> list[str]:
    counts: dict[str, int] = defaultdict(int)
    result: list[str] = []
    for key in keys:
        counts[key] += 1
        if counts[key] == 1:
            result.append(key)
        else:
            result.append(f"{key}_{counts[key]}")
    return result


def row_to_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {k: clean_value(v) for k, v in row.items()}


def dumps_payload(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str)


def discover_files(source: str) -> list[Path]:
    if source == "nsa":
        return sorted(
            p
            for p in (IQVIA_ROOT / "NSA").iterdir()
            if p.suffix.lower() in {".csv", ".xlsx", ".xls"}
        )
    if source == "csd":
        return sorted(
            p
            for p in (IQVIA_ROOT / "CSD").rglob("*")
            if p.suffix.lower() in {".xlsx", ".xls"}
        )
    if source == "chso":
        return sorted(
            p
            for p in (IQVIA_ROOT / "CHSO").iterdir()
            if p.suffix.lower() in {".xlsx", ".xls"}
        )
    raise ValueError(f"unknown source: {source}")


def quarter_from_month(month: int) -> int | None:
    return {3: 1, 6: 2, 9: 3, 12: 4}.get(month)


def parse_yyyymm(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    m = re.search(r"(\d{4})[./-](\d{1,2})", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

    m = re.search(r"(\d{1,2})/(\d{4})", text)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"

    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"

    for m in re.finditer(r"(?:^|[^A-Za-z])([A-Za-z]+)\.?\s*'?(\d{2,4})(?=$|[^0-9])", text):
        month = MONTH_NAME_TO_NUM.get(m.group(1).lower())
        if not month:
            continue
        if month:
            year = int(m.group(2))
            if year < 100:
                year += 2000
            return f"{year:04d}-{month:02d}"

    return None


def parse_period_from_filename(path: Path) -> str | None:
    return parse_yyyymm(path.stem)


def nsa_period_columns(headers: Iterable[str]) -> dict[str, dict[str, str]]:
    periods: dict[str, dict[str, str]] = defaultdict(dict)
    for header in headers:
        match = re.match(r"^(\d{1,2})/(\d{4})_(.+)$", str(header).strip())
        if not match:
            continue
        month = int(match.group(1))
        if quarter_from_month(month) is None:
            continue
        period = f"{int(match.group(2)):04d}-{month:02d}"
        periods[period][match.group(3)] = header
    return dict(periods)


def normalize_nsa_audit_desc(row: dict[str, Any]) -> Any:
    return row.get("AUDIT DESC") or row.get("AUDIT CODE")


def iter_nsa_csv(path: Path, chunk_size: int = 2000) -> Iterator[dict[str, Any]]:
    for chunk in pd.read_csv(path, encoding="utf-8-sig", chunksize=chunk_size, dtype=object):
        periods = nsa_period_columns(chunk.columns)
        static_cols = [c for c in chunk.columns if not re.match(r"^\d{1,2}/\d{4}_", str(c))]
        for offset, row in chunk.iterrows():
            raw = row_to_payload(row.to_dict())
            static = {k: raw.get(k) for k in static_cols}
            source_row_no = int(offset) + 2
            for period, metric_cols in periods.items():
                period_values = {metric: clean_value(raw.get(col)) for metric, col in metric_cols.items()}
                if all(v is None for v in period_values.values()):
                    continue
                year, month = period.split("-")
                quarter = quarter_from_month(int(month))
                if quarter is None:
                    continue
                yield {
                    "source_file": path.name,
                    "sheet_name": "CSV",
                    "source_row_no": source_row_no,
                    "audit_code": clean_value(raw.get("AUDIT CODE")),
                    "audit_desc": clean_value(normalize_nsa_audit_desc(raw)),
                    "mfr_code": clean_value(raw.get("MFR CODE")),
                    "mfr_name": clean_value(raw.get("MFR NAME")),
                    "period_yyyy": int(year),
                    "period_quarter": quarter,
                    "period_label": f"{year}Q{quarter}",
                    "payload": dumps_payload(
                        {
                            "__source": "NSA",
                            "__raw_period": period,
                            "__period_metric_columns": metric_cols,
                            "static": static,
                            "period_values": period_values,
                        }
                    ),
                    "source_master_version": None,
                }


def iter_nsa_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows)
        except StopIteration:
            continue
        headers = dedupe_keys([clean_key(v, idx) for idx, v in enumerate(headers_raw)])
        periods = nsa_period_columns(headers)
        static_cols = [c for c in headers if not re.match(r"^\d{1,2}/\d{4}_", c)]
        for row_no, values in enumerate(rows, start=2):
            if all(v is None for v in values):
                continue
            raw = row_to_payload(dict(zip(headers, values)))
            static = {k: raw.get(k) for k in static_cols}
            for period, metric_cols in periods.items():
                period_values = {metric: clean_value(raw.get(col)) for metric, col in metric_cols.items()}
                if all(v is None for v in period_values.values()):
                    continue
                year, month = period.split("-")
                quarter = quarter_from_month(int(month))
                if quarter is None:
                    continue
                yield {
                    "source_file": path.name,
                    "sheet_name": ws.title,
                    "source_row_no": row_no,
                    "audit_code": clean_value(raw.get("AUDIT CODE")),
                    "audit_desc": clean_value(normalize_nsa_audit_desc(raw)),
                    "mfr_code": clean_value(raw.get("MFR CODE")),
                    "mfr_name": clean_value(raw.get("MFR NAME")),
                    "period_yyyy": int(year),
                    "period_quarter": quarter,
                    "period_label": f"{year}Q{quarter}",
                    "payload": dumps_payload(
                        {
                            "__source": "NSA",
                            "__raw_period": period,
                            "__period_metric_columns": metric_cols,
                            "static": static,
                            "period_values": period_values,
                        }
                    ),
                    "source_master_version": None,
                }


def chso_period_columns(headers: Iterable[str]) -> dict[str, dict[str, str]]:
    periods: dict[str, dict[str, str]] = defaultdict(dict)
    for header in headers:
        text = str(header).strip()
        parts = text.split("\n")
        if len(parts) >= 2:
            metric = "\n".join(parts[:-1]).strip()
            raw_period = parts[-1].strip()
        else:
            match = re.match(r"^(.+?)\s+(\d{1,2}/\d{4})$", text)
            if not match:
                continue
            metric = match.group(1).strip()
            raw_period = match.group(2).strip()
        period = parse_yyyymm(raw_period)
        if not period:
            continue
        if metric:
            periods[period][metric] = header
    return dict(periods)


def iter_chso(path: Path) -> Iterator[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows)
        except StopIteration:
            continue
        headers = dedupe_keys([clean_key(v, idx) for idx, v in enumerate(headers_raw)])
        periods = chso_period_columns(headers)
        period_cols = {col for metric_cols in periods.values() for col in metric_cols.values()}
        static_cols = [c for c in headers if c not in period_cols]
        for row_no, values in enumerate(rows, start=2):
            if all(v is None for v in values):
                continue
            raw = row_to_payload(dict(zip(headers, values)))
            static = {k: raw.get(k) for k in static_cols}
            for period, metric_cols in periods.items():
                period_values = {metric: clean_value(raw.get(col)) for metric, col in metric_cols.items()}
                if all(v is None for v in period_values.values()):
                    continue
                yield {
                    "source_file": path.name,
                    "sheet_name": ws.title,
                    "source_row_no": row_no,
                    "period_yyyymm": period,
                    "payload": dumps_payload(
                        {
                            "__source": "CHSO",
                            "__raw_period": period,
                            "__period_metric_columns": metric_cols,
                            "static": static,
                            "period_values": period_values,
                        }
                    ),
                    "source_master_version": None,
                }


def find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    candidates = []
    for idx, row in enumerate(rows):
        values = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(values) < 2:
            continue
        score = len(values)
        if any(v in {"JW Channel", "Related date", "PRODUCT NAME", "Product name"} for v in values):
            score += 20
        if any("Rank" in v or "date" in v.lower() or "Product" in v for v in values):
            score += 5
        candidates.append((score, idx))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def metadata_from_rows(rows: list[tuple[Any, ...]]) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in rows[:8]:
        first = clean_value(row[0]) if row else None
        if isinstance(first, str) and ":" in first:
            key, value = first.split(":", 1)
            meta[key.strip()] = value.strip()
    return meta


def derive_channel(sheet_name: str, row: dict[str, Any], meta: dict[str, str]) -> str | None:
    for key in ("JW Channel", "VISIT LOCATION", "Hospital Type"):
        value = clean_value(row.get(key))
        if value:
            return str(value)[:64]
    if meta.get("Hospital Type"):
        return meta["Hospital Type"][:64]
    normalized = sheet_name.upper().replace(" ", "")
    for token in ["GH+SHPPI+CPPI", "GH+SHPPI", "SHPPI", "CPPI", "GH", "TOTAL"]:
        if token in normalized:
            return token[:64]
    return None


def derive_region(row: dict[str, Any]) -> str | None:
    for key in ("Region", "REGION", "VISIT LOCATION"):
        value = clean_value(row.get(key))
        if value:
            return str(value)[:64]
    return None


def derive_csd_period(path: Path, row: dict[str, Any], meta: dict[str, str]) -> str | None:
    file_period = parse_period_from_filename(path)
    if file_period:
        return file_period
    for key in ("Related date", "Meeting date", "RELATED DATE"):
        period = parse_yyyymm(row.get(key))
        if period:
            return period
    for key, value in row.items():
        if "date" in key.lower():
            period = parse_yyyymm(value)
            if period:
                return period
    for value in meta.values():
        period = parse_yyyymm(value)
        if period:
            return period
    return None


def iter_csd(path: Path) -> Iterator[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        if all(all(v is None for v in row) for row in rows):
            continue
        header_idx = find_header_row(rows[:12])
        if header_idx is None:
            continue
        meta = metadata_from_rows(rows[:header_idx])
        headers_raw = rows[header_idx]
        headers = dedupe_keys([clean_key(v, idx) for idx, v in enumerate(headers_raw)])
        for row_no, values in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            if all(v is None for v in values):
                continue
            raw = row_to_payload(dict(zip(headers, values)))
            period = derive_csd_period(path, raw, meta)
            payload = {
                "__source": "CSD",
                "__header_row": header_idx + 1,
                "__metadata": meta,
                "row": raw,
            }
            yield {
                "source_file": path.name,
                "sheet_name": ws.title,
                "source_row_no": row_no,
                "period_yyyymm": period,
                "channel": derive_channel(ws.title, raw, meta),
                "region": derive_region(raw),
                "payload": dumps_payload(payload),
                "source_master_version": None,
            }


def source_sheets_for_resume(source: str, files: list[Path]) -> list[SourceSheet]:
    if source == "nsa":
        result: list[SourceSheet] = []
        for path in files:
            if path.suffix.lower() == ".csv":
                result.append(SourceSheet(path, "CSV"))
            else:
                wb = openpyxl.load_workbook(path, read_only=True)
                result.extend(SourceSheet(path, ws) for ws in wb.sheetnames)
        return result
    if source in {"csd", "chso"}:
        result = []
        for path in files:
            wb = openpyxl.load_workbook(path, read_only=True)
            result.extend(SourceSheet(path, ws) for ws in wb.sheetnames)
        return result
    raise ValueError(source)


def loaded_sheet_keys(conn: pymysql.connections.Connection, table: str) -> set[tuple[str, str | None]]:
    with conn.cursor() as cur:
        cur.execute(f"SELECT DISTINCT source_file, sheet_name FROM {table}")
        return {(row[0], row[1]) for row in cur.fetchall()}


def batch_insert(
    conn: pymysql.connections.Connection,
    target_table: str,
    records: Iterable[dict[str, Any]],
    batch_size: int,
    max_batch_bytes: int = 8_000_000,
) -> int:
    sql = INSERT_SQL[target_table]
    total = 0
    batch: list[dict[str, Any]] = []
    batch_bytes = 0
    with conn.cursor() as cursor:
        for record in records:
            batch.append(record)
            batch_bytes += len(str(record.get("payload", "")).encode("utf-8")) + 1024
            if len(batch) >= batch_size or batch_bytes >= max_batch_bytes:
                cursor.executemany(sql, batch)
                conn.commit()
                total += len(batch)
                print(f"  [{target_table}] {total:,} rows committed", flush=True)
                batch.clear()
                batch_bytes = 0
        if batch:
            cursor.executemany(sql, batch)
            conn.commit()
            total += len(batch)
            print(f"  [{target_table}] {total:,} rows committed", flush=True)
    return total


def iter_records(source: str, path: Path) -> Iterator[dict[str, Any]]:
    if source == "nsa":
        if path.suffix.lower() == ".csv":
            yield from iter_nsa_csv(path)
        else:
            yield from iter_nsa_xlsx(path)
    elif source == "csd":
        yield from iter_csd(path)
    elif source == "chso":
        yield from iter_chso(path)
    else:
        raise ValueError(source)


def table_for_source(source: str) -> str:
    return {"nsa": NSA_TABLE, "csd": CSD_TABLE, "chso": CHSO_TABLE}[source]


def dry_run(source: str, files: list[Path], out_path: Path | None) -> None:
    lines: list[str] = [f"# IQVIA {source.upper()} Dry Run", ""]
    lines.append(f"- files: {len(files)}")
    sample_records: list[dict[str, Any]] = []
    total_preview = 0
    errors: list[str] = []

    for path in files[:1] if len(files) > 1 else files:
        lines.append(f"## File: `{path}`")
        try:
            if source == "nsa":
                periods = set()
                metrics = set()
                if path.suffix.lower() == ".csv":
                    df = pd.read_csv(path, nrows=3, encoding="utf-8-sig", dtype=object)
                    headers = list(df.columns)
                else:
                    df = pd.read_excel(path, nrows=3, dtype=object)
                    headers = list(df.columns)
                for period, metric_cols in nsa_period_columns(headers).items():
                    periods.add(period)
                    metrics.update(metric_cols)
                lines.append(f"- period count: {len(periods)}")
                lines.append(f"- period range: {min(periods) if periods else None} .. {max(periods) if periods else None}")
                lines.append(f"- metrics: {sorted(metrics)}")
                lines.append(f"- first columns: {headers[:20]}")
            elif source == "chso":
                df = pd.read_excel(path, nrows=3, dtype=object)
                periods = set()
                metrics = set()
                for period, metric_cols in chso_period_columns(df.columns).items():
                    periods.add(period)
                    metrics.update(metric_cols)
                lines.append(f"- period count: {len(periods)}")
                lines.append(f"- period range: {min(periods) if periods else None} .. {max(periods) if periods else None}")
                lines.append(f"- metrics: {sorted(metrics)}")
                lines.append(f"- first columns: {list(df.columns[:20])}")
            else:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                lines.append(f"- sheets: {wb.sheetnames}")
                for ws in wb.worksheets[:5]:
                    rows = list(ws.iter_rows(max_row=12, values_only=True))
                    header_idx = find_header_row(rows)
                    meta = metadata_from_rows(rows[: header_idx or 0])
                    lines.append(f"  - {ws.title}: header_row={None if header_idx is None else header_idx + 1}, meta={meta}")

            for record in iter_records(source, path):
                sample_records.append(record)
                total_preview += 1
                if len(sample_records) >= 3:
                    break
            lines.append(f"- preview records generated before stop: {total_preview}")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{path}: {exc}")
            lines.append(f"- ERROR: {exc}")

    lines.append("")
    lines.append("## Sample records")
    for record in sample_records[:3]:
        preview = dict(record)
        payload = preview.pop("payload", "")
        preview["payload_preview"] = payload[:800]
        lines.append("```json")
        lines.append(json.dumps(preview, ensure_ascii=False, indent=2, default=str))
        lines.append("```")

    if errors:
        lines.extend(["", "## Errors", *[f"- {e}" for e in errors]])

    text = "\n".join(lines) + "\n"
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text)
    print(text)


def load_source(source: str, files: list[Path], batch_size: int, dry: bool = False) -> LoadStats:
    stats = LoadStats()
    table = table_for_source(source)
    if dry:
        dry_run(source, files, None)
        return stats

    conn = connect()
    loaded = loaded_sheet_keys(conn, table)
    try:
        for path in files:
            stats.files += 1
            print(f"reading {source.upper()} {path}", flush=True)
            if source in {"nsa"} and path.suffix.lower() == ".csv":
                sheet_keys = [("CSV",)]
            else:
                wb = openpyxl.load_workbook(path, read_only=True)
                sheet_keys = [(s,) for s in wb.sheetnames]
            if all((path.name, sheet_name) in loaded for (sheet_name,) in sheet_keys):
                print(f"  skip already loaded file: {path.name}", flush=True)
                continue
            try:
                count = batch_insert(conn, table, iter_records(source, path), batch_size)
                stats.rows += count
                stats.sheets += len(sheet_keys)
                print(f"  loaded {path.name}: {count:,} rows", flush=True)
            except Exception as exc:  # noqa: BLE001
                conn.rollback()
                message = f"{path}: {exc}"
                stats.errors.append(message)
                print(f"  ERROR {message}", flush=True)
    finally:
        conn.close()
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", choices=["nsa", "csd", "chso"], help="source to load")
    parser.add_argument("--all", action="store_true", help="load all sources")
    parser.add_argument("--file", type=Path, help="load or dry-run one file")
    parser.add_argument("--dry-run", action="store_true", help="inspect schema without writing")
    parser.add_argument("--batch-size", type=int, default=10000)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.all and not args.source:
        print("ERROR: provide --source or --all", file=sys.stderr)
        return 2

    sources = ["nsa", "chso", "csd"] if args.all else [args.source]
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    all_errors: list[str] = []
    for source in sources:
        files = [args.file] if args.file else discover_files(source)
        files = [p for p in files if p is not None]
        if args.dry_run:
            out_path = AUDIT_DIR / f"dry_run_{source}.md"
            dry_run(source, files, out_path)
            continue
        stats = load_source(source, files, args.batch_size)
        print(
            f"SUMMARY {source.upper()}: files={stats.files}, sheets={stats.sheets}, "
            f"rows={stats.rows:,}, errors={len(stats.errors)}",
            flush=True,
        )
        all_errors.extend(stats.errors)

    if all_errors:
        print("ERRORS:", file=sys.stderr)
        for error in all_errors:
            print(f"- {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
