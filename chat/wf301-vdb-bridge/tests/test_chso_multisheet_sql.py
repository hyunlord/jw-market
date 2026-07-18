from __future__ import annotations

from hashlib import sha256
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from src import main, xlsx_sql_route
from src.xlsx_preprocessor import XlsxPreprocessError, extract_xlsx_chunks
from src.xlsx_sql_route import (
    FileSqlRouteConfig,
    SheetSqlProfile,
    classify_workbook_profiles,
    _fast_sheet_features,
    inspect_xlsx_for_sql,
    load_sql_sheet,
    logical_names_for_profiles,
    workbook_storage_route,
)


def _profile(index: int, name: str, *, columns: int = 8) -> SheetSqlProfile:
    return SheetSqlProfile(
        sheet_index=index,
        sheet_name=name,
        sheet_path=f"xl/worksheets/sheet{index}.xml",
        row_count=6_395,
        column_count=columns,
        used_cell_count=51_117,
        formula_cell_count=0,
        merged_range_count=0,
    )


def _compact_profile(
    index: int,
    name: str,
    *,
    rows: int,
    columns: int,
    used_cells: int,
) -> SheetSqlProfile:
    return SheetSqlProfile(
        sheet_index=index,
        sheet_name=name,
        sheet_path=f"xl/worksheets/sheet{index}.xml",
        row_count=rows,
        column_count=columns,
        used_cell_count=used_cells,
        formula_cell_count=0,
        merged_range_count=0,
    )


def _config() -> FileSqlRouteConfig:
    return FileSqlRouteConfig(
        enabled=True,
        min_rows=1_000,
        min_columns=8,
        max_columns=1_900,
        min_used_cells=20_000,
        min_density=0.10,
        max_merged_ranges=0,
    )


def test_fast_sql_profile_matches_dense_sheet_shape_and_used_cells(
    tmp_path: Path,
) -> None:
    path = tmp_path / "dense-profile.xlsx"
    sheet_xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="A1:C3"/>
  <sheetData>
    <row r="1" spans="1:3"><c r="A1"><v>1</v></c><c r="B1"><v>2</v></c><c r="C1"><is><t>x</t></is></c></row>
    <row r="2" spans="1:3"><c r="A2"><v>3</v></c><c r="B2"/><c r="C2"><v>4</v></c></row>
    <row r="3" spans="1:3"><c r="A3"><v>5</v></c><c r="B3"><v>6</v></c><c r="C3"><v>7</v></c></row>
  </sheetData>
</worksheet>"""
    with ZipFile(path, "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)

    with ZipFile(path) as archive:
        features = _fast_sheet_features(archive, "xl/worksheets/sheet1.xml")

    assert features is not None
    assert features.row_count == 3
    assert features.column_count == 3
    assert features.used_cell_count == 8
    assert features.formula_cell_count == 0
    assert features.merged_range_count == 0


def test_fast_sql_profile_falls_back_when_formula_or_span_proof_is_missing(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unsafe-profile.xlsx"
    formula_xml = b"""<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<dimension ref="A1:B2"/><sheetData><row r="1" spans="1:2"><c r="A1"><f>1+1</f><v>2</v></c></row>
<row r="2" spans="1:2"><c r="B2"><v>3</v></c></row></sheetData></worksheet>"""
    missing_span_xml = formula_xml.replace(b"<f>1+1</f>", b"").replace(
        b' spans="1:2"', b""
    )
    with ZipFile(path, "w") as archive:
        archive.writestr("xl/worksheets/formula.xml", formula_xml)
        archive.writestr("xl/worksheets/missing-span.xml", missing_span_xml)

    with ZipFile(path) as archive:
        assert _fast_sheet_features(archive, "xl/worksheets/formula.xml") is None
        assert _fast_sheet_features(archive, "xl/worksheets/missing-span.xml") is None


def test_fast_sql_profile_falls_back_when_row_or_dimension_proof_is_invalid(
    tmp_path: Path,
) -> None:
    path = tmp_path / "invalid-profile.xlsx"
    missing_row_xml = b"""<worksheet><dimension ref="A1:B3"/><sheetData>
<row r="1" spans="1:2"><c r="A1"><v>1</v></c></row>
<row r="3" spans="1:2"><c r="B3"><v>2</v></c></row></sheetData></worksheet>"""
    stale_span_xml = missing_row_xml.replace(
        b'<dimension ref="A1:B3"', b'<dimension ref="A1:B2"'
    ).replace(b'<row r="3" spans="1:2"', b'<row r="2" spans="1:3"')
    invalid_dimension_xml = missing_row_xml.replace(b"A1:B3", b"A1:\xff3")
    with ZipFile(path, "w") as archive:
        archive.writestr("xl/worksheets/missing-row.xml", missing_row_xml)
        archive.writestr("xl/worksheets/stale-span.xml", stale_span_xml)
        archive.writestr("xl/worksheets/invalid-dimension.xml", invalid_dimension_xml)

    with ZipFile(path) as archive:
        assert _fast_sheet_features(archive, "xl/worksheets/missing-row.xml") is None
        assert _fast_sheet_features(archive, "xl/worksheets/stale-span.xml") is None
        assert (
            _fast_sheet_features(archive, "xl/worksheets/invalid-dimension.xml") is None
        )


def test_large_sql_inspection_prefers_proven_fast_profile(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Large Dense"
    sheet.append(["brand", "sales"])
    sheet.append(["LIVALO", 100])
    path = tmp_path / "large-dense.xlsx"
    workbook.save(path)
    fast_profile = _compact_profile(
        1,
        "Large Dense",
        rows=2,
        columns=2,
        used_cells=4,
    )
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=fast_profile.row_count,
            column_count=fast_profile.column_count,
            used_cell_count=fast_profile.used_cell_count,
            formula_cell_count=fast_profile.formula_cell_count,
            merged_range_count=fast_profile.merged_range_count,
        ),
    )
    monkeypatch.setattr(
        xlsx_sql_route,
        "_sheet_features_streaming",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("structured fallback must not run")
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())

    assert decision.route == "sql"
    assert decision.profiles[0].audit_dict() == fast_profile.audit_dict()


def test_proven_dense_sql_sheet_reuses_validated_xml_for_header_and_rows(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dense SQL"
    sheet.append(["brand", "sales", "note"])
    sheet.append(["LIVALO & Co", 100, None])
    sheet.append(["LIPITOR", None, "comparison"])
    path = tmp_path / "dense-sql.xlsx"
    workbook.save(path)

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=3,
            column_count=3,
            used_cell_count=7,
            formula_cell_count=0,
            merged_range_count=0,
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())
    monkeypatch.setattr(
        xlsx_sql_route,
        "_iter_sheet_rows",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("validated dense XML must not use the generic row decoder")
        ),
    )

    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert list(sql_sheet.rows()) == [
        ("LIVALO & Co", "100", None),
        ("LIPITOR", None, "comparison"),
    ]


def test_proven_dense_sql_sheet_reuses_content_proof_without_rescanning_cells(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    source_path = tmp_path / "source.xlsx"
    path = tmp_path / "dense-proof.xlsx"
    workbook.save(source_path)
    sheet_xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="A1:B2"/><sheetData>
    <row r="1" spans="1:2"><c r="A1" t="inlineStr"><is><t>brand</t></is></c><c r="B1" t="inlineStr"><is><t>sales</t></is></c></row>
    <row r="2" spans="1:2"><c r="A2" t="inlineStr"><is><t>LIVALO</t></is></c><c r="B2" t="n"><v>100</v></c></row>
  </sheetData>
</worksheet>"""
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = (
                sheet_xml
                if item.filename == "xl/worksheets/sheet1.xml"
                else source.read(item.filename)
            )
            target.writestr(item, payload)

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    decision = inspect_xlsx_for_sql(path, _config())
    profile = decision.selected_sheets[0]
    dense_cell_pattern = xlsx_sql_route._DENSE_VALUE_CELL_RE

    class PatternSpy:
        def __init__(self) -> None:
            self.calls = 0

        def finditer(self, raw: bytes):
            self.calls += 1
            return dense_cell_pattern.finditer(raw)

    pattern_spy = PatternSpy()
    monkeypatch.setattr(xlsx_sql_route, "_DENSE_VALUE_CELL_RE", pattern_spy)

    sql_sheet = load_sql_sheet(path, profile)

    assert profile.dense_xml_sha256 == sha256(sheet_xml).hexdigest()
    assert list(sql_sheet.rows()) == [("LIVALO", "100")]
    assert pattern_spy.calls == 2


def test_proven_dense_sql_sheet_rechecks_content_proof_after_file_change(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    source_path = tmp_path / "source.xlsx"
    path = tmp_path / "dense-change.xlsx"
    workbook.save(source_path)
    sheet_xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="A1:B2"/><sheetData>
    <row r="1" spans="1:2"><c r="A1" t="inlineStr"><is><t>brand</t></is></c><c r="B1" t="inlineStr"><is><t>sales</t></is></c></row>
    <row r="2" spans="1:2"><c r="A2" t="inlineStr"><is><t>LIVALO</t></is></c><c r="B2" t="n"><v>100</v></c></row>
  </sheetData>
</worksheet>"""
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = (
                sheet_xml
                if item.filename == "xl/worksheets/sheet1.xml"
                else source.read(item.filename)
            )
            target.writestr(item, payload)

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    decision = inspect_xlsx_for_sql(path, _config())
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = (
                sheet_xml.replace(b">100<", b">101<")
                if item.filename == "xl/worksheets/sheet1.xml"
                else source.read(item.filename)
            )
            target.writestr(item, payload)

    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert sql_sheet.dense_sheet_xml is None
    assert list(sql_sheet.rows()) == [("LIVALO", "101")]


def test_proven_dense_sql_sheet_falls_back_before_yielding_unsupported_xml(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["brand", "sales"])
    sheet.append(["LIVALO", 100])
    source_path = tmp_path / "source.xlsx"
    path = tmp_path / "unsupported-cell.xlsx"
    workbook.save(source_path)
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(
                    b't="n"><v>100</v>',
                    b't="n" custom="1"><v>100</v>',
                )
            target.writestr(item, payload)

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=2,
            column_count=2,
            used_cell_count=4,
            formula_cell_count=0,
            merged_range_count=0,
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())
    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert sql_sheet.dense_sheet_xml is None
    assert list(sql_sheet.rows()) == [("LIVALO", "100")]


def test_proven_dense_sql_sheet_rejects_cells_hidden_in_xml_comments(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["brand", "sales"])
    sheet.append(["LIVALO", 100])
    source_path = tmp_path / "source.xlsx"
    path = tmp_path / "commented-cell.xlsx"
    workbook.save(source_path)
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(
                    b'<c r="B2" t="n"><v>100</v></c>',
                    b'<!--<c r="B2" t="n"><v>999</v></c>-->',
                )
            target.writestr(item, payload)

    with ZipFile(path) as archive:
        assert _fast_sheet_features(archive, "xl/worksheets/sheet1.xml") is None

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=2,
            column_count=2,
            used_cell_count=4,
            formula_cell_count=0,
            merged_range_count=0,
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())
    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert sql_sheet.dense_sheet_xml is None
    assert list(sql_sheet.rows()) == [("LIVALO", None)]


def test_proven_dense_sql_sheet_rejects_invalid_utf8(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    workbook.active.append(["brand", "sales"])
    workbook.active.append(["LIVALO", 100])
    source_path = tmp_path / "source.xlsx"
    path = tmp_path / "invalid-utf8.xlsx"
    workbook.save(source_path)
    with ZipFile(source_path) as source, ZipFile(path, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"LIVALO", b"LIV\xffLO")
            target.writestr(item, payload)

    with ZipFile(path) as archive:
        assert _fast_sheet_features(archive, "xl/worksheets/sheet1.xml") is None

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=2,
            column_count=2,
            used_cell_count=4,
            formula_cell_count=0,
            merged_range_count=0,
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())
    with pytest.raises(XlsxPreprocessError, match="xlsx SQL header read failed"):
        load_sql_sheet(path, decision.selected_sheets[0])


def test_proven_dense_sql_sheet_respects_raw_xml_memory_limit(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    workbook.active.append(["brand", "sales"])
    workbook.active.append(["LIVALO", 100])
    path = tmp_path / "bounded-dense.xlsx"
    workbook.save(path)

    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MIN_XML_BYTES", 0)
    monkeypatch.setattr(xlsx_sql_route, "FAST_SQL_PROFILE_MAX_XML_BYTES", 10**9)
    monkeypatch.setattr(xlsx_sql_route, "DENSE_SQL_MAX_XML_BYTES", 1)
    monkeypatch.setattr(
        xlsx_sql_route,
        "_fast_sheet_features",
        lambda _archive, _sheet_path: xlsx_sql_route.SheetFeatures(
            row_count=2,
            column_count=2,
            used_cell_count=4,
            formula_cell_count=0,
            merged_range_count=0,
        ),
    )

    decision = inspect_xlsx_for_sql(path, _config())
    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert sql_sheet.dense_sheet_xml is None
    assert list(sql_sheet.rows()) == [("LIVALO", "100")]


def test_chso_eight_column_dense_sheet_is_sql_candidate() -> None:
    decision = classify_workbook_profiles(
        (_profile(1, "LIVALO Market"), _profile(2, "Seven Columns", columns=7)),
        _config(),
    )

    assert decision.route == "sql"
    assert [sheet.sheet_name for sheet in decision.selected_sheets] == ["LIVALO Market"]


def test_chso_wide_sheet_preserves_all_252_columns_including_last_measure(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sell Out Standard"
    headers = ["AUDIT DESC", "MFR NAME KOR", "PRODUCT NAME KOR"]
    headers.extend(f"COLUMN {index}" for index in range(4, 252))
    headers.append("VALUES LC SI PRICE\n1/2026")
    sheet.append(headers)
    sheet.append(["Sell_Out", "DONG-A", "LIVALO", *range(4, 253)])
    path = tmp_path / "chso-wide.xlsx"
    workbook.save(path)

    decision = inspect_xlsx_for_sql(path, _config())
    sql_sheet = load_sql_sheet(path, decision.selected_sheets[0])

    assert decision.route == "sql"
    assert len(sql_sheet.columns) == 252
    assert sql_sheet.columns[:3] == (
        "AUDIT DESC",
        "MFR NAME KOR",
        "PRODUCT NAME KOR",
    )
    assert sql_sheet.columns[-1] == "VALUES LC SI PRICE\n1/2026"
    assert len(next(sql_sheet.rows())) == 252


def test_compact_tabular_sheets_are_sql_candidates_but_prose_sheet_stays_vdb() -> None:
    decision = classify_workbook_profiles(
        (
            _compact_profile(1, "Questions", rows=15, columns=10, used_cells=141),
            _compact_profile(2, "Coverage", rows=27, columns=5, used_cells=135),
            _compact_profile(3, "Notes", rows=27, columns=1, used_cells=22),
        ),
        _config(),
    )

    assert decision.route == "sql"
    assert [sheet.sheet_name for sheet in decision.selected_sheets] == [
        "Questions",
        "Coverage",
    ]


def test_default_column_floor_is_eight(monkeypatch) -> None:
    monkeypatch.delenv("FILE_SQL_ROUTE_MIN_COLUMNS", raising=False)

    assert FileSqlRouteConfig.from_env().min_columns == 8


def test_selected_sql_sheets_are_excluded_from_residual_vdb_chunks(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sql_sheet = workbook.active
    sql_sheet.title = "LIVALO Market"
    sql_sheet.append(["brand", "sales"])
    sql_sheet.append(["LIVALO", 100])
    residual_sheet = workbook.create_sheet("Overview")
    residual_sheet.append(["metric", "value"])
    residual_sheet.append(["share", "3.81%"])
    path = tmp_path / "mixed.xlsx"
    workbook.save(path)

    chunks = extract_xlsx_chunks(
        path,
        exclude_sheet_names=frozenset({"LIVALO Market"}),
    )

    assert chunks
    assert all("LIVALO Market" not in chunk for chunk in chunks)
    assert any("Overview" in chunk and "3.81%" in chunk for chunk in chunks)


def test_all_sql_sheets_skip_redundant_residual_workbook_scan(
    tmp_path: Path, monkeypatch
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sell Out Standard"
    sheet.append(["brand", "sales"])
    sheet.append(["LIVALO", 100])
    path = tmp_path / "all-sql.xlsx"
    workbook.save(path)
    document = main.TempDocument(
        temp_document_id=1,
        file_name=path.name,
        file_path=str(path),
    )

    def fail_if_scanned(_path: Path) -> bool:
        raise AssertionError(
            "all-SQL workbook must not be scanned for residual VDB chunks"
        )

    monkeypatch.setattr(main, "should_stream_xlsx_chunks", fail_if_scanned)

    collection, texts, notes, file_size = main._load_local_xlsx_texts(
        document,
        exclude_sheet_names=frozenset({"Sell Out Standard"}),
        known_sheet_names=frozenset({"Sell Out Standard"}),
    )

    assert collection == "local_xlsx_preprocessor"
    assert texts == []
    assert notes == ["SQL 시트 제외 후 VDB 잔여 청킹: Sell Out Standard"]
    assert file_size == path.stat().st_size


def test_logical_names_are_readable_and_collision_safe() -> None:
    names = logical_names_for_profiles(
        (
            _profile(1, "LIVALO Market"),
            _profile(2, "LIVALO-Market"),
            _profile(3, "PPI Market"),
        ),
        scope_prefix="doc_42",
    )

    assert names == (
        "doc_42_livalo_market",
        "doc_42_livalo_market_2",
        "doc_42_ppi_market",
    )


def test_workbook_storage_route_preserves_both_sql_and_vdb() -> None:
    assert workbook_storage_route(has_sql=True, vdb_chunk_count=1_246) == "hybrid"
    assert workbook_storage_route(has_sql=True, vdb_chunk_count=0) == "sql"
    assert workbook_storage_route(has_sql=False, vdb_chunk_count=1_246) == "vdb"


def test_hybrid_document_exposes_sql_sources_while_remaining_vdb_searchable() -> None:
    rows = [
        {
            "document_id": 42,
            "file_name": "channel-dynamics.xlsx",
            "storage_route": "hybrid",
            "sql_tables": [
                {
                    "logical_name": "doc_42_livalo_market",
                    "sheet_name": "LIVALO Market",
                    "row_count": 6_395,
                    "column_count": 8,
                }
            ],
        }
    ]

    sources = main._sql_sources_from_rows(rows)
    vdb_rows = [row for row in rows if row.get("storage_route") != "sql"]

    assert [source.logical_name for source in sources] == ["doc_42_livalo_market"]
    assert vdb_rows == rows
