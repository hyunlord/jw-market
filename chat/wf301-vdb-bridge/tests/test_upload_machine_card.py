from pathlib import Path
from zipfile import ZipFile

import fitz
from openpyxl import Workbook

from src.upload_machine_card import inspect_upload_machine_card


def test_xlsx_card_reports_observed_sheet_dimensions(tmp_path: Path) -> None:
    path = tmp_path / "wide.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sell Out Standard"
    sheet.cell(row=12_269, column=252, value="last")
    workbook.create_sheet("Lookup").cell(row=4, column=3, value="value")
    workbook.save(path)

    card = inspect_upload_machine_card(path, "wide.xlsx")

    assert card.file_type == "xlsx"
    assert card.size_bytes == path.stat().st_size
    assert card.sheet_count == 2
    assert [(item.name, item.row_count, item.column_count) for item in card.sheets] == [
        ("Sell Out Standard", 12_269, 252),
        ("Lookup", 4, 3),
    ]
    assert card.page_count is None
    assert card.slide_count is None


def test_pdf_card_reports_page_count_without_reading_page_text(tmp_path: Path) -> None:
    path = tmp_path / "report.pdf"
    document = fitz.open()
    document.new_page()
    document.new_page()
    document.save(path)
    document.close()

    card = inspect_upload_machine_card(path, "report.pdf")

    assert card.file_type == "pdf"
    assert card.page_count == 2
    assert card.sheets == ()


def test_pptx_card_counts_slides_from_package_members(tmp_path: Path) -> None:
    path = tmp_path / "brief.pptx"
    with ZipFile(path, "w") as archive:
        archive.writestr("ppt/slides/slide1.xml", "<slide />")
        archive.writestr("ppt/slides/slide2.xml", "<slide />")
        archive.writestr("ppt/notesSlides/notesSlide1.xml", "<notes />")

    card = inspect_upload_machine_card(path, "brief.pptx")

    assert card.file_type == "pptx"
    assert card.slide_count == 2
    assert card.page_count is None


def test_broken_package_returns_safe_minimal_card(tmp_path: Path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not-a-workbook")

    card = inspect_upload_machine_card(path, "broken.xlsx")

    assert card.file_type == "xlsx"
    assert card.size_bytes == len(b"not-a-workbook")
    assert card.sheet_count is None
    assert card.sheets == ()
