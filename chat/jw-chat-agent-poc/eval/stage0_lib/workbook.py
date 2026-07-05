from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .model import ScoredRow

FIXED_HEADERS = (
    "질문",
    "카테고리",
    "골드기준",
)


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _apply_table(sheet, name: str) -> None:
    table = Table(displayName=name, ref=sheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def _set_widths(sheet) -> None:
    widths = {
        "A": 34,
        "B": 18,
        "C": 44,
    }
    for column in range(4, sheet.max_column + 1, 4):
        widths[get_column_letter(column)] = 80
        widths[get_column_letter(column + 1)] = 13
        widths[get_column_letter(column + 2)] = 13
        widths[get_column_letter(column + 3)] = 44
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_scores(sheet) -> None:
    fills = {
        "O": PatternFill("solid", fgColor="D9EAD3"),
        "X": PatternFill("solid", fgColor="F4CCCC"),
        "NA": PatternFill("solid", fgColor="FFF2CC"),
    }
    for score_column in range(5, sheet.max_column + 1, 4):
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=score_column)
            fill = fills.get(str(cell.value))
            if fill is not None:
                cell.fill = fill
            sheet.cell(row=row, column=score_column + 1).alignment = Alignment(horizontal="center")


def _write_summary(workbook: Workbook, rows: list[ScoredRow]) -> None:
    sheet = workbook.create_sheet("summary")
    sheet.append(["카테고리", "문항수", "숫자 O", "숫자 X", "숫자 NA", "정성평균"])
    grouped: dict[str, list[ScoredRow]] = defaultdict(list)
    for row in rows:
        grouped[row.question.category].append(row)
    for category, category_rows in sorted(grouped.items()):
        numeric = [row.numeric_accuracy for row in category_rows]
        avg = sum(row.qualitative_score for row in category_rows) / len(category_rows)
        sheet.append(
            [
                category,
                len(category_rows),
                numeric.count("O"),
                numeric.count("X"),
                numeric.count("NA"),
                round(avg, 2),
            ]
        )
    _style_header(sheet)
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18


def _write_pl_slot(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("PL 추가 슬롯")
    sheet.append(["id", "category", "question", "gold_note", "expected_behavior", "gold_keys"])
    sheet.append(
        [
            "PL001",
            "PL 추가",
            "여기에 PL 실제 질문을 추가",
            "기대 동작 또는 cache 기준",
            "manual",
            "[]",
        ]
    )
    _style_header(sheet)
    widths = [14, 16, 42, 44, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _version_headers(version: str) -> list[str]:
    return [f"{version}_답변", f"{version}_숫자정확", f"{version}_정성점수", f"{version}_비고"]


def _scored_cells(row: ScoredRow) -> list[Any]:
    observations = "; ".join(
        f"{item.label}={item.value}" for item in row.gold_observations
    )
    note = row.note if not observations else f"{row.note}; gold={observations}"
    return [row.answer, row.numeric_accuracy, row.qualitative_score, note]


def _saved_cells(row: dict[str, Any] | None) -> list[Any]:
    if row is None:
        return ["", "NA", "", "previous row missing"]
    return [
        row.get("answer", ""),
        row.get("numeric_accuracy", "NA"),
        row.get("qualitative_score", ""),
        row.get("note", ""),
    ]


def write_workbook(
    path: Path,
    rows: list[ScoredRow],
    *,
    version: str = "v1",
    previous_blocks: tuple[tuple[str, dict[str, dict[str, Any]]], ...] = (),
) -> None:
    """Write the Stage 0 Excel workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"evalset_{version}"
    headers = list(FIXED_HEADERS)
    for previous_version, _previous_rows in previous_blocks:
        headers.extend(_version_headers(previous_version))
    headers.extend(_version_headers(version))
    sheet.append(headers)
    for row in rows:
        row_cells = [row.question.question, row.question.category, row.question.gold_note]
        for _previous_version, previous_rows in previous_blocks:
            row_cells.extend(_saved_cells(previous_rows.get(row.question.question_id)))
        row_cells.extend(_scored_cells(row))
        sheet.append(row_cells)
    _style_header(sheet)
    _set_widths(sheet)
    _style_scores(sheet)
    _apply_table(sheet, "Stage0Baseline")
    thin = Side(style="thin", color="D9E2F3")
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    _write_summary(workbook, rows)
    _write_pl_slot(workbook)
    workbook.save(path)
