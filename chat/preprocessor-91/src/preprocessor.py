"""
지원 확장자: .pdf, .ppt, .pptx, .xlsx
VLM/Docling/Audio/HWP 의존성 제거.

xlsx 전처리 개선 (Data Monitoring 양식 대응):
  1. 시트 분류(Triage): 정형 테이블 / 보고서형 레이아웃 / 빈 시트 자동 판별
  2. 병합 셀 unmerge + fill-down (좌상단 값을 병합 범위 전체에 전파)
  3. 헤더 행 자동 감지 + 다중(2단) 헤더 평탄화 ("상반기_매출" 식 결합)
  4. 헤더 위 제목/설명 행 → 표 컨텍스트로 보존
  5. Markdown 테이블 직렬화 (dict repr 제거)
  6. 행 경계 기준 청킹 + 모든 청크에 컨텍스트 헤더(문서/시트/표제목/컬럼) 반복
  7. 하이퍼링크 → Markdown 링크 보존, 숨김 시트 제외, 이미지 → media_files 메타데이터

pdf 전처리 개선 (1단계 — 의존성 최소, PyMuPDF4LLM 기반):
  1. flat text 대신 PyMuPDF4LLM으로 헤딩/표/리스트가 보존된 Markdown 추출
  2. 헤딩 경계 우선 청킹 (섹션 블록 단위 packing, 크기 초과 시에만 라인 분할)
  3. 표 분할 시 Markdown 헤더 행(컬럼명+구분선)을 다음 청크에 반복
  4. 모든 청크에 [DA] 컨텍스트 헤더(문서명/페이지/헤딩 경로) 반복 → 자기완결적 청크
  5. chunk_bboxes에 실제 bbox 채움 (page.search_for 기반, 0~1 정규화 좌표)
     → 출처 뷰어에서 청크 위치 하이라이트 가능 (PDF는 뷰어 페이지와 1:1 대응)
  6. ppt/pptx도 PDF 변환 후 동일 경로로 처리 (변환 실패 시 레거시 폴백)
  7. 실패 시 기존 PyMuPDF flat text + RecursiveCharacterTextSplitter 경로로 자동 폴백
"""

from __future__ import annotations

import asyncio
import fcntl
import gc
import json
import logging
import multiprocessing
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Type

from fastapi import Request
from pydantic import BaseModel

_log = logging.getLogger(__name__)

# ============================================================================
# Constants
# ============================================================================

SUPPORTED_EXTS = {".pdf", ".ppt", ".pptx", ".xlsx"}

# xlsx 시트 분류 파라미터
HEADER_SCAN_ROWS = 10          # 헤더 후보를 탐색할 상단 행 수
HEADER_MIN_FILL_RATIO = 0.6    # 헤더 행으로 인정할 최소 채움 비율 (사용 열 대비)
REPORT_MERGED_RATIO = 0.15     # 병합 셀 비율이 이 이상이면 보고서형으로 분류
MAX_TITLE_LINES = 5            # 표 위 제목/설명으로 보존할 최대 행 수

# pdf 청킹/좌표 파라미터
PDF_CTX_RESERVE = 150          # 컨텍스트 헤더용으로 청크 예산에서 예약할 문자 수
PDF_MAX_BBOX_PER_CHUNK = 8     # 청크당 저장할 최대 bbox 수
PDF_BBOX_SEARCH_TRUNC = 60     # search_for에 사용할 라인 최대 길이 (하이픈/개행 오탐 방지)
PDF_BBOX_MIN_LEN = 4           # 검색 후보로 쓸 최소 문자열 길이
PDF_GRAPHICS_LIMIT = 20000     # 페이지당 벡터 그래픽 상한 (초과 시 그래픽 분석 스킵 — 병적 페이지 방어)
PDF_PROGRESS_EVERY = 10        # N페이지마다 진행 로그 + gc
PDF_TEXT_LAYER_MIN_CHARS = max(
    int(
        os.environ.get(
            "PDF_TEXT_LAYER_MIN_CHARS",
            os.environ.get("PDF_OCR_TEXT_MIN_NONSPACE", "20"),
        )
    ),
    0,
)
# Backward-compatible name for deployments that still carry the old setting.
PDF_OCR_TEXT_MIN_NONSPACE = PDF_TEXT_LAYER_MIN_CHARS
PDF_OCR_LANGUAGES = os.environ.get("PDF_OCR_LANGUAGES", "eng+kor").strip() or "eng+kor"
PDF_MARKDOWN_CHUNK_SIZE = max(int(os.environ.get("PDF_MARKDOWN_CHUNK_SIZE", "2400")), 200)
PDF_MARKDOWN_CHUNK_OVERLAP = max(int(os.environ.get("PDF_MARKDOWN_CHUNK_OVERLAP", "100")), 0)
PREPROC_LARGE_PDF_MIN_PAGES = max(int(os.environ.get("PREPROC_LARGE_PDF_MIN_PAGES", "50")), 1)
PREPROC_LARGE_PDF_MIN_BYTES = max(int(os.environ.get("PREPROC_LARGE_PDF_MIN_BYTES", str(5 * 1024 * 1024))), 1)
PREPROC_LARGE_PDF_LOCK_PATH = os.environ.get(
    "PREPROC_LARGE_PDF_LOCK_PATH", "/tmp/preprocessor-91-large-pdf.lock"
)


# ============================================================================
# Models & Exceptions
# ============================================================================

class GenOSVectorMeta(BaseModel):
    class Config:
        extra = "allow"

    text: Optional[str] = None
    n_char: Optional[int] = None
    n_word: Optional[int] = None
    n_line: Optional[int] = None
    i_page: Optional[int] = None
    e_page: Optional[int] = None
    i_chunk_on_page: Optional[int] = None
    n_chunk_of_page: Optional[int] = None
    i_chunk_on_doc: Optional[int] = None
    n_chunk_of_doc: Optional[int] = None
    n_page: Optional[int] = None
    reg_date: Optional[str] = None
    chunk_bboxes: Optional[str] = None
    media_files: Optional[str] = None


class GenosServiceException(Exception):
    def __init__(self, error_code: str, error_msg: Optional[str] = None, msg_params: Optional[dict] = None):
        self.code = 1
        self.error_code = error_code
        self.error_msg = error_msg or "GenOS Service Exception"
        self.msg_params = msg_params or {}

    def __repr__(self):
        return f"{type(self).__name__}(code={self.code!r}, errMsg={self.error_msg!r})"


class PdfPageExtractionError(RuntimeError):
    """Raised when any PDF page cannot be extracted completely."""


@dataclass(frozen=True)
class PdfPageResult:
    page_index: int
    markdown: str
    used_ocr: bool
    elapsed_s: float
    error: Optional[str] = None


# ============================================================================
# Utilities
# ============================================================================

def _effective_cpu_count(
    cpu_max_path: Path = Path("/sys/fs/cgroup/cpu.max"),
    cpu_quota_path: Path = Path("/sys/fs/cgroup/cpu/cpu.cfs_quota_us"),
    cpu_period_path: Path = Path("/sys/fs/cgroup/cpu/cpu.cfs_period_us"),
) -> int:
    """Return the usable integer CPU count, respecting cgroup v2/v1 quotas."""
    host_count = max(os.cpu_count() or 1, 1)
    try:
        quota_text, period_text = cpu_max_path.read_text(encoding="ascii").split()[:2]
        if quota_text != "max":
            return max(1, min(host_count, int(quota_text) // int(period_text)))
        return host_count
    except (OSError, ValueError, IndexError, ZeroDivisionError):
        pass

    try:
        quota = int(cpu_quota_path.read_text(encoding="ascii").strip())
        period = int(cpu_period_path.read_text(encoding="ascii").strip())
        if quota > 0:
            return max(1, min(host_count, quota // period))
    except (OSError, ValueError, ZeroDivisionError):
        pass
    return host_count


def _page_worker_count(cpu_count: Optional[int] = None) -> int:
    available = max(cpu_count or _effective_cpu_count(), 1)
    configured = os.environ.get("PREPROC_PAGE_WORKERS")
    if configured is None:
        return available
    try:
        return max(1, min(int(configured), available))
    except ValueError:
        _log.warning("[pdf-md] PREPROC_PAGE_WORKERS=%r invalid; using %d", configured, available)
        return available


def _is_large_pdf(page_count: int, file_size: int) -> bool:
    return page_count >= PREPROC_LARGE_PDF_MIN_PAGES or file_size >= PREPROC_LARGE_PDF_MIN_BYTES


class LargePdfGate:
    """Cross-process capacity-one gate shared by all gunicorn workers in a pod."""

    def __init__(self, lock_path: Path | str = PREPROC_LARGE_PDF_LOCK_PATH):
        self.lock_path = Path(lock_path)
        self._lock_file = None

    def __enter__(self) -> "LargePdfGate":
        self.lock_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock_file = self.lock_path.open("a+")
        started = time.monotonic()
        _log.info("[pdf-md] large PDF gate waiting | lock=%s", self.lock_path)
        fcntl.flock(self._lock_file.fileno(), fcntl.LOCK_EX)
        _log.info("[pdf-md] large PDF gate acquired | wait=%.2fs", time.monotonic() - started)
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        if self._lock_file is None:
            return
        try:
            fcntl.flock(self._lock_file.fileno(), fcntl.LOCK_UN)
        finally:
            self._lock_file.close()
            self._lock_file = None
            _log.info("[pdf-md] large PDF gate released")


_PAGE_WORKER_DOC = None
_PAGE_WORKER_PYMUPDF4LLM = None
_PAGE_WORKER_HDR_INFO = None


def _set_page_worker_thread_limits(thread_count: Optional[int] = None) -> int:
    if thread_count is None:
        try:
            thread_count = int(os.environ.get("PREPROC_PAGE_WORKER_THREADS", "1"))
        except ValueError:
            thread_count = 1
    bounded = max(int(thread_count), 1)
    for variable in (
        "OMP_NUM_THREADS",
        "OMP_THREAD_LIMIT",
        "OPENBLAS_NUM_THREADS",
        "MKL_NUM_THREADS",
        "NUMEXPR_NUM_THREADS",
        "VECLIB_MAXIMUM_THREADS",
        "BLIS_NUM_THREADS",
    ):
        os.environ[variable] = str(bounded)
    return bounded


def _page_worker_init(file_path: str) -> None:
    global _PAGE_WORKER_DOC, _PAGE_WORKER_PYMUPDF4LLM, _PAGE_WORKER_HDR_INFO
    _set_page_worker_thread_limits()
    import fitz
    import pymupdf4llm

    _PAGE_WORKER_DOC = fitz.open(file_path)
    _PAGE_WORKER_PYMUPDF4LLM = pymupdf4llm
    try:
        _PAGE_WORKER_HDR_INFO = pymupdf4llm.IdentifyHeaders(_PAGE_WORKER_DOC)
    except Exception:
        _PAGE_WORKER_HDR_INFO = None


def _should_ocr_page(native_text: str) -> bool:
    return len("".join(native_text.split())) < PDF_TEXT_LAYER_MIN_CHARS


def _page_markdown_items(pymupdf4llm, doc, pno: int, hdr_info=None) -> Tuple[list, bool]:
    native_text = doc[pno].get_text("text") or ""
    use_ocr = _should_ocr_page(native_text)
    kwargs_try = [
        {"page_chunks": True, "graphics_limit": PDF_GRAPHICS_LIMIT,
         "show_progress": False, "use_ocr": use_ocr, "ocr_language": PDF_OCR_LANGUAGES},
        {"page_chunks": True, "graphics_limit": PDF_GRAPHICS_LIMIT,
         "use_ocr": use_ocr, "ocr_language": PDF_OCR_LANGUAGES},
        {"page_chunks": True, "use_ocr": use_ocr, "ocr_language": PDF_OCR_LANGUAGES},
    ]
    for kwargs in kwargs_try:
        if hdr_info is not None:
            kwargs = {**kwargs, "hdr_info": hdr_info}
        try:
            return pymupdf4llm.to_markdown(doc, pages=[pno], **kwargs), use_ocr
        except TypeError:
            continue
    return pymupdf4llm.to_markdown(
        doc,
        pages=[pno],
        page_chunks=True,
        use_ocr=use_ocr,
        ocr_language=PDF_OCR_LANGUAGES,
    ), use_ocr


def _extract_pdf_page(page_index: int) -> PdfPageResult:
    started = time.perf_counter()
    try:
        if _PAGE_WORKER_DOC is None or _PAGE_WORKER_PYMUPDF4LLM is None:
            raise RuntimeError("page worker not initialized")
        items, used_ocr = _page_markdown_items(
            _PAGE_WORKER_PYMUPDF4LLM,
            _PAGE_WORKER_DOC,
            page_index,
            _PAGE_WORKER_HDR_INFO,
        )
        item = items[0] if items else {}
        markdown = (item.get("text") or "") if isinstance(item, dict) else str(item)
        return PdfPageResult(page_index, markdown, used_ocr, time.perf_counter() - started)
    except Exception as exc:
        return PdfPageResult(
            page_index,
            "",
            False,
            time.perf_counter() - started,
            error=f"{type(exc).__name__}: {exc}",
        )


def _ordered_page_results(results: List[PdfPageResult], expected_pages: int) -> List[PdfPageResult]:
    ordered = sorted(results, key=lambda item: item.page_index)
    actual_pages = [item.page_index for item in ordered]
    if actual_pages != list(range(expected_pages)):
        raise PdfPageExtractionError(
            f"PDF page result mismatch: expected {expected_pages}, got indexes {actual_pages}"
        )
    failures = [item for item in ordered if item.error]
    if failures:
        first = failures[0]
        raise PdfPageExtractionError(
            f"PDF page {first.page_index + 1} extraction failed: {first.error}"
        )
    return ordered


def _extract_pages_parallel(
    file_path: str,
    page_count: int,
    worker_count: int,
    executor_factory: Type[ProcessPoolExecutor] = ProcessPoolExecutor,
) -> List[PdfPageResult]:
    context = multiprocessing.get_context("spawn")
    with executor_factory(
        max_workers=worker_count,
        mp_context=context,
        initializer=_page_worker_init,
        initargs=(file_path,),
    ) as executor:
        results = list(executor.map(_extract_pdf_page, range(page_count), chunksize=1))
    return _ordered_page_results(results, expected_pages=page_count)

async def assert_cancelled(request: Request):
    if await request.is_disconnected():
        raise GenosServiceException(1, "Cancelled")


def install_packages(packages):
    """
    누락 패키지 설치.
    1) GENOS_PKG_DIR 환경변수(사전 배치한 wheel 디렉토리)가 있으면 오프라인 설치 우선
       — egress 차단 환경에서 온라인 pip의 네트워크 타임아웃 대기를 피하기 위함
    2) 온라인 pip (PIP_INDEX_URL 환경변수로 사내 미러 지정 가능)
    실패 시 stderr를 로그에 남기고 예외 발생 (호출측 폴백 경로가 처리).
    """
    for package in packages:
        try:
            __import__(package)
            continue
        except ImportError:
            pass

        _log.warning(f"{package} 패키지가 없습니다. 설치를 시도합니다.")

        cmds: List[List[str]] = []
        wheel_dir = os.environ.get("GENOS_PKG_DIR", "")
        if wheel_dir and os.path.isdir(wheel_dir):
            cmds.append([sys.executable, "-m", "pip", "install",
                         "--no-index", "--find-links", wheel_dir, package])
        cmds.append([sys.executable, "-m", "pip", "install", package])

        last_err = ""
        for cmd in cmds:
            proc = subprocess.run(cmd, capture_output=True, text=True)
            if proc.returncode == 0:
                _log.info(f"[install] 성공: {package}")
                break
            last_err = (proc.stderr or proc.stdout or "").strip()[-800:]
            _log.warning(f"[install] 실패: {' '.join(cmd)}\n{last_err}")
        else:
            raise GenosServiceException(1, f"Package install failed: {package} | {last_err}")


def setup_logging(level_num: int):
    level_map = {5: "DEBUG", 4: "INFO", 3: "WARNING", 2: "ERROR", 1: "CRITICAL", 0: "NOLOG"}
    level_name = level_map.get(level_num, "INFO")
    if level_name == "NOLOG":
        logging.disable(logging.CRITICAL)
        return
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler()],
    )
    logging.getLogger().setLevel(level)


def convert_to_pdf(file_path: str) -> Optional[str]:
    """LibreOffice로 PDF 변환. 실패 시 None 반환."""
    try:
        in_path = Path(file_path).resolve()
        out_dir = in_path.parent
        pdf_path = in_path.with_suffix(".pdf")

        env = os.environ.copy()
        env.setdefault("LANG", "C.UTF-8")
        env.setdefault("LC_ALL", "C.UTF-8")

        ext = in_path.suffix.lower()
        convert_arg_map = {
            ".ppt": "pdf:impress_pdf_Export",
            ".pptx": "pdf:impress_pdf_Export",
        }
        convert_arg = convert_arg_map.get(ext, "pdf")

        # 비ASCII 파일명 처리
        try:
            in_path.name.encode("ascii")
            candidates = [in_path]
            tmp_dir = None
        except UnicodeEncodeError:
            tmp_dir = Path(tempfile.mkdtemp())
            ascii_name = (
                unicodedata.normalize("NFKD", in_path.stem)
                .encode("ascii", "ignore")
                .decode("ascii")
                or "file"
            )
            ascii_copy = tmp_dir / f"{ascii_name}{in_path.suffix}"
            shutil.copy2(in_path, ascii_copy)
            candidates = [ascii_copy, in_path]

        for cand in candidates:
            cmd = ["soffice", "--headless", "--convert-to", convert_arg, "--outdir", str(out_dir), str(cand)]
            proc = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=300)
            if proc.returncode == 0 and pdf_path.exists():
                if tmp_dir:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                return str(pdf_path)
            _log.debug(f"[convert_to_pdf] stderr: {proc.stderr.strip()}")

        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        return None
    except Exception as e:
        _log.warning(f"PDF conversion failed: {e}")
        return None


def _get_pdf_path(file_path: str) -> str:
    base, _ = os.path.splitext(file_path)
    return base + ".pdf"


def _get_real_file_type(file_path: str) -> str:
    try:
        with open(file_path, "rb") as f:
            header = f.read(8)
        if header.startswith(b"%PDF-"):
            return "pdf"
        if header.startswith(b"\x89PNG"):
            return "png"
        if header.startswith(b"\xff\xd8\xff"):
            return "jpg"
    except Exception:
        pass
    return os.path.splitext(file_path)[-1].lower().lstrip(".")


# ============================================================================
# Excel 파싱 계층 (openpyxl 기반)
# ============================================================================

@dataclass
class ParsedSheet:
    """시트 하나의 파싱 결과."""
    name: str
    index: int                              # 1-based 시트 순번 (가시 시트 기준)
    kind: str = "empty"                     # "table" | "report" | "empty"
    title_lines: List[str] = field(default_factory=list)   # 헤더 위 제목/설명 행
    headers: List[str] = field(default_factory=list)        # 평탄화된 컬럼명
    rows: List[List[str]] = field(default_factory=list)     # 데이터 행 (table) 또는 레이아웃 행 (report)
    images: List[str] = field(default_factory=list)         # 임베디드 이미지 식별자
    n_merged: int = 0


def _format_cell_value(value: Any) -> str:
    """셀 값을 검색 친화적 문자열로 정규화."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # 시각이 자정이면 날짜만 표기
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        # 12.0 → "12", 불필요한 지수 표기 방지
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value).strip()


def _escape_md(text: str) -> str:
    """Markdown 테이블 셀 이스케이프."""
    return text.replace("|", "\\|").replace("\n", " ").strip()


class ExcelWorkbookParser:
    """
    openpyxl 기반 워크북 파서.

    - 숨김 시트 제외
    - 병합 셀 fill-down (좌상단 값을 범위 전체에 전파)
    - 하이퍼링크 → Markdown 링크
    - 헤더 감지 / 다중 헤더 평탄화 / 시트 분류
    """

    def __init__(self, file_path: str):
        install_packages(["openpyxl"])
        from openpyxl import load_workbook

        self.file_path = file_path
        _wb = load_workbook(file_path, data_only=True)
        self.sheets: List[ParsedSheet] = self._parse_all(_wb)
        _wb.close()
        del _wb
        gc.collect()

    # ------------------------------------------------------------------

    def _parse_all(self, wb) -> List[ParsedSheet]:
        parsed: List[ParsedSheet] = []
        visible_idx = 0
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                _log.info(f"[xlsx] 숨김 시트 제외: {ws.title}")
                continue
            visible_idx += 1
            parsed.append(self._parse_sheet(ws, visible_idx))
        return parsed

    # ------------------------------------------------------------------

    def _parse_sheet(self, ws, index: int) -> ParsedSheet:
        sheet = ParsedSheet(name=ws.title, index=index)

        # --- 임베디드 이미지 수집 (VLM 없이 식별자만 메타데이터로 보존) ---
        try:
            for i, img in enumerate(getattr(ws, "_images", []) or []):
                anchor = getattr(getattr(img, "anchor", None), "_from", None)
                pos = f"r{anchor.row + 1}c{anchor.col + 1}" if anchor else f"idx{i}"
                sheet.images.append(f"{ws.title}:{pos}")
        except Exception as e:
            _log.debug(f"[xlsx] 이미지 수집 실패({ws.title}): {e}")

        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return sheet

        # --- 병합 셀 맵: (row, col) → 좌상단 값 ---
        merged_fill: Dict[Tuple[int, int], Any] = {}
        merged_ranges = list(ws.merged_cells.ranges)
        sheet.n_merged = len(merged_ranges)
        for rng in merged_ranges:
            top_left = ws.cell(rng.min_row, rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_fill[(r, c)] = top_left

        # --- 그리드 구성 (하이퍼링크 → Markdown 링크) ---
        grid: List[List[str]] = []
        for r, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col), start=1):
            row_vals: List[str] = []
            for c, cell in enumerate(row_cells, start=1):
                raw = merged_fill.get((r, c), cell.value)
                text = _format_cell_value(raw)
                link = getattr(cell, "hyperlink", None)
                if link is not None and getattr(link, "target", None):
                    label = text or link.target
                    text = f"[{label}]({link.target})"
                row_vals.append(text)
            grid.append(row_vals)

        # --- 완전히 빈 행/열 트리밍 ---
        grid = self._trim_grid(grid)
        if not grid:
            return sheet

        # --- 분류: 헤더 감지 우선, 병합 비율은 '데이터 영역' 기준으로 판단 ---
        # (헤더/제목 영역의 병합은 정상 테이블에서도 흔하므로 분류 근거에서 제외)
        header_idx, headers = self._detect_header(grid)

        if header_idx is None:
            sheet.kind = "report" if any(v for row in grid for v in row) else "empty"
            sheet.rows = grid
            return sheet

        # 2단 헤더 여부를 먼저 판정 → 데이터 시작 행(워크시트 1-based)을 정확히 계산
        # (grid는 leading 행을 제거하지 않으므로 grid[i] ↔ 워크시트 행 i+1)
        two_tier = self._is_second_header_row(grid, header_idx)
        data_first_row = (header_idx + 1) + (2 if two_tier else 1)

        data_merged = sum(1 for rng in merged_ranges if rng.min_row >= data_first_row)
        n_data_rows = max(len(grid) - (data_first_row - 1), 1)
        data_merged_ratio = data_merged / n_data_rows

        data_rows_ = [r for r in grid[data_first_row - 1:] if any(v for v in r)]
        data_fill = (
            sum(1 for r in data_rows_ for v in r if v)
            / max(sum(len(r) for r in data_rows_), 1)
        )

        if data_merged_ratio > REPORT_MERGED_RATIO or data_fill < 0.25 or len(data_rows_) < 1:
            # 보고서형: 행 단위 레이아웃 텍스트로 보존
            sheet.kind = "report"
            sheet.rows = grid
            return sheet

        sheet.kind = "table"

        def _dedup_consecutive(values: List[str]) -> List[str]:
            out: List[str] = []
            for v in values:
                if v and (not out or out[-1] != v):
                    out.append(v)
            return out

        # 병합 fill-down으로 반복된 제목 값은 1회만 보존
        sheet.title_lines = [
            " ".join(_dedup_consecutive([v for v in row if v]))
            for row in grid[:header_idx]
            if any(v for v in row)
        ][:MAX_TITLE_LINES]

        if two_tier:
            second = grid[header_idx + 1]
            headers = [
                f"{h}_{s}".strip("_") if s and s != h else (h or s)
                for h, s in zip(headers, second)
            ]
        sheet.headers = self._dedupe_headers(headers)
        data_start = data_first_row - 1  # grid 0-based

        sheet.rows = [row for row in grid[data_start:] if any(v for v in row)]
        return sheet

    # ------------------------------------------------------------------

    @staticmethod
    def _trim_grid(grid: List[List[str]]) -> List[List[str]]:
        # 빈 행 제거는 하지 않고(레이아웃 보존), 끝쪽 완전 빈 행/열만 제거
        while grid and not any(v for v in grid[-1]):
            grid.pop()
        if not grid:
            return grid
        n_col = len(grid[0])
        last_used = 0
        for row in grid:
            for c in range(n_col - 1, -1, -1):
                if row[c]:
                    last_used = max(last_used, c + 1)
                    break
        return [row[:last_used] for row in grid]

    @staticmethod
    def _detect_header(grid: List[List[str]]) -> Tuple[Optional[int], List[str]]:
        """상단 HEADER_SCAN_ROWS 내에서 헤더 행 탐색."""
        n_col = max((len(r) for r in grid), default=0)
        for i, row in enumerate(grid[:HEADER_SCAN_ROWS]):
            filled = [v for v in row if v]
            if len(filled) < 2:
                continue
            # 병합된 제목 행(모든 셀 동일 값)은 헤더가 아님
            if len(set(filled)) < 2:
                continue
            fill_ratio = len(filled) / max(n_col, 1)
            # 숫자만으로 이뤄진 행은 헤더가 아님 (데이터 행일 가능성)
            numeric = sum(1 for v in filled if v.replace(".", "", 1).replace("-", "", 1).isdigit())
            if fill_ratio >= HEADER_MIN_FILL_RATIO and numeric / len(filled) < 0.5:
                return i, [v if v else f"col{j+1}" for j, v in enumerate(row)]
        return None, []

    @staticmethod
    def _is_second_header_row(grid: List[List[str]], header_idx: int) -> bool:
        if header_idx + 2 >= len(grid):
            return False
        second, third = grid[header_idx + 1], grid[header_idx + 2]
        sec_filled = [v for v in second if v]
        thr_filled = [v for v in third if v]
        if not sec_filled or not thr_filled:
            return False
        sec_numeric = sum(1 for v in sec_filled if v.replace(".", "", 1).replace("-", "", 1).isdigit()) / len(sec_filled)
        thr_numeric = sum(1 for v in thr_filled if v.replace(".", "", 1).replace("-", "", 1).isdigit()) / len(thr_filled)
        # 2행은 문자열 위주, 3행은 숫자 위주면 2단 헤더로 판단
        return sec_numeric < 0.3 and thr_numeric >= 0.5

    @staticmethod
    def _dedupe_headers(headers: List[str]) -> List[str]:
        seen: Dict[str, int] = {}
        out = []
        for h in headers:
            h = h or "col"
            if h in seen:
                seen[h] += 1
                out.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                out.append(h)
        return out


# ============================================================================
# TabularLoader (xlsx) — 구조 보존 Markdown 직렬화 + 행 경계 청킹
# ============================================================================

class TabularLoader:
    """
    xlsx → 시트 분류 → Markdown 직렬화 → 행 경계 청킹.

    모든 청크는 자기완결적(self-contained):
      [DA] 문서/시트/표 제목/컬럼 정보를 담은 컨텍스트 헤더 + Markdown 헤더 행이
      매 청크에 반복되므로, 청크 단독으로도 어떤 표의 어느 컬럼 값인지 식별 가능.
    """

    def __init__(self, file_path: str, chunk_size: int = 1000, chunk_overlap: int = 100):
        self.file_path = file_path
        self.file_name = Path(file_path).name
        self.chunk_size = max(int(chunk_size), 200)
        self.chunk_overlap = max(int(chunk_overlap), 0)
        self.parser = ExcelWorkbookParser(file_path)

    # ------------------------------------------------------------------
    # 직렬화
    # ------------------------------------------------------------------

    def _context_header(self, sheet: ParsedSheet) -> str:
        parts = [f"[DA] 문서: {self.file_name} | 시트: {sheet.name}"]
        if sheet.title_lines:
            parts.append(f"표 제목: {' / '.join(sheet.title_lines)}")
        if sheet.headers:
            parts.append(f"컬럼: {', '.join(sheet.headers)}")
        if sheet.images:
            parts.append(f"(이미지 {len(sheet.images)}개 포함 — 텍스트 미추출)")
        return "\n".join(parts)

    @staticmethod
    def _md_header_block(headers: List[str]) -> str:
        head = "| " + " | ".join(_escape_md(h) for h in headers) + " |"
        sep = "| " + " | ".join("---" for _ in headers) + " |"
        return head + "\n" + sep

    @staticmethod
    def _md_row(row: List[str], n_col: int) -> str:
        padded = list(row) + [""] * (n_col - len(row))
        return "| " + " | ".join(_escape_md(v) for v in padded) + " |"

    # ------------------------------------------------------------------
    # 청킹
    # ------------------------------------------------------------------

    def _chunk_table_sheet(self, sheet: ParsedSheet) -> List[str]:
        """행 경계 기준 청킹. 컨텍스트+헤더는 매 청크 반복, 행 단위 overlap."""
        ctx = self._context_header(sheet)
        md_head = self._md_header_block(sheet.headers)
        base = ctx + "\n\n" + md_head + "\n"
        base_len = len(base)
        n_col = len(sheet.headers)

        row_lines = [self._md_row(r, n_col) for r in sheet.rows]

        # 전체가 한 청크에 들어가면 그대로
        total_len = base_len + sum(len(l) + 1 for l in row_lines)
        if total_len <= self.chunk_size:
            return [base + "\n".join(row_lines)] if row_lines else [ctx]

        chunks: List[str] = []
        buf: List[str] = []
        buf_len = 0
        budget = max(self.chunk_size - base_len, 200)  # 행 텍스트에 쓸 수 있는 예산

        def flush():
            nonlocal buf, buf_len
            if buf:
                chunks.append(base + "\n".join(buf))
                # 행 단위 overlap: 마지막 행들 중 overlap 예산 내의 것을 다음 청크로 이월
                carry: List[str] = []
                carry_len = 0
                for line in reversed(buf):
                    if carry_len + len(line) + 1 > self.chunk_overlap:
                        break
                    carry.insert(0, line)
                    carry_len += len(line) + 1
                buf = carry
                buf_len = carry_len

        for line in row_lines:
            # 단일 행이 예산을 초과하는 극단적 케이스는 강제 분할
            if len(line) + 1 > budget and not buf:
                chunks.append(base + line[:budget])
                continue
            if buf_len + len(line) + 1 > budget:
                flush()
            buf.append(line)
            buf_len += len(line) + 1

        if buf and (not chunks or chunks[-1] != base + "\n".join(buf)):
            chunks.append(base + "\n".join(buf))
        return chunks

    def _chunk_report_sheet(self, sheet: ParsedSheet) -> List[str]:
        """보고서형: 행 레이아웃을 'A | B | C' 라인으로 보존, 라인 경계 청킹."""
        ctx = self._context_header(sheet)
        lines: List[str] = []
        prev_blank = False
        for row in sheet.rows:
            filled = [v for v in row if v]
            if not filled:
                if not prev_blank:
                    lines.append("")  # 문단 경계 유지 (연속 빈 행은 1개로)
                prev_blank = True
                continue
            prev_blank = False
            # 중복 값(병합 fill-down 결과) 연속 반복 축약
            dedup: List[str] = []
            for v in filled:
                if not dedup or dedup[-1] != v:
                    dedup.append(v)
            lines.append(" | ".join(dedup))

        base = ctx + "\n\n"
        budget = max(self.chunk_size - len(base), 200)

        chunks: List[str] = []
        buf: List[str] = []
        buf_len = 0
        for line in lines:
            if buf_len + len(line) + 1 > budget and buf:
                chunks.append(base + "\n".join(buf))
                buf, buf_len = [], 0
            buf.append(line)
            buf_len += len(line) + 1
        if buf and any(l.strip() for l in buf):
            chunks.append(base + "\n".join(buf))
        return chunks or [ctx]

    # ------------------------------------------------------------------
    # GenOSVectorMeta 변환
    # ------------------------------------------------------------------

    def return_vectormeta_format(self) -> Optional[List[GenOSVectorMeta]]:
        now_iso = datetime.now().isoformat(timespec="seconds") + "Z"
        vectors: List[GenOSVectorMeta] = []
        chunk_on_doc = 0

        for sheet in self.parser.sheets:
            if sheet.kind == "empty" and not sheet.images:
                continue

            if sheet.kind == "table":
                chunk_texts = self._chunk_table_sheet(sheet)
            elif sheet.kind == "report":
                chunk_texts = self._chunk_report_sheet(sheet)
            else:  # empty지만 이미지는 있는 시트
                chunk_texts = [self._context_header(sheet)]

            # GenOS 출처 뷰어가 media_files를 실제 파일 경로로 해석하므로
            # 경로가 아닌 값(시트:셀 식별자 등)을 넣으면 뷰어에서
            # 'Invalid file_path format' 오류 발생 → 플레이스홀더 유지.
            # 이미지 존재 여부는 청크 본문 컨텍스트 헤더에 텍스트로 포함됨.
            media = "."
            n_chunk_of_page = len(chunk_texts)

            for i_on_page, text in enumerate(chunk_texts):
                if not text.strip():
                    continue
                vectors.append(
                    GenOSVectorMeta.model_validate({
                        "text": text,
                        "n_char": len(text),
                        "n_word": len(text.split()),
                        "n_line": len(text.splitlines()),
                        # 뷰어가 로드할 PDF의 페이지와 시트 순번은 1:1 대응이
                        # 아니므로 페이지는 안전하게 1로 고정하고,
                        # 시트 출처는 별도 필드(sheet_name/sheet_index)와
                        # 청크 본문 컨텍스트 헤더로 보존한다.
                        "i_page": 1,
                        "e_page": 1,
                        "n_page": 1,
                        "i_chunk_on_page": i_on_page,
                        "n_chunk_of_page": n_chunk_of_page,
                        "i_chunk_on_doc": chunk_on_doc,
                        "n_chunk_of_doc": 0,  # 아래에서 일괄 갱신
                        "reg_date": now_iso,
                        "chunk_bboxes": ".",
                        "media_files": media,
                        "sheet_name": sheet.name,
                        "sheet_index": sheet.index,
                    })
                )
                chunk_on_doc += 1

        total = len(vectors)
        for v in vectors:
            v.n_chunk_of_doc = total
        return vectors if vectors else None


# ============================================================================
# PdfMarkdownLoader (pdf / 변환된 ppt·pptx) — 1단계 고도화
# ============================================================================

@dataclass
class PdfChunk:
    """PDF 청크 하나 (컨텍스트 헤더 포함 최종 텍스트 + 출처 좌표)."""
    text: str
    page: int                                            # 1-based
    bboxes: List[Dict[str, float]] = field(default_factory=list)  # 정규화(0~1) l/t/r/b


class PdfMarkdownLoader:
    """
    PyMuPDF4LLM 기반 PDF 로더 (의존성: pymupdf4llm — 기존 fitz의 경량 확장).

    파이프라인:
      1. pymupdf4llm.to_markdown(page_chunks=True) → 페이지별 Markdown
         (폰트 크기 기반 헤딩 추론, 표 → Markdown 테이블, 리스트 보존)
      2. 헤딩(#) 경계로 섹션 블록 분리, 문서 전체 헤딩 스택으로 경로 추적
      3. 섹션 블록 단위 packing → chunk_size 초과 블록만 라인 경계 분할
         (표 분할 시 헤더 행 반복, 라인 단위 overlap)
      4. 청크마다 [DA] 컨텍스트 헤더(문서명 | 페이지 | 헤딩 경로) 프리픽스
      5. page.search_for로 청크 본문 라인의 실제 bbox 수집 → chunk_bboxes

    페이지 경계는 넘지 않는다(청크 ↔ 페이지 1:1 매핑 유지 → 뷰어 하이라이트 정확성 우선).
    """

    def __init__(self, file_path: str, chunk_size: int = 1000, chunk_overlap: int = 100,
                 display_name: Optional[str] = None):
        # 런타임 pip 설치 시도하지 않음 (distroless 이미지 — pip 없음).
        # pymupdf4llm은 실행 이미지에 사전 설치되어 있어야 하며(1.3.8.1-*-fv 이후 포함),
        # 미설치 시 ImportError → 호출측(_process_pdf_like)이 레거시 경로로 폴백.
        try:
            import pymupdf4llm  # noqa: F401
        except ImportError:
            _log.warning("[pdf-md] pymupdf4llm 미설치 — 실행 이미지에 사전 설치 필요 (레거시 폴백)")
            raise

        self.file_path = file_path
        self.file_name = display_name or Path(file_path).name
        self.chunk_size = max(int(chunk_size), 200)
        self.chunk_overlap = max(int(chunk_overlap), 0)

        # 검증용 (verify_pdf에서 사용)
        self.headings: List[Tuple[int, int, str]] = []   # (page, level, title)
        self.last_chunks: List[PdfChunk] = []
        self.stats: Dict[str, Any] = {}
        self.timing: Dict[str, float] = {}
        self._hdr_info = None  # IdentifyHeaders 결과 (문서당 1회 계산)

    # ------------------------------------------------------------------
    # Entry
    # ------------------------------------------------------------------

    def return_vectormeta_format(self) -> Optional[List[GenOSVectorMeta]]:
        import time
        import fitz
        import pymupdf4llm

        doc = fitz.open(self.file_path)
        large_pdf_gate: Optional[LargePdfGate] = None
        try:
            n_page = doc.page_count
            file_size = os.path.getsize(self.file_path)
            if _is_large_pdf(page_count=n_page, file_size=file_size):
                large_pdf_gate = LargePdfGate()
                large_pdf_gate.__enter__()
                _log.info(
                    "[pdf-md] large PDF serialized | pages=%d bytes=%d",
                    n_page,
                    file_size,
                )

            # 헤딩 폰트 분석은 문서당 1회만 수행 (페이지 단위 호출 시 중복 방지 +
            # 문서 전체 폰트 분포 기준의 일관된 헤딩 레벨 판정)
            try:
                self._hdr_info = pymupdf4llm.IdentifyHeaders(doc)
            except Exception as e:
                _log.debug(f"[pdf-md] IdentifyHeaders 실패(페이지별 판정으로 폴백): {e}")
                self._hdr_info = None

            heading_stack: List[Tuple[int, str]] = []  # (level, title) — 문서 전체에서 유지
            chunks: List[PdfChunk] = []
            pages_with_text = 0
            self.ocr_page_count = 0
            t_extract = t_bbox = t_chunk = 0.0
            worker_cpu_s = 0.0
            worker_count = _page_worker_count()

            extract_started = time.perf_counter()
            if worker_count > 1 and n_page > 1:
                _log.info(
                    "[pdf-md] page extraction pool | workers=%d pages=%d start_method=spawn",
                    worker_count,
                    n_page,
                )
                page_results = _extract_pages_parallel(
                    self.file_path,
                    page_count=n_page,
                    worker_count=worker_count,
                )
            else:
                serial_results: List[PdfPageResult] = []
                for pno in range(n_page):
                    page_started = time.perf_counter()
                    try:
                        items, used_ocr = _page_markdown_items(
                            pymupdf4llm, doc, pno, self._hdr_info
                        )
                        item = items[0] if items else {}
                        markdown = (
                            (item.get("text") or "")
                            if isinstance(item, dict)
                            else str(item)
                        )
                        serial_results.append(
                            PdfPageResult(
                                pno,
                                markdown,
                                used_ocr,
                                time.perf_counter() - page_started,
                            )
                        )
                    except Exception as exc:
                        serial_results.append(
                            PdfPageResult(
                                pno,
                                "",
                                False,
                                time.perf_counter() - page_started,
                                error=f"{type(exc).__name__}: {exc}",
                            )
                        )
                page_results = _ordered_page_results(serial_results, expected_pages=n_page)
            t_extract = time.perf_counter() - extract_started
            worker_cpu_s = sum(item.elapsed_s for item in page_results)
            self.ocr_page_count = sum(1 for item in page_results if item.used_ocr)

            # Markdown extraction is parallel; heading/chunk/bbox assembly stays ordered
            # so output contracts remain deterministic and page-local.
            for page_result in page_results:
                pno = page_result.page_index
                page_no = pno + 1
                md = page_result.markdown
                if not md.strip():
                    _log.debug(f"[pdf-md] p.{page_no}: 텍스트 없음 (스캔/이미지 페이지 가능)")
                    continue
                pages_with_text += 1

                t0 = time.perf_counter()
                blocks = self._split_blocks(md, heading_stack, page_no=page_no)
                packed = self._pack_blocks(blocks)
                t_chunk += time.perf_counter() - t0

                # bbox 검색용 TextPage는 페이지당 1회만 생성해 재사용
                # (search_for는 기본적으로 호출마다 페이지 텍스트를 재파싱함)
                t0 = time.perf_counter()
                page_obj = doc[pno]
                textpage = None
                try:
                    textpage = page_obj.get_textpage()
                except Exception:
                    pass
                for heading_path, body in packed:
                    ctx = self._context_header(page_no, heading_path)
                    bboxes = self._find_bboxes(page_obj, body, textpage=textpage)
                    chunks.append(PdfChunk(text=ctx + "\n\n" + body, page=page_no, bboxes=bboxes))
                del textpage
                t_bbox += time.perf_counter() - t0

                if page_no % PDF_PROGRESS_EVERY == 0:
                    _log.info(f"[pdf-md] 진행 {page_no}/{n_page} 페이지 | 청크 {len(chunks)}개"
                              f" | 추출벽시계 {t_extract:.1f}s bbox {t_bbox:.1f}s")
                    gc.collect()

            _log.info(f"[pdf-md] 파싱 완료 | {n_page}p → 청크 {len(chunks)}개"
                      f" | OCR {self.ocr_page_count}/{n_page}p"
                      f" | workers {worker_count}"
                      f" | 추출벽시계 {t_extract:.1f}s / worker합 {worker_cpu_s:.1f}s"
                      f" / bbox {t_bbox:.1f}s / 청킹 {t_chunk:.2f}s")

            if not chunks:
                # 텍스트 레이어가 전혀 없는 스캔본 → 파이프라인이 죽지 않도록 플레이스홀더 1개
                _log.warning(f"[pdf-md] 텍스트 미추출 (스캔 문서 추정): {self.file_name}")
                chunks = [PdfChunk(
                    text=(f"[DA] 문서: {self.file_name}\n\n"
                          "[텍스트 미추출] 텍스트 레이어가 없는 스캔 문서로 추정됩니다. (OCR 미지원 환경)"),
                    page=1,
                )]

            # --- 검증 통계 (verify_pdf에서 사용, 운영 경로 동작에는 영향 없음) ---
            self.timing = {
                "extract_s": round(t_extract, 2),
                "extract_worker_s": round(worker_cpu_s, 2),
                "bbox_s": round(t_bbox, 2),
                "chunking_s": round(t_chunk, 3),
                "page_workers": worker_count,
            }
            self.last_chunks = chunks
            sizes = [len(c.text) for c in chunks]
            with_section = sum(1 for c in chunks if "\n섹션: " in c.text.split("\n\n")[0])
            with_bbox = sum(1 for c in chunks if c.bboxes)
            level_counts: Dict[int, int] = defaultdict(int)
            for _, lv, _t in self.headings:
                level_counts[lv] += 1
            self.stats = {
                "file": self.file_name,
                "n_page": n_page,
                "pages_with_text": pages_with_text,
                "pages_without_text": n_page - pages_with_text,
                "n_heading": len(self.headings),
                "heading_levels": dict(sorted(level_counts.items())),
                "n_chunk": len(chunks),
                "chunk_size_min": min(sizes),
                "chunk_size_avg": round(sum(sizes) / len(sizes)),
                "chunk_size_max": max(sizes),
                "section_coverage": round(with_section / len(chunks) * 100, 1),
                "bbox_coverage": round(with_bbox / len(chunks) * 100, 1),
                "bbox_avg_per_chunk": round(sum(len(c.bboxes) for c in chunks) / len(chunks), 1),
            }

            return self._to_vectors(chunks, n_page)
        finally:
            if large_pdf_gate is not None:
                large_pdf_gate.__exit__(*sys.exc_info())
            doc.close()
            gc.collect()

    # ------------------------------------------------------------------
    # Markdown → 섹션 블록
    # ------------------------------------------------------------------

    def _page_markdown(self, pymupdf4llm, doc, pno: int) -> list:
        """
        페이지 1장 Markdown 추출.
        graphics_limit: 벡터 그래픽이 비정상적으로 많은 페이지(차트 수천 개 등)는
        표 감지 분석이 폭주해 시간/메모리를 잡아먹으므로 상한 초과 시 해당 페이지의
        그래픽 분석을 건너뛴다. 구버전 pymupdf4llm은 이 인자가 없으므로 폴백.
        """
        items, use_ocr = _page_markdown_items(pymupdf4llm, doc, pno, self._hdr_info)
        if use_ocr:
            self.ocr_page_count = getattr(self, "ocr_page_count", 0) + 1
        return items

    _HEADING_RE = re.compile(r"^(#{1,6})\s+(.+)$")

    def _split_blocks(self, md: str, heading_stack: List[Tuple[int, str]],
                      page_no: int = 0) -> List[Tuple[str, List[str]]]:
        """
        페이지 Markdown을 헤딩 경계로 블록화.
        반환: [(heading_path, lines)], heading_stack은 페이지를 넘어 갱신됨.
        """
        blocks: List[Tuple[str, List[str]]] = []
        cur_lines: List[str] = []
        cur_path = " > ".join(t for _, t in heading_stack)

        def flush():
            nonlocal cur_lines
            if cur_lines and any(l.strip() for l in cur_lines):
                blocks.append((cur_path, cur_lines))
            cur_lines = []

        for line in md.splitlines():
            m = self._HEADING_RE.match(line.strip())
            if m:
                flush()
                level = len(m.group(1))
                title = m.group(2).strip().strip("*_ ")
                self.headings.append((page_no, level, title))  # 검증 통계용
                # 같은/하위 레벨 헤딩 pop 후 push
                heading_stack[:] = [(lv, t) for lv, t in heading_stack if lv < level]
                heading_stack.append((level, title))
                cur_path = " > ".join(t for _, t in heading_stack)
                cur_lines = [line]  # 헤딩 라인 자체도 본문에 포함 (검색성)
            else:
                cur_lines.append(line)
        flush()
        return blocks

    # ------------------------------------------------------------------
    # 블록 packing / 분할
    # ------------------------------------------------------------------

    def _pack_blocks(self, blocks: List[Tuple[str, List[str]]]) -> List[Tuple[str, str]]:
        """섹션 블록들을 예산 내에서 병합. 초과 블록만 라인 경계 분할."""
        budget = max(self.chunk_size - PDF_CTX_RESERVE, 200)
        results: List[Tuple[str, str]] = []
        buf: List[str] = []
        buf_path = ""
        buf_len = 0

        def flush():
            nonlocal buf, buf_len
            if buf and any(l.strip() for l in buf):
                results.append((buf_path, "\n".join(buf).strip("\n")))
            buf, buf_len = [], 0

        for path, lines in blocks:
            block_len = sum(len(l) + 1 for l in lines)

            if block_len > budget:
                flush()
                for part in self._split_long_block(lines, budget):
                    results.append((path, "\n".join(part).strip("\n")))
                continue

            if buf_len + block_len > budget:
                flush()
            if not buf:
                buf_path = path
            buf.extend(lines)
            buf_len += block_len
        flush()
        return results

    _TABLE_SEP_RE = re.compile(r"^\s*\|[\s:\-|]+\|\s*$")

    def _split_long_block(self, lines: List[str], budget: int) -> List[List[str]]:
        """예산 초과 블록을 라인 경계 분할. 표 헤더 행 반복 + 라인 overlap."""
        parts: List[List[str]] = []
        buf: List[str] = []
        buf_len = 0
        table_header: List[str] = []
        prev_line = ""

        for line in lines:
            # 표 헤더(컬럼 행 + 구분선) 추적
            if self._TABLE_SEP_RE.match(line) and prev_line.strip().startswith("|"):
                table_header = [prev_line, line]

            # 극단: 한 줄이 예산 초과 → 예산 단위로 강제 분할 (overlap 적용)
            if len(line) + 1 > budget:
                head: Optional[List[str]] = None
                if buf and any(l.strip() for l in buf):
                    if buf_len < budget // 3:
                        head = buf  # 짧은 선행 라인(헤딩 등)은 첫 조각에 병합
                    else:
                        parts.append(buf)
                buf, buf_len = [], 0

                head_len = sum(len(l) + 1 for l in head) if head else 0
                i, first = 0, True
                while i < len(line):
                    cap = max(budget - (head_len if first else 0), 100)
                    piece = line[i:i + cap]
                    parts.append((head + [piece]) if (first and head) else [piece])
                    if i + cap >= len(line):
                        break
                    i += max(cap - self.chunk_overlap, cap // 2, 100)
                    first = False
                prev_line = line
                continue

            if buf_len + len(line) + 1 > budget and buf:
                parts.append(buf)
                # 라인 단위 overlap 이월
                carry: List[str] = []
                carry_len = 0
                for l in reversed(buf):
                    if carry_len + len(l) + 1 > self.chunk_overlap:
                        break
                    carry.insert(0, l)
                    carry_len += len(l) + 1
                buf = list(carry)
                # 표 중간에서 잘렸으면 다음 청크에 표 헤더 반복
                if line.strip().startswith("|") and table_header and table_header[0] not in buf:
                    buf = table_header + buf
                buf_len = sum(len(l) + 1 for l in buf)

            buf.append(line)
            buf_len += len(line) + 1
            prev_line = line

        if buf and any(l.strip() for l in buf):
            parts.append(buf)
        return parts

    # ------------------------------------------------------------------
    # 컨텍스트 헤더 / bbox
    # ------------------------------------------------------------------

    def _context_header(self, page_no: int, heading_path: str) -> str:
        parts = [f"[DA] 문서: {self.file_name} | p.{page_no}"]
        if heading_path:
            parts.append(f"섹션: {heading_path}")
        return "\n".join(parts)

    @classmethod
    def _searchable_lines(cls, body: str) -> List[str]:
        """청크 본문에서 search_for에 쓸 plain 문자열 후보 추출 (md 문법 제거)."""
        out: List[str] = []
        for line in body.splitlines():
            s = line.strip()
            if not s or cls._TABLE_SEP_RE.match(s):
                continue
            if s.startswith("|"):
                # 표 행은 셀 단위로 검색 (렌더링 좌표와 md 행 문자열이 다르므로)
                for c in s.strip("|").split("|"):
                    c = c.strip().strip("*_` ")
                    if len(c) >= PDF_BBOX_MIN_LEN:
                        out.append(c)
                continue
            s = re.sub(r"^#{1,6}\s+", "", s)
            s = re.sub(r"^([-*+]|\d+\.)\s+", "", s)
            s = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", s)  # 링크 → 라벨
            s = s.replace("**", "").replace("`", "").strip("*_ ")
            if len(s) >= PDF_BBOX_MIN_LEN:
                out.append(s)
        return out

    def _find_bboxes(self, page, body: str, textpage=None) -> List[Dict[str, float]]:
        """대표 라인들을 페이지에서 검색해 bbox 수집 (0~1 정규화).
        textpage를 넘기면 search_for의 페이지 텍스트 재파싱을 생략해 크게 빨라진다."""
        if page is None:
            return []
        w = float(page.rect.width) or 1.0
        h = float(page.rect.height) or 1.0

        rects = []
        # 긴 라인일수록 오탐 확률이 낮으므로 길이 내림차순으로 시도
        for s in sorted(self._searchable_lines(body), key=len, reverse=True):
            needle = s[:PDF_BBOX_SEARCH_TRUNC].strip()
            if len(needle) < PDF_BBOX_MIN_LEN:
                continue
            try:
                found = page.search_for(needle, textpage=textpage)
            except TypeError:
                found = page.search_for(needle)  # 구버전 호환
            except Exception:
                continue
            if found:
                rects.append(found[0])
            if len(rects) >= PDF_MAX_BBOX_PER_CHUNK:
                break

        return [
            {
                "l": round(r.x0 / w, 4),
                "t": round(r.y0 / h, 4),
                "r": round(r.x1 / w, 4),
                "b": round(r.y1 / h, 4),
            }
            for r in rects
        ]

    # ------------------------------------------------------------------
    # GenOSVectorMeta 변환
    # ------------------------------------------------------------------

    def _to_vectors(self, chunks: List[PdfChunk], n_page: int) -> Optional[List[GenOSVectorMeta]]:
        now_iso = datetime.now().isoformat(timespec="seconds") + "Z"
        total = len(chunks)

        page_counts: Dict[int, int] = defaultdict(int)
        for c in chunks:
            page_counts[c.page] += 1
        seen_on_page: Dict[int, int] = defaultdict(int)

        vectors: List[GenOSVectorMeta] = []
        for i, c in enumerate(chunks):
            # chunk_bboxes 직렬화 포맷: [{"page": n, "bbox": {"l","t","r","b"}}] (0~1 정규화)
            # ※ 뷰어 스펙이 절대 좌표(pt)나 다른 키를 요구하면 여기만 수정하면 됨.
            if c.bboxes:
                bbox_str = json.dumps(
                    [{"page": c.page, "bbox": b} for b in c.bboxes],
                    ensure_ascii=False,
                )
            else:
                bbox_str = "."

            vectors.append(
                GenOSVectorMeta.model_validate({
                    "text": c.text,
                    "n_char": len(c.text),
                    "n_word": len(c.text.split()),
                    "n_line": len(c.text.splitlines()),
                    "i_page": c.page,
                    "e_page": c.page,
                    "n_page": n_page,
                    "i_chunk_on_page": seen_on_page[c.page],
                    "n_chunk_of_page": page_counts[c.page],
                    "i_chunk_on_doc": i,
                    "n_chunk_of_doc": total,
                    "reg_date": now_iso,
                    "chunk_bboxes": bbox_str,
                    "media_files": ".",
                })
            )
            seen_on_page[c.page] += 1

        return vectors if vectors else None


# ============================================================================
# DocumentProcessor
# ============================================================================

class DocumentProcessor:
    """
    포탈 첨부 파일 전처리기.
    지원: .pdf, .ppt, .pptx, .xlsx
    """

    def __init__(self):
        self.page_chunk_counts: Dict[int, int] = defaultdict(int)

    # ------------------------------------------------------------------
    # Backup / Restore
    # ------------------------------------------------------------------

    def _handle_backup_restore(self, file_path: str, backup_path: str, restore_only: bool) -> bool:
        """
        backup_path 기반 백업/복원 처리.
        반환값: restore_only=True 이면 True (호출자가 조기 종료해야 함).
        """
        if backup_path:
            orig_exists   = os.path.exists(file_path)
            backup_exists = os.path.exists(backup_path)

            if not orig_exists and backup_exists:
                try:
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    shutil.copy2(backup_path, file_path)
                    _log.info(f"[restore] 완료 | {backup_path} → {file_path}")
                except Exception as e:
                    _log.warning(f"[restore] 실패: {e}")

            elif orig_exists and not backup_exists:
                try:
                    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
                    shutil.copy2(file_path, backup_path)
                    _log.info(f"[backup] 완료 | {os.path.basename(file_path)} → {backup_path}")
                except Exception as e:
                    _log.warning(f"[backup] 실패: {e}")

            elif backup_exists:
                _log.info(f"[backup] 이미 존재 | {os.path.basename(backup_path)}")

            else:
                _log.warning(f"[backup/restore] 원본·backup 모두 없음 | {file_path}")

        return restore_only

    # ------------------------------------------------------------------
    # Chunk Cache
    # ------------------------------------------------------------------

    @staticmethod
    def _chunk_cache_path(backup_path: str) -> str:
        base, _ = os.path.splitext(backup_path)
        return base + "_chunks.json"

    def _save_chunks_cache(self, backup_path: str, chunks: List[Dict[str, Any]]) -> None:
        cache_path = self._chunk_cache_path(backup_path)
        try:
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(chunks, f, ensure_ascii=False)
            _log.info(f"[cache] 저장 완료 | {len(chunks)}개 → {cache_path}")
        except Exception as e:
            _log.warning(f"[cache] 저장 실패: {e}")

    def _load_chunks_cache(self, backup_path: str) -> Optional[List[Dict[str, Any]]]:
        cache_path = self._chunk_cache_path(backup_path)
        if not os.path.exists(cache_path):
            return None
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                chunks = json.load(f)
            _log.info(f"[cache] 히트 | {len(chunks)}개 ← {cache_path}")
            return chunks
        except Exception as e:
            _log.warning(f"[cache] 읽기 실패 (재처리): {e}")
            return None

    # ------------------------------------------------------------------
    # Loader (레거시 폴백 경로)
    # ------------------------------------------------------------------

    def _get_loader(self, file_path: str):
        from langchain_community.document_loaders import PyMuPDFLoader, UnstructuredPowerPointLoader

        ext = os.path.splitext(file_path)[-1].lower()
        real_type = _get_real_file_type(file_path)

        if ext.lstrip(".") != real_type and real_type == "pdf":
            return PyMuPDFLoader(file_path)

        if ext == ".pdf":
            return PyMuPDFLoader(file_path)

        if ext in (".ppt", ".pptx"):
            pdf_path = convert_to_pdf(file_path)
            if pdf_path and os.path.exists(pdf_path):
                return PyMuPDFLoader(pdf_path)
            return UnstructuredPowerPointLoader(file_path)

        raise ValueError(f"Unsupported document extension: {ext}")

    # ------------------------------------------------------------------
    # Load / Split / Compose (레거시 폴백 경로)
    # ------------------------------------------------------------------

    def _load_documents(self, file_path: str) -> list:
        from langchain_core.documents import Document
        loader = self._get_loader(file_path)
        try:
            return loader.load()
        except Exception as e:
            _log.warning(f"Primary loader failed: {e}")
            pdf_candidate = _get_pdf_path(file_path) if not file_path.endswith(".pdf") else file_path
            if os.path.exists(pdf_candidate):
                return self._load_pdf_fitz(pdf_candidate)
            return [Document(page_content=f"[로드 실패] {Path(file_path).name}", metadata={"source": file_path, "page": 0})]

    @staticmethod
    def _load_pdf_fitz(file_path: str) -> list:
        import fitz
        from langchain_core.documents import Document
        docs = []
        with fitz.open(file_path) as pdf:
            for i in range(len(pdf)):
                text = pdf.load_page(i).get_text("text") or f"[empty_page] page={i + 1}"
                docs.append(Document(page_content=text.strip(), metadata={"source": file_path, "page": i}))
        return docs or [Document(page_content="[empty]", metadata={"source": file_path, "page": 0})]

    def _split_documents(self, documents: list, chunk_size: int = 1000, chunk_overlap: int = 100) -> list:
        from langchain_text_splitters import RecursiveCharacterTextSplitter
        self.page_chunk_counts.clear()

        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=max(chunk_size, 1),
            chunk_overlap=max(chunk_overlap, 0),
        )
        chunks = text_splitter.split_documents(documents)
        chunks = [c for c in chunks if c.page_content and c.page_content.strip()]

        if not chunks:
            raise GenosServiceException(1, "Empty document")

        for chunk in chunks:
            page = chunk.metadata.get("page", 0)
            self.page_chunk_counts[page] += 1

        return chunks

    def _compose_vectors(self, file_path: str, chunks: list) -> List[GenOSVectorMeta]:
        ext = os.path.splitext(file_path)[-1].lower()

        if ext in (".ppt", ".pptx"):
            pdf_path = _get_pdf_path(file_path)
            if os.path.exists(pdf_path):
                try:
                    subprocess.run(["rm", pdf_path], check=True)
                except Exception:
                    pass

        global_metadata = {
            "n_chunk_of_doc": len(chunks),
            "n_page": max((c.metadata.get("page", 0) for c in chunks), default=0) + 1,
            "reg_date": datetime.now().isoformat(timespec="seconds") + "Z",
        }

        current_page = None
        chunk_idx_on_page = 0
        vectors: List[GenOSVectorMeta] = []

        for chunk_idx, chunk in enumerate(chunks):
            page = chunk.metadata.get("page", 0) + 1  # 1-based

            if page != current_page:
                current_page = page
                chunk_idx_on_page = 0

            text = chunk.page_content
            vectors.append(
                GenOSVectorMeta.model_validate({
                    "text": text,
                    "n_char": len(text),
                    "n_word": len(text.split()),
                    "n_line": len(text.splitlines()),
                    "i_page": page,
                    "e_page": page,
                    "i_chunk_on_page": chunk_idx_on_page,
                    "n_chunk_of_page": self.page_chunk_counts.get(page - 1, 1),
                    "i_chunk_on_doc": chunk_idx,
                    **global_metadata,
                })
            )
            chunk_idx_on_page += 1

        return vectors

    # ------------------------------------------------------------------
    # PDF-like 처리 (신규 1단계 경로 + 레거시 폴백)
    # ------------------------------------------------------------------

    async def _process_pdf_like(
        self, request: Request, file_path: str, ext: str,
        chunk_size: int, chunk_overlap: int,
    ) -> List[Dict[str, Any]]:
        """
        .pdf / .ppt / .pptx 공통 처리.
        1) ppt/pptx는 PDF 변환 후 동일 경로 사용
        2) PdfMarkdownLoader(PyMuPDF4LLM) 우선 시도
        3) 실패 시 레거시(PyMuPDF flat text + RecursiveCharacterTextSplitter) 폴백
        """
        pdf_path = file_path
        converted_pdf: Optional[str] = None

        if ext in (".ppt", ".pptx"):
            # LibreOffice 변환(최대 300초 동기 블로킹) → 스레드로 격리
            conv = await asyncio.to_thread(convert_to_pdf, file_path)
            if conv and os.path.exists(conv):
                pdf_path = conv
                converted_pdf = conv
            else:
                _log.warning(f"[pdf-md] {ext} → PDF 변환 실패, 레거시 경로 사용")

        result: Optional[List[Dict[str, Any]]] = None

        if pdf_path != file_path or _get_real_file_type(pdf_path) == "pdf":
            try:
                loader = PdfMarkdownLoader(
                    pdf_path,
                    chunk_size=PDF_MARKDOWN_CHUNK_SIZE,
                    chunk_overlap=PDF_MARKDOWN_CHUNK_OVERLAP,
                    display_name=Path(file_path).name,  # 컨텍스트 헤더에는 원본 파일명 표기
                )
                _log.info(
                    f"[pdf-md] 청킹 설정 | size={PDF_MARKDOWN_CHUNK_SIZE}"
                    f" overlap={PDF_MARKDOWN_CHUNK_OVERLAP}"
                )
                # 대용량 문서 파싱은 수 분간 CPU를 점유하는 동기 작업 →
                # 이벤트 루프를 막으면 헬스체크 미응답으로 liveness 재시작이 발생하므로
                # 반드시 워커 스레드에서 실행한다.
                vectors = await asyncio.to_thread(loader.return_vectormeta_format)
                if vectors:
                    result = [self._to_dict(v) for v in vectors]
                    _log.info(f"[pdf-md] Markdown 경로 성공 | 청크 {len(result)}개")
            except PdfPageExtractionError as e:
                _log.error("[pdf-md] page extraction incomplete; refusing partial result: %s", e)
                raise GenosServiceException("PDF_PAGE_EXTRACTION_FAILED", str(e)) from e
            except Exception as e:
                _log.warning(f"[pdf-md] Markdown 추출 실패 → 레거시 경로 폴백: {e}")

        await assert_cancelled(request)

        if result is None:
            documents = await asyncio.to_thread(self._load_documents, file_path)
            await assert_cancelled(request)
            chunks = await asyncio.to_thread(
                self._split_documents, documents, chunk_size, chunk_overlap
            )
            await assert_cancelled(request)
            vectors = self._compose_vectors(file_path, chunks)
            result = [self._to_dict(v) for v in vectors]

        # ppt/pptx 변환 산출물 정리
        if converted_pdf and os.path.exists(converted_pdf):
            try:
                os.remove(converted_pdf)
            except Exception:
                pass

        return result

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------

    async def __call__(self, request: Request, file_path: str, **kwargs) -> List[Dict[str, Any]]:
        setup_logging(kwargs.get("log_level", 4))

        backup_path  = kwargs.get("backup_path", "")
        restore_only = kwargs.get("restore_only", False)

        if self._handle_backup_restore(file_path, backup_path, restore_only):
            return []

        # 캐시 확인: backup_path 기반으로 저장된 청크가 있으면 재처리 없이 반환
        if backup_path:
            cached = self._load_chunks_cache(backup_path)
            if cached is not None:
                return cached

        _log.info(f"file_path: {file_path}")
        _log.info(f"kwargs: {kwargs}")

        ext = os.path.splitext(file_path)[-1].lower()
        if ext not in SUPPORTED_EXTS:
            raise ValueError(f"Unsupported extension: {ext}. Supported: {sorted(SUPPORTED_EXTS)}")

        chunk_size    = int(kwargs.get("chunk_size", 1000))
        chunk_overlap = int(kwargs.get("chunk_overlap", 100))

        # ── xlsx ──────────────────────────────────────────────
        if ext == ".xlsx":
            # 워크북 파싱(대형 시트에서 수십 초~분 단위 동기 작업)도 스레드로 격리
            loader = await asyncio.to_thread(
                TabularLoader, file_path, chunk_size=chunk_size, chunk_overlap=chunk_overlap
            )
            await assert_cancelled(request)
            vectors = await asyncio.to_thread(loader.return_vectormeta_format)
            if not vectors:
                raise GenosServiceException(1, "Empty xlsx")
            result = [self._to_dict(v) for v in vectors]
            if backup_path:
                self._save_chunks_cache(backup_path, result)
            return result

        # ── pdf / ppt / pptx ─────────────────────────────────
        result = await self._process_pdf_like(request, file_path, ext, chunk_size, chunk_overlap)
        if backup_path:
            self._save_chunks_cache(backup_path, result)
        return result

    @staticmethod
    def _to_dict(model: GenOSVectorMeta) -> Dict[str, Any]:
        return model.model_dump() if hasattr(model, "model_dump") else model.dict()


# ============================================================================
# 검증 도구 (verify) — 실 문서 헤딩 인식률 · bbox 정확도 확인
# ============================================================================

# 검증 결과 해석 가이드:
#   section_coverage  헤딩 경로가 붙은 청크 비율. 낮으면(< 30%) pymupdf4llm이 해당 문서의
#                     헤딩을 폰트 크기로 구분하지 못한 것 → 레거시 대비 이득은 표/컨텍스트 헤더뿐.
#   bbox_coverage     bbox가 1개 이상 잡힌 청크 비율. 낮으면 뷰어 하이라이트가 자주 비게 됨.
#   *_bbox_preview.pdf  bbox를 사각형으로 그린 사본 — 하이라이트 위치가 실제 본문과 맞는지 육안 확인.

def verify_pdf(file_path: str, chunk_size: int = 1000, chunk_overlap: int = 100,
               out_dir: Optional[str] = None) -> Dict[str, Any]:
    """
    실 PDF 문서로 신규 경로 품질 검증.
    - 헤딩 인식: 레벨별 개수, 샘플 제목, 청크 섹션 커버리지
    - bbox: 커버리지, 청크당 평균 개수, 시각 확인용 preview PDF 생성
    산출물: {stem}_verify_report.json, {stem}_bbox_preview.pdf
    """
    import fitz

    setup_logging(4)
    in_path = Path(file_path).resolve()
    out = Path(out_dir) if out_dir else in_path.parent
    out.mkdir(parents=True, exist_ok=True)

    loader = PdfMarkdownLoader(str(in_path), chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    loader.return_vectormeta_format()
    stats = dict(loader.stats)
    stats["timing"] = loader.timing
    stats["heading_samples"] = [
        {"page": p, "level": lv, "title": t[:60]} for p, lv, t in loader.headings[:15]
    ]

    # --- bbox preview PDF: 청크별 bbox를 색상 사각형으로 표시 ---
    palette = [(0.85, 0.2, 0.2), (0.2, 0.45, 0.85), (0.15, 0.6, 0.35),
               (0.85, 0.55, 0.1), (0.55, 0.25, 0.75)]
    preview_path = out / f"{in_path.stem}_bbox_preview.pdf"
    doc = fitz.open(str(in_path))
    try:
        for i, chunk in enumerate(loader.last_chunks):
            if not chunk.bboxes or not (1 <= chunk.page <= doc.page_count):
                continue
            page = doc[chunk.page - 1]
            w, h = page.rect.width, page.rect.height
            color = palette[i % len(palette)]
            first = True
            for b in chunk.bboxes:
                rect = fitz.Rect(b["l"] * w, b["t"] * h, b["r"] * w, b["b"] * h)
                page.draw_rect(rect, color=color, width=0.8)
                if first:
                    page.insert_text(fitz.Point(rect.x0, max(rect.y0 - 3, 8)),
                                     f"#{i}", fontsize=7, color=color)
                    first = False
        doc.save(str(preview_path))
    finally:
        doc.close()

    report_path = out / f"{in_path.stem}_verify_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

    # --- 요약 출력 ---
    print(f"\n=== 검증 결과: {stats['file']} ===")
    print(f"페이지: {stats['n_page']} (텍스트 있음 {stats['pages_with_text']}"
          f" / 없음 {stats['pages_without_text']})")
    print(f"헤딩: {stats['n_heading']}개, 레벨 분포 {stats['heading_levels']}")
    print(f"청크: {stats['n_chunk']}개, 크기 min/avg/max ="
          f" {stats['chunk_size_min']}/{stats['chunk_size_avg']}/{stats['chunk_size_max']}")
    print(f"섹션 커버리지: {stats['section_coverage']}%  |  "
          f"bbox 커버리지: {stats['bbox_coverage']}% (청크당 평균 {stats['bbox_avg_per_chunk']}개)")
    if stats["heading_samples"]:
        print("헤딩 샘플:")
        for hs in stats["heading_samples"][:8]:
            print(f"  p.{hs['page']} {'#' * hs['level']} {hs['title']}")
    else:
        print("헤딩 샘플: 없음 — 이 문서에서는 헤딩 추론이 동작하지 않음")
    print(f"\n산출물:\n  {report_path}\n  {preview_path}")
    print("→ preview PDF를 열어 사각형이 실제 본문 위치와 일치하는지 육안 확인\n")
    return stats


# ============================================================================
# Module Entry Point
# ============================================================================

if __name__ == "__main__":
    # 사용법:
    #   검증:  python preprocessor.py verify <file.pdf> [chunk_size] [chunk_overlap] [out_dir]
    #   처리:  python preprocessor.py run <file.(pdf|pptx|xlsx)>
    args = sys.argv[1:]

    if len(args) >= 2 and args[0] == "verify":
        verify_pdf(
            args[1],
            chunk_size=int(args[2]) if len(args) > 2 else 1000,
            chunk_overlap=int(args[3]) if len(args) > 3 else 100,
            out_dir=args[4] if len(args) > 4 else None,
        )
    elif len(args) >= 2 and args[0] == "run":
        from unittest.mock import MagicMock

        async def _run(path: str):
            processor = DocumentProcessor()
            mock_request = MagicMock()

            async def _not_disconnected():
                return False

            mock_request.is_disconnected = _not_disconnected
            result = await processor(mock_request, path)
            print(f"Processed {len(result)} chunks")
            for r in result[:3]:
                print(f"  - page={r.get('i_page')} bbox={str(r.get('chunk_bboxes'))[:60]}")
                print(f"    {r.get('text', '')[:150]}...")

        asyncio.run(_run(args[1]))
    else:
        print("usage:\n"
              "  python preprocessor.py verify <file.pdf> [chunk_size] [chunk_overlap] [out_dir]\n"
              "  python preprocessor.py run <file>")
